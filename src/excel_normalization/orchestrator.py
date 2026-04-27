"""Orchestrator for coordinating normalization across all worksheets.

This module provides the NormalizationOrchestrator class that coordinates
processing across all worksheets in a workbook. It creates instances of all
field processors and calls them in order for each worksheet.
"""

import logging
from typing import Dict, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .io_layer.excel_reader import ExcelReader
from .io_layer.excel_writer import ExcelWriter
from .engines.name_engine import NameEngine
from .engines.text_processor import TextProcessor
from .engines.gender_engine import GenderEngine
from .engines.date_engine import DateEngine
from .engines.identifier_engine import IdentifierEngine
from .processing.name_processor import NameFieldProcessor
from .processing.gender_processor import GenderFieldProcessor
from .processing.date_processor import DateFieldProcessor
from .processing.identifier_processor import IdentifierFieldProcessor
from .data_types import FieldKey
from .json_exporter import JsonExporter
from .export.export_engine import ExportEngine
class NormalizationOrchestrator:
    """Coordinates processing across all worksheets in a workbook.

    This class creates instances of all field processors (NameFieldProcessor,
    GenderFieldProcessor, DateFieldProcessor, IdentifierFieldProcessor) and
    calls them in order for each worksheet. It tracks corrected column positions
    using FieldKey enum.

    The processing order for each worksheet is:
    1. Names (first name, last name, father's name)
    2. Gender
    3. Dates (birth date, entry date)
    4. Identifiers (Israeli ID and passport)

    Attributes:
        reader: ExcelReader instance for reading data
        writer: ExcelWriter instance for writing data
        name_processor: NameFieldProcessor instance
        gender_processor: GenderFieldProcessor instance
        date_processor: DateFieldProcessor instance
        identifier_processor: IdentifierFieldProcessor instance
        corrected_columns: Dict tracking corrected column positions by (sheet_name, field_key)
    """

    def __init__(self) -> None:
        """Initialize the orchestrator with all processors and engines."""
        self.logger = logging.getLogger(__name__)

        # Create I/O layer instances
        self.reader = ExcelReader()
        self.writer = ExcelWriter()

        # Create engine instances
        text_processor = TextProcessor()
        name_engine = NameEngine(text_processor)
        gender_engine = GenderEngine()
        date_engine = DateEngine()
        identifier_engine = IdentifierEngine()

        # Create field processor instances
        self.name_processor = NameFieldProcessor(self.reader, self.writer, name_engine)
        self.gender_processor = GenderFieldProcessor(self.reader, self.writer, gender_engine)
        self.date_processor = DateFieldProcessor(self.reader, self.writer, date_engine)
        self.identifier_processor = IdentifierFieldProcessor(self.reader, self.writer, identifier_engine)

        # Track corrected column positions: (sheet_name, field_key) -> column_number
        self.corrected_columns: Dict[Tuple[str, FieldKey], int] = {}

        self.logger.debug("NormalizationOrchestrator initialized")

    def normalize_workbook(self, file_path: str) -> None:
        """Process all worksheets in the workbook using the legacy direct-Excel path.

        .. deprecated::
            This method modifies the workbook in-place using the old processor
            approach (find headers → insert corrected columns → write values).
            Prefer :meth:`process_workbook_json` for all new usage, which uses
            the clean JSON-based pipeline and never modifies the original file.

        Args:
            file_path: Path to the Excel workbook file (modified in-place)

        Requirements: 1.1, 1.2, 1.5, 19.1
        """
        import os
        import shutil
        from pathlib import Path

        self.logger.info(f"Loading workbook: {file_path}")

        # Determine file extension and appropriate loading parameters
        file_ext = Path(file_path).suffix.lower()
        is_macro_enabled = file_ext in ['.xlsm', '.xltm', '.xlam']

        try:
            # Load the workbook with appropriate parameters
            # data_only=False ensures formulas are preserved
            # keep_vba=True only for macro-enabled files to avoid format issues
            # keep_links=False to avoid external link corruption
            workbook = load_workbook(
                file_path,
                data_only=False,
                keep_vba=is_macro_enabled,
                keep_links=False
            )
            self.logger.info(f"Workbook loaded successfully. Found {len(workbook.worksheets)} worksheet(s)")
            
            # Log warning if file contains VBA but we're not preserving it
            if not is_macro_enabled and file_ext == '.xlsx':
                self.logger.debug("Loading .xlsx file without VBA preservation (standard behavior)")
                
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {e}")
            raise

        # Create backup before processing
        backup_path = None
        try:
            backup_path = f"{file_path}.backup"
            shutil.copy2(file_path, backup_path)
            self.logger.debug(f"Created backup at: {backup_path}")
        except Exception as e:
            self.logger.warning(f"Failed to create backup: {e}")
            # Continue without backup

        # Process each worksheet
        worksheet_count = 0
        processing_errors = []
        
        for worksheet in workbook.worksheets:
            try:
                self.process_worksheet(worksheet)
                worksheet_count += 1
            except Exception as e:
                error_msg = f"Failed to process worksheet '{worksheet.title}': {e}"
                self.logger.error(error_msg, exc_info=True)
                processing_errors.append(error_msg)
                # Continue with next worksheet
                continue

        self.logger.info(f"Successfully processed {worksheet_count} worksheet(s)")
        
        if processing_errors:
            self.logger.warning(f"Encountered {len(processing_errors)} worksheet processing errors")

        # Save the modified workbook
        try:
            self.logger.info(f"Saving workbook to: {file_path}")
            
            # For macro-enabled files, ensure we're saving with the correct format
            if is_macro_enabled:
                self.logger.debug("Saving macro-enabled workbook")
                # openpyxl automatically handles .xlsm format when keep_vba=True
                workbook.save(file_path)
            else:
                # For regular .xlsx files, standard save
                workbook.save(file_path)
            
            self.logger.info("Workbook saved successfully")
            
            # Remove backup if save was successful
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                    self.logger.debug("Backup removed after successful save")
                except Exception as e:
                    self.logger.warning(f"Failed to remove backup: {e}")
                    
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            
            # Attempt to restore from backup
            if backup_path and os.path.exists(backup_path):
                try:
                    shutil.copy2(backup_path, file_path)
                    self.logger.info(f"Restored original file from backup")
                except Exception as restore_error:
                    self.logger.error(f"Failed to restore from backup: {restore_error}")
            
            raise
    def export_raw_json(self, input_excel_path: str, output_json_path: str) -> None:
        """Export raw JSON dataset from Excel workbook.

        Extracts data from Excel workbook and exports it to JSON format without
        any normalization. This preserves the exact original values from the Excel file.

        Args:
            input_excel_path: Path to input Excel workbook
            output_json_path: Path for output JSON file

        Raises:
            FileNotFoundError: If input file does not exist
            ValueError: If workbook has no valid sheets
            IOError: If file operations fail

        Requirements: 11.6
        """
        self.logger.info(f"Exporting raw JSON: {input_excel_path} -> {output_json_path}")

        try:
            # Create ExcelToJsonExtractor instance
            from .io_layer.excel_to_json_extractor import ExcelToJsonExtractor

            extractor = ExcelToJsonExtractor(
                excel_reader=self.reader,
                skip_empty_rows=False,
                handle_formulas=True,
                preserve_types=True
            )

            # Extract workbook to JSON
            self.logger.info(f"Extracting workbook to JSON from '{input_excel_path}'")
            workbook_dataset = extractor.extract_workbook_to_json(input_excel_path)

            if not workbook_dataset.sheets:
                raise ValueError(f"No valid sheets found in workbook '{input_excel_path}'")

            self.logger.info(
                f"Extracted {len(workbook_dataset.sheets)} sheet(s) with "
                f"{sum(len(sheet.rows) for sheet in workbook_dataset.sheets)} total rows"
            )

            # Export to JSON
            from .json_exporter import JsonExporter

            exporter = JsonExporter(indent=2, ensure_ascii=False)
            exporter.export_workbook_to_json(workbook_dataset, output_json_path)

            self.logger.info(f"Raw JSON exported successfully to '{output_json_path}'")

        except FileNotFoundError as e:
            self.logger.error(f"Input file not found: {e}")
            raise
        except ValueError as e:
            self.logger.error(f"Validation error: {e}")
            raise
        except Exception as e:
            self.logger.error(f"Unexpected error exporting raw JSON: {e}", exc_info=True)
            raise IOError(f"Failed to export raw JSON: {e}")

    def export_normalized_json(self, input_excel_path: str, output_json_path: str) -> None:
        """Export normalized JSON dataset from Excel workbook.

        Extracts data from Excel workbook, applies normalization engines, and exports
        the result to JSON format. The output includes both original and corrected values.

        Args:
            input_excel_path: Path to input Excel workbook
            output_json_path: Path for output JSON file

        Raises:
            FileNotFoundError: If input file does not exist
            ValueError: If workbook has no valid sheets
            IOError: If file operations fail

        Requirements: 11.6
        """
        self.logger.info(f"Exporting normalized JSON: {input_excel_path} -> {output_json_path}")

        try:
            # Create ExcelToJsonExtractor instance
            from .io_layer.excel_to_json_extractor import ExcelToJsonExtractor

            extractor = ExcelToJsonExtractor(
                excel_reader=self.reader,
                skip_empty_rows=False,
                handle_formulas=True,
                preserve_types=True
            )

            # Extract workbook to JSON
            self.logger.info(f"Extracting workbook to JSON from '{input_excel_path}'")
            workbook_dataset = extractor.extract_workbook_to_json(input_excel_path)

            if not workbook_dataset.sheets:
                raise ValueError(f"No valid sheets found in workbook '{input_excel_path}'")

            self.logger.info(
                f"Extracted {len(workbook_dataset.sheets)} sheet(s) with "
                f"{sum(len(sheet.rows) for sheet in workbook_dataset.sheets)} total rows"
            )

            # Create NormalizationPipeline instance reusing engines from __init__
            from .processing.normalization_pipeline import NormalizationPipeline

            pipeline = NormalizationPipeline(
                name_engine=self.name_processor.name_engine,
                gender_engine=self.gender_processor.gender_engine,
                date_engine=self.date_processor.date_engine,
                identifier_engine=self.identifier_processor.identifier_engine,
                apply_name_normalization_enabled=True,
                apply_gender_normalization_enabled=True,
                apply_date_normalization_enabled=True,
                apply_identifier_normalization_enabled=True
            )

            # Normalize all sheet datasets
            self.logger.info(f"Normalizing {len(workbook_dataset.sheets)} sheet(s)")
            normalized_sheets = []

            for idx, sheet_dataset in enumerate(workbook_dataset.sheets, start=1):
                self.logger.info(
                    f"Normalizing sheet {idx}/{len(workbook_dataset.sheets)}: '{sheet_dataset.sheet_name}'"
                )

                try:
                    normalized_sheet = pipeline.normalize_dataset(sheet_dataset)
                    normalized_sheets.append(normalized_sheet)

                    # Log normalization statistics
                    stats = normalized_sheet.get_metadata("normalization_statistics", {})
                    success_rate = stats.get("success_rate", 1.0) * 100
                    self.logger.info(
                        f"Sheet '{sheet_dataset.sheet_name}' normalized: "
                        f"{stats.get('total_rows', 0)} rows, "
                        f"{success_rate:.1f}% success rate"
                    )

                except Exception as e:
                    self.logger.error(
                        f"Failed to normalize sheet '{sheet_dataset.sheet_name}': {e}",
                        exc_info=True
                    )
                    # Continue with other sheets
                    continue

            if not normalized_sheets:
                raise ValueError("No sheets were successfully normalized")

            # Update workbook dataset with normalized sheets
            workbook_dataset.sheets = normalized_sheets

            # Export to JSON
            from .json_exporter import JsonExporter

            exporter = JsonExporter(indent=2, ensure_ascii=False)
            exporter.export_workbook_to_json(workbook_dataset, output_json_path)

            self.logger.info(f"Normalized JSON exported successfully to '{output_json_path}'")

        except FileNotFoundError as e:
            self.logger.error(f"Input file not found: {e}")
            raise
        except ValueError as e:
            self.logger.error(f"Validation error: {e}")
            raise
        except Exception as e:
            self.logger.error(f"Unexpected error exporting normalized JSON: {e}", exc_info=True)
            raise IOError(f"Failed to export normalized JSON: {e}")

    def export_vba_parity_workbook_from_json(self, input_excel_path: str, output_excel_path: Optional[str] = None) -> str:
        """Run JSON normalization pipeline then create VBA-parity export workbook.

        Pipeline:
            Excel workbook -> ExcelToJsonExtractor -> NormalizationPipeline -> ExportEngine
        """
        from .io_layer.excel_to_json_extractor import ExcelToJsonExtractor
        from .processing.normalization_pipeline import NormalizationPipeline

        # Extract workbook into JSON datasets
        extractor = ExcelToJsonExtractor(
            excel_reader=self.reader,
            skip_empty_rows=False,
            handle_formulas=True,
            preserve_types=True,
        )
        workbook_dataset = extractor.extract_workbook_to_json(input_excel_path)

        # Normalize datasets using the existing engines (pure logic)
        pipeline = NormalizationPipeline(
            name_engine=self.name_processor.name_engine,
            gender_engine=self.gender_processor.gender_engine,
            date_engine=self.date_processor.date_engine,
            identifier_engine=self.identifier_processor.identifier_engine,
            apply_name_normalization_enabled=True,
            apply_gender_normalization_enabled=True,
            apply_date_normalization_enabled=True,
            apply_identifier_normalization_enabled=True,
        )

        normalized_sheets = []
        for sheet in workbook_dataset.sheets:
            normalized_sheets.append(pipeline.normalize_dataset(sheet))
        workbook_dataset.sheets = normalized_sheets

        # Determine output path (default: Desktop with _Export suffix, like VBA)
        if output_excel_path is None:
            from pathlib import Path as _Path
            src = _Path(input_excel_path)
            desktop = _Path.home() / "Desktop"
            base = src.stem + "_Export"
            ext = src.suffix.lower()
            if ext not in [".xlsx", ".xlsm"]:
                ext = ".xlsx"

            candidate = desktop / f"{base}{ext}"
            i = 1
            while candidate.exists():
                candidate = desktop / f"{base} ({i}){ext}"
                i += 1
            output_excel_path = str(candidate)

        engine = ExportEngine()
        # When exporting from JSON, tracking dict is unavailable by default, so
        # MosadID/SugMosad/MisparDiraBeMosad remain empty (VBA parity).
        return engine.export_from_normalized_dataset(workbook_dataset, output_excel_path, corrected_columns_by_sheet=None)

    def export_vba_parity_workbook_from_processors(
        self, input_excel_path: str, output_excel_path: Optional[str] = None
    ) -> str:
        """Run the Excel processors then export via worksheet-based engine."""
        from pathlib import Path as _Path
        from openpyxl import load_workbook

        src = _Path(input_excel_path)
        file_ext = src.suffix.lower()
        is_macro_enabled = file_ext in [".xlsm", ".xltm", ".xlam"]

        # Reset tracking state for this run
        self.corrected_columns = {}

        wb = load_workbook(
            input_excel_path,
            data_only=False,
            keep_vba=is_macro_enabled,
            keep_links=False,
        )

        # Run processors (in-place on the in-memory workbook)
        for ws in wb.worksheets:
            self.process_worksheet(ws)

        # Build corrected_columns_by_sheet in the required structure:
        # corrected_columns_by_sheet[sheet][fieldKey] = col
        corrected_columns_by_sheet: dict[str, dict[str, int]] = {}
        for (sheet_name, field_key), col in self.corrected_columns.items():
            corrected_columns_by_sheet.setdefault(sheet_name, {})[field_key.value] = col

        # Refresh mapping from the augmented workbook headers to avoid stale
        # column positions after multiple insertions. This preserves VBA behavior
        # (fallback DetectCorrectedColumns) while keeping Mosad fields tracking-only.
        engine = ExportEngine()
        for spec in engine.SOURCE_SHEET_SPECS:
            if spec.source_sheet_name not in wb.sheetnames:
                continue
            ws = wb[spec.source_sheet_name]
            header_row = engine.detect_header_row(ws)
            if header_row == 0:
                continue
            detected = engine.detect_corrected_columns(ws, header_row)

            tracked = corrected_columns_by_sheet.get(spec.source_sheet_name, {})
            mosad_only = {k: v for k, v in tracked.items() if k in ["MosadID", "SugMosad", "MisparDiraBeMosad"]}

            # Build final mapping: detected + tracked mosad fields
            corrected_columns_by_sheet[spec.source_sheet_name] = dict(detected)
            corrected_columns_by_sheet[spec.source_sheet_name].update(mosad_only)

        # Determine output path (Desktop, _Export suffix, (N) collision)
        if output_excel_path is None:
            desktop = _Path.home() / "Desktop"
            base = src.stem + "_Export"
            ext = src.suffix.lower()
            if ext not in [".xlsx", ".xlsm"]:
                ext = ".xlsx"

            candidate = desktop / f"{base}{ext}"
            i = 1
            while candidate.exists():
                candidate = desktop / f"{base} ({i}){ext}"
                i += 1
            output_excel_path = str(candidate)

        engine = ExportEngine()
        return engine.export_from_augmented_workbook(wb, corrected_columns_by_sheet, output_excel_path)

    def export_raw_and_normalized_json(self, input_excel_path: str) -> tuple[str, str]:
        """Export both raw and normalized JSON datasets from Excel workbook.

        Convenience method that exports both raw and normalized JSON files using
        automatically generated filenames based on the input Excel file path.

        Args:
            input_excel_path: Path to input Excel workbook

        Returns:
            Tuple of (raw_json_path, normalized_json_path)

        Raises:
            FileNotFoundError: If input file does not exist
            ValueError: If workbook has no valid sheets
            IOError: If file operations fail

        Requirements: 11.6
        """
        from .json_exporter import generate_output_filenames

        # Generate output filenames
        raw_json_path, normalized_json_path = generate_output_filenames(input_excel_path)

        self.logger.info(
            f"Exporting raw and normalized JSON from '{input_excel_path}'\n"
            f"  Raw JSON: {raw_json_path}\n"
            f"  Normalized JSON: {normalized_json_path}"
        )

        # Export raw JSON
        self.export_raw_json(input_excel_path, raw_json_path)

        # Export normalized JSON
        self.export_normalized_json(input_excel_path, normalized_json_path)

        return raw_json_path, normalized_json_path



    def process_worksheet(self, worksheet: Worksheet) -> None:
        """Process a single worksheet in order: names, gender, dates, identifiers.

        Calls each field processor in the specified order:
        1. Names (first name, last name, father's name)
        2. Gender
        3. Dates (birth date, entry date)
        4. Identifiers (Israeli ID and passport)

        After processing, tracks the corrected column positions for later retrieval.

        Args:
            worksheet: The worksheet to process

        Requirements: 1.3, 19.2
        """
        sheet_name = worksheet.title
        self.logger.info(f"Processing worksheet: '{sheet_name}'")

        # Step 1: Process names (first, last, father)
        self.logger.debug(f"[{sheet_name}] Processing names...")
        self.name_processor.process_field(worksheet)
        self.reader.invalidate_cache(worksheet)
        self._track_name_columns(sheet_name)

        # Step 2: Process gender
        # Re-scan headers on the current worksheet state so that column positions
        # reflect any columns inserted by the name processor (VBA parity: each
        # processor always works on the live sheet, not a stale snapshot).
        self.logger.debug(f"[{sheet_name}] Processing gender...")
        self.reader.invalidate_cache(worksheet)
        self.gender_processor.process_field(worksheet)
        self._track_gender_columns(sheet_name)

        # Step 3: Process dates (birth, entry)
        self.logger.debug(f"[{sheet_name}] Processing dates...")
        self.reader.invalidate_cache(worksheet)
        self.date_processor.process_field(worksheet)
        self.reader.invalidate_cache(worksheet)
        self._track_date_columns(sheet_name)
        # Cross-validate entry date against birth date (VBA parity)
        self._validate_entry_vs_birth(worksheet)
        # Clear after tracking to avoid leaking state across sheets
        self.date_processor.date_fields = {}

        # Step 4: Process identifiers (ID and passport)
        self.logger.debug(f"[{sheet_name}] Processing identifiers...")
        self.reader.invalidate_cache(worksheet)
        self.identifier_processor.process_field(worksheet)
        self._track_identifier_columns(sheet_name)

        self.logger.info(f"Completed processing worksheet: '{sheet_name}'")

    def get_corrected_column(self, sheet_name: str, field_key: FieldKey) -> Optional[int]:
        """Retrieve the column number for a corrected field.

        Args:
            sheet_name: Name of the worksheet
            field_key: FieldKey enum value identifying the field

        Returns:
            Column number (1-based) if found, None otherwise

        Requirements: 20.3, 20.4
        """
        return self.corrected_columns.get((sheet_name, field_key))

    def _validate_entry_vs_birth(self, worksheet: Worksheet) -> None:
        """Cross-validate entry dates against birth dates (VBA parity).

        After both birth and entry date columns are written, scan each row and
        flag entry dates that precede the birth date by appending a warning to
        the entry status cell.
        """
        from .data_types import DateFieldType
        from datetime import date as _date

        birth_groups = self.date_processor.date_fields.get(DateFieldType.BIRTH_DATE, [])
        entry_groups = self.date_processor.date_fields.get(DateFieldType.ENTRY_DATE, [])

        if not birth_groups or not entry_groups:
            return

        birth_info = birth_groups[0]
        entry_info = entry_groups[0]

        # Need corrected columns to be present
        for key in ("corrected_year_col", "corrected_month_col", "corrected_day_col", "corrected_status_col"):
            if key not in birth_info or key not in entry_info:
                return

        start_row = birth_info["sub_header_row"] + 1
        end_row = max(birth_info["last_row"], entry_info["last_row"])

        for row_idx in range(start_row, end_row + 1):
            try:
                by = worksheet.cell(row=row_idx, column=birth_info["corrected_year_col"]).value
                bm = worksheet.cell(row=row_idx, column=birth_info["corrected_month_col"]).value
                bd = worksheet.cell(row=row_idx, column=birth_info["corrected_day_col"]).value
                ey = worksheet.cell(row=row_idx, column=entry_info["corrected_year_col"]).value
                em = worksheet.cell(row=row_idx, column=entry_info["corrected_month_col"]).value
                ed = worksheet.cell(row=row_idx, column=entry_info["corrected_day_col"]).value

                if not all([by, bm, bd, ey, em, ed]):
                    continue

                birth_date = _date(int(by), int(bm), int(bd))
                entry_date = _date(int(ey), int(em), int(ed))

                if entry_date < birth_date:
                    status_col = entry_info["corrected_status_col"]
                    existing = worksheet.cell(row=row_idx, column=status_col).value or ""
                    warning = "תאריך כניסה לפני תאריך לידה"
                    if warning not in str(existing):
                        new_status = f"{existing} | {warning}".strip(" |") if existing else warning
                        worksheet.cell(row=row_idx, column=status_col).value = new_status
                        self.writer.format_cell(
                            worksheet, row_idx, status_col,
                            bg_color=self.writer.PINK_ERROR, bold=True
                        )
            except Exception:
                continue

    def _track_name_columns(self, sheet_name: str) -> None:
        """Track corrected column positions for name fields.

        Args:
            sheet_name: Name of the worksheet

        Requirements: 20.1, 20.2
        """
        # Track first name column
        if "first_name" in self.name_processor.corrected_columns:
            self.corrected_columns[(sheet_name, FieldKey.SHEM_PRATI)] = self.name_processor.corrected_columns[
                "first_name"
            ]

        # Track last name column
        if "last_name" in self.name_processor.corrected_columns:
            self.corrected_columns[(sheet_name, FieldKey.SHEM_MISHPAHA)] = self.name_processor.corrected_columns[
                "last_name"
            ]

        # Track father name column
        if "father_name" in self.name_processor.corrected_columns:
            self.corrected_columns[(sheet_name, FieldKey.SHEM_HAAV)] = self.name_processor.corrected_columns[
                "father_name"
            ]

    def _track_gender_columns(self, sheet_name: str) -> None:
        """Track corrected column positions for gender field.

        Args:
            sheet_name: Name of the worksheet

        Requirements: 20.1, 20.2
        """
        if hasattr(self.gender_processor, "corrected_col") and self.gender_processor.corrected_col is not None:
            self.corrected_columns[(sheet_name, FieldKey.MIN)] = self.gender_processor.corrected_col

    def _track_date_columns(self, sheet_name: str) -> None:
        """Track corrected column positions for date fields.

        Args:
            sheet_name: Name of the worksheet

        Requirements: 20.1, 20.2
        """
        from .data_types import DateFieldType

        # Track birth date columns
        if DateFieldType.BIRTH_DATE in self.date_processor.date_fields:
            birth_date_groups = self.date_processor.date_fields[DateFieldType.BIRTH_DATE]
            if birth_date_groups:
                birth_date_info = birth_date_groups[0]

                if "corrected_year_col" in birth_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.SHNAT_LIDA)] = birth_date_info["corrected_year_col"]

                if "corrected_month_col" in birth_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.HODESH_LIDA)] = birth_date_info["corrected_month_col"]

                if "corrected_day_col" in birth_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.YOM_LIDA)] = birth_date_info["corrected_day_col"]

        # Track entry date columns
        if DateFieldType.ENTRY_DATE in self.date_processor.date_fields:
            entry_date_groups = self.date_processor.date_fields[DateFieldType.ENTRY_DATE]
            if entry_date_groups:
                entry_date_info = entry_date_groups[0]

                if "corrected_year_col" in entry_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.SHNAT_KNISA)] = entry_date_info["corrected_year_col"]

                if "corrected_month_col" in entry_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.HODESH_KNISA)] = entry_date_info["corrected_month_col"]

                if "corrected_day_col" in entry_date_info:
                    self.corrected_columns[(sheet_name, FieldKey.YOM_KNISA)] = entry_date_info["corrected_day_col"]

    def _track_identifier_columns(self, sheet_name: str) -> None:
        """Track corrected column positions for identifier fields.

        Args:
            sheet_name: Name of the worksheet

        Requirements: 20.1, 20.2
        """
        # Track corrected ID column
        if (
            hasattr(self.identifier_processor, "corrected_id_col")
            and self.identifier_processor.corrected_id_col is not None
        ):
            self.corrected_columns[(sheet_name, FieldKey.MISPAR_ZEHUT)] = self.identifier_processor.corrected_id_col

        # Track corrected passport column
        if (
            hasattr(self.identifier_processor, "corrected_passport_col")
            and self.identifier_processor.corrected_passport_col is not None
        ):
            self.corrected_columns[(sheet_name, FieldKey.DARKON)] = self.identifier_processor.corrected_passport_col

    def process_workbook_json(self, input_excel_path: str, output_excel_path: str) -> None:
        """Process workbook and write VBA-parity output to a new file.

        Loads the original workbook, runs all field processors (names, gender,
        dates, identifiers) which insert corrected columns immediately to the
        right of each original column — matching the VBA FieldProcessor output
        structure exactly — then saves the augmented workbook to output_excel_path.

        The original file is never modified.

        Args:
            input_excel_path: Path to input Excel workbook (read-only)
            output_excel_path: Path for output Excel workbook (augmented copy)

        Raises:
            FileNotFoundError: If input file does not exist
            IOError: If file operations fail

        Requirements: 20.1-20.6
        """
        import shutil
        from pathlib import Path as _Path

        self.logger.info(f"Starting VBA-parity pipeline: {input_excel_path} -> {output_excel_path}")

        # Reset tracking state for this run
        self.corrected_columns = {}

        file_ext = _Path(input_excel_path).suffix.lower()
        is_macro_enabled = file_ext in [".xlsm", ".xltm", ".xlam"]

        # ---------------------------------------------------------------
        # FORMATTING PRESERVATION: start from a byte-for-byte file clone.
        #
        # openpyxl's load→save round-trip is lossy: it silently drops
        # conditional formatting ranges, column widths on shifted columns,
        # rich-text, sparklines, named ranges, print settings, and any
        # feature it does not model.  Starting from a real filesystem copy
        # means the output ZIP already contains every original entry; openpyxl
        # only rewrites the worksheet XML parts it actually touches, so all
        # untouched formatting survives intact.
        # ---------------------------------------------------------------
        _Path(output_excel_path).parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(input_excel_path, output_excel_path)
            self.logger.debug(f"Cloned source to output: {output_excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to clone source workbook: {e}")
            raise

        try:
            workbook = load_workbook(
                output_excel_path,          # open the CLONE, not the source
                data_only=False,
                keep_vba=is_macro_enabled,
                keep_links=False,
            )
            self.logger.info(f"Workbook loaded: {len(workbook.worksheets)} worksheet(s)")
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {e}")
            raise

        # VBA parity pre-clean: remove numeric helper row under detected headers.
        # This must run before any processor inserts columns so that row alignment
        # matches the VBA system exactly.
        for ws in workbook.worksheets:
            try:
                # Unmerge header-area horizontal merges before column insertions.
                # openpyxl insert_cols() does not update merged-cell ranges, so a
                # new column inserted inside a merge silently inherits the merge
                # and its header text is hidden.  Unmerging first gives every
                # header cell its own independent cell reference.
                self._unmerge_header_area(ws, max_row=20)
                if self._remove_numeric_helper_row(ws):
                    self.reader.invalidate_cache(ws)
            except Exception:
                continue

        # Export raw JSON alongside output (non-blocking — errors are logged only)
        workbook_dataset = None
        try:
            from .io_layer.excel_to_json_extractor import ExcelToJsonExtractor
            from .json_exporter import JsonExporter as _JsonExporter

            extractor = ExcelToJsonExtractor(
                excel_reader=self.reader,
                skip_empty_rows=False,
                handle_formulas=True,
                preserve_types=True,
            )
            workbook_dataset = extractor.extract_workbook_to_json(input_excel_path)
            _output_dir = _Path(output_excel_path).parent
            _exporter = _JsonExporter(indent=2, ensure_ascii=False)
            _exporter.export_workbook_to_json(workbook_dataset, str(_output_dir / "raw_dataset.json"))
            self.logger.info("Raw JSON exported")
        except Exception as e:
            self.logger.warning(f"Raw JSON export skipped: {e}")

        # Process each worksheet using field processors (VBA-style column insertion)
        for worksheet in workbook.worksheets:
            try:
                self.process_worksheet(worksheet)
            except Exception as e:
                self.logger.error(f"Failed to process worksheet '{worksheet.title}': {e}", exc_info=True)

        # Export normalized JSON and print console summary (non-blocking)
        try:
            if workbook_dataset is not None:
                from .processing.normalization_pipeline import NormalizationPipeline
                from .json_exporter import JsonExporter as _JsonExporter2

                pipeline = NormalizationPipeline(
                    name_engine=self.name_processor.name_engine,
                    gender_engine=self.gender_processor.gender_engine,
                    date_engine=self.date_processor.date_engine,
                    identifier_engine=self.identifier_processor.identifier_engine,
                    apply_name_normalization_enabled=True,
                    apply_gender_normalization_enabled=True,
                    apply_date_normalization_enabled=True,
                    apply_identifier_normalization_enabled=True,
                )

                total_rows = 0
                name_mods = gender_mods = date_mods = id_mods = 0

                normalized_sheets = []
                for sheet in workbook_dataset.sheets:
                    norm_sheet = pipeline.normalize_dataset(sheet)
                    normalized_sheets.append(norm_sheet)
                    stats = norm_sheet.get_metadata("normalization_statistics", {})
                    total_rows += stats.get("total_rows", 0)

                    for row in norm_sheet.rows:
                        for field in ["first_name", "last_name", "father_name"]:
                            orig = row.get(field)
                            corr = row.get(f"{field}_corrected")
                            if orig is not None and corr is not None and str(orig).strip() != str(corr).strip():
                                name_mods += 1
                        orig_g = row.get("gender")
                        corr_g = row.get("gender_corrected")
                        if orig_g is not None and corr_g is not None and str(orig_g).strip() != str(corr_g).strip():
                            gender_mods += 1
                        for prefix in ["birth", "entry"]:
                            for sub in ["year", "month", "day"]:
                                orig_d = row.get(f"{prefix}_{sub}")
                                corr_d = row.get(f"{prefix}_{sub}_corrected")
                                if orig_d is not None and corr_d is not None and str(orig_d).strip() != str(corr_d).strip():
                                    date_mods += 1
                                    break
                        orig_id = row.get("id_number")
                        corr_id = row.get("id_number_corrected")
                        if orig_id is not None and corr_id is not None and str(orig_id).strip() != str(corr_id).strip():
                            id_mods += 1

                workbook_dataset.sheets = normalized_sheets
                _output_dir2 = _Path(output_excel_path).parent
                _exporter2 = _JsonExporter2(indent=2, ensure_ascii=False)
                _exporter2.export_workbook_to_json(workbook_dataset, str(_output_dir2 / "normalized_dataset.json"))
                self.logger.info("Normalized JSON exported")

                # Console summary
                print("\n" + "=" * 60)
                print("Normalization Summary")
                print("=" * 60)
                print(f"Total rows processed : {total_rows}")
                print(f"NameEngine changes   : {name_mods}")
                print(f"GenderEngine changes : {gender_mods}")
                print(f"DateEngine changes   : {date_mods}")
                print(f"IdentifierEngine chg : {id_mods}")
                print("=" * 60)
                print("Verification: compare raw_dataset.json vs normalized_dataset.json")
                print("=" * 60 + "\n")
        except Exception as e:
            self.logger.warning(f"Normalized JSON export/summary skipped: {e}")

        # Save augmented workbook to output path (original is untouched)
        try:
            workbook.save(output_excel_path)
            self.logger.info(f"Workbook saved to: {output_excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            raise IOError(f"Failed to save workbook: {e}")

    # ------------------------------------------------------------------
    # VBA parity helpers
    # ------------------------------------------------------------------

    def _detect_header_row_for_vba_helper_row(self, ws: Worksheet) -> int:
        """Mirror VBA DetectHeaderRow: row 1..20 with >=3 '*- מתוקן*' cells."""
        max_row = min(20, ws.max_row or 0)
        max_col = ws.max_column or 0
        for r in range(1, max_row + 1):
            match_count = 0
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                if "- מתוקן" in str(v):
                    match_count += 1
                    if match_count >= 3:
                        return r
        return 0

    def _remove_numeric_helper_row(self, ws: Worksheet) -> bool:
        """VBA parity: delete numeric helper row immediately under headers.

        Mirrors VBA MainRunner.RemoveNumericHelperRow:
        - Detect header row using '*- מתוקן*' count >= 3 (rows 1..20).
        - Inspect checkRow = headerRow + 1 across columns 1..lastCol where
          lastCol is last non-empty column on headerRow.
        - If every non-empty cell is numeric and < 100, delete that row.
        """
        header_row = self._detect_header_row_for_vba_helper_row(ws)
        if header_row == 0:
            return False

        check_row = header_row + 1
        if check_row > (ws.max_row or 0):
            return False

        # lastCol: last non-empty cell on the header row
        last_col = 0
        for c in range(ws.max_column or 0, 0, -1):
            v = ws.cell(row=header_row, column=c).value
            if v is not None and str(v).strip() != "":
                last_col = c
                break
        if last_col == 0:
            return False

        non_helper_found = False
        for c in range(1, last_col + 1):
            val = ws.cell(row=check_row, column=c).value
            if str(val or "").strip() == "":
                continue
            # VBA: non-numeric text => not a helper row
            try:
                # accept numeric strings/floats/ints like VBA IsNumeric
                num = float(str(val).strip())
                if int(num) >= 100:
                    non_helper_found = True
                    break
            except Exception:
                non_helper_found = True
                break

        if not non_helper_found:
            ws.delete_rows(check_row, 1)
            return True

        return False

    def _unmerge_header_area(self, ws: Worksheet, max_row: int = 20) -> None:
        """Unmerge merged cells in the header area to preserve VBA-like insert behavior.

        The A.xlsx template uses horizontally merged header blocks; after inserting
        columns, openpyxl can keep those merges and thereby prevent distinct header
        values in the shifted columns. Excel/VBA output expects discrete header cells.
        """
        to_unmerge = []
        for rng in list(ws.merged_cells.ranges):
            if rng.max_row <= max_row:
                to_unmerge.append(rng)

        # Excel behavior: when unmerging, formatting is effectively preserved across
        # all cells in the former merged block. openpyxl keeps most style only on
        # the top-left cell; propagate it to the whole block before unmerging.
        import copy as _copy

        for rng in to_unmerge:
            try:
                tl = ws.cell(row=rng.min_row, column=rng.min_col)
                # Capture style facets from the top-left cell before unmerging.
                # We copy each facet individually because openpyxl's _style is an
                # internal index reference, not a standalone style object.
                tl_font = _copy.copy(tl.font)
                tl_fill = _copy.copy(tl.fill)
                tl_alignment = _copy.copy(tl.alignment)
                tl_border = _copy.copy(tl.border)
                tl_numfmt = tl.number_format

                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        if r == rng.min_row and c == rng.min_col:
                            continue  # top-left already has the style
                        cell = ws.cell(row=r, column=c)
                        try:
                            cell.font = _copy.copy(tl_font)
                            cell.fill = _copy.copy(tl_fill)
                            cell.alignment = _copy.copy(tl_alignment)
                            cell.border = _copy.copy(tl_border)
                            cell.number_format = tl_numfmt
                        except Exception:
                            pass
            except Exception:
                pass

            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                continue
