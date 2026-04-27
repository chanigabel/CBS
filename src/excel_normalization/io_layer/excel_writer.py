"""Excel writing operations for the normalization system.

This module provides the ExcelWriter class which encapsulates all openpyxl
write operations. It isolates Excel I/O from business logic.
"""

import logging
import os
from pathlib import Path
from typing import Any, List, Optional
import copy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font

from ..data_types import SheetDataset, WorkbookDataset, JsonRow

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles writing data to Excel worksheets.

    This class encapsulates all openpyxl write operations, providing a clean
    interface for the processing layer. It never performs business logic,
    only data writing and formatting.
    """

    # Pink highlight color for changed cells: RGB(255, 199, 206)
    # Use ARGB to match Excel/VBA-exported xlsx (FF = opaque).
    PINK_HIGHLIGHT = "FFFFC7CE"

    # Yellow highlight color for age warnings: RGB(255, 230, 150)
    YELLOW_HIGHLIGHT = "FFFFE696"

    # Pink background for date errors: RGB(255, 200, 200)
    PINK_ERROR = "FFFFC8C8"

    def _copy_cell_style_facets(self, src, dst) -> None:
        """Copy style facets that Excel typically propagates on column insertion."""
        try:
            dst.fill = copy.copy(src.fill)
        except Exception:
            pass
        try:
            dst.font = copy.copy(src.font)
        except Exception:
            pass
        try:
            dst.border = copy.copy(src.border)
        except Exception:
            pass
        try:
            dst.alignment = copy.copy(src.alignment)
        except Exception:
            pass
        try:
            dst.protection = copy.copy(src.protection)
        except Exception:
            pass
        try:
            dst.number_format = src.number_format
        except Exception:
            pass

    def prepare_output_column(self, worksheet: Worksheet, after_col: int, header_text: str, header_row: int) -> int:
        """Insert a single corrected column immediately to the right of after_col.

        VBA parity: matches FieldProcessor.ProcessSimpleField / PrepareOutputColumn,
        inserting the column only if the exact header is not already present.
        """

        output_col = after_col + 1

        existing = worksheet.cell(row=header_row, column=output_col).value

        if (existing or "").strip() != header_text:
            # Capture style of the column to the left (Excel/VBA behavior: inserted
            # columns inherit the formatting of the column they are inserted next to).
            max_row = worksheet.max_row or 1

            worksheet.insert_cols(output_col)

            # Shift ColumnDimension widths so existing widths follow their columns.
            self._shift_column_dimensions(worksheet, output_col, 1)

            # Apply propagated styles to the newly inserted column.
            for r in range(1, max_row + 1):
                left = worksheet.cell(row=r, column=after_col)
                inserted = worksheet.cell(row=r, column=output_col)
                self._copy_cell_style_facets(left, inserted)

            worksheet.cell(row=header_row, column=output_col).value = header_text

        return output_col

    def insert_output_columns(
        self,
        worksheet: Worksheet,
        after_col: int,
        count: int,
        header_row: int,
        headers: List[str],
    ) -> None:
        """Insert a block of output columns after a given column (VBA InsertOutputColumns).

        This mirrors the VBA ExcelWriter.InsertOutputColumns used for date and
        identifier fields:
        - Inserts `count` columns starting at after_col + 1
        - Writes the provided headers into the specified header_row
        - Shifts ColumnDimension widths so existing column widths are preserved
        """
        if count <= 0:
            return

        start_col = after_col + 1

        # Capture left-column styles once; VBA/Excel copies formatting from the
        # column immediately to the left of the insertion point.
        max_row = worksheet.max_row or 1

        worksheet.insert_cols(start_col, count)

        # Shift ColumnDimension entries so widths follow their original columns.
        # openpyxl's insert_cols() does NOT update ColumnDimension keys.
        self._shift_column_dimensions(worksheet, start_col, count)

        # Apply propagated style to each inserted column.
        for col in range(start_col, start_col + count):
            for r in range(1, max_row + 1):
                left = worksheet.cell(row=r, column=after_col)
                inserted = worksheet.cell(row=r, column=col)
                self._copy_cell_style_facets(left, inserted)

        for offset, header in enumerate(headers):
            col = start_col + offset
            worksheet.cell(row=header_row, column=col).value = header

    @staticmethod
    def _shift_column_dimensions(worksheet: Worksheet, insert_at: int, count: int) -> None:
        """Shift ColumnDimension width/hidden entries right after a column insertion.

        openpyxl stores ColumnDimension by column letter key. When insert_cols()
        is called, the cell data shifts but the ColumnDimension dict does not.
        This method rebuilds the dict so widths follow their original columns.
        """
        from openpyxl.utils import get_column_letter, column_index_from_string

        existing = {
            column_index_from_string(k): v
            for k, v in list(worksheet.column_dimensions.items())
        }

        new_dims: dict = {}
        for col_idx, dim in existing.items():
            if col_idx >= insert_at:
                new_dims[col_idx + count] = dim
            else:
                new_dims[col_idx] = dim

        worksheet.column_dimensions.clear()
        for col_idx, dim in new_dims.items():
            letter = get_column_letter(col_idx)
            worksheet.column_dimensions[letter] = dim

    def write_column_array(self, worksheet: Worksheet, col: int, start_row: int, values: List[Any]) -> None:
        """Write array to column.

        Writes a list of values to a column starting at the specified row.
        This enables efficient array-based writing instead of individual
        cell operations.

        Args:
            worksheet: The worksheet to write to
            col: Column number (1-based)
            start_row: Starting row number (1-based)
            values: List of values to write
        """
        for i, value in enumerate(values):
            row = start_row + i
            worksheet.cell(row=row, column=col).value = value

    def write_cell_value(self, worksheet: Worksheet, row: int, col: int, value: Any) -> None:
        """Write single cell value.

        Args:
            worksheet: The worksheet to write to
            row: Row number (1-based)
            col: Column number (1-based)
            value: Value to write to the cell
        """
        worksheet.cell(row=row, column=col).value = value

    def format_cell(
        self,
        worksheet: Worksheet,
        row: int,
        col: int,
        bg_color: Optional[str] = None,
        bold: bool = False,
        number_format: Optional[str] = None,
    ) -> None:
        """Apply formatting to cell.

        Applies background color, bold font, and/or number format to a cell.

        Args:
            worksheet: The worksheet containing the cell
            row: Row number (1-based)
            col: Column number (1-based)
            bg_color: Background color in hex format (e.g., "FFC7CE" for pink)
            bold: Whether to make the font bold
            number_format: Excel number format string (e.g., "0" for integer)
        """
        cell = worksheet.cell(row=row, column=col)

        # Apply background color
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        # Apply bold font
        if bold:
            cell.font = Font(bold=True)

        # Apply number format
        if number_format:
            cell.number_format = number_format

    def set_column_format(self, worksheet: Worksheet, col: int, format_string: str, start_row: int = 1) -> None:
        """Apply a number format string to an entire column.

        VBA parity intent: ws.Columns(col).NumberFormat = "0".

        Note: openpyxl does not support a true column-level number_format that
        Excel will always honor; applying it per-cell across the used range is
        the most reliable equivalent.
        """
        if col < 1:
            return

        # Important parity detail:
        # VBA applies a column number format, but in the saved workbook Excel often
        # only persists formats for cells in/near the used range. If we apply the
        # format to every row up to ws.max_row, openpyxl will persist it widely,
        # causing style diffs vs the VBA-produced file.
        #
        # Apply the format only to the worksheet's "effective used range" for this column.
        # This matches how Excel typically persists formats in saved files.
        if start_row < 1:
            start_row = 1

        last_used = start_row
        for r in range(worksheet.max_row or 1, 0, -1):
            v = worksheet.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                last_used = max(start_row, r)
                break

        for row_idx in range(start_row, last_used + 1):
            # Only persist formats on cells that are part of the used range band:
            # header cell and any populated data cell.
            cell = worksheet.cell(row=row_idx, column=col)
            if row_idx == start_row or (cell.value is not None and str(cell.value).strip() != ""):
                cell.number_format = format_string

    def highlight_changed_cells(
        self, worksheet: Worksheet, col: int, start_row: int, original_values: List[Any], corrected_values: List[Any]
    ) -> None:
        """Apply pink highlight to cells where values differ.

        Compares original and corrected values and applies pink background
        (RGB 255, 199, 206) to cells where the corrected value differs from
        the original value.

        Args:
            worksheet: The worksheet containing the cells
            col: Column number of corrected values (1-based)
            start_row: Starting row number (1-based)
            original_values: List of original values
            corrected_values: List of corrected values
        """
        # Pre-create a default fill to clear formatting when unchanged (VBA sets Pattern=None).
        clear_fill = PatternFill()

        for i, (original, corrected) in enumerate(zip(original_values, corrected_values)):
            # VBA parity: compare corrected against Trim(original) — whitespace-only
            # differences must NOT trigger a highlight.
            # Also handle numeric equality: int 1 == str "1" should not highlight.
            original_trimmed = str(original).strip() if original is not None else ""
            corrected_str = str(corrected).strip() if corrected is not None else ""

            # Normalize numeric representations: "1.0" == "1"
            def _normalize_num(s: str) -> str:
                try:
                    f = float(s)
                    if f == int(f):
                        return str(int(f))
                    return s
                except (ValueError, OverflowError):
                    return s

            original_norm = _normalize_num(original_trimmed)
            corrected_norm = _normalize_num(corrected_str)

            row = start_row + i

            # Apply pink highlight only when the corrected value differs from Trim(original)
            if corrected_norm != original_norm:
                self.format_cell(worksheet, row, col, bg_color=self.PINK_HIGHLIGHT)
            else:
                # VBA parity: explicitly clear highlight if there is no semantic change.
                worksheet.cell(row=row, column=col).fill = clear_fill



class JsonToExcelWriter:
    """Handles exporting JSON datasets back to Excel format.
    
    This class exports SheetDataset and WorkbookDataset instances to Excel files,
    writing both original and corrected values in side-by-side columns.
    
    Column Layout:
        For each field in the dataset, two columns are created:
        - Column N: Original field (e.g., "first_name")
        - Column N+1: Corrected field (e.g., "first_name_corrected")
    
    Features:
        - Writes header row with field names
        - Writes data rows with original and corrected values
        - Applies basic formatting to headers (bold)
        - Handles None values gracefully
        - Supports single-sheet and multi-sheet export
    
    Requirements:
        - Validates: Requirements 15.1-15.8, 20.4
    """
    
    def __init__(self, column_width: int = 15, apply_header_formatting: bool = True):
        """Initialize JsonToExcelWriter with configuration options.
        
        Args:
            column_width: Default width for columns (default: 15)
            apply_header_formatting: Whether to apply bold formatting to headers (default: True)
        """
        self.column_width = column_width
        self.apply_header_formatting = apply_header_formatting
    
    def _validate_dataset(self, dataset: SheetDataset) -> None:
        """Validate dataset structure before export.
        
        Args:
            dataset: SheetDataset to validate
        
        Raises:
            ValueError: If dataset structure is invalid
        
        Requirements:
            - Validates: Requirements 18.1-18.4
        """
        if not dataset.validate():
            raise ValueError(f"Invalid dataset structure for sheet '{dataset.sheet_name}'")
        
        if not dataset.rows:
            logger.warning(f"Sheet '{dataset.sheet_name}' has no data rows")
        
        # Validate that all rows have consistent field structure
        if dataset.rows:
            first_row_fields = set(dataset.rows[0].keys())
            for idx, row in enumerate(dataset.rows[1:], start=2):
                row_fields = set(row.keys())
                if row_fields != first_row_fields:
                    logger.warning(
                        f"Row {idx} in sheet '{dataset.sheet_name}' has inconsistent fields. "
                        f"Expected: {first_row_fields}, Got: {row_fields}"
                    )
    
    def _validate_workbook_dataset(self, workbook_dataset: WorkbookDataset) -> None:
        """Validate workbook dataset structure before export.
        
        Args:
            workbook_dataset: WorkbookDataset to validate
        
        Raises:
            ValueError: If workbook dataset structure is invalid
        
        Requirements:
            - Validates: Requirements 18.1-18.4
        """
        if not workbook_dataset.validate():
            raise ValueError(f"Invalid workbook dataset structure for '{workbook_dataset.source_file}'")
        
        if not workbook_dataset.sheets:
            raise ValueError(f"Workbook dataset '{workbook_dataset.source_file}' has no sheets")
        
        # Validate each sheet
        for sheet in workbook_dataset.sheets:
            self._validate_dataset(sheet)
    
    def _validate_output_path(self, output_path: str) -> None:
        """Validate output path is writable.
        
        Args:
            output_path: Path for output Excel file
        
        Raises:
            ValueError: If output path is invalid
            PermissionError: If output path is not writable
        
        Requirements:
            - Validates: Requirements 18.1-18.4
        """
        if not output_path:
            raise ValueError("Output path cannot be empty")
        
        # Convert to Path object for easier manipulation
        path = Path(output_path)
        
        # Check if parent directory exists
        parent_dir = path.parent
        if not parent_dir.exists():
            raise ValueError(f"Output directory does not exist: {parent_dir}")
        
        # Check if parent directory is writable
        if not os.access(parent_dir, os.W_OK):
            raise PermissionError(f"Output directory is not writable: {parent_dir}")
        
        # Check if file already exists and is writable
        if path.exists() and not os.access(path, os.W_OK):
            raise PermissionError(f"Output file exists but is not writable: {output_path}")
    
    def _cleanup_partial_file(self, output_path: str) -> None:
        """Clean up partial file if export fails.
        
        Args:
            output_path: Path to partial file to clean up
        
        Requirements:
            - Validates: Requirements 18.1-18.4
        """
        try:
            path = Path(output_path)
            if path.exists():
                path.unlink()
                logger.info(f"Cleaned up partial file: {output_path}")
        except Exception as e:
            logger.error(f"Failed to clean up partial file {output_path}: {e}")
    
    def create_header_row(self, worksheet: Worksheet, field_names: List[str]) -> None:
        """Create header row with original and corrected column names.
        
        Writes header row with field names, creating two columns for each field:
        one for the original value and one for the corrected value.
        
        Args:
            worksheet: Worksheet to write to
            field_names: List of field names from JSON (without "_corrected" suffix)
        
        Requirements:
            - Validates: Requirements 15.4-15.6
        """
        col_idx = 1
        
        for field_name in field_names:
            # Skip fields that already have "_corrected" suffix
            if field_name.endswith("_corrected"):
                continue
            
            # Write original field name
            cell = worksheet.cell(row=1, column=col_idx, value=field_name)
            if self.apply_header_formatting:
                cell.font = Font(bold=True)
            worksheet.column_dimensions[self._get_column_letter(col_idx)].width = self.column_width
            col_idx += 1
            
            # Write corrected field name
            corrected_name = f"{field_name}_corrected"
            cell = worksheet.cell(row=1, column=col_idx, value=corrected_name)
            if self.apply_header_formatting:
                cell.font = Font(bold=True)
            worksheet.column_dimensions[self._get_column_letter(col_idx)].width = self.column_width
            col_idx += 1
    
    def write_json_row(self, worksheet: Worksheet, row_num: int, 
                       json_row: JsonRow, field_names: List[str]) -> None:
        """Write a single JSON row to Excel worksheet.
        
        Writes both original and corrected values for each field in the row.
        Handles None values by writing empty cells.
        
        Args:
            worksheet: Worksheet to write to
            row_num: Row number to write to (1-based)
            json_row: Dictionary with field values
            field_names: Ordered list of field names (without "_corrected" suffix)
        
        Requirements:
            - Validates: Requirements 15.7
        """
        col_idx = 1
        
        for field_name in field_names:
            # Skip fields that already have "_corrected" suffix
            if field_name.endswith("_corrected"):
                continue
            
            # Write original value
            original_value = json_row.get(field_name)
            worksheet.cell(row=row_num, column=col_idx, value=original_value)
            col_idx += 1
            
            # Write corrected value
            corrected_name = f"{field_name}_corrected"
            corrected_value = json_row.get(corrected_name)
            worksheet.cell(row=row_num, column=col_idx, value=corrected_value)
            col_idx += 1
    
    def write_dataset_to_excel(self, dataset: SheetDataset, output_path: str) -> None:
        """Write a single sheet dataset to Excel file.
        
        Creates a new Excel workbook with one worksheet containing the dataset.
        Writes header row and all data rows with original and corrected values.
        
        Args:
            dataset: SheetDataset with corrected values
            output_path: Path for output Excel file
        
        Raises:
            ValueError: If dataset structure is invalid or output path is invalid
            PermissionError: If output path is not writable
            IOError: If file write operation fails
        
        Requirements:
            - Validates: Requirements 15.1-15.8, 18.1-18.4
        """
        # Validate dataset structure
        logger.info(f"Validating dataset for sheet '{dataset.sheet_name}'")
        self._validate_dataset(dataset)
        
        # Validate output path
        logger.info(f"Validating output path: {output_path}")
        self._validate_output_path(output_path)
        
        workbook = None
        try:
            # Create new workbook
            logger.info(f"Creating workbook for sheet '{dataset.sheet_name}'")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = dataset.sheet_name
            
            # Get field names (only original fields, not corrected)
            field_names = [name for name in dataset.field_names if not name.endswith("_corrected")]
            
            if not field_names:
                raise ValueError(f"No valid field names found in dataset '{dataset.sheet_name}'")
            
            # Write header row
            logger.info(f"Writing header row with {len(field_names)} fields")
            self.create_header_row(worksheet, field_names)
            
            # Write data rows
            logger.info(f"Writing {len(dataset.rows)} data rows")
            for idx, json_row in enumerate(dataset.rows, start=2):
                try:
                    self.write_json_row(worksheet, idx, json_row, field_names)
                except Exception as e:
                    logger.error(f"Error writing row {idx}: {e}")
                    raise IOError(f"Failed to write row {idx} in sheet '{dataset.sheet_name}': {e}")
            
            # Save workbook
            logger.info(f"Saving workbook to: {output_path}")
            workbook.save(output_path)
            logger.info(f"Successfully exported dataset to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error exporting dataset to {output_path}: {e}")
            # Clean up partial file
            self._cleanup_partial_file(output_path)
            raise
    
    def write_workbook_to_excel(self, workbook_dataset: WorkbookDataset, 
                                output_path: str) -> None:
        """Write multiple sheet datasets to Excel workbook.
        
        Creates a new Excel workbook with multiple worksheets, one for each
        SheetDataset in the WorkbookDataset. Each worksheet contains header
        row and data rows with original and corrected values.
        
        Args:
            workbook_dataset: WorkbookDataset with all sheets
            output_path: Path for output Excel file
        
        Raises:
            ValueError: If workbook dataset structure is invalid or output path is invalid
            PermissionError: If output path is not writable
            IOError: If file write operation fails
        
        Requirements:
            - Validates: Requirements 15.1-15.8, 16.3-16.4, 18.1-18.4
        """
        # Validate workbook dataset structure
        logger.info(f"Validating workbook dataset with {len(workbook_dataset.sheets)} sheets")
        self._validate_workbook_dataset(workbook_dataset)
        
        # Validate output path
        logger.info(f"Validating output path: {output_path}")
        self._validate_output_path(output_path)
        
        workbook = None
        try:
            # Create new workbook
            logger.info("Creating new workbook")
            workbook = Workbook()
            
            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]
            
            # Process each sheet dataset
            for sheet_idx, sheet_dataset in enumerate(workbook_dataset.sheets, start=1):
                try:
                    logger.info(f"Processing sheet {sheet_idx}/{len(workbook_dataset.sheets)}: '{sheet_dataset.sheet_name}'")
                    
                    # Create worksheet
                    worksheet = workbook.create_sheet(title=sheet_dataset.sheet_name)
                    
                    # Get field names (only original fields, not corrected)
                    field_names = [name for name in sheet_dataset.field_names 
                                  if not name.endswith("_corrected")]
                    
                    if not field_names:
                        logger.warning(f"No valid field names found in sheet '{sheet_dataset.sheet_name}', skipping")
                        continue
                    
                    # Write header row
                    logger.info(f"Writing header row with {len(field_names)} fields")
                    self.create_header_row(worksheet, field_names)
                    
                    # Write data rows
                    logger.info(f"Writing {len(sheet_dataset.rows)} data rows")
                    for idx, json_row in enumerate(sheet_dataset.rows, start=2):
                        try:
                            self.write_json_row(worksheet, idx, json_row, field_names)
                        except Exception as e:
                            logger.error(f"Error writing row {idx} in sheet '{sheet_dataset.sheet_name}': {e}")
                            raise IOError(f"Failed to write row {idx} in sheet '{sheet_dataset.sheet_name}': {e}")
                    
                    logger.info(f"Successfully processed sheet '{sheet_dataset.sheet_name}'")
                    
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_dataset.sheet_name}': {e}")
                    raise IOError(f"Failed to process sheet '{sheet_dataset.sheet_name}': {e}")
            
            # Save workbook
            logger.info(f"Saving workbook to: {output_path}")
            workbook.save(output_path)
            logger.info(f"Successfully exported workbook with {len(workbook_dataset.sheets)} sheets to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error exporting workbook to {output_path}: {e}")
            # Clean up partial file
            self._cleanup_partial_file(output_path)
            raise
    
    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """Convert column index to Excel column letter.
        
        Args:
            col_idx: Column index (1-based)
        
        Returns:
            Column letter (e.g., 1 -> 'A', 27 -> 'AA')
        """
        return get_column_letter(col_idx)
