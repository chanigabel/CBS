"""ExportEngine: create the final VBA-style export workbook.

This module replicates the VBA ExportEngine behavior:
- Creates a new workbook with sheets: DayarimYahidim, MeshkeyBayt, AnasheyTzevet
- Writes standardized headers in row 1
- Exports only rows that pass IsValidDataRow
- Export loop starts at headerRow + 2 (skipping the corrected sub-header row)
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, List, Any, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..data_types import WorkbookDataset, SheetDataset, JsonRow


CorrectedColumnsBySheet = Dict[str, Dict[str, int]]


@dataclass(frozen=True)
class ExportSheetSpec:
    source_sheet_name: str
    target_sheet_name: str
    include_dira: bool


class ExportEngine:
    """Create and populate the export workbook with VBA-parity layout."""

    def __init__(self) -> None:
        # Cache header detection results to avoid repeated scanning on large workbooks.
        # Keys are based on worksheet identity and header_row argument.
        self._header_row_cache_by_ws: Dict[int, int] = {}
        self._corrected_columns_cache_by_ws_and_header: Dict[Tuple[int, int], Dict[str, int]] = {}

        # Optional internal traceability for debugging parity mismatches.
        # When enabled, this is populated with per-exported-row metadata but is never
        # written to the output workbook.
        self.last_export_trace: List[Dict[str, Any]] = []

    # Exact source worksheet names used by the VBA system
    SOURCE_SHEET_SPECS: List[ExportSheetSpec] = [
        ExportSheetSpec("דיירים יחידים", "DayarimYahidim", False),
        ExportSheetSpec("מתגוררים במשקי בית", "MeshkeyBayt", True),
        ExportSheetSpec("אנשי צוות ובני משפחותיהם", "AnasheyTzevet", True),
    ]

    # Fixed header order (must match spec exactly, including casing)
    HEADERS_NO_DIRA = [
        "MosadID",
        "SugMosad",
        "ShemPrati",
        "ShemMishpaha",
        "ShemHaAv",
        "MisparZehut",
        "Darkon",
        "Min",
        "ShnatLida",
        "HodeshLida",
        "YomLida",
        "shnatknisa",
        "Hodeshknisa",
        "YomKnisa",
    ]

    HEADERS_WITH_DIRA = [
        "MosadID",
        "SugMosad",
        "MisparDiraBeMosad",
        "ShemPrati",
        "ShemMishpaha",
        "ShemHaAv",
        "MisparZehut",
        "Darkon",
        "Min",
        "ShnatLida",
        "HodeshLida",
        "YomLida",
        "shnatknisa",
        "Hodeshknisa",
        "YomKnisa",
    ]

    # ------------------------------------------------------------------
    # VBA helper functions (worksheet-based)
    # ------------------------------------------------------------------

    def detect_header_row(self, ws: Worksheet) -> int:
        """DetectHeaderRow: scan rows 1–20 for row with >= 3 '*- מתוקן*' cells."""
        ws_id = id(ws)
        if ws_id in self._header_row_cache_by_ws:
            return self._header_row_cache_by_ws[ws_id]

        max_row = min(20, ws.max_row or 0)
        max_col = ws.max_column or 0

        for r in range(1, max_row + 1):
            count = 0
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                if "- מתוקן" in str(v):
                    count += 1
                    if count >= 3:
                        self._header_row_cache_by_ws[ws_id] = r
                        return r

        self._header_row_cache_by_ws[ws_id] = 0
        return 0

    def detect_corrected_columns(self, ws: Worksheet, header_row: int) -> Dict[str, int]:
        """DetectCorrectedColumns: fallback mapping based on header text patterns."""
        cache_key = (id(ws), header_row)
        if cache_key in self._corrected_columns_cache_by_ws_and_header:
            return dict(self._corrected_columns_cache_by_ws_and_header[cache_key])

        max_col = ws.max_column or 0
        mapping: Dict[str, int] = {}

        # helper to access parent header (row above) for date subheaders
        def parent_text(col: int) -> str:
            if header_row <= 1:
                return ""
            pv = ws.cell(row=header_row - 1, column=col).value
            return str(pv or "")

        def nearest_left_header_text(col: int) -> str:
            """Find nearest non-empty header to the left on header_row."""
            for cc in range(col, 0, -1):
                v = ws.cell(row=header_row, column=cc).value
                if v is not None and str(v).strip() != "":
                    return str(v)
            return ""

        for c in range(1, max_col + 1):
            header_text = str(ws.cell(row=header_row, column=c).value or "")
            # In some layouts (including our Python processor output), date corrected headers
            # may be written on the row immediately below the main header row. For robust
            # detection, include the next row's text in matching.
            next_row_text = str(ws.cell(row=header_row + 1, column=c).value or "") if header_row + 1 <= (ws.max_row or 0) else ""
            combined_text = f"{header_text} {next_row_text}"

            # Names
            if ("שם פרטי" in header_text) and ("- מתוקן" in header_text or "ShemPrati" not in mapping):
                mapping["ShemPrati"] = c
            if ("שם משפחה" in header_text) and ("- מתוקן" in header_text or "ShemMishpaha" not in mapping):
                mapping["ShemMishpaha"] = c
            if ("שם האב" in header_text) and ("- מתוקן" in header_text or "ShemHaAv" not in mapping):
                mapping["ShemHaAv"] = c

            # Identifiers
            if (("ת.ז" in header_text) or ("זהות" in header_text)) and ("- מתוקן" in header_text or "MisparZehut" not in mapping):
                mapping["MisparZehut"] = c
            if ("דרכון" in header_text) and ("- מתוקן" in header_text or "Darkon" not in mapping):
                mapping["Darkon"] = c

            # Gender
            if ("מין" in header_text) and ("- מתוקן" in header_text or "Min" not in mapping):
                mapping["Min"] = c

            # Dates: match the VBA spec patterns, plus a pragmatic fallback when
            # the sub-header is only "שנה - מתוקן"/"חודש - מתוקן"/"יום - מתוקן".
            # If the main header row does not contain לידה/כניסה for these columns,
            # use the nearest non-empty header to the left as a parent context.
            parent_context = parent_text(c)
            if parent_context.strip() == "":
                parent_context = nearest_left_header_text(c)

            if "שנה" in combined_text:
                if ("לידה" in combined_text) and "ShnatLida" not in mapping:
                    mapping["ShnatLida"] = c
                elif ("כניסה" in combined_text) and "shnatknisa" not in mapping:
                    mapping["shnatknisa"] = c
                else:
                    pt = parent_context
                    if "לידה" in pt and "ShnatLida" not in mapping:
                        mapping["ShnatLida"] = c
                    if "כניסה" in pt and "shnatknisa" not in mapping:
                        mapping["shnatknisa"] = c

            if "חודש" in combined_text:
                if ("לידה" in combined_text) and "HodeshLida" not in mapping:
                    mapping["HodeshLida"] = c
                elif ("כניסה" in combined_text) and "Hodeshknisa" not in mapping:
                    mapping["Hodeshknisa"] = c
                else:
                    pt = parent_context
                    if "לידה" in pt and "HodeshLida" not in mapping:
                        mapping["HodeshLida"] = c
                    if "כניסה" in pt and "Hodeshknisa" not in mapping:
                        mapping["Hodeshknisa"] = c

            if "יום" in combined_text:
                if ("לידה" in combined_text) and "YomLida" not in mapping:
                    mapping["YomLida"] = c
                elif ("כניסה" in combined_text) and "YomKnisa" not in mapping:
                    mapping["YomKnisa"] = c
                else:
                    pt = parent_context
                    if "לידה" in pt and "YomLida" not in mapping:
                        mapping["YomLida"] = c
                    if "כניסה" in pt and "YomKnisa" not in mapping:
                        mapping["YomKnisa"] = c

        self._corrected_columns_cache_by_ws_and_header[cache_key] = dict(mapping)
        return mapping

    def determine_last_row_from_mapped_columns(self, ws: Worksheet, source_dict: Dict[str, int]) -> int:
        """DetermineLastRowFromMappedColumns: max last-used row across mapped columns."""
        max_row = 0
        for col in source_dict.values():
            max_row = max(max_row, self._last_used_row_in_column(ws, col))
        return max_row

    def is_valid_data_row(self, ws: Worksheet, source_dict: Dict[str, int], row_num: int) -> bool:
        """IsValidDataRow: export if ANY one of key personal fields is non-empty."""
        for key in ["ShemPrati", "ShemMishpaha", "ShemHaAv", "MisparZehut", "Darkon"]:
            col = source_dict.get(key)
            if not col:
                continue
            val = ws.cell(row=row_num, column=col).value
            if str(val or "").strip() != "":
                return True
        return False

    def find_target_column(self, ws: Worksheet, header_name: str) -> int:
        """FindTargetColumn: locate exact header name in row 1."""
        max_col = ws.max_column or 0
        for c in range(1, max_col + 1):
            if ws.cell(row=1, column=c).value == header_name:
                return c
        return 0

    def _last_used_row_in_column(self, ws: Worksheet, col: int) -> int:
        for r in range(ws.max_row or 0, 0, -1):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                return r
        return 0

    # ------------------------------------------------------------------
    # Workbook creation and JSON-based export (Step 6 pipeline target)
    # ------------------------------------------------------------------

    def write_headers(self, ws: Worksheet, include_dira: bool) -> None:
        headers = self.HEADERS_WITH_DIRA if include_dira else self.HEADERS_NO_DIRA
        for idx, name in enumerate(headers, start=1):
            ws.cell(row=1, column=idx).value = name

    def create_export_workbook(self) -> Workbook:
        wb = Workbook()
        # Remove default sheet
        if wb.sheetnames:
            wb.remove(wb[wb.sheetnames[0]])

        for spec in self.SOURCE_SHEET_SPECS:
            ws = wb.create_sheet(spec.target_sheet_name)
            self.write_headers(ws, include_dira=spec.include_dira)

        return wb

    # ------------------------------------------------------------------
    # Worksheet-based export (VBA parity with tracking dictionary)
    # ------------------------------------------------------------------

    def export_from_augmented_workbook(
        self,
        wb_source,
        corrected_columns_by_sheet: CorrectedColumnsBySheet,
        output_path: str,
        debug_trace: bool = False,
    ) -> str:
        """Export from an augmented workbook (after running processors).

        This replicates VBA ExportSheet behavior:
        - DetectHeaderRow scans 1..20 for >=3 '*- מתוקן*' headers
        - Prefer corrected_columns_by_sheet for mapping; fallback to DetectCorrectedColumns
        - Export loop starts at headerRow + 2
        - Row validity uses ANY ONE of key fields (if present in mapping)
        - MosadID/SugMosad/MisparDiraBeMosad are written ONLY if present in the tracking dict
        - The three main sheets always appear first in fixed order; any remaining
          source sheets are appended after them in their original relative order.
        """
        wb_export = self.create_export_workbook()
        self.last_export_trace = []

        for spec in self.SOURCE_SHEET_SPECS:
            if spec.source_sheet_name not in wb_source.sheetnames:
                continue

            ws_source = wb_source[spec.source_sheet_name]
            ws_target = wb_export[spec.target_sheet_name]

            self._export_sheet_from_worksheet(
                ws_source=ws_source,
                ws_target=ws_target,
                corrected_columns_by_sheet=corrected_columns_by_sheet,
                include_dira=spec.include_dira,
                debug_trace=debug_trace,
            )

        # Append any extra sheets (not in SOURCE_SHEET_SPECS) in their original order.
        known_source_names = {spec.source_sheet_name for spec in self.SOURCE_SHEET_SPECS}
        for sheet_name in wb_source.sheetnames:
            if sheet_name in known_source_names:
                continue
            ws_src = wb_source[sheet_name]
            ws_dst = wb_export.create_sheet(sheet_name)
            for row in ws_src.iter_rows():
                for cell in row:
                    ws_dst.cell(row=cell.row, column=cell.column).value = cell.value

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb_export.save(output_path)
        return output_path

    def _export_sheet_from_worksheet(
        self,
        ws_source: Worksheet,
        ws_target: Worksheet,
        corrected_columns_by_sheet: CorrectedColumnsBySheet,
        include_dira: bool,
        debug_trace: bool = False,
    ) -> None:
        header_row = self.detect_header_row(ws_source)
        if header_row == 0:
            return

        tracked = corrected_columns_by_sheet.get(ws_source.title) or {}

        # Prefer tracking dict if it has entries; else fallback to header-based detection
        if tracked:
            # Tracking columns may become stale after later insertions (Excel-style shifting).
            # To preserve VBA behavior ("always re-reading positions after insertions"),
            # merge the tracked mapping with a fresh header-based detection.
            detected = self.detect_corrected_columns(ws_source, header_row)
            # Start with detected, then apply tracked entries that still look valid.
            source_dict = dict(detected)

            def tracked_col_looks_valid(col: int) -> bool:
                v1 = ws_source.cell(row=header_row, column=col).value
                v2 = ws_source.cell(row=header_row + 1, column=col).value if header_row + 1 <= (ws_source.max_row or 0) else None
                t1 = str(v1 or "")
                t2 = str(v2 or "")
                return ("- מתוקן" in t1) or ("- מתוקן" in t2)

            for k, col in tracked.items():
                if tracked_col_looks_valid(col):
                    source_dict[k] = col
        else:
            source_dict = self.detect_corrected_columns(ws_source, header_row)

        # VBA parity: MosadID/SugMosad/MisparDiraBeMosad are ONLY available if tracked.
        if not tracked:
            source_dict.pop("MosadID", None)
            source_dict.pop("SugMosad", None)
            source_dict.pop("MisparDiraBeMosad", None)
        else:
            # Even when tracked exists, do NOT fill these from fallback detection.
            # Keep them only if present in tracked.
            for k in ["MosadID", "SugMosad", "MisparDiraBeMosad"]:
                if k not in tracked:
                    source_dict.pop(k, None)

        last_row = self.determine_last_row_from_mapped_columns(ws_source, source_dict)
        if last_row <= header_row:
            return

        out_row = 2
        start_row = header_row + 2  # critical: skip corrected sub-header row

        headers = self.HEADERS_WITH_DIRA if include_dira else self.HEADERS_NO_DIRA

        for r in range(start_row, last_row + 1):
            if not self.is_valid_data_row(ws_source, source_dict, r):
                continue

            for col_idx, field_key in enumerate(headers, start=1):
                source_col = source_dict.get(field_key)
                if source_col:
                    ws_target.cell(row=out_row, column=col_idx).value = ws_source.cell(row=r, column=source_col).value
                else:
                    # Missing mappings remain empty
                    ws_target.cell(row=out_row, column=col_idx).value = ""

            if debug_trace:
                self.last_export_trace.append(
                    {
                        "source_sheet_name": ws_source.title,
                        "source_row_index": r,
                        "target_sheet_name": ws_target.title,
                        "target_row_index": out_row,
                    }
                )

            out_row += 1

    def export_from_normalized_dataset(
        self,
        workbook_dataset: WorkbookDataset,
        output_path: str,
        corrected_columns_by_sheet: Optional[CorrectedColumnsBySheet] = None,
        debug_trace: bool = False,
    ) -> str:
        """Create the final VBA export workbook from normalized JSON datasets.

        NOTE: For VBA parity, MosadID/SugMosad/MisparDiraBeMosad should only be
        populated if they exist in the tracking dictionary. When exporting from
        JSON, pass corrected_columns_by_sheet if you want those fields filled;
        otherwise they remain empty (matching VBA behavior when tracking is absent).
        """
        wb = self.create_export_workbook()
        self.last_export_trace = []

        for spec in self.SOURCE_SHEET_SPECS:
            src = workbook_dataset.get_sheet_by_name(spec.source_sheet_name)
            if src is None:
                continue
            tgt = wb[spec.target_sheet_name]
            tracked = (corrected_columns_by_sheet or {}).get(spec.source_sheet_name) or {}
            self._export_sheet_from_json(
                src,
                tgt,
                include_dira=spec.include_dira,
                tracked_keys=tracked,
                debug_trace=debug_trace,
            )

        # Append any extra sheets (not in SOURCE_SHEET_SPECS) in their original order.
        known_source_names = {spec.source_sheet_name for spec in self.SOURCE_SHEET_SPECS}
        for sheet in workbook_dataset.sheets:
            if sheet.sheet_name in known_source_names:
                continue
            ws_extra = wb.create_sheet(sheet.sheet_name)
            # Write field names as header row
            for col_idx, field_name in enumerate(sheet.field_names, start=1):
                ws_extra.cell(row=1, column=col_idx).value = field_name
            # Write data rows
            for row_idx, row in enumerate(sheet.rows, start=2):
                for col_idx, field_name in enumerate(sheet.field_names, start=1):
                    ws_extra.cell(row=row_idx, column=col_idx).value = row.get(field_name)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _export_sheet_from_json(
        self,
        source: SheetDataset,
        target_ws: Worksheet,
        include_dira: bool,
        tracked_keys: Optional[Dict[str, int]] = None,
        debug_trace: bool = False,
    ) -> None:
        """Export a single source SheetDataset into a target export worksheet."""
        out_row = 2

        for idx, row in enumerate(source.rows, start=1):
            export_row = self._map_row_to_export_fields(
                row,
                include_dira=include_dira,
                allow_mosad_fields=bool(tracked_keys),
            )
            if not self._is_valid_export_row(export_row):
                continue

            # Write in the exact header order
            headers = self.HEADERS_WITH_DIRA if include_dira else self.HEADERS_NO_DIRA
            for col_idx, header_name in enumerate(headers, start=1):
                target_ws.cell(row=out_row, column=col_idx).value = export_row.get(header_name, "")

            if debug_trace:
                self.last_export_trace.append(
                    {
                        "source_sheet_name": source.sheet_name,
                        "source_row_index": idx,
                        "target_sheet_name": target_ws.title,
                        "target_row_index": out_row,
                    }
                )

            out_row += 1

    def _is_valid_export_row(self, export_row: Dict[str, Any]) -> bool:
        for key in ["ShemPrati", "ShemMishpaha", "ShemHaAv", "MisparZehut", "Darkon"]:
            if str(export_row.get(key, "") or "").strip() != "":
                return True
        return False

    def _map_row_to_export_fields(
        self,
        row: JsonRow,
        include_dira: bool,
        allow_mosad_fields: bool,
    ) -> Dict[str, Any]:
        """Map the project's JSON fields to the VBA export headers."""
        def pick(*keys: str) -> Any:
            for k in keys:
                if k in row:
                    v = row.get(k)
                    if v is not None and v != "":
                        return v
            # If all are empty/missing, return empty string for export
            return ""

        mosad_id = pick("MosadID", "mosad_id") if allow_mosad_fields else ""
        sug_mosad = pick("SugMosad", "sug_mosad") if allow_mosad_fields else ""

        mapped: Dict[str, Any] = {
            "MosadID": mosad_id,
            "SugMosad": sug_mosad,
            "ShemPrati": pick("first_name_corrected", "first_name"),
            "ShemMishpaha": pick("last_name_corrected", "last_name"),
            "ShemHaAv": pick("father_name_corrected", "father_name"),
            "MisparZehut": pick("id_number_corrected", "id_number"),
            "Darkon": pick("passport_corrected", "passport"),
            "Min": pick("gender_corrected", "gender"),
            "ShnatLida": pick("birth_year_corrected", "birth_year"),
            "HodeshLida": pick("birth_month_corrected", "birth_month"),
            "YomLida": pick("birth_day_corrected", "birth_day"),
            "shnatknisa": pick("entry_year_corrected", "entry_year"),
            "Hodeshknisa": pick("entry_month_corrected", "entry_month"),
            "YomKnisa": pick("entry_day_corrected", "entry_day"),
        }

        if include_dira:
            mapped["MisparDiraBeMosad"] = pick("MisparDiraBeMosad", "dira", "apartment") if allow_mosad_fields else ""

        return mapped

