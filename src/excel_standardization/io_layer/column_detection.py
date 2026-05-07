"""Column/date detection helpers for ExcelReader."""

from __future__ import annotations

import re

from typing import Dict, List, Optional, Set

from openpyxl.worksheet.worksheet import Worksheet

from ..data_types import ColumnHeaderInfo, DateFieldType, DateGroup, TableRegion

HEBREW_YEAR = "שנה"
HEBREW_MONTH = "חודש"
HEBREW_DAY = "יום"
HEBREW_DATE = "תאריך"
HEBREW_BIRTH = "לידה"
HEBREW_ENTRY = "כניסה"


def detect_date_groups(reader, worksheet: Worksheet, table_region: TableRegion) -> Dict[DateFieldType, DateGroup]:
    """Detect split date groups for birth/entry independently per field."""
    groups: Dict[DateFieldType, DateGroup] = {}

    if table_region.header_rows < 2:
        return groups

    header_row = table_region.start_row
    subheader_row = header_row + 1

    effective_end_col = table_region.end_col
    for c in range(worksheet.max_column or 0, 0, -1):
        v = worksheet.cell(row=subheader_row, column=c).value
        if v is not None and str(v).strip() != "":
            effective_end_col = max(effective_end_col, c)
            break

    def find_parent_col(keyword_list: List[str]) -> Optional[int]:
        for c in range(table_region.start_col, effective_end_col + 1):
            v = worksheet.cell(row=header_row, column=c).value
            if v is None or str(v).strip() == "":
                continue
            norm = reader._normalize_text(str(v).strip())
            if any(reader._normalize_text(k) in norm for k in keyword_list):
                return c
        return None

    birth_parent_col = find_parent_col(reader.FIELD_KEYWORDS.get("birth_date", []))
    entry_parent_col = find_parent_col(reader.FIELD_KEYWORDS.get("entry_date", []))

    for field_type, parent_col, other_parent_col in [
        (DateFieldType.BIRTH_DATE, birth_parent_col, entry_parent_col),
        (DateFieldType.ENTRY_DATE, entry_parent_col, birth_parent_col),
    ]:
        if parent_col is None:
            continue

        if other_parent_col is not None and other_parent_col > parent_col:
            hard_stop_primary = other_parent_col - 1
            hard_stop_fallback = other_parent_col
        else:
            hard_stop_primary = effective_end_col
            hard_stop_fallback = effective_end_col

        year_col = month_col = day_col = 0
        found_count = 0
        for c in range(parent_col + 1, hard_stop_primary + 1):
            v = worksheet.cell(row=subheader_row, column=c).value
            if v is None or str(v).strip() == "":
                if found_count > 0 and c < parent_col + 10:
                    continue
                elif found_count > 0:
                    break
                else:
                    continue
            txt = str(v).strip()
            if txt == HEBREW_YEAR and year_col == 0:
                year_col = c; found_count += 1
            elif txt == HEBREW_MONTH and month_col == 0:
                month_col = c; found_count += 1
            elif txt == HEBREW_DAY and day_col == 0:
                day_col = c; found_count += 1
            elif found_count > 0:
                break
            if found_count == 3:
                break

        if not (year_col and month_col and day_col):
            year_col = month_col = day_col = 0
            for c in range(parent_col, hard_stop_fallback + 1):
                v = worksheet.cell(row=subheader_row, column=c).value
                if v is None or str(v).strip() == "":
                    continue
                txt = str(v).strip()
                if txt == HEBREW_YEAR and year_col == 0:
                    year_col = c
                elif txt == HEBREW_MONTH and month_col == 0:
                    month_col = c
                elif txt == HEBREW_DAY and day_col == 0:
                    day_col = c
                if year_col and month_col and day_col:
                    break

        if year_col and month_col and day_col:
            groups[field_type] = DateGroup(
                year_col=year_col,
                month_col=month_col,
                day_col=day_col,
                main_col=parent_col,
                field_type=field_type,
            )

    return groups


def match_field(reader, normalized_text: str) -> Optional[str]:
    best_match = None
    best_match_length = 0

    for field_name, keywords in reader.FIELD_KEYWORDS.items():
        for keyword in keywords:
            if keyword in normalized_text and len(keyword) > best_match_length:
                best_match = field_name
                best_match_length = len(keyword)

    return best_match


def detect_date_subcolumns(
    reader,
    worksheet: Worksheet,
    start_col: int,
    subheader_row: int,
    max_col: int,
) -> Dict[str, int]:
    date_columns = {}

    parent_row = subheader_row - 1
    search_start_col = start_col
    search_end_col = min(start_col + 5, max_col + 1)

    if parent_row >= 1:
        if reader._is_merged_cell(worksheet, parent_row, start_col):
            merge_range = reader._get_merged_cell_range(worksheet, parent_row, start_col)
            if merge_range:
                search_start_col = merge_range[2]
                search_end_col = min(merge_range[3] + 1, max_col + 1)

                parent_cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value
                if parent_cell_value:
                    parent_normalized = reader._normalize_text(str(parent_cell_value))
                    date_keywords = [HEBREW_DATE, HEBREW_BIRTH, HEBREW_ENTRY, 'date', 'birth', 'entry']
                    if not any(kw in parent_normalized for kw in date_keywords):
                        search_start_col = start_col
                        search_end_col = min(start_col + 5, max_col + 1)
        else:
            parent_cell_value = worksheet.cell(row=parent_row, column=start_col).value
            if parent_cell_value:
                parent_normalized = reader._normalize_text(str(parent_cell_value))
                date_keywords = [HEBREW_DATE, HEBREW_BIRTH, HEBREW_ENTRY, 'date', 'birth', 'entry']
                if any(kw in parent_normalized for kw in date_keywords):
                    search_end_col = min(start_col + 5, max_col + 1)

    for col_idx in range(search_start_col, search_end_col):
        cell_value = worksheet.cell(row=subheader_row, column=col_idx).value

        if (cell_value is None or str(cell_value).strip() == "") and reader._is_merged_cell(worksheet, subheader_row, col_idx):
            merge_range = reader._get_merged_cell_range(worksheet, subheader_row, col_idx)
            if merge_range:
                cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

        if cell_value is None:
            continue

        normalized = reader._normalize_text(str(cell_value))

        if any(kw in normalized for kw in [HEBREW_YEAR, 'year']):
            if 'year' not in date_columns:
                date_columns['year'] = col_idx
        elif any(kw in normalized for kw in [HEBREW_MONTH, 'month']):
            if 'month' not in date_columns:
                date_columns['month'] = col_idx
        elif any(kw in normalized for kw in [HEBREW_DAY, 'day']):
            if 'day' not in date_columns:
                date_columns['day'] = col_idx

    if len(date_columns) == 3:
        return date_columns
    return {}

def detect_columns(reader, worksheet: Worksheet) -> Dict[str, ColumnHeaderInfo]:
    """Detect all relevant columns in the worksheet using intelligent table detection.

        This method:
        1. Detects the table region
        2. Identifies column headers using keyword matching
        3. Handles multi-row headers (e.g., date groups)
        4. Returns a mapping of field names to column information

        Args:
            worksheet: The worksheet to analyze

        Returns:
            Dictionary mapping field names to ColumnHeaderInfo
        """
    if True:
        # Check cache
        ws_id = id(worksheet)
        if ws_id in reader._column_mapping_cache:
            return reader._column_mapping_cache[ws_id]

        # Detect table region
        table_region = reader.detect_table_region(worksheet)
        if table_region is None:
            reader._column_mapping_cache[ws_id] = {}
            return {}

        column_mapping: Dict[str, ColumnHeaderInfo] = {}
        processed_merged_cols = set()  # Track columns already processed as part of merged cells
        # Track every col_idx that was handled by the keyword-matching loop
        # (including date group parent headers that produce sub-columns rather
        # than a direct mapping entry).  These must be excluded from the
        # passthrough pass so they don't appear as duplicate raw columns.
        keyword_handled_cols: Set[int] = set()

        # Scan header row(s) for columns
        header_row = table_region.start_row
        subheader_row = header_row + 1 if table_region.header_rows == 2 else None

        # Deterministic date grouping (birth/entry)
        date_groups = reader.detect_date_groups(worksheet, table_region)

        for col_idx in range(table_region.start_col, table_region.end_col + 1):
            # Skip if this column was already processed as part of a merged cell
            if col_idx in processed_merged_cols:
                continue

            # Get header cell text
            header_cell = worksheet.cell(row=header_row, column=col_idx)
            cell_value = header_cell.value

            # Handle merged cells - get value from top-left cell and mark all spanned columns
            if (cell_value is None or str(cell_value).strip() == "") and reader._is_merged_cell(worksheet, header_row, col_idx):
                merge_range = reader._get_merged_cell_range(worksheet, header_row, col_idx)
                if merge_range:
                    cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value
                    for merged_col in range(merge_range[2], merge_range[3] + 1):
                        processed_merged_cols.add(merged_col)
            elif reader._is_merged_cell(worksheet, header_row, col_idx):
                merge_range = reader._get_merged_cell_range(worksheet, header_row, col_idx)
                if merge_range:
                    for merged_col in range(merge_range[2], merge_range[3] + 1):
                        processed_merged_cols.add(merged_col)

            if cell_value is None:
                continue

            header_text = str(cell_value).strip()

            if reader._should_ignore_column(header_text):
                continue

            normalized_header = reader._normalize_text(header_text)
            matched_field = reader._match_field(normalized_header)

            if matched_field:
                keyword_handled_cols.add(col_idx)

                if matched_field in ["birth_date", "entry_date"] and subheader_row:
                    group_type = DateFieldType.BIRTH_DATE if matched_field == "birth_date" else DateFieldType.ENTRY_DATE
                    group = date_groups.get(group_type)
                    if group:
                        prefix = "birth" if matched_field == "birth_date" else "entry"
                        column_mapping[f"{prefix}_year"] = ColumnHeaderInfo(
                            col=group.year_col,
                            header_row=subheader_row,
                            last_row=table_region.end_row,
                            header_text=str(worksheet.cell(row=subheader_row, column=group.year_col).value or ""),
                        )
                        column_mapping[f"{prefix}_month"] = ColumnHeaderInfo(
                            col=group.month_col,
                            header_row=subheader_row,
                            last_row=table_region.end_row,
                            header_text=str(worksheet.cell(row=subheader_row, column=group.month_col).value or ""),
                        )
                        column_mapping[f"{prefix}_day"] = ColumnHeaderInfo(
                            col=group.day_col,
                            header_row=subheader_row,
                            last_row=table_region.end_row,
                            header_text=str(worksheet.cell(row=subheader_row, column=group.day_col).value or ""),
                        )
                    else:
                        column_mapping[matched_field] = ColumnHeaderInfo(
                            col=col_idx,
                            header_row=header_row,
                            last_row=table_region.end_row,
                            header_text=header_text,
                        )
                else:
                    column_mapping[matched_field] = ColumnHeaderInfo(
                        col=col_idx,
                        header_row=header_row,
                        last_row=table_region.end_row,
                        header_text=header_text,
                    )

        # ---------------------------------------------------------------
        # Passthrough pass: add every column whose header did NOT match a
        # keyword so that no Excel column is silently dropped.
        # ---------------------------------------------------------------
        already_mapped_cols: Set[int] = {info.col for info in column_mapping.values()}

        for col_idx in range(table_region.start_col, table_region.end_col + 1):
            if col_idx in already_mapped_cols or col_idx in processed_merged_cols or col_idx in keyword_handled_cols:
                continue

            cell_value = worksheet.cell(row=header_row, column=col_idx).value

            if (cell_value is None or str(cell_value).strip() == "") and reader._is_merged_cell(worksheet, header_row, col_idx):
                merge_range = reader._get_merged_cell_range(worksheet, header_row, col_idx)
                if merge_range:
                    cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

            if cell_value is None or str(cell_value).strip() == "":
                continue

            header_text = str(cell_value).strip()

            if reader._should_ignore_column(header_text):
                continue

            safe_key = re.sub(r'[^\w\u0590-\u05FF]+', '_', header_text).strip('_') or f"col_{col_idx}"
            if safe_key in column_mapping:
                safe_key = f"{safe_key}_{col_idx}"

            column_mapping[safe_key] = ColumnHeaderInfo(
                col=col_idx,
                header_row=header_row,
                last_row=table_region.end_row,
                header_text=header_text,
            )

        # ---------------------------------------------------------------
        # Sub-header pass (two-row header layout only):
        # When header_rows == 2, some regular fields (e.g. שם פרטי, שם משפחה,
        # שם האב) may live exclusively on the sub-header row while the top
        # header row has empty cells in those columns.
        #
        # ALL שנה/חודש/יום cells on the sub-header row are excluded to prevent
        # phantom `year`/`month`/`day` columns in the mapping.
        # ---------------------------------------------------------------
        if subheader_row is not None:
            already_mapped_cols = {info.col for info in column_mapping.values()}

            # Exclude all date-component sub-header columns
            _date_component_cols: Set[int] = set()
            for dg in date_groups.values():
                _date_component_cols.update([dg.year_col, dg.month_col, dg.day_col])
            _eff_end = table_region.end_col
            for _c in range(worksheet.max_column or 0, 0, -1):
                _v = worksheet.cell(row=subheader_row, column=_c).value
                if _v is not None and str(_v).strip() != "":
                    _eff_end = max(_eff_end, _c)
                    break
            for _c in range(table_region.start_col, _eff_end + 1):
                _v = worksheet.cell(row=subheader_row, column=_c).value
                if _v is not None and str(_v).strip() in ("שנה", "חודש", "יום"):
                    _date_component_cols.add(_c)

            for col_idx in range(table_region.start_col, table_region.end_col + 1):
                if col_idx in already_mapped_cols or col_idx in _date_component_cols:
                    continue

                cell_value = worksheet.cell(row=subheader_row, column=col_idx).value

                if (cell_value is None or str(cell_value).strip() == "") and reader._is_merged_cell(worksheet, subheader_row, col_idx):
                    merge_range = reader._get_merged_cell_range(worksheet, subheader_row, col_idx)
                    if merge_range:
                        cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

                if cell_value is None or str(cell_value).strip() == "":
                    continue

                header_text = str(cell_value).strip()

                if reader._should_ignore_column(header_text):
                    continue

                # Skip cells that look like data values rather than headers.
                if reader._looks_like_data_value(cell_value):
                    continue

                normalized_header = reader._normalize_text(header_text)
                matched_field = reader._match_field(normalized_header)

                if matched_field:
                    if matched_field not in column_mapping:
                        column_mapping[matched_field] = ColumnHeaderInfo(
                            col=col_idx,
                            header_row=subheader_row,
                            last_row=table_region.end_row,
                            header_text=header_text,
                        )
                else:
                    safe_key = re.sub(r'[^\w\u0590-\u05FF]+', '_', header_text).strip('_') or f"col_{col_idx}"
                    if safe_key in column_mapping:
                        safe_key = f"{safe_key}_{col_idx}"
                    column_mapping[safe_key] = ColumnHeaderInfo(
                        col=col_idx,
                        header_row=subheader_row,
                        last_row=table_region.end_row,
                        header_text=header_text,
                    )

        # Sort the final mapping by physical Excel column number so that
        # field_names (built from list(column_mapping.keys())) reflects the
        # true left-to-right worksheet column order regardless of which pass
        # (keyword, passthrough, sub-header) inserted each entry.
        sorted_mapping: Dict[str, ColumnHeaderInfo] = dict(
            sorted(column_mapping.items(), key=lambda kv: kv[1].col)
        )

        reader._column_mapping_cache[ws_id] = sorted_mapping
        return sorted_mapping

