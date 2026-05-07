"""Merged-cell helpers for ExcelReader."""

from __future__ import annotations

from typing import Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet


def is_merged_cell(worksheet: Worksheet, row: int, col: int) -> bool:
    """Check if a cell is part of a merged range."""
    try:
        cell = worksheet.cell(row=row, column=col)
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False
    except Exception:
        return False


def get_merged_cell_range(
    worksheet: Worksheet,
    row: int,
    col: int,
) -> Optional[Tuple[int, int, int, int]]:
    """Get the boundaries of a merged cell range."""
    try:
        cell = worksheet.cell(row=row, column=col)
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return (
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )
        return None
    except Exception:
        return None
