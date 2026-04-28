"""Excel to JSON extraction component.

This module provides the ExcelToJsonExtractor class that converts Excel worksheets
to JSON format using the detected column mappings from ExcelReader.

The extractor handles:
- Single and multi-row headers
- Empty cells (stored as None)
- Formula cells (extracts calculated values)
- Multiple worksheets in a workbook
- Error handling for extraction failures

Requirements:
    - Validates: Requirements 10.1, 20.2
"""

import logging
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

from ..data_types import (
    ColumnHeaderInfo,
    JsonRow,
    SheetDataset,
    TableRegion,
    WorkbookDataset,
)
from .excel_reader import ExcelReader

logger = logging.getLogger(__name__)


class ExcelToJsonExtractor:
    """Extracts data from Excel worksheets and converts to JSON format.
    
    This class is responsible for converting Excel rows to JSON dictionaries
    based on column mappings detected by ExcelReader. It preserves original
    values exactly as they appear in Excel and handles various cell types.
    
    The extractor uses ExcelReader as a dependency to detect headers and
    table regions, then extracts data rows into JSON format.
    
    Configuration Options:
        skip_empty_rows: If True, skip rows where all values are None/empty
        handle_formulas: If True, extract calculated values from formula cells
        preserve_types: If True, preserve Excel data types (numbers, dates)
        max_scan_rows: Maximum rows to scan for headers (default: 30)
    
    Example:
        reader = ExcelReader()
        extractor = ExcelToJsonExtractor(
            excel_reader=reader,
            skip_empty_rows=True,
            handle_formulas=True
        )
        
        # Extract single sheet
        workbook = load_workbook("data.xlsx")
        sheet = workbook.active
        dataset = extractor.extract_sheet_to_json(sheet)
        
        # Extract entire workbook
        workbook_dataset = extractor.extract_workbook_to_json("data.xlsx")
    
    Requirements:
        - Validates: Requirements 10.1, 20.2
    """
    
    def __init__(
        self,
        excel_reader: ExcelReader,
        skip_empty_rows: bool = False,
        handle_formulas: bool = True,
        preserve_types: bool = True,
        max_scan_rows: int = 30,
    ) -> None:
        """Initialize the ExcelToJsonExtractor.
        
        Args:
            excel_reader: ExcelReader instance for detecting headers and columns
            skip_empty_rows: If True, skip rows where all values are None/empty
            handle_formulas: If True, extract calculated values from formula cells
            preserve_types: If True, preserve Excel data types (numbers, dates)
            max_scan_rows: Maximum rows to scan for headers (default: 30)
        """
        self.excel_reader = excel_reader
        self.skip_empty_rows = skip_empty_rows
        self.handle_formulas = handle_formulas
        self.preserve_types = preserve_types
        self.max_scan_rows = max_scan_rows
    
    def extract_row_to_json(
        self,
        worksheet: Worksheet,
        row_num: int,
        column_mapping: Dict[str, ColumnHeaderInfo],
    ) -> JsonRow:
        """Extract a single row and convert to JSON dictionary.
        
        Reads cell values from the specified row for each field in the column
        mapping and creates a JSON dictionary with field names as keys.
        
        Handles errors gracefully:
        - Invalid cell values are stored as None with a warning
        - Merged cells are handled by extracting the top-left cell value
        - Formula errors are logged and stored as None
        
        Args:
            worksheet: The worksheet to extract from
            row_num: Row number to extract (1-based)
            column_mapping: Field name to column mapping from detect_columns
        
        Returns:
            Dictionary with field names as keys and cell values as values
        
        Requirements:
            - Validates: Requirements 10.2-10.7, 18.1-18.4
        """
        json_row: JsonRow = {}
        
        for field_name, col_info in column_mapping.items():
            try:
                # Read cell value using ExcelReader's method
                cell_value = self.excel_reader.read_cell_value(
                    worksheet, row_num, col_info.col
                )
                
                # Always get the cell object for merged-cell and formula checks
                cell = worksheet.cell(row=row_num, column=col_info.col)

                # Handle formulas: extract calculated value
                if self.handle_formulas:
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_value = cell.value
                    
                    # Check for formula errors
                    if isinstance(cell_value, str) and cell_value.startswith('#'):
                        logger.warning(
                            f"Formula error in sheet '{worksheet.title}', "
                            f"row {row_num}, column {col_info.col}, field '{field_name}': {cell_value}"
                        )
                        cell_value = None

                    # Formula string (starts with '=') means data_only=True returned
                    # the formula text instead of the computed value — treat as None.
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        logger.debug(
                            f"Formula cell (unevaluated) in sheet '{worksheet.title}', "
                            f"row {row_num}, col {col_info.col}, field '{field_name}' — stored as None"
                        )
                        cell_value = None
                
                # Handle merged cells - check if this cell is part of a merged range
                if cell.coordinate in worksheet.merged_cells:
                    logger.debug(
                        f"Merged cell detected in sheet '{worksheet.title}', "
                        f"row {row_num}, column {col_info.col}, field '{field_name}'"
                    )
                    # openpyxl automatically returns the value from the top-left cell
                    # of a merged range, so we can use cell_value as-is
                
                # Store value in JSON row
                # Empty cells are stored as None
                json_row[field_name] = cell_value
                
            except Exception as e:
                # Handle any unexpected errors during cell extraction
                logger.warning(
                    f"Error extracting cell value in sheet '{worksheet.title}', "
                    f"row {row_num}, column {col_info.col}, field '{field_name}': {e}"
                )
                # Store None for cells that cannot be read
                json_row[field_name] = None
        
        return json_row
    
    def extract_sheet_to_json(
        self,
        worksheet: Worksheet,
        column_mapping: Optional[Dict[str, ColumnHeaderInfo]] = None,
        table_region: Optional[TableRegion] = None,
    ) -> SheetDataset:
        """Extract all data rows from worksheet and convert to JSON format.
        
        Uses ExcelReader to detect headers and table region if not provided,
        then extracts all data rows into JSON format.
        
        Handles errors gracefully:
        - Missing headers result in an empty dataset with error metadata
        - Invalid rows are skipped with warnings
        - Extraction continues even if some cells fail
        
        Args:
            worksheet: The worksheet to extract from
            column_mapping: Field name to column mapping (auto-detected if None)
            table_region: Table boundaries and header info (auto-detected if None)
        
        Returns:
            SheetDataset with raw JSON rows and metadata
        
        Requirements:
            - Validates: Requirements 10.1-10.5, 11.1-11.5, 18.1-18.4
        """
        try:
            # Detect columns and table region if not provided
            if column_mapping is None:
                column_mapping = self.excel_reader.detect_columns(worksheet)
            
            if table_region is None:
                table_region = self.excel_reader.detect_table_region(
                    worksheet, max_scan_rows=self.max_scan_rows
                )
            
            # Handle case where no valid headers found
            if not column_mapping or table_region is None:
                logger.warning(
                    f"No valid headers found in sheet '{worksheet.title}'. "
                    f"Sheet will be skipped."
                )
                return SheetDataset(
                    sheet_name=worksheet.title,
                    header_row=0,
                    header_rows_count=0,
                    field_names=[],
                    rows=[],
                    metadata={
                        "error": "No valid headers found",
                        "skipped": True,
                    },
                )
            
            # Extract field names from column mapping
            field_names = list(column_mapping.keys())
            logger.info(
                f"Extracting sheet '{worksheet.title}' with {len(field_names)} fields: {field_names}"
            )
            
            # Extract all data rows
            rows: List[JsonRow] = []
            skipped_rows = 0
            error_rows = 0
            
            for row_num in range(table_region.data_start_row, table_region.end_row + 1):
                try:
                    json_row = self.extract_row_to_json(worksheet, row_num, column_mapping)
                    
                    # Skip empty rows if configured
                    if self.skip_empty_rows:
                        if all(value is None or value == "" for value in json_row.values()):
                            skipped_rows += 1
                            continue
                    
                    rows.append(json_row)
                    
                except Exception as e:
                    # Log error and continue with next row
                    logger.warning(
                        f"Error extracting row {row_num} in sheet '{worksheet.title}': {e}. "
                        f"Row will be skipped."
                    )
                    error_rows += 1
                    continue
            
            # Determine date field structure for metadata
            date_field_structure = {}
            if "birth_date" in field_names:
                date_field_structure["birth_date"] = "single"
            elif "birth_year" in field_names:
                date_field_structure["birth_date"] = "split"
            
            if "entry_date" in field_names:
                date_field_structure["entry_date"] = "single"
            elif "entry_year" in field_names:
                date_field_structure["entry_date"] = "split"
            
            logger.info(
                f"Successfully extracted {len(rows)} rows from sheet '{worksheet.title}' "
                f"(skipped {skipped_rows} empty rows, {error_rows} error rows)"
            )
            
            # Create SheetDataset with metadata
            dataset = SheetDataset(
                sheet_name=worksheet.title,
                header_row=table_region.start_row,
                header_rows_count=table_region.header_rows,
                field_names=field_names,
                rows=rows,
                metadata={
                    "total_rows": len(rows),
                    "skipped_rows": skipped_rows,
                    "error_rows": error_rows,
                    "date_field_structure": date_field_structure,
                    "data_start_row": table_region.data_start_row,
                    "data_end_row": table_region.end_row,
                },
            )
            
            return dataset
            
        except Exception as e:
            # Handle any unexpected errors during sheet extraction
            logger.error(
                f"Unexpected error extracting sheet '{worksheet.title}': {e}. "
                f"Sheet will be skipped."
            )
            return SheetDataset(
                sheet_name=worksheet.title,
                header_row=0,
                header_rows_count=0,
                field_names=[],
                rows=[],
                metadata={
                    "error": str(e),
                    "skipped": True,
                },
            )
    
    def extract_workbook_to_json(self, workbook_path: str) -> WorkbookDataset:
        """Extract all worksheets from a workbook to JSON format.
        
        Opens the workbook and processes each worksheet independently,
        creating a SheetDataset for each one. Sheets with no valid headers
        are skipped with warnings logged.
        
        Handles errors gracefully:
        - Invalid workbook files raise clear exceptions
        - Sheets with no headers are skipped with warnings
        - Sheet extraction errors are logged and the sheet is skipped
        - Processing continues even if some sheets fail
        
        Args:
            workbook_path: Path to Excel file
        
        Returns:
            WorkbookDataset containing all sheet datasets
        
        Raises:
            FileNotFoundError: If workbook file does not exist
            InvalidFileException: If file is not a valid Excel file
        
        Requirements:
            - Validates: Requirements 16.1-16.6, 18.1-18.4
        """
        try:
            logger.info(f"Loading workbook from '{workbook_path}'")
            workbook = load_workbook(workbook_path, data_only=True)
            logger.info(f"Workbook loaded with {len(workbook.worksheets)} sheets")
            
        except FileNotFoundError:
            logger.error(f"Workbook file not found: '{workbook_path}'")
            raise
        except InvalidFileException as e:
            logger.error(f"Invalid Excel file: '{workbook_path}': {e}")
            raise
        except Exception as e:
            logger.error(f"Error loading workbook '{workbook_path}': {e}")
            raise
        
        sheets: List[SheetDataset] = []
        skipped_sheets: List[str] = []
        
        for worksheet in workbook.worksheets:
            try:
                logger.info(f"Processing sheet '{worksheet.title}'")
                dataset = self.extract_sheet_to_json(worksheet)
                
                # Check if sheet was skipped due to no headers or errors
                if dataset.get_metadata("skipped", False):
                    skipped_sheets.append(worksheet.title)
                    error_msg = dataset.get_metadata("error", "Unknown error")
                    logger.warning(
                        f"Sheet '{worksheet.title}' skipped: {error_msg}"
                    )
                else:
                    sheets.append(dataset)
                    logger.info(
                        f"Sheet '{worksheet.title}' processed successfully: "
                        f"{len(dataset.rows)} rows extracted"
                    )
                    
            except Exception as e:
                # Log error and skip this sheet
                logger.error(
                    f"Unexpected error processing sheet '{worksheet.title}': {e}. "
                    f"Sheet will be skipped."
                )
                skipped_sheets.append(worksheet.title)
                continue
        
        logger.info(
            f"Workbook extraction complete: {len(sheets)} sheets processed, "
            f"{len(skipped_sheets)} sheets skipped"
        )
        
        if skipped_sheets:
            logger.warning(f"Skipped sheets: {', '.join(skipped_sheets)}")
        
        # Create WorkbookDataset with metadata
        workbook_dataset = WorkbookDataset(
            source_file=workbook_path,
            sheets=sheets,
            metadata={
                "total_sheets": len(workbook.worksheets),
                "processed_sheets": len(sheets),
                "skipped_sheets": skipped_sheets,
            },
        )
        
        return workbook_dataset
