"""Sheet access helpers for ExcelReader."""

from __future__ import annotations

from typing import Any, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from ..data_types import ColumnHeaderInfo

HEBREW_CORRECTED = "\u05de\u05ea\u05d5\u05e7\u05df"


def find_header(reader, worksheet: Worksheet, search_terms: List[str], normalize_linebreaks: bool = False) -> Optional[ColumnHeaderInfo]:
    """Find column by exact text matching (xlPart equivalent)."""
    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value

            if cell_value is None:
                continue

            cell_text = str(cell_value)

            if normalize_linebreaks:
                cell_text = cell_text.replace("\r\n", "\n").replace("\r", "\n")

            for search_term in search_terms:
                search_text = search_term
                if normalize_linebreaks:
                    search_text = search_text.replace("\\r\\n", "\n").replace("\\r", "\n").replace("\\n", "\n")

                if search_text in cell_text:
                    if HEBREW_CORRECTED in cell_text or "corrected" in cell_text.lower():
                        continue

                    last_row = reader.get_last_row(worksheet, col_idx)
                    return ColumnHeaderInfo(col=col_idx, header_row=row_idx, last_row=last_row, header_text=cell_text)

    return None


def read_column_array(reader, worksheet: Worksheet, col: int, start_row: int, end_row: int) -> List[Any]:
    """Read column data as array."""
    values = []
    for row_idx in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        values.append(cell.value)
    return values


def read_cell_value(reader, worksheet: Worksheet, row: int, col: int) -> Any:
    """Read single cell value."""
    return worksheet.cell(row=row, column=col).value


def get_last_row(reader, worksheet: Worksheet, col: int) -> int:
    """Find last non-empty row in column."""
    for row_idx in range(worksheet.max_row, 0, -1):
        cell = worksheet.cell(row=row_idx, column=col)
        if cell.value is not None and str(cell.value).strip() != "":
            return row_idx

    return 0
