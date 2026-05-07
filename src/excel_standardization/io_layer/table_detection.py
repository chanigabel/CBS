"""Table detection helpers for ExcelReader."""

from __future__ import annotations

from typing import Any, Callable, Dict, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from ..data_types import TableRegion


def detect_table_region(
    worksheet: Worksheet,
    max_scan_rows: int,
    table_region_cache: Dict[int, Optional[TableRegion]],
    normalize_text: Callable[[str], str],
    score_header_row: Callable[[Worksheet, int, int], int],
    score_subheader_row: Callable[[Worksheet, int, int], int],
    is_merged_cell: Callable[[Worksheet, int, int], bool],
    get_merged_cell_range: Callable[[Worksheet, int, int], Optional[Tuple[int, int, int, int]]],
) -> Optional[TableRegion]:
    ws_id = id(worksheet)
    if ws_id in table_region_cache:
        return table_region_cache[ws_id]

    max_row = min(max_scan_rows, worksheet.max_row)
    max_col = worksheet.max_column

    row_scores = []
    for row_idx in range(1, max_row + 1):
        score = score_header_row(worksheet, row_idx, max_col)
        row_scores.append((row_idx, score))

    if not row_scores:
        table_region_cache[ws_id] = None
        return None

    row_scores.sort(key=lambda x: x[1], reverse=True)
    best_row, best_score = row_scores[0]

    if best_score < 3:
        table_region_cache[ws_id] = None
        return None

    header_rows = 1
    header_start_row = best_row

    is_subheader = False
    subheader_keywords = ['שנה', 'חודש', 'יום', 'year', 'month', 'day']
    for col_idx in range(1, min(max_col + 1, 50)):
        cell_value = worksheet.cell(row=best_row, column=col_idx).value
        if cell_value:
            normalized = normalize_text(str(cell_value))
            if any(kw in normalized for kw in subheader_keywords):
                is_subheader = True
                break

    if is_subheader and best_row > 1:
        parent_row = best_row - 1
        parent_score = score_header_row(worksheet, parent_row, max_col)
        if parent_score >= 2:
            header_rows = 2
            header_start_row = parent_row
        else:
            for col_idx in range(1, min(max_col + 1, 50)):
                if is_merged_cell(worksheet, parent_row, col_idx):
                    mr = get_merged_cell_range(worksheet, parent_row, col_idx)
                    if mr and mr[3] > mr[2]:
                        pv = worksheet.cell(row=mr[0], column=mr[2]).value
                        if pv and any(
                            kw in normalize_text(str(pv))
                            for kw in ['תאריך', 'לידה', 'כניסה', 'date', 'birth', 'entry']
                        ):
                            header_rows = 2
                            header_start_row = parent_row
                            break

    if best_row < max_row and header_rows == 1:
        next_row_score = score_subheader_row(worksheet, best_row + 1, max_col)
        if next_row_score >= 2:
            header_rows = 2

    data_start_row = header_start_row + header_rows

    start_col, end_col = find_table_columns(
        worksheet,
        header_start_row,
        max_col,
        is_merged_cell,
        get_merged_cell_range,
    )

    if header_rows == 2:
        sub_start, sub_end = find_table_columns(
            worksheet,
            header_start_row + 1,
            max_col,
            is_merged_cell,
            get_merged_cell_range,
        )
        start_col = min(start_col, sub_start)
        end_col = max(end_col, sub_end)

    if is_column_index_row(worksheet, data_start_row, start_col, end_col):
        data_start_row += 1

    end_row = find_table_end_row(
        worksheet,
        data_start_row,
        start_col,
        end_col,
        is_merged_cell,
        get_merged_cell_range,
    )

    table_region = TableRegion(
        start_row=header_start_row,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col,
        header_rows=header_rows,
        data_start_row=data_start_row,
    )

    table_region_cache[ws_id] = table_region
    return table_region


def score_header_row(
    worksheet: Worksheet,
    row_idx: int,
    max_col: int,
    normalize_text: Callable[[str], str],
    contains_field_keyword: Callable[[str], bool],
    is_merged_cell: Callable[[Worksheet, int, int], bool],
    get_merged_cell_range: Callable[[Worksheet, int, int], Optional[Tuple[int, int, int, int]]],
) -> int:
    score = 0
    non_empty_count = 0
    text_count = 0
    keyword_matches = 0

    for col_idx in range(1, min(max_col + 1, 50)):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value

        if (cell_value is None or str(cell_value).strip() == "") and is_merged_cell(worksheet, row_idx, col_idx):
            merge_range = get_merged_cell_range(worksheet, row_idx, col_idx)
            if merge_range:
                cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

        if cell_value is None or str(cell_value).strip() == "":
            continue

        non_empty_count += 1
        cell_text = str(cell_value).strip()

        if not cell_text.replace(".", "").replace(",", "").isdigit():
            text_count += 1

        normalized = normalize_text(cell_text)
        if contains_field_keyword(normalized):
            keyword_matches += 1

    if non_empty_count >= 3:
        score += 2
    if non_empty_count >= 5:
        score += 1

    if text_count >= non_empty_count * 0.7:
        score += 2

    score += keyword_matches * 2
    return score


def score_subheader_row(
    worksheet: Worksheet,
    row_idx: int,
    max_col: int,
    normalize_text: Callable[[str], str],
    is_merged_cell: Callable[[Worksheet, int, int], bool],
    get_merged_cell_range: Callable[[Worksheet, int, int], Optional[Tuple[int, int, int, int]]],
) -> int:
    score = 0
    subheader_keywords = ['שנה', 'חודש', 'יום', 'year', 'month', 'day']
    date_keywords = ['תאריך', 'לידה', 'כניסה', 'date', 'birth', 'entry']

    parent_row = row_idx - 1
    if parent_row < 1:
        return 0

    matched_subheaders = 0
    valid_parent_child_pairs = 0

    for col_idx in range(1, min(max_col + 1, 50)):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value

        if (cell_value is None or str(cell_value).strip() == "") and is_merged_cell(worksheet, row_idx, col_idx):
            merge_range = get_merged_cell_range(worksheet, row_idx, col_idx)
            if merge_range:
                cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

        if cell_value is None:
            continue

        normalized = normalize_text(str(cell_value))
        has_subheader_keyword = any(keyword in normalized for keyword in subheader_keywords)

        if has_subheader_keyword:
            matched_subheaders += 1

            if is_merged_cell(worksheet, parent_row, col_idx):
                merge_range = get_merged_cell_range(worksheet, parent_row, col_idx)
                if merge_range:
                    start_col, end_col = merge_range[2], merge_range[3]
                    if end_col > start_col:
                        parent_cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value
                        if parent_cell_value:
                            parent_normalized = normalize_text(str(parent_cell_value))
                            if any(kw in parent_normalized for kw in date_keywords):
                                valid_parent_child_pairs += 1
                                score += 3
                            else:
                                score += 1
                        else:
                            score += 1
                    else:
                        score += 1
                else:
                    score += 1
            else:
                parent_cell_value = worksheet.cell(row=parent_row, column=col_idx).value
                if parent_cell_value:
                    parent_normalized = normalize_text(str(parent_cell_value))
                    if any(kw in parent_normalized for kw in date_keywords):
                        valid_parent_child_pairs += 1
                        score += 2
                    else:
                        score += 1
                else:
                    score += 1

    if valid_parent_child_pairs >= 2:
        score += 2

    if matched_subheaders >= 3:
        score += 1

    return score


def is_column_index_row(
    worksheet: Worksheet,
    row_idx: int,
    start_col: int,
    end_col: int,
) -> bool:
    values = []
    for col_idx in range(start_col, end_col + 1):
        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
        if cell_value is None:
            continue
        if isinstance(cell_value, float):
            if cell_value != int(cell_value):
                continue
            cell_value = int(cell_value)
        if not isinstance(cell_value, int):
            continue
        if cell_value < 1 or cell_value > end_col:
            return False
        values.append(cell_value)

    if len(values) < 3:
        return False
    if len(values) != len(set(values)):
        return False
    return True


def find_table_columns(
    worksheet: Worksheet,
    header_row: int,
    max_col: int,
    is_merged_cell: Callable[[Worksheet, int, int], bool],
    get_merged_cell_range: Callable[[Worksheet, int, int], Optional[Tuple[int, int, int, int]]],
) -> Tuple[int, int]:
    start_col = 1
    end_col = max_col

    for col_idx in range(1, max_col + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if (cell_value is None or str(cell_value).strip() == "") and is_merged_cell(worksheet, header_row, col_idx):
            merge_range = get_merged_cell_range(worksheet, header_row, col_idx)
            if merge_range:
                cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

        if cell_value is not None and str(cell_value).strip() != "":
            start_col = col_idx
            break

    for col_idx in range(max_col, 0, -1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if (cell_value is None or str(cell_value).strip() == "") and is_merged_cell(worksheet, header_row, col_idx):
            merge_range = get_merged_cell_range(worksheet, header_row, col_idx)
            if merge_range:
                cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

        if cell_value is not None and str(cell_value).strip() != "":
            end_col = col_idx
            break

    return start_col, end_col


def find_table_end_row(
    worksheet: Worksheet,
    data_start_row: int,
    start_col: int,
    end_col: int,
    is_merged_cell: Callable[[Worksheet, int, int], bool],
    get_merged_cell_range: Callable[[Worksheet, int, int], Optional[Tuple[int, int, int, int]]],
) -> int:
    max_row = worksheet.max_row
    last_data_row = data_start_row

    for row_idx in range(data_start_row, max_row + 1):
        has_data = False
        for col_idx in range(start_col, end_col + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value

            if (cell_value is None or str(cell_value).strip() == "") and is_merged_cell(worksheet, row_idx, col_idx):
                merge_range = get_merged_cell_range(worksheet, row_idx, col_idx)
                if merge_range:
                    cell_value = worksheet.cell(row=merge_range[0], column=merge_range[2]).value

            if cell_value is not None and str(cell_value).strip() != "":
                has_data = True
                break

        if has_data:
            last_data_row = row_idx
        elif row_idx > last_data_row + 5:
            break

    return last_data_row
