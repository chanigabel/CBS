from typing import Dict, Any, List, Optional
from openpyxl.worksheet.worksheet import Worksheet

from .field_processor import FieldProcessor
from ..io_layer.excel_reader import ExcelReader
from ..io_layer.excel_writer import ExcelWriter
from ..engines.date_engine import DateEngine
from ..data_types import DateFormatPattern, DateFieldType


class DateFieldProcessor(FieldProcessor):

    def __init__(self, reader: ExcelReader, writer: ExcelWriter, date_engine: DateEngine):
        super().__init__(reader, writer)
        self.date_engine = date_engine
        # For each DateFieldType keep a list of groups (to support multiple
        # date groups per sheet, mirroring VBA's Find/FindNext loop).
        self.date_fields: Dict[DateFieldType, List[Dict[str, Any]]] = {}

    # ----------------------------------------------------
    # HEADER DETECTION
    # ----------------------------------------------------

    def find_headers(self, worksheet: Worksheet) -> bool:

        # reset state (critical for multi-sheet workbooks)
        self.date_fields = {DateFieldType.BIRTH_DATE: [], DateFieldType.ENTRY_DATE: []}

        # Discover all birth-date groups (VBA-style Find/FindNext semantics).
        birth_terms = ["תאריך לידה"]
        self._collect_date_groups(worksheet, birth_terms, DateFieldType.BIRTH_DATE)

        # Discover all entry-date groups.
        entry_terms = ["תאריך כניסה למוסד"]
        self._collect_date_groups(worksheet, entry_terms, DateFieldType.ENTRY_DATE)

        # True if we found at least one group of any type
        return any(self.date_fields[field_type] for field_type in self.date_fields)

    def _collect_date_groups(
        self,
        worksheet: Worksheet,
        search_terms: List[str],
        field_type: DateFieldType,
    ) -> None:
        """Collect all date groups for a given header text.

        This mirrors VBA's pattern of:
            Set headerCell = Cells.Find(...)
            Do
                'process this headerCell
                Set headerCell = Cells.FindNext(headerCell)
            Loop While ...

        Here we approximate that behavior by scanning the sheet and collecting
        all cells whose value contains the search term, ordered top-to-bottom,
        left-to-right.
        """
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0

        groups: List[Dict[str, Any]] = []

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                val = worksheet.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                text = val.strip()
                if not text:
                    continue
                if any(term in text for term in search_terms):
                    # Treat this cell as a main header for a date group.
                    sub = self._find_date_sub_headers(worksheet, c, r)
                    if not sub:
                        continue
                    groups.append(
                        {
                            "main_header": type("HeaderInfo", (), {"col": c, "header_row": r, "text": text}),
                            "year_col": sub["year_col"],
                            "month_col": sub["month_col"],
                            "day_col": sub["day_col"],
                            "sub_header_row": sub["sub_header_row"],
                            "last_row": sub["last_row"],
                        }
                    )

        if groups:
            self.date_fields[field_type] = groups
    # ----------------------------------------------------

    def _find_date_sub_headers(self, worksheet: Worksheet, main_col: int, main_header_row: int):

        sub_row = main_header_row + 1

        year_col = None
        month_col = None
        day_col = None

        for offset in range(0, 10):

            col = main_col + offset

            if col > worksheet.max_column:
                break

            val = self.reader.read_cell_value(worksheet, sub_row, col)

            if val is None:
                continue

            text = str(val).strip()

            if text == "שנה" and year_col is None:
                year_col = col

            elif text == "חודש" and month_col is None:
                month_col = col

            elif text == "יום" and day_col is None:
                day_col = col

        if year_col and month_col and day_col:
            # VBA parity: lastRow is the MAX of last-used rows across the main and split columns.
            last_row = max(
                self.reader.get_last_row(worksheet, main_col),
                self.reader.get_last_row(worksheet, year_col),
                self.reader.get_last_row(worksheet, month_col),
                self.reader.get_last_row(worksheet, day_col),
            )

            return {
                "year_col": year_col,
                "month_col": month_col,
                "day_col": day_col,
                "sub_header_row": sub_row,
                "last_row": last_row,
            }

        return None

    # ----------------------------------------------------
    # OUTPUT COLUMNS
    # ----------------------------------------------------

    def prepare_output_columns(self, worksheet: Worksheet):

        # Process each group by re-scanning the live worksheet state before
        # each insertion — mirroring VBA's Find/FindNext which always operates
        # on the current sheet, never on cached positions.
        for field_type, groups in self.date_fields.items():
            for info in groups:
                # Re-locate the main header on the live worksheet using the
                # anchor stored at find_headers time (main header row + text).
                main_header_row = info["main_header"].header_row
                main_col = info["main_header"].col
                main_text = info["main_header"].text

                # Re-find the current column of this main header by scanning
                # its known row for the stored text (handles any prior shifts).
                live_main_col = self._find_col_by_text(worksheet, main_header_row, main_text)
                if live_main_col is None:
                    # Fallback: use stored col (should not happen on a clean sheet)
                    live_main_col = main_col

                # Re-detect sub-headers from the live worksheet state.
                sub = self._find_date_sub_headers(worksheet, live_main_col, main_header_row)
                if sub is None:
                    continue

                # Update info with live positions so process_data reads correctly.
                info["main_header"].col = live_main_col
                info["year_col"] = sub["year_col"]
                info["month_col"] = sub["month_col"]
                info["day_col"] = sub["day_col"]
                info["sub_header_row"] = sub["sub_header_row"]
                info["last_row"] = sub["last_row"]

                base_col = sub["day_col"]
                header_row = sub["sub_header_row"]

                # Idempotency guard: only insert when corrected header is not
                # already present immediately to the right of the day column.
                existing = worksheet.cell(row=header_row, column=base_col + 1).value
                if isinstance(existing, str) and existing.strip() == "שנה - מתוקן":
                    info["corrected_year_col"] = base_col + 1
                    info["corrected_month_col"] = base_col + 2
                    info["corrected_day_col"] = base_col + 3
                    info["corrected_status_col"] = base_col + 4
                    continue

                self.writer.insert_output_columns(
                    worksheet,
                    after_col=base_col,
                    count=4,
                    header_row=header_row,
                    headers=["שנה - מתוקן", "חודש - מתוקן", "יום - מתוקן", "סטטוס תאריך"],
                )

                info["corrected_year_col"] = base_col + 1
                info["corrected_month_col"] = base_col + 2
                info["corrected_day_col"] = base_col + 3
                info["corrected_status_col"] = base_col + 4

                self.writer.set_column_format(worksheet, info["corrected_year_col"], "0", start_row=header_row)
                self.writer.set_column_format(worksheet, info["corrected_month_col"], "0", start_row=header_row)
                self.writer.set_column_format(worksheet, info["corrected_day_col"], "0", start_row=header_row)

    def _find_col_by_text(self, worksheet: Worksheet, row: int, text: str) -> Optional[int]:
        """Scan a row for a cell whose value contains text; return its column."""
        for c in range(1, (worksheet.max_column or 0) + 1):
            val = worksheet.cell(row=row, column=c).value
            if isinstance(val, str) and text in val.strip():
                return c
        return None

    # ----------------------------------------------------
    # PATTERN DETECTION
    # ----------------------------------------------------

    def detect_date_format_pattern(self, date_values: List[Any]) -> DateFormatPattern:

        ddmm = 0
        mmdd = 0

        for value in date_values:

            if value is None:
                continue

            text = str(value).strip()

            if "/" not in text and "." not in text and "-" not in text:
                continue

            text = text.replace(".", "/").replace("-", "/")

            parts = text.split("/")

            if len(parts) < 2:
                continue

            try:

                first = int(parts[0])
                second = int(parts[1])

                if first > 12 and second <= 12:
                    ddmm += 1

                elif second > 12 and first <= 12:
                    mmdd += 1

            except Exception:
                continue

        if mmdd > ddmm:
            return DateFormatPattern.MMDD

        return DateFormatPattern.DDMM

    # ----------------------------------------------------
    # PROCESSING
    # ----------------------------------------------------

    def process_data(self, worksheet: Worksheet):

        for field_type, groups in self.date_fields.items():
            for info in groups:
                self._process_date_field(worksheet, info, field_type)

    # ----------------------------------------------------

    def _normalize_split_value(self, val):
        """Normalize a split date component value to a string or None."""
        if val is None:
            return None

        # Pass through numeric types directly — DateEngine handles int/float
        if isinstance(val, (int, float)):
            if val != val:  # NaN check
                return None
            return val

        txt = str(val).strip()

        if txt == "" or txt.lower() in ("none", "null", "nan"):
            return None

        return txt

    # ----------------------------------------------------

    def _process_date_field(self, worksheet: Worksheet, info: Dict[str, Any], field_type: DateFieldType):

        year_col = info["year_col"]
        month_col = info["month_col"]
        day_col = info["day_col"]

        start_row = info["sub_header_row"] + 1
        end_row = info["last_row"]

        year_vals = self.reader.read_column_array(worksheet, year_col, start_row, end_row)
        month_vals = self.reader.read_column_array(worksheet, month_col, start_row, end_row)
        day_vals = self.reader.read_column_array(worksheet, day_col, start_row, end_row)

        main_vals = self.reader.read_column_array(
            worksheet,
            info["main_header"].col,
            start_row,
            end_row,
        )

        pattern = self.detect_date_format_pattern(main_vals)

        row_count = max(len(year_vals), len(month_vals), len(day_vals), len(main_vals))

        # ------------------------------------------------------------------
        # Pass 1: initial per-row completion
        # ------------------------------------------------------------------
        results = []
        for i in range(row_count):
            y = year_vals[i] if i < len(year_vals) else None
            m = month_vals[i] if i < len(month_vals) else None
            d = day_vals[i] if i < len(day_vals) else None
            main = main_vals[i] if i < len(main_vals) else None

            y = self._normalize_split_value(y)
            m = self._normalize_split_value(m)
            d = self._normalize_split_value(d)

            result = self.date_engine.parse_date(y, m, d, main, pattern, field_type)
            results.append(result)

        # ------------------------------------------------------------------
        # Pass 2: list-level one-way majority correction (birth dates only)
        #
        # Rule: if the majority of auto-completed years landed in the 1900s
        # and only a minority landed in the 2000s, flip those 2000s outliers
        # to their 1900s equivalents.  The reverse (flipping 1900s to 2000s)
        # is intentionally never done.
        #
        # Only auto-completed years (year_was_auto_completed=True) are
        # eligible for correction; explicitly written 4-digit years are
        # never touched.
        # ------------------------------------------------------------------
        if field_type == DateFieldType.BIRTH_DATE:
            results = self._apply_majority_century_correction(results)

        # ------------------------------------------------------------------
        # Collect output arrays
        # ------------------------------------------------------------------
        corrected_years = []
        corrected_months = []
        corrected_days = []
        status = []

        for result in results:
            corrected_years.append(result.year or "")
            corrected_months.append(result.month or "")
            corrected_days.append(result.day or "")
            status.append(result.status_text)

        self.writer.write_column_array(
            worksheet,
            info["corrected_year_col"],
            start_row,
            corrected_years,
        )

        self.writer.write_column_array(
            worksheet,
            info["corrected_month_col"],
            start_row,
            corrected_months,
        )

        self.writer.write_column_array(
            worksheet,
            info["corrected_day_col"],
            start_row,
            corrected_days,
        )

        # Apply integer number format ("0") to corrected year/month/day columns
        # (column-level intent; implemented via writer helper).
        self.writer.set_column_format(worksheet, info["corrected_year_col"], "0", start_row=info["sub_header_row"])
        self.writer.set_column_format(worksheet, info["corrected_month_col"], "0", start_row=info["sub_header_row"])
        self.writer.set_column_format(worksheet, info["corrected_day_col"], "0", start_row=info["sub_header_row"])

        self.writer.write_column_array(
            worksheet,
            info["corrected_status_col"],
            start_row,
            status,
        )

        # Status-cell formatting rules:
        # - Ignore empty status
        # - Trim text before evaluating
        # - "גיל מעל" => yellow + bold
        # - otherwise => pink + bold
        for offset, msg in enumerate(status):
            text = str(msg).strip() if msg is not None else ""
            if not text:
                continue

            row_idx = start_row + offset

            if "גיל מעל" in text:
                self.writer.format_cell(
                    worksheet,
                    row_idx,
                    info["corrected_status_col"],
                    bg_color=self.writer.YELLOW_HIGHLIGHT,
                    bold=True,
                )
            else:
                self.writer.format_cell(
                    worksheet,
                    row_idx,
                    info["corrected_status_col"],
                    bg_color=self.writer.PINK_ERROR,
                    bold=True,
                )

    # ----------------------------------------------------
    # MAJORITY CENTURY CORRECTION
    # ----------------------------------------------------

    def _apply_majority_century_correction(
        self, results: List
    ) -> List:
        """One-way list-level correction: flip auto-completed 2000s outliers
        to 1900s when the majority of auto-completed years are in the 1900s.

        Rules:
        - Only auto-completed years (year_was_auto_completed=True) are
          considered and eligible for correction.
        - Explicitly written 4-digit years are never touched.
        - If majority of auto-completed years are 1900s → flip 2000s outliers
          to 1900s equivalents (subtract 100).
        - The reverse (flipping 1900s to 2000s) is intentionally never done.
        - After flipping, re-run validate_business_rules so status is correct.
        """
        from ..data_types import DateFieldType

        # Collect auto-completed years that have a valid year value
        auto_1900s = sum(
            1 for r in results
            if r.year_was_auto_completed and r.year is not None and 1900 <= r.year <= 1999
        )
        auto_2000s = sum(
            1 for r in results
            if r.year_was_auto_completed and r.year is not None and 2000 <= r.year <= 2099
        )

        total_auto = auto_1900s + auto_2000s
        if total_auto == 0:
            return results

        # Only correct when 1900s are the strict majority
        if auto_1900s <= auto_2000s:
            return results

        # Flip auto-completed 2000s outliers to 1900s
        corrected = []
        for r in results:
            if (
                r.year_was_auto_completed
                and r.year is not None
                and 2000 <= r.year <= 2099
            ):
                new_year = r.year - 100  # e.g. 2026 → 1926
                new_result = self.date_engine._validate_date(new_year, r.month, r.day)
                new_result.year_was_auto_completed = True
                new_result = self.date_engine.validate_business_rules(
                    new_result, DateFieldType.BIRTH_DATE
                )
                corrected.append(new_result)
            else:
                corrected.append(r)

        return corrected