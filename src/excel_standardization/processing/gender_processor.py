"""Gender field processor for gender standardization.

This module provides the GenderFieldProcessor class that processes the gender
field in Excel worksheets. It handles the specific gender header format with
line breaks and normalizes gender values to 1 (male) or 2 (female).
"""

from typing import Optional, List, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from .field_processor import FieldProcessor
from ..io_layer.excel_reader import ExcelReader
from ..io_layer.excel_writer import ExcelWriter
from ..engines.gender_engine import GenderEngine
from ..data_types import ColumnHeaderInfo


class GenderFieldProcessor(FieldProcessor):
    """Process gender field with line break standardization.

    This processor handles gender standardization using GenderEngine. It searches
    for the specific gender header format "מין\\n1=זכר\\n2+נקבה" with line break
    standardization to handle various line break variants.

    Attributes:
        reader: ExcelReader instance for reading data
        writer: ExcelWriter instance for writing data
        gender_engine: GenderEngine instance for gender standardization
        gender_info: Header info for gender column
        corrected_col: Column number for corrected gender values
    """

    def __init__(self, reader: ExcelReader, writer: ExcelWriter, gender_engine: GenderEngine):
        """Initialize the gender field processor.

        Args:
            reader: ExcelReader instance for reading data
            writer: ExcelWriter instance for writing data
            gender_engine: GenderEngine instance for gender standardization
        """
        super().__init__(reader, writer)
        self.gender_engine = gender_engine
        # VBA parity: ProcessGender iterates over all headers found (FindAllHeaders).
        self.gender_headers: List[ColumnHeaderInfo] = []
        self.corrected_cols: List[int] = []

    def find_headers(self, worksheet: Worksheet) -> bool:
        """Find gender headers with VBA-style multiplicity.

        Resets per-sheet state before scanning so the processor can be reused.

        VBA parity:
        - The VBA system uses FindAllHeaders(ws, "מין") (xlWhole) and processes each hit.
        - We therefore treat a gender header as a cell whose trimmed value is
          exactly "מין", excluding any already-corrected headers that contain "מתוקן".

        For robustness on sheets that encode the gender header as a specific
        multi-line label, we also accept the exact value "מין\\n1=זכר\\n2+נקבה"
        after standardizing line breaks.

        Args:
            worksheet: The worksheet to search for headers

        Returns:
            True if gender header was found, False otherwise

        Requirements: 7.1, 7.2
        """
        # Reset per-sheet state
        self.gender_headers = []
        self.corrected_cols = []

        max_row = min(worksheet.max_row or 0, 30)
        max_col = worksheet.max_column or 0
        multiline_exact = "מין\n1=זכר\n2+נקבה"

        found: List[Tuple[int, int, str]] = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                val = worksheet.cell(row=r, column=c).value
                if val is None:
                    continue
                cell_text = str(val)
                norm = cell_text.replace("\r\n", "\n").replace("\r", "\n").strip()
                if "מתוקן" in norm:
                    continue
                if norm == "מין" or norm == multiline_exact:
                    found.append((r, c, cell_text))

        found.sort(key=lambda x: (x[0], x[1]))
        for r, c, cell_text in found:
            last_row = self.reader.get_last_row(worksheet, c)
            self.gender_headers.append(
                ColumnHeaderInfo(col=c, header_row=r, last_row=last_row, header_text=cell_text)
            )

        return len(self.gender_headers) > 0

    def prepare_output_columns(self, worksheet: Worksheet) -> None:
        """Insert 'מין - מתוקן' column.

        Inserts a corrected column immediately after the original gender column
        with the header text "מין - מתוקן".

        Args:
            worksheet: The worksheet to modify

        Requirements: 7.3
        """
        if not self.gender_headers:
            return

        corrected_header = "מין - מתוקן"
        for idx, info in enumerate(self.gender_headers):
            # Ensure base header text remains (Excel insertions preserve values; openpyxl
            # can drop header text in some layouts). VBA expects the original "מין"
            # header to remain adjacent to "מין - מתוקן".
            base_cell = worksheet.cell(row=info.header_row, column=info.col)
            if (base_cell.value is None) or (str(base_cell.value).strip() == ""):
                base_cell.value = "מין"

            corrected_col = self.writer.prepare_output_column(
                worksheet, info.col, corrected_header, info.header_row
            )
            self.corrected_cols.append(corrected_col)

            # Column insertions shift all columns to the right; keep subsequent
            # header positions aligned to the live worksheet state.
            inserted_at = info.col + 1
            for j in range(idx + 1, len(self.gender_headers)):
                if self.gender_headers[j].col >= inserted_at:
                    self.gender_headers[j].col += 1

    def process_data(self, worksheet: Worksheet) -> None:
        """Normalize gender values using GenderEngine.

        Reads original gender values, applies standardization using GenderEngine
        (converting to 1 for male or 2 for female), writes corrected values,
        and applies pink highlighting where values differ.

        Note: When we insert a column AFTER the original column, the original
        data stays in its original position. We don't need to adjust the column
        number for reading.

        Args:
            worksheet: The worksheet to process

        Requirements: 7.8
        """
        if not self.gender_headers or not self.corrected_cols:
            return

        for info, corrected_col in zip(self.gender_headers, self.corrected_cols):
            start_row = info.header_row + 1
            end_row = info.last_row

            original_values = self.reader.read_column_array(worksheet, info.col, start_row, end_row)

            corrected_values = []
            for value in original_values:
                normalized = self.gender_engine.normalize_gender(value)
                corrected_values.append(normalized)

            self.writer.write_column_array(worksheet, corrected_col, start_row, corrected_values)
            self.writer.highlight_changed_cells(worksheet, corrected_col, start_row, original_values, corrected_values)
