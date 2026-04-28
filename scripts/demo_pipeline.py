"""Demo script showing the complete JSON-based Excel standardization pipeline.

This script demonstrates the end-to-end flow:
1. Excel (read-only) → Raw JSON extraction
2. Raw JSON → standardization → Normalized JSON
3. Output: Two JSON files (raw and normalized)

CRITICAL: The original Excel file is NEVER modified.
"""

from openpyxl import Workbook
from src.excel_standardization.io_layer import ExcelReader, ExcelToJsonExtractor
from src.excel_standardization.processing import standardizationPipeline
from src.excel_standardization.json_exporter import JsonExporter, generate_output_filenames
from src.excel_standardization.engines import (
    NameEngine, GenderEngine, DateEngine, IdentifierEngine, TextProcessor
)


def create_sample_excel(filename: str):
    """Create a sample Excel file with data to normalize."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    ws['A1'] = 'שם פרטי'  # First Name
    ws['B1'] = 'שם משפחה'  # Last Name
    ws['C1'] = 'מין'  # Gender
    ws['D1'] = 'ת.ז'  # ID Number
    
    # Sample data with issues to normalize
    ws['A2'] = 'יוסי 123'  # Name with numbers
    ws['B2'] = 'כהן  '  # Name with extra spaces
    ws['C2'] = 'ז'  # Hebrew gender
    ws['D2'] = '123456789'
    
    ws['A3'] = '  שרה'  # Name with leading spaces
    ws['B3'] = 'לוי'
    ws['C3'] = 'נ'  # Hebrew gender
    ws['D3'] = '987654321'
    
    wb.save(filename)
    print(f"✓ Created sample Excel file: {filename}")


def main():
    """Run the complete pipeline demo."""
    print("=" * 60)
    print("JSON-Based Excel standardization Pipeline Demo")
    print("=" * 60)
    print()
    print("CRITICAL: Original Excel file is NEVER modified!")
    print("Output: Raw JSON + Normalized JSON files")
    print()
    
    # Step 1: Create sample Excel file
    input_file = "demo_input.xlsx"
    create_sample_excel(input_file)
    print()
    
    # Generate output filenames
    raw_json_path, normalized_json_path = generate_output_filenames(input_file)
    print(f"Output files will be:")
    print(f"  - Raw JSON: {raw_json_path}")
    print(f"  - Normalized JSON: {normalized_json_path}")
    print()
    
    # Step 2: Extract Excel to JSON (READ-ONLY)
    print("Step 1: Extracting Excel to JSON (read-only)...")
    reader = ExcelReader()
    extractor = ExcelToJsonExtractor(excel_reader=reader)
    workbook_dataset = extractor.extract_workbook_to_json(input_file)
    
    print(f"  ✓ Extracted {len(workbook_dataset.sheets)} sheet(s)")
    for sheet in workbook_dataset.sheets:
        print(f"    - {sheet.sheet_name}: {len(sheet.rows)} rows, {len(sheet.field_names)} fields")
    print()
    
    # Step 3: Show raw JSON data
    print("Step 2: Raw JSON data (before standardization):")
    sheet = workbook_dataset.sheets[0]
    for i, row in enumerate(sheet.rows[:2], 1):
        print(f"  Row {i}:")
        for field, value in row.items():
            print(f"    {field}: {repr(value)}")
    print()
    
    # Step 4: Export raw JSON
    print("Step 3: Exporting raw JSON...")
    exporter = JsonExporter()
    exporter.export_workbook_to_json(workbook_dataset, raw_json_path)
    print(f"  ✓ Exported raw JSON to: {raw_json_path}")
    print()
    
    # Step 5: Normalize the data
    print("Step 4: standardizing data...")
    pipeline = standardizationPipeline(
        name_engine=NameEngine(TextProcessor()),
        gender_engine=GenderEngine(),
        date_engine=DateEngine(),
        identifier_engine=IdentifierEngine()
    )
    
    normalized_sheets = []
    for sheet in workbook_dataset.sheets:
        normalized_sheet = pipeline.normalize_dataset(sheet)
        normalized_sheets.append(normalized_sheet)
    
    print(f"  ✓ Normalized {len(normalized_sheets)} sheet(s)")
    print()
    
    # Step 6: Show normalized JSON data
    print("Step 5: Normalized JSON data (after standardization):")
    normalized_sheet = normalized_sheets[0]
    for i, row in enumerate(normalized_sheet.rows[:2], 1):
        print(f"  Row {i}:")
        for field, value in row.items():
            if field.endswith('_corrected'):
                print(f"    {field}: {repr(value)} ← CORRECTED")
            else:
                print(f"    {field}: {repr(value)}")
    print()
    
    # Step 7: Export normalized JSON
    print("Step 6: Exporting normalized JSON...")
    workbook_dataset.sheets = normalized_sheets
    exporter.export_workbook_to_json(workbook_dataset, normalized_json_path)
    print(f"  ✓ Exported normalized JSON to: {normalized_json_path}")
    print()
    
    # Summary
    print("=" * 60)
    print("Pipeline Complete!")
    print("=" * 60)
    print()
    print("Summary:")
    print(f"  Input (READ-ONLY):  {input_file}")
    print(f"  Output Raw JSON:    {raw_json_path}")
    print(f"  Output Normalized:  {normalized_json_path}")
    print()
    print("✓ Original Excel file was NOT modified")
    print("✓ Two JSON files created with extracted and normalized data")
    print()
    print("standardizations applied:")
    print("  ✓ Names: Removed numbers, trimmed spaces")
    print("  ✓ Gender: Converted Hebrew codes to numeric (ז→2, נ→1)")
    print("  ✓ IDs: Validated format")
    print()
    print("You can now:")
    print("  1. Review the raw JSON to see extracted data")
    print("  2. Review the normalized JSON to see corrections")
    print("  3. Use the JSON files for further processing")


if __name__ == "__main__":
    main()
