from __future__ import annotations

from pathlib import Path
from typing import Any, List, Tuple

from openpyxl import load_workbook


def detect_header_row(ws) -> Tuple[int, List[Tuple[int, int]]]:
    max_row = min(20, ws.max_row or 0)
    max_col = ws.max_column or 0
    details: List[Tuple[int, int]] = []
    for r in range(1, max_row + 1):
        count = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if "- מתוקן" in str(v):
                count += 1
        details.append((r, count))
        if count >= 3:
            return r, details
    return 0, details


def _ns(v: Any) -> str:
    return str(v or "").strip()


def find_exact(ws, text: str, max_rows: int = 40) -> List[Tuple[int, int, str]]:
    hits: List[Tuple[int, int, str]] = []
    max_row = min(max_rows, ws.max_row or 0)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip() == text:
                hits.append((r, c, str(v)))
    return hits


def find_contains(ws, needle: str, max_rows: int = 40) -> List[Tuple[int, int, str]]:
    hits: List[Tuple[int, int, str]] = []
    max_row = min(max_rows, ws.max_row or 0)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if needle in s:
                hits.append((r, c, s))
    return hits


def main() -> None:
    base = Path(".").resolve()
    paths = [
        base / "A.xlsx",
        base / "python_vs_vba_python.xlsx",
        base / "vba_auto.xlsx",
    ]

    for p in paths:
        if not p.exists():
            raise SystemExit(f"Missing: {p}")

    for p in paths:
        wb = load_workbook(p, data_only=False)
        print("=" * 80)
        print(p.name)
        print("sheets:", wb.sheetnames)
        if "דיירים יחידים" not in wb.sheetnames:
            print("missing sheet: דיירים יחידים")
            continue
        ws = wb["דיירים יחידים"]
        hdr, details = detect_header_row(ws)
        print("detect_header_row:", hdr)
        print("row->count:", details)
        print("exact 'מין' hits:", [(r, c) for (r, c, _t) in find_exact(ws, "מין")])
        print("exact 'מין - מתוקן' hits:", [(r, c) for (r, c, _t) in find_exact(ws, "מין - מתוקן")])
        print("exact multiline hits:", [(r, c) for (r, c, _t) in find_exact(ws, "מין\\n1=זכר\\n2+נקבה")])
        print("contains 'מין' hits:", [(r, c, t.replace('\\r', '\\\\r').replace('\\n', '\\\\n')) for (r, c, t) in find_contains(ws, "מין")][:10])
        print("contains 'תאריך לידה' hits:", [(r, c) for (r, c, _t) in find_contains(ws, "תאריך לידה")])
        print("contains 'תאריך כניסה' hits:", [(r, c) for (r, c, _t) in find_contains(ws, "תאריך כניסה")])
        # Dump a small window around the VBA gender header location for this workbook.
        row = 14
        cols = list(range(1, 26))
        window = [(c, _ns(ws.cell(row=row, column=c).value)) for c in cols]
        print("row14 col1-25 values:", window)
        # Merged ranges that intersect header row 14 around identifier/gender area
        try:
            relevant = []
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= 14 <= rng.max_row and not (rng.max_col < 8 or rng.min_col > 16):
                    relevant.append(str(rng))
            print("merged_ranges_row14_col8-16:", relevant)
        except Exception as e:
            print("merged_ranges_row14_col8-16: error", e)


if __name__ == "__main__":
    main()

