from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm_str(v: Any) -> str:
    return str(v or "").strip()


def _detect_header_row(ws: Worksheet) -> int:
    """VBA DetectHeaderRow: rows 1..20 with >=3 cells like '*- מתוקן*'."""
    max_row = min(20, ws.max_row or 0)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        match_count = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if "- מתוקן" in str(v):
                match_count += 1
                if match_count >= 3:
                    return r
    return 0


def _is_numeric_helper_row(ws: Worksheet, header_row: int) -> bool:
    """Mirror VBA RemoveNumericHelperRow predicate (row below header)."""
    if header_row <= 0:
        return False
    check_row = header_row + 1
    if check_row > (ws.max_row or 0):
        return False

    # lastCol on header row
    last_col = 0
    for c in range(ws.max_column or 0, 0, -1):
        if _norm_str(ws.cell(row=header_row, column=c).value) != "":
            last_col = c
            break
    if last_col == 0:
        return False

    non_helper_found = False
    for c in range(1, last_col + 1):
        val = ws.cell(row=check_row, column=c).value
        if _norm_str(val) == "":
            continue
        try:
            num = float(str(val).strip())
            if int(num) >= 100:
                non_helper_found = True
                break
        except Exception:
            non_helper_found = True
            break

    return not non_helper_found


@dataclass(frozen=True)
class HeaderHit:
    row: int
    col: int
    text: str


def _find_all_headers_exact(ws: Worksheet, expected: str, max_rows: int = 40) -> List[HeaderHit]:
    hits: List[HeaderHit] = []
    max_row = min(ws.max_row or 0, max_rows)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            txt = str(v)
            if txt.strip() == expected:
                hits.append(HeaderHit(r, c, txt))
    hits.sort(key=lambda h: (h.row, h.col))
    return hits


def _find_first_header_contains(ws: Worksheet, needle: str, max_rows: int = 40) -> Optional[HeaderHit]:
    max_row = min(ws.max_row or 0, max_rows)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            txt = str(v)
            if needle in txt:
                return HeaderHit(r, c, txt)
    return None


def _style_sig(ws: Worksheet, r: int, c: int) -> Tuple[Optional[str], Optional[bool], str]:
    cell = ws.cell(row=r, column=c)

    # Fill color (rgb) if present
    rgb = None
    try:
        fill = cell.fill
        if fill is not None and getattr(fill, "patternType", None) is not None:
            color = fill.start_color
            cand = getattr(color, "rgb", None)
            rgb = cand if isinstance(cand, str) else None
    except Exception:
        rgb = None

    bold = None
    try:
        bold = bool(cell.font.bold) if cell.font and cell.font.bold is not None else False
    except Exception:
        bold = None

    numfmt = ""
    try:
        numfmt = cell.number_format or ""
    except Exception:
        numfmt = ""

    return rgb, bold, numfmt


def _col_letter(n: int) -> str:
    # local helper to avoid importing utils
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def compare_structural(py_path: Path, vba_path: Path, *, max_style_rows: int = 250, max_style_cols: int = 80) -> int:
    wb_py = load_workbook(py_path, data_only=False)
    wb_vba = load_workbook(vba_path, data_only=False)

    mismatches: List[str] = []

    def emit(line: str) -> None:
        mismatches.append(line)

    all_sheets = sorted(set(wb_py.sheetnames) | set(wb_vba.sheetnames))

    for sheet in all_sheets:
        if sheet not in wb_py.sheetnames:
            emit(f"SHEET | {sheet} | missing_in_python")
            continue
        if sheet not in wb_vba.sheetnames:
            emit(f"SHEET | {sheet} | missing_in_vba")
            continue

        ws_py = wb_py[sheet]
        ws_vba = wb_vba[sheet]

        # 1) Header row detection + helper row predicate parity
        hdr_py = _detect_header_row(ws_py)
        hdr_vba = _detect_header_row(ws_vba)
        if hdr_py != hdr_vba:
            emit(f"HEADER_ROW | {sheet} | python={hdr_py} vba={hdr_vba}")

        if hdr_py and hdr_vba:
            helper_py = _is_numeric_helper_row(ws_py, hdr_py)
            helper_vba = _is_numeric_helper_row(ws_vba, hdr_vba)
            if helper_py != helper_vba:
                emit(f"HELPER_ROW | {sheet} | python_is_helper={helper_py} vba_is_helper={helper_vba} | row_py={hdr_py+1} row_vba={hdr_vba+1}")

        # 2) Corrected column adjacency / insertion anchor checks (header-level)
        # Names (corrected immediately to the right of original)
        for base in ["שם פרטי", "שם משפחה", "שם האב"]:
            base_py = _find_first_header_contains(ws_py, base)
            base_vba = _find_first_header_contains(ws_vba, base)
            corr_text = f"{base} - מתוקן"
            corr_py = _find_first_header_contains(ws_py, corr_text)
            corr_vba = _find_first_header_contains(ws_vba, corr_text)
            if base_py and corr_py:
                if not (corr_py.row == base_py.row and corr_py.col == base_py.col + 1):
                    emit(
                        f"COL_ADJACENCY | {sheet} | field={base} | python: base@{_col_letter(base_py.col)}{base_py.row} corrected@{_col_letter(corr_py.col)}{corr_py.row}"
                    )
            if base_vba and corr_vba:
                if not (corr_vba.row == base_vba.row and corr_vba.col == base_vba.col + 1):
                    emit(
                        f"COL_ADJACENCY | {sheet} | field={base} | vba: base@{_col_letter(base_vba.col)}{base_vba.row} corrected@{_col_letter(corr_vba.col)}{corr_vba.row}"
                    )

        # Gender: repeated header handling (count) + corrected adjacency (per hit)
        gender_hits_py = _find_all_headers_exact(ws_py, "מין")
        gender_hits_vba = _find_all_headers_exact(ws_vba, "מין")
        if len(gender_hits_py) != len(gender_hits_vba):
            emit(f"GENDER_HEADERS_COUNT | {sheet} | python={len(gender_hits_py)} vba={len(gender_hits_vba)}")

        corr_gender_py = _find_all_headers_exact(ws_py, "מין - מתוקן")
        corr_gender_vba = _find_all_headers_exact(ws_vba, "מין - מתוקן")
        if len(corr_gender_py) != len(corr_gender_vba):
            emit(f"GENDER_CORRECTED_COUNT | {sheet} | python={len(corr_gender_py)} vba={len(corr_gender_vba)}")

        # Adjacency per header occurrence (best-effort by ordering)
        for idx, hit in enumerate(gender_hits_py):
            if idx < len(corr_gender_py):
                corr = corr_gender_py[idx]
                if not (corr.row == hit.row and corr.col == hit.col + 1):
                    emit(
                        f"GENDER_ADJACENCY | {sheet} | python idx={idx} | base@{_col_letter(hit.col)}{hit.row} corrected@{_col_letter(corr.col)}{corr.row}"
                    )
        for idx, hit in enumerate(gender_hits_vba):
            if idx < len(corr_gender_vba):
                corr = corr_gender_vba[idx]
                if not (corr.row == hit.row and corr.col == hit.col + 1):
                    emit(
                        f"GENDER_ADJACENCY | {sheet} | vba idx={idx} | base@{_col_letter(hit.col)}{hit.row} corrected@{_col_letter(corr.col)}{corr.row}"
                    )

        # Identifier block anchor: corrected headers should start immediately after passport col
        pass_py = _find_first_header_contains(ws_py, "דרכון")
        pass_vba = _find_first_header_contains(ws_vba, "דרכון")
        id_block_py = _find_first_header_contains(ws_py, "ת.ז. - מתוקן")
        id_block_vba = _find_first_header_contains(ws_vba, "ת.ז. - מתוקן")
        if pass_py and id_block_py:
            if not (id_block_py.row == pass_py.row and id_block_py.col == pass_py.col + 1):
                emit(
                    f"IDENT_ANCHOR | {sheet} | python passport@{_col_letter(pass_py.col)}{pass_py.row} first_corrected@{_col_letter(id_block_py.col)}{id_block_py.row}"
                )
        if pass_vba and id_block_vba:
            if not (id_block_vba.row == pass_vba.row and id_block_vba.col == pass_vba.col + 1):
                emit(
                    f"IDENT_ANCHOR | {sheet} | vba passport@{_col_letter(pass_vba.col)}{pass_vba.row} first_corrected@{_col_letter(id_block_vba.col)}{id_block_vba.row}"
                )

        # 3) Formatting/layout comparisons (styles only, same coordinates)
        # Limit scan to control noise; this targets structural formatting differences.
        max_r = min(max_style_rows, max(ws_py.max_row or 0, ws_vba.max_row or 0))
        max_c = min(max_style_cols, max(ws_py.max_column or 0, ws_vba.max_column or 0))

        style_mismatch_count = 0
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                s_py = _style_sig(ws_py, r, c)
                s_vba = _style_sig(ws_vba, r, c)
                if s_py != s_vba:
                    # Only report differences that are likely to be meaningful:
                    # fill color, bold, number format
                    emit(
                        f"STYLE | {sheet} | { _col_letter(c)}{r} | python(fill,bold,fmt)={s_py} vba(fill,bold,fmt)={s_vba}"
                    )
                    style_mismatch_count += 1
                    if style_mismatch_count >= 200:
                        emit(f"STYLE | {sheet} | capped_at=200")
                        r = max_r  # break outer loops
                        break
            else:
                continue
            break

    print(f"# Structural compare: python={py_path.name} vs vba={vba_path.name}")
    for line in mismatches:
        print(line)
    print(f"# TOTAL structural mismatches: {len(mismatches)}")
    return len(mismatches)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python compare_workbooks_structural.py <python_output.xlsx> <vba_output.xlsx>")
        raise SystemExit(1)

    py = Path(sys.argv[1]).resolve()
    vba = Path(sys.argv[2]).resolve()
    if not py.exists():
        raise SystemExit(f"Python workbook not found: {py}")
    if not vba.exists():
        raise SystemExit(f"VBA workbook not found: {vba}")

    raise SystemExit(0 if compare_structural(py, vba) == 0 else 2)

