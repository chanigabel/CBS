from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


def find_date_groups(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    header_texts = ["תאריך לידה", "תאריך כניסה למוסד"]
    groups: List[Dict[str, Any]] = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            text = v.strip()
            if not text:
                continue
            if any(ht in text for ht in header_texts):
                # sub-header row is r+1
                sub_r = r + 1
                year_col = month_col = day_col = None
                for offset in range(0, 10):
                    col = c + offset
                    if col > max_col:
                        break
                    sv = ws.cell(row=sub_r, column=col).value
                    if not isinstance(sv, str):
                        continue
                    st = sv.strip()
                    if st == "שנה" and year_col is None:
                        year_col = col
                    elif st == "חודש" and month_col is None:
                        month_col = col
                    elif st == "יום" and day_col is None:
                        day_col = col
                if year_col and month_col and day_col:
                    # corrected headers: look to the right of day_col on same sub-header row
                    cy = ws.cell(row=sub_r, column=day_col + 1).value
                    cm = ws.cell(row=sub_r, column=day_col + 2).value
                    cd = ws.cell(row=sub_r, column=day_col + 3).value
                    cs = ws.cell(row=sub_r, column=day_col + 4).value
                    groups.append(
                        {
                            "header_text": text,
                            "main_row": r,
                            "main_col": c,
                            "year_col": year_col,
                            "month_col": month_col,
                            "day_col": day_col,
                            "sub_header_row": sub_r,
                            "corrected_headers": (
                                cy,
                                cm,
                                cd,
                                cs,
                            ),
                            "corrected_cols": (
                                day_col + 1,
                                day_col + 2,
                                day_col + 3,
                                day_col + 4,
                            ),
                        }
                    )
    return groups


def main() -> None:
    base_dir = Path(".").resolve()
    vba_path = base_dir / "vba_auto.xlsx"
    py_path = base_dir / "python_auto.xlsx"

    for label, p in [("VBA", vba_path), ("PY", py_path)]:
        print(f"=== {label} date groups in {p.name} ===")
        groups = find_date_groups(p)
        if not groups:
            print("  (no groups found)")
            continue
        for g in groups:
            print(
                f"- {g['header_text']} @ main(row={g['main_row']}, col={g['main_col']}); "
                f"Y/M/D cols=({g['year_col']},{g['month_col']},{g['day_col']}); "
                f"sub_header_row={g['sub_header_row']}; "
                f"corrected_headers={g['corrected_headers']}; "
                f"corrected_cols={g['corrected_cols']}"
            )
        print()


if __name__ == "__main__":
    main()

