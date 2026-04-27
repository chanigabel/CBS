from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


def load(path: Path):
    return load_workbook(path, data_only=True)


def normalize_value(v: Any) -> Any:
    # Treat empty string and None as equivalent blank
    if v == "":
        return None
    return v


def compare_workbooks(py_path: Path, vba_path: Path) -> None:
    wb_py = load(py_path)
    wb_vba = load(vba_path)

    print(f"# Comparing Python={py_path.name} vs VBA={vba_path.name}")
    print(f"Sheets (Python): {wb_py.sheetnames}")
    print(f"Sheets (VBA):    {wb_vba.sheetnames}")
    print()

    all_sheets = sorted(set(wb_py.sheetnames) | set(wb_vba.sheetnames))

    total_mismatches: int = 0
    per_sheet_counts: Dict[str, int] = {}

    for sheet in all_sheets:
        if sheet not in wb_vba.sheetnames:
            print(f"SHEET MISMATCH | sheet={sheet} | type=missing_in_vba")
            continue
        if sheet not in wb_py.sheetnames:
            print(f"SHEET MISMATCH | sheet={sheet} | type=missing_in_python")
            continue

        ws_py = wb_py[sheet]
        ws_vba = wb_vba[sheet]

        max_row = max(ws_py.max_row, ws_vba.max_row)
        max_col = max(ws_py.max_column, ws_vba.max_column)

        sheet_mismatches = 0

        for r in range(1, max_row + 1):
            row_py_empty = all(
                normalize_value(ws_py.cell(row=r, column=c).value) is None
                for c in range(1, ws_py.max_column + 1)
            )
            row_vba_empty = all(
                normalize_value(ws_vba.cell(row=r, column=c).value) is None
                for c in range(1, ws_vba.max_column + 1)
            )

            if row_py_empty and not row_vba_empty:
                print(
                    f"ROW MISMATCH | sheet={sheet} | row={r} | type=missing_row_in_python"
                )
                sheet_mismatches += 1
                total_mismatches += 1
                continue
            if row_vba_empty and not row_py_empty:
                print(
                    f"ROW MISMATCH | sheet={sheet} | row={r} | type=extra_row_in_python"
                )
                sheet_mismatches += 1
                total_mismatches += 1
                continue

            for c in range(1, max_col + 1):
                v_py = normalize_value(ws_py.cell(row=r, column=c).value)
                v_vba = normalize_value(ws_vba.cell(row=r, column=c).value)
                if v_py != v_vba:
                    diff_type = "wrong_normalized_value"
                    if v_py is None and v_vba is not None:
                        diff_type = "empty_in_python_filled_in_vba"
                    elif v_py is not None and v_vba is None:
                        diff_type = "filled_in_python_empty_in_vba"

                    print(
                        "CELL MISMATCH | "
                        f"sheet={sheet} | row={r} | col={c} | "
                        f"type={diff_type} | vba={v_vba!r} | python={v_py!r}"
                    )
                    sheet_mismatches += 1
                    total_mismatches += 1

        per_sheet_counts[sheet] = sheet_mismatches

    print()
    print("# MISMATCH SUMMARY")
    for sheet, count in per_sheet_counts.items():
        print(f"{sheet}: {count} mismatches")
    print(f"TOTAL mismatches: {total_mismatches}")


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python compare_workbooks.py <python_output.xlsx> <vba_output.xlsx>")
        sys.exit(1)

    py_path = Path(sys.argv[1]).resolve()
    vba_path = Path(sys.argv[2]).resolve()

    if not py_path.exists():
        print(f"Python workbook not found: {py_path}")
        sys.exit(1)
    if not vba_path.exists():
        print(f"VBA workbook not found: {vba_path}")
        sys.exit(1)

    compare_workbooks(py_path, vba_path)

