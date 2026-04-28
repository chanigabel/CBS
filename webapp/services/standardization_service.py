"""standardizationService: runs the standardization pipeline on a session's workbook."""

import logging
from typing import List, Optional

from fastapi import HTTPException
from openpyxl import load_workbook as _lw

from src.excel_standardization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.processing.standardization_pipeline import standardizationPipeline
from src.excel_standardization.engines.name_engine import NameEngine
from src.excel_standardization.engines.gender_engine import GenderEngine
from src.excel_standardization.engines.date_engine import DateEngine
from src.excel_standardization.engines.identifier_engine import IdentifierEngine
from src.excel_standardization.engines.text_processor import TextProcessor
from src.excel_standardization.data_types import SheetDataset

from webapp.models.responses import StandardizeResponse, PerSheetStat
from webapp.services.session_service import SessionService
from webapp.services.mosad_id_scanner import scan_mosad_id

logger = logging.getLogger(__name__)


class standardizationService:
    """Runs the standardization pipeline on a session's working copy."""

    def __init__(self, session_service: SessionService) -> None:
        self.session_service = session_service

    def standardize(self, session_id: str, sheet_name: Optional[str] = None) -> StandardizeResponse:
        """Run standardization on the session's working copy.

        If *sheet_name* is given, only that sheet is (re-)standardized and the
        rest of the in-memory dataset is left untouched.  This is the fast path
        used by the UI.  When *sheet_name* is None all loaded sheets are
        processed (kept for CLI / batch compatibility).
        """
        record = self.session_service.get(session_id)

        pipeline = self._build_pipeline()
        extractor = ExcelToJsonExtractor(
            excel_reader=ExcelReader(),
            skip_empty_rows=False,
            handle_formulas=True,
            preserve_types=True,
        )

        if record.workbook_dataset is None:
            # Auto-load all sheets from disk when no prior sheet access has occurred.
            try:
                wbd = extractor.extract_workbook_to_json(record.working_copy_path)
                self.session_service.update(session_id, workbook_dataset=wbd)
                record = self.session_service.get(session_id)
            except Exception as exc:
                logger.error(f"Failed to load workbook for standardization: {exc}", exc_info=True)
                raise HTTPException(
                    status_code=500,
                    detail="No workbook data available. Please load a sheet first.",
                )

        # Determine which sheets to extract fresh from disk and normalize.
        if sheet_name is not None:
            # Fast path: single sheet only.
            try:
                wb = _lw(record.working_copy_path, data_only=True)
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    raise HTTPException(
                        status_code=404,
                        detail=f"Sheet '{sheet_name}' not found.",
                    )
                ws = wb[sheet_name]
                fresh = extractor.extract_sheet_to_json(ws)
                # Preserve MosadID from existing metadata or re-scan.
                existing = record.workbook_dataset.get_sheet_by_name(sheet_name)
                mosad_id = (
                    existing.get_metadata("MosadID")
                    if existing is not None
                    else None
                ) or scan_mosad_id(ws)
                if mosad_id is not None:
                    fresh.set_metadata("MosadID", mosad_id)
                sheets_to_normalize = [fresh]
                wb.close()
            except HTTPException:
                raise
            except Exception as exc:
                logger.error(f"Failed to extract sheet '{sheet_name}': {exc}", exc_info=True)
                raise HTTPException(
                    status_code=500,
                    detail="Failed to read the working copy for standardization.",
                )
        else:
            # Full path: re-extract all sheets and re-scan for MosadID.
            try:
                wb = _lw(record.working_copy_path, data_only=True)
                sheets_to_normalize = []
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    fresh = extractor.extract_sheet_to_json(ws)
                    existing = record.workbook_dataset.get_sheet_by_name(sname)
                    mosad_id = (
                        existing.get_metadata("MosadID")
                        if existing is not None
                        else None
                    ) or scan_mosad_id(ws)
                    if mosad_id is not None:
                        fresh.set_metadata("MosadID", mosad_id)
                    sheets_to_normalize.append(fresh)
                wb.close()
            except Exception as exc:
                logger.error(f"Failed to extract workbook: {exc}", exc_info=True)
                raise HTTPException(
                    status_code=500,
                    detail="Failed to read the working copy for standardization.",
                )

        # Normalize
        normalized_sheets: List[SheetDataset] = []
        per_sheet_stats: List[PerSheetStat] = []
        failed_sheets: List[str] = []

        for sheet in sheets_to_normalize:
            try:
                norm = pipeline.normalize_dataset(sheet)
                normalized_sheets.append(norm)
                stats = norm.get_metadata("standardization_statistics", {})
                per_sheet_stats.append(PerSheetStat(
                    sheet_name=sheet.sheet_name,
                    rows=stats.get("total_rows", len(norm.rows)),
                    success_rate=stats.get("success_rate", 1.0),
                ))
                logger.info(f"Sheet '{sheet.sheet_name}' standardized: "
                            f"{per_sheet_stats[-1].rows} rows")
            except Exception as exc:
                logger.error(f"Failed to normalize sheet '{sheet.sheet_name}': {exc}",
                             exc_info=True)
                failed_sheets.append(sheet.sheet_name)

        if not normalized_sheets:
            raise HTTPException(
                status_code=500,
                detail=f"standardization failed for all sheets: {', '.join(failed_sheets)}",
            )

        # Merge normalized sheets back into the session dataset.
        # For single-sheet standardization, replace only that sheet.
        norm_by_name = {s.sheet_name: s for s in normalized_sheets}
        updated_sheets = []
        for existing in record.workbook_dataset.sheets:
            if existing.sheet_name in norm_by_name:
                updated_sheets.append(norm_by_name.pop(existing.sheet_name))
            else:
                updated_sheets.append(existing)
        # Any newly normalized sheets not previously in the dataset
        updated_sheets.extend(norm_by_name.values())
        record.workbook_dataset.sheets = updated_sheets

        # F-01: Replay manual edits that were recorded before this standardization.
        # record.edits stores {(sheet_name, row_uid, field_name): value} for every
        # PATCH /cell call.  Re-applying them here ensures that manual corrections
        # survive a re-normalize without requiring the user to redo them.
        if record.edits:
            for (edit_sheet, edit_row_uid, edit_field), edit_value in record.edits.items():
                sheet_obj = record.workbook_dataset.get_sheet_by_name(edit_sheet)
                if sheet_obj is None:
                    continue
                for row in sheet_obj.rows:
                    if row.get("_row_uid") == edit_row_uid and edit_field in row:
                        row[edit_field] = edit_value
                        break
            logger.debug(
                f"Replayed {len(record.edits)} manual edit(s) after standardization "
                f"for session {session_id}"
            )

        self.session_service.update(session_id, status="standardized")

        total_rows = sum(s.rows for s in per_sheet_stats)
        logger.info(f"standardization complete for session {session_id}: "
                    f"{len(normalized_sheets)} sheets, {total_rows} total rows")

        return StandardizeResponse(
            session_id=session_id,
            status="standardized",
            sheets_processed=len(normalized_sheets),
            total_rows=total_rows,
            per_sheet_stats=per_sheet_stats,
        )

    # Backward-compatible alias
    def normalize(self, session_id: str, sheet_name: Optional[str] = None) -> StandardizeResponse:
        return self.standardize(session_id, sheet_name=sheet_name)

    def _build_pipeline(self) -> standardizationPipeline:
        """Build a fresh standardizationPipeline with all four engines."""
        tp = TextProcessor()
        return standardizationPipeline(
            name_engine=NameEngine(tp),
            gender_engine=GenderEngine(),
            date_engine=DateEngine(),
            identifier_engine=IdentifierEngine(),
            apply_name_standardization_enabled=True,
            apply_gender_standardization_enabled=True,
            apply_date_standardization_enabled=True,
            apply_identifier_standardization_enabled=True,
        )
