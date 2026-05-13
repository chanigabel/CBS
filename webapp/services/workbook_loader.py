"""Central workbook loader dispatch for supported Excel formats."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_standardization.io_layer import xls_reader
from webapp.services.mosad_id_scanner import scan_mosad_id

logger = logging.getLogger(__name__)

ALLOWED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


class WorkbookLoadError(ValueError):
    """Raised when a workbook cannot be opened through the approved loader."""


def workbook_suffix(path: str | Path) -> str:
    return Path(path).suffix.lower()


def ensure_supported_workbook(path_or_name: str | Path) -> str:
    """Return the normalized suffix or raise for unsupported workbook types."""
    suffix = workbook_suffix(path_or_name)
    if suffix not in ALLOWED_WORKBOOK_EXTENSIONS:
        raise WorkbookLoadError(
            f"File format not supported. Please upload a .xlsx, .xlsm, or .xls file. Got: '{suffix}'"
        )
    return suffix


def _extractor() -> ExcelToJsonExtractor:
    return ExcelToJsonExtractor(
        excel_reader=ExcelReader(),
        skip_empty_rows=False,
        handle_formulas=True,
        preserve_types=True,
    )


def get_workbook_sheet_names(path: str | Path) -> List[str]:
    """Read sheet names using the same extension policy as upload/export."""
    suffix = ensure_supported_workbook(path)
    try:
        if suffix == ".xls":
            names = xls_reader.get_xls_sheet_names(str(path))
        else:
            wb = load_workbook(str(path), data_only=True, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
        if not names:
            raise WorkbookLoadError("Workbook has no sheets.")
        return names
    except WorkbookLoadError:
        raise
    except ValueError as exc:
        if suffix == ".xls":
            raise WorkbookLoadError(xls_reader.XLS_ERROR_HE) from exc
        raise WorkbookLoadError("The workbook could not be opened.") from exc
    except Exception as exc:
        logger.warning("workbook_sheet_names_failed", exc_info=True)
        raise WorkbookLoadError("The workbook could not be opened.") from exc


def extract_workbook_dataset(path: str | Path) -> WorkbookDataset:
    """Extract a full WorkbookDataset using the approved loader for the suffix."""
    suffix = ensure_supported_workbook(path)
    try:
        if suffix == ".xls":
            return xls_reader.extract_xls_to_workbook_dataset(str(path))
        return _extractor().extract_workbook_to_json(str(path))
    except Exception as exc:
        logger.error("workbook_extract_failed", exc_info=True)
        raise WorkbookLoadError("Failed to read workbook data.") from exc


def extract_sheet_dataset(path: str | Path, sheet_name: str) -> SheetDataset:
    """Extract one sheet using the approved loader for the suffix."""
    suffix = ensure_supported_workbook(path)
    try:
        if suffix == ".xls":
            return xls_reader.extract_xls_sheet_to_dataset(str(path), sheet_name)
        wb = load_workbook(str(path), data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise KeyError(sheet_name)
        ws = wb[sheet_name]
        sheet = _extractor().extract_sheet_to_json(ws)
        mosad_id = scan_mosad_id(ws)
        if mosad_id is not None:
            sheet.set_metadata("MosadID", mosad_id)
        wb.close()
        return sheet
    except KeyError:
        raise
    except ValueError as exc:
        if suffix == ".xls":
            raise WorkbookLoadError(xls_reader.XLS_ERROR_HE) from exc
        raise WorkbookLoadError(f"Failed to read sheet '{sheet_name}'.") from exc
    except Exception as exc:
        logger.error("sheet_extract_failed", exc_info=True)
        raise WorkbookLoadError(f"Failed to read sheet '{sheet_name}'.") from exc


def sheet_exists(path: str | Path, sheet_name: str) -> bool:
    try:
        return sheet_name in get_workbook_sheet_names(path)
    except WorkbookLoadError:
        return False
