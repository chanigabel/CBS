"""Workbook writer for the export service."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment

from src.excel_standardization.export.excel_safe import safe_cell_value, safe_sheet_title
from webapp.services.export_rows import build_row_export_view, resolve_sug_mosad_for_sheet, visible_rows
from webapp.services.export_schema import EXPORT_MAPPING, canonical_sheet_name, headers_for_sheet

logger = logging.getLogger(__name__)


# כותב בפועל את Workbook היצוא מתוך ה־Dataset והסכמה הפעילה.
def write_export_workbook(record, output_path: Path, workbook_factory=Workbook) -> Tuple[int, Dict[str, int]]:
    """Write the active workbook dataset to ``output_path`` and return counts."""
    wb = workbook_factory()
    if wb.sheetnames:
        wb.remove(wb[wb.sheetnames[0]])

    rows_exported = 0
    rows_exported_by_sheet: Dict[str, int] = {}

    for sheet_dataset in record.workbook_dataset.sheets:
        export_name = canonical_sheet_name(sheet_dataset.sheet_name)
        ws = wb.create_sheet(title=safe_sheet_title(export_name, wb.sheetnames))
        ws.sheet_view.rightToLeft = True
        schema = headers_for_sheet(export_name)

        for col_idx, header in enumerate(schema, start=1):
            cell = ws.cell(row=1, column=col_idx, value=safe_cell_value(header))
            cell.alignment = Alignment(horizontal="right")

        data_rows, _ui_cols = visible_rows(sheet_dataset)
        active_mosad_type = record.mosad_types[0] if record.mosad_types else ""
        scoped_type = resolve_sug_mosad_for_sheet(
            record.sug_mosad_configs,
            sheet_dataset.sheet_name,
            active_mosad_type,
        )

        out_row = 2
        sheet_rows_exported = 0
        for row in data_rows:
            export_row = build_row_export_view(
                row,
                mosad_id=record.mosad_id or "",
                scoped_sug_mosad=scoped_type,
            )
            for col_idx, header in enumerate(schema, start=1):
                json_key = EXPORT_MAPPING.get(header)
                if json_key is None:
                    continue
                v = export_row.get(json_key)
                if v is not None and v != "":
                    ws.cell(row=out_row, column=col_idx, value=safe_cell_value(v))
            out_row += 1
            rows_exported += 1
            sheet_rows_exported += 1
        rows_exported_by_sheet[sheet_dataset.sheet_name] = sheet_rows_exported

    wb.save(str(output_path))
    logger.debug("Workbook written to %s", output_path)
    return rows_exported, rows_exported_by_sheet
