"""WorkbookService: serves workbook summary and sheet data from session state."""

import logging
import uuid
from fastapi import HTTPException

from src.excel_normalization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_normalization.io_layer.excel_reader import ExcelReader
from webapp.models.responses import SheetDataResponse, SheetSummary, WorkbookSummary
from webapp.services.session_service import SessionService
from webapp.services.mosad_id_scanner import scan_mosad_id
from webapp.services.derived_columns import apply_derived_columns

logger = logging.getLogger(__name__)


def _is_numeric_like(value) -> bool:
    """Return True if *value* is numeric or a string that represents a number.

    Accepts int, float, and strings whose stripped form parses as int or float.
    Rejects empty strings and anything that cannot be parsed as a number.
    """
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip())
        return True
    except (ValueError, TypeError):
        return False


class WorkbookService:
    """Provides workbook summary and sheet data from in-memory session state."""

    def __init__(self, session_service: SessionService) -> None:
        self.session_service = session_service

    def _ensure_sheet_loaded(self, record, sheet_name: str) -> None:
        """Lazily extract a single sheet from disk if not yet in the dataset."""
        from src.excel_normalization.data_types import WorkbookDataset
        from openpyxl import load_workbook as _lw

        # If the sheet is already in memory, nothing to do.
        if record.workbook_dataset is not None:
            if record.workbook_dataset.get_sheet_by_name(sheet_name) is not None:
                return
            # Dataset is loaded but sheet isn't in it — check if it exists on disk
            # before attempting extraction; if not, raise 404 immediately.
            try:
                _wb_check = _lw(record.working_copy_path, data_only=True, read_only=True)
                exists_on_disk = sheet_name in _wb_check.sheetnames
                _wb_check.close()
            except Exception:
                exists_on_disk = False
            if not exists_on_disk:
                raise HTTPException(
                    status_code=404,
                    detail=f"Sheet '{sheet_name}' not found in this workbook.",
                )

        extractor = ExcelToJsonExtractor(
            excel_reader=ExcelReader(),
            skip_empty_rows=False,
            handle_formulas=True,
            preserve_types=True,
        )

        try:
            wb = _lw(record.working_copy_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise HTTPException(
                    status_code=404,
                    detail=f"Sheet '{sheet_name}' not found in this workbook.",
                )
            ws = wb[sheet_name]
            sheet_dataset = extractor.extract_sheet_to_json(ws)
            # Scan for MosadID label/value pair outside the main table.
            mosad_id = scan_mosad_id(ws)
            if mosad_id is not None:
                sheet_dataset.set_metadata("MosadID", mosad_id)
            wb.close()
        except HTTPException:
            raise
        except Exception as exc:
            logger.error(f"Failed to extract sheet '{sheet_name}': {exc}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Failed to read sheet '{sheet_name}' from the workbook.",
            )

        if record.workbook_dataset is None:
            # First load — create the dataset with just this sheet.
            # Preserve sheet order by reading all sheet names from the file.
            try:
                _wb2 = _lw(record.working_copy_path, data_only=True, read_only=True)
                all_names = _wb2.sheetnames
                _wb2.close()
            except Exception:
                all_names = [sheet_name]

            record.workbook_dataset = WorkbookDataset(
                source_file=record.working_copy_path,
                sheets=[sheet_dataset],
                metadata={"sheet_names": list(all_names)},
            )
        else:
            # Add the newly loaded sheet to the existing dataset.
            record.workbook_dataset.sheets.append(sheet_dataset)

    def get_summary(self, session_id: str) -> WorkbookSummary:
        """Return a summary of all sheets in the workbook."""
        record = self.session_service.get(session_id)

        # If no sheet has been loaded yet, read sheet names from the file
        # without doing a full extraction.
        if record.workbook_dataset is None:
            try:
                from openpyxl import load_workbook as _lw
                _wb = _lw(record.working_copy_path, data_only=True, read_only=True)
                names = _wb.sheetnames
                _wb.close()
            except Exception as exc:
                raise HTTPException(
                    status_code=500,
                    detail="Workbook data is not available for this session.",
                )
            sheets = [
                SheetSummary(sheet_name=n, row_count=0, field_names=[])
                for n in names
            ]
            return WorkbookSummary(session_id=session_id, sheets=sheets)

        sheets = [
            SheetSummary(
                sheet_name=sheet.sheet_name,
                row_count=sheet.get_row_count(),
                field_names=sheet.get_field_names(),
            )
            for sheet in record.workbook_dataset.sheets
        ]
        return WorkbookSummary(session_id=session_id, sheets=sheets)

    def get_sheet_data(self, session_id: str, sheet_name: str) -> SheetDataResponse:
        """Return all rows for a specific sheet.

        Args:
            session_id: UUID string of the session
            sheet_name: Name of the sheet to retrieve

        Returns:
            SheetDataResponse with field_names and rows

        Raises:
            HTTPException 404: If session or sheet not found
        """
        record = self.session_service.get(session_id)

        # Lazily extract the requested sheet if not yet in memory.
        self._ensure_sheet_loaded(record, sheet_name)

        sheet = record.workbook_dataset.get_sheet_by_name(sheet_name)
        if sheet is None:
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{sheet_name}' not found in this workbook.",
            )

        # Assign stable _row_uid to every row that doesn't have one yet.
        # This must happen on the original sheet.rows list so the UID persists
        # across multiple calls.
        for row in sheet.rows:
            if "_row_uid" not in row:
                row["_row_uid"] = uuid.uuid4().hex

        # Build display_columns preserving the original Excel left-to-right column
        # order exactly.  The rule is simple:
        #
        #   For each field in field_names (already in Excel column order):
        #     1. Place the original field.
        #     2. Place its _corrected column immediately after (if present in rows).
        #     3. If this field is the LAST member of a known status group that
        #        appears in field_names, place the group's status column next.
        #
        # "Last member of the group in field_names" is determined by scanning
        # field_names left-to-right — no hardcoded reverse priority, no semantic
        # reordering.  Whatever field appears rightmost in the actual Excel sheet
        # is the anchor.

        original_fields = list(sheet.field_names)

        # Collect all non-metadata keys that actually appear in the rows.
        # Exclude ALL underscore-prefixed internal keys (e.g. _birth_year_auto_completed,
        # _entry_year_auto_completed, _normalization_failures) — these are pipeline
        # implementation details and must never appear in the UI payload.
        seen: set = set()
        all_row_keys: list = []
        for row in sheet.rows:
            for k in row.keys():
                if k not in seen and not k.startswith("_"):
                    seen.add(k)
                    all_row_keys.append(k)

        # Status groups: map each status key to the set of original field names
        # that belong to its group.
        _STATUS_GROUPS: dict = {
            "identifier_status": {"id_number", "passport"},
            "birth_date_status":  {"birth_year", "birth_month", "birth_day", "birth_date"},
            "entry_date_status":  {"entry_year", "entry_month", "entry_day", "entry_date"},
        }

        # For each status key, find which field in original_fields is the
        # rightmost (last in Excel order) member of that group.
        # That field becomes the sole anchor — the status is emitted after its
        # _corrected column.
        #
        # Special case: when the source has a plain date column (e.g. birth_date)
        # the pipeline now writes structured year/month/day corrected fields
        # (birth_year_corrected, birth_month_corrected, birth_day_corrected)
        # instead of birth_date_corrected.  If the naive anchor key is not in
        # `seen`, fall back to the last structured corrected field for that group.
        _DATE_STRUCTURED_FALLBACK: dict = {
            "birth_date_corrected": ["birth_day_corrected", "birth_month_corrected", "birth_year_corrected"],
            "entry_date_corrected": ["entry_day_corrected", "entry_month_corrected", "entry_year_corrected"],
        }

        _anchor_to_status: dict = {}  # corrected_key -> status_key
        for status_key, group_members in _STATUS_GROUPS.items():
            if status_key not in seen:
                continue
            # Walk original_fields left-to-right; keep updating the anchor so
            # the last group member wins.
            anchor_orig = None
            for f in original_fields:
                if f in group_members:
                    anchor_orig = f
            if anchor_orig is not None:
                anchor_corrected = f"{anchor_orig}_corrected"
                # If the naive corrected key doesn't exist in seen (plain date
                # source → structured output), resolve to the last structured
                # corrected field that does exist.
                if anchor_corrected not in seen and anchor_corrected in _DATE_STRUCTURED_FALLBACK:
                    for fallback in _DATE_STRUCTURED_FALLBACK[anchor_corrected]:
                        if fallback in seen:
                            anchor_corrected = fallback
                            break
                _anchor_to_status[anchor_corrected] = status_key

        # Date field groups — these are emitted as a block:
        # [all source fields] [all corrected fields] [status]
        # rather than interleaved original+corrected per field.
        _DATE_GROUPS: list = [
            {
                "source_fields": ["birth_year", "birth_month", "birth_day", "birth_date"],
                "corrected_fields": ["birth_year_corrected", "birth_month_corrected", "birth_day_corrected"],
                "status_key": "birth_date_status",
            },
            {
                "source_fields": ["entry_year", "entry_month", "entry_day", "entry_date"],
                "corrected_fields": ["entry_year_corrected", "entry_month_corrected", "entry_day_corrected"],
                "status_key": "entry_date_status",
            },
        ]

        # Build a set of all columns that belong to a date group so we can
        # skip them in the generic per-field loop and emit them as a block.
        _date_group_cols: set = set()
        for dg in _DATE_GROUPS:
            _date_group_cols.update(dg["source_fields"])
            _date_group_cols.update(dg["corrected_fields"])
            _date_group_cols.add(dg["status_key"])

        display_columns: list = []
        placed: set = set()

        # Track which date groups have been emitted (keyed by status_key)
        _date_groups_emitted: set = set()

        for orig in original_fields:
            # Check if this field belongs to a date group
            owning_group = None
            for dg in _DATE_GROUPS:
                if orig in dg["source_fields"]:
                    owning_group = dg
                    break

            if owning_group is not None:
                sk = owning_group["status_key"]
                if sk not in _date_groups_emitted:
                    _date_groups_emitted.add(sk)
                    # Emit all source fields in this group that appear in original_fields
                    # (in their original Excel column order)
                    for src in original_fields:
                        if src in owning_group["source_fields"] and src not in placed:
                            display_columns.append(src)
                            placed.add(src)
                    # Emit all corrected fields in order
                    for cf in owning_group["corrected_fields"]:
                        if cf in seen and cf not in placed:
                            display_columns.append(cf)
                            placed.add(cf)
                    # Emit status
                    if sk in seen and sk not in placed:
                        display_columns.append(sk)
                        placed.add(sk)
                continue  # already handled (or will be handled when group is first seen)

            # Non-date field: emit original then corrected then status
            if orig not in placed:
                display_columns.append(orig)
                placed.add(orig)

            corrected = f"{orig}_corrected"
            if corrected in seen and corrected not in placed:
                display_columns.append(corrected)
                placed.add(corrected)

            # If this corrected column is the anchor for a status key, emit it now.
            status_key = _anchor_to_status.get(corrected)
            if status_key and status_key in seen and status_key not in placed:
                display_columns.append(status_key)
                placed.add(status_key)

        # Append any remaining keys not yet placed (unexpected extras).
        for k in all_row_keys:
            if k not in placed:
                display_columns.append(k)
                placed.add(k)

        # Build the clean row list, stripping ALL underscore-prefixed internal keys
        # EXCEPT _row_uid which is the stable identifier for each row.
        _KEEP_INTERNAL = {"_row_uid"}
        clean_rows = []
        for row in sheet.rows:
            clean_row = {k: v for k, v in row.items()
                         if not k.startswith("_") or k in _KEEP_INTERNAL}
            clean_rows.append(clean_row)

        original_field_set = set(original_fields)

        # Drop completely empty rows — rows where every original-column cell is
        # None, empty string, or whitespace-only.  Checked against original
        # source columns only (same rule as the numeric helper-row check below)
        # so that generated _corrected / _status columns don't keep a blank row
        # visible after normalization.
        clean_rows = [
            row for row in clean_rows
            if any(
                v is not None and str(v).strip() != ""
                for k, v in row.items()
                if k in original_field_set
            )
        ]

        # If the first displayed row is a numbers-only helper row (e.g. a
        # column-index row like 1, 2, 3 … that some Excel forms include), hide
        # it from the UI.  The check is value-based and looks only at the
        # original source columns (not at _corrected or _status columns added
        # by the normalization pipeline, which would corrupt the check after
        # normalization runs).  Every non-empty original-column cell must be
        # numeric-like, and at least one must be non-empty.
        if clean_rows:
            first = clean_rows[0]
            non_empty_original = [
                v for k, v in first.items()
                if k in original_field_set
                and v is not None
                and str(v).strip() != ""
            ]
            if non_empty_original and all(_is_numeric_like(v) for v in non_empty_original):
                clean_rows = clean_rows[1:]

        # Inject serial number and MosadID derived columns (shared with export).
        # Prefer session-level mosad_id over sheet metadata (user may have set it
        # in the institution bar after upload).
        session_mosad_id = record.mosad_id or None
        meta_mosad_id = session_mosad_id or sheet.get_metadata("MosadID")
        clean_rows, display_columns = apply_derived_columns(
            rows=clean_rows,
            field_names=original_fields,
            display_columns=display_columns,
            meta_mosad_id=meta_mosad_id,
        )

        # Inject SugMosad (institution type) from session into every row so it
        # appears in the UI grid — mirroring what the export pipeline does.
        # Uses the first user-entered mosad_type value (active default).
        active_mosad_type = record.mosad_types[0] if record.mosad_types else None
        if active_mosad_type:
            for row in clean_rows:
                if not row.get("SugMosad"):
                    row["SugMosad"] = active_mosad_type
            # Insert SugMosad into display_columns immediately after MosadID,
            # or at position 1 if MosadID is not present.
            if "SugMosad" not in display_columns:
                try:
                    insert_pos = display_columns.index("MosadID") + 1
                except ValueError:
                    insert_pos = 1
                display_columns.insert(insert_pos, "SugMosad")

        return SheetDataResponse(
            sheet_name=sheet_name,
            field_names=display_columns,
            rows=clean_rows,
        )
