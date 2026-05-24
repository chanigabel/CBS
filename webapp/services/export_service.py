"""Export service facade for writing the normalized workbook to disk."""

from __future__ import annotations

import logging
import os
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook

from webapp.services.export_rows import build_export_filename as _build_export_filename
from webapp.services.export_rows import resolve_sug_mosad_for_sheet as _resolve_sug_mosad_for_sheet
from webapp.services.export_rows import visible_rows
from webapp.services.export_schema import EXPORT_MAPPING, canonical_sheet_name, headers_for_sheet
from webapp.services.export_writer import write_export_workbook
from webapp.services.processing_report_service import ProcessingReportService
from webapp.services.session_service import SessionService

logger = logging.getLogger(__name__)


# שירות יצוא שמייצר קובץ Excel להורדה מתוך ה־WorkbookDataset שב־session.
class ExportService:
    """Writes the current in-memory WorkbookDataset to an Excel file for download."""

    # מקבל session/output/report כדי לכתוב קובץ ולעדכן דוח עיבוד.
    def __init__(
        self,
        session_service: SessionService,
        output_dir: Path,
        processing_report_service: ProcessingReportService | None = None,
    ) -> None:
        self.session_service = session_service
        self.output_dir = output_dir
        self.processing_report_service = (
            processing_report_service or ProcessingReportService(session_service)
        )

    # מוודא שיש Dataset, כותב קובץ יצוא ומחזיר את נתיב ההורדה.
    def export(self, session_id: str) -> Path:
        """Export the session's workbook using the fixed export schema."""
        logger.info(
            "export_requested",
            extra={"event": "export_requested", "session_id": session_id},
        )
        record = self.session_service.get(session_id)

        # If the session was already standardized, allow manual edits (e.g.
        # user-corrected *_corrected fields) without blocking export. Keep a
        # warning in the log when the working dataset is dirty, but do not
        # prevent export after a successful standardization run.
        if record.working_dataset_dirty:
            logger.info(
                "export_dirty_dataset_allowed",
                extra={
                    "event": "export_dirty_dataset_allowed",
                    "session_id": session_id,
                    "working_dataset_dirty": True,
                },
            )

        workbook_dataset = record.workbook_dataset
        if (
            record.status != "standardized"
            or workbook_dataset is None
            or not getattr(workbook_dataset, "sheets", None)
        ):
            raise HTTPException(
                status_code=409,
                detail="Run Standardization before exporting. Export uses the latest successful Standardization result.",
            )

        output_filename = _build_export_filename(record)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(
            "export_started",
            extra={
                "event": "export_started",
                "session_id": session_id,
                "output_filename": output_filename,
                "sheet_count": len(record.workbook_dataset.sheets),
                "row_count": sum(len(sheet.rows) for sheet in record.workbook_dataset.sheets),
            },
        )

        original_stem = Path(record.original_filename).stem
        for old_file in self.output_dir.glob(f"{original_stem}_standardized_*.xlsx"):
            try:
                old_file.unlink()
                logger.debug("Removed previous export: %s", old_file.name)
            except Exception as exc:
                logger.warning("Could not remove old export file %s: %s", old_file, exc)

        output_path = self.output_dir / output_filename
        temp_path = output_path.with_name(f"{output_path.name}.tmp")
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                logger.warning("Could not remove stale temp export file %s", temp_path, exc_info=True)

        try:
            rows_exported, rows_exported_by_sheet = write_export_workbook(
                record,
                temp_path,
                workbook_factory=Workbook,
            )
            if output_path.exists():
                try:
                    output_path.unlink()
                except Exception:
                    logger.warning("Could not remove previous export file %s", output_path, exc_info=True)
            os.replace(temp_path, output_path)
            self.processing_report_service.finalize_export_details(
                session_id,
                record=record,
                rows_exported_by_sheet=rows_exported_by_sheet,
                output_filename=output_path.name,
            )
            logger.info(
                "export_successful",
                extra={
                    "event": "export_successful",
                    "session_id": session_id,
                    "output_filename": output_path.name,
                    "rows_exported": rows_exported,
                },
            )
        except Exception as exc:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    logger.warning("Could not remove failed temp export file %s", temp_path, exc_info=True)
            self.processing_report_service.add_error(session_id, "Export failed.")
            logger.error(
                "export_failed",
                exc_info=True,
                extra={
                    "event": "export_failed",
                    "session_id": session_id,
                    "error_type": type(exc).__name__,
                },
            )
            raise HTTPException(
                status_code=500,
                detail="Export failed. Please try again. Your session data is preserved.",
            )

        return output_path

    def export_sheet(self, session_id: str, sheet_name: str) -> Path:
        """Export a single sheet from the session's workbook.
        
        Args:
            session_id: Session UUID
            sheet_name: Name of sheet to export
            
        Returns:
            Path to the exported file
            
        Raises:
            HTTPException if sheet not found or export fails
        """
        logger.info(
            "single_sheet_export_requested",
            extra={
                "event": "single_sheet_export_requested",
                "session_id": session_id,
                "sheet_name": sheet_name,
            },
        )
        record = self.session_service.get(session_id)
        
        workbook_dataset = record.workbook_dataset
        if (
            record.status != "standardized"
            or workbook_dataset is None
            or not getattr(workbook_dataset, "sheets", None)
        ):
            raise HTTPException(
                status_code=409,
                detail="Run Standardization before exporting. Export uses the latest successful Standardization result.",
            )
        
        # Find the sheet
        sheet_dataset = workbook_dataset.get_sheet_by_name(sheet_name)
        if sheet_dataset is None:
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{sheet_name}' not found in this workbook.",
            )
        
        # Build filename for single sheet
        stem = Path(record.original_filename).stem
        sheet_safe_name = sheet_name.replace(" ", "_").replace("/", "_")
        output_filename = f"{stem}_{sheet_safe_name}_standardized.xlsx"
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(
            "single_sheet_export_started",
            extra={
                "event": "single_sheet_export_started",
                "session_id": session_id,
                "sheet_name": sheet_name,
                "output_filename": output_filename,
            },
        )
        
        output_path = self.output_dir / output_filename
        temp_path = output_path.with_name(f"{output_path.name}.tmp")
        
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                logger.warning("Could not remove stale temp export file %s", temp_path, exc_info=True)
        
        try:
            # Write workbook with only the selected sheet
            from openpyxl import Workbook
            wb = Workbook()
            if wb.sheetnames:
                wb.remove(wb[wb.sheetnames[0]])
            
            from webapp.services.export_rows import build_row_export_view, resolve_sug_mosad_for_sheet, visible_rows
            from webapp.services.export_schema import EXPORT_MAPPING, canonical_sheet_name, headers_for_sheet
            from src.excel_standardization.export.excel_safe import safe_cell_value, safe_sheet_title
            from openpyxl.styles import Alignment
            
            export_name = canonical_sheet_name(sheet_dataset.sheet_name)
            ws = wb.create_sheet(title=safe_sheet_title(export_name, wb.sheetnames))
            ws.sheet_view.rightToLeft = True
            schema = headers_for_sheet(export_name)
            
            for col_idx, header in enumerate(schema, start=1):
                cell = ws.cell(row=1, column=col_idx, value=safe_cell_value(header))
                cell.alignment = Alignment(horizontal="right")
            
            data_rows, _ui_cols = visible_rows(sheet_dataset)
            active_mosad_type = record.mosad_types[0] if record.mosad_types else ""
            scoped_type = resolve_sug_mosad_for_sheet(
                record.sug_mosad_configs,
                sheet_dataset.sheet_name,
                active_mosad_type,
            )
            
            out_row = 2
            sheet_rows_exported = 0
            for row in data_rows:
                export_row = build_row_export_view(
                    row,
                    mosad_id=record.mosad_id or "",
                    scoped_sug_mosad=scoped_type,
                )
                for col_idx, header in enumerate(schema, start=1):
                    json_key = EXPORT_MAPPING.get(header)
                    if json_key is None:
                        continue
                    v = export_row.get(json_key)
                    if v is not None and v != "":
                        ws.cell(row=out_row, column=col_idx, value=safe_cell_value(v))
                out_row += 1
                sheet_rows_exported += 1
            
            wb.save(str(temp_path))
            
            if output_path.exists():
                try:
                    output_path.unlink()
                except Exception:
                    logger.warning("Could not remove previous export file %s", output_path, exc_info=True)
            
            os.replace(temp_path, output_path)
            
            logger.info(
                "single_sheet_export_successful",
                extra={
                    "event": "single_sheet_export_successful",
                    "session_id": session_id,
                    "sheet_name": sheet_name,
                    "output_filename": output_path.name,
                    "rows_exported": sheet_rows_exported,
                },
            )
            
        except Exception as exc:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    logger.warning("Could not remove failed temp export file %s", temp_path, exc_info=True)
            
            logger.error(
                "single_sheet_export_failed",
                exc_info=True,
                extra={
                    "event": "single_sheet_export_failed",
                    "session_id": session_id,
                    "sheet_name": sheet_name,
                    "error_type": type(exc).__name__,
                },
            )
            raise HTTPException(
                status_code=500,
                detail=f"Export of sheet '{sheet_name}' failed. Please try again.",
            )
        
        return output_path
canonical_sheet_name = canonical_sheet_name
headers_for_sheet = headers_for_sheet
EXPORT_MAPPING = EXPORT_MAPPING
_build_export_filename = _build_export_filename
_resolve_sug_mosad_for_sheet = _resolve_sug_mosad_for_sheet
visible_rows = visible_rows
