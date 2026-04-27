"""mosad_id_scanner: scan an openpyxl worksheet for a MosadID label/value pair.

Looks for a cell whose text matches known Hebrew/English labels for
"institution identifier" (e.g. ``מספר מזהה מוסד``, ``מספר מוסד``,
``institution id``, …).  When found, the value in the immediately adjacent
cell (right or left) is returned as the MosadID string.

The scan is intentionally cheap: it only reads cells, never modifies the
worksheet, and stops at the first match.
"""

import unicodedata
import logging
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Label patterns (after NFC + lower + whitespace-collapse).
# Any cell whose normalised text *contains* one of these fragments is treated
# as the MosadID label cell.
# ---------------------------------------------------------------------------
_LABEL_FRAGMENTS = [
    # Hebrew variants
    "מספר מזהה מוסד",
    "מספר מוסד",
    "מזהה מוסד",
    "קוד מוסד",
    "מס מוסד",
    "מס' מוסד",
    # English variants
    "institution id",
    "mosad id",
    "mosadid",
    "institution identifier",
    "institution number",
]


def _norm(text: str) -> str:
    """NFC-normalise, lower-case, collapse whitespace."""
    return unicodedata.normalize("NFC", " ".join(text.split())).lower()


_NORM_FRAGMENTS = [_norm(f) for f in _LABEL_FRAGMENTS]


def _is_label_cell(value) -> bool:
    """Return True if *value* looks like a MosadID label."""
    if value is None:
        return False
    n = _norm(str(value))
    return any(frag in n for frag in _NORM_FRAGMENTS)


def _coerce_value(raw) -> Optional[str]:
    """Return a non-empty string from a raw cell value, or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


def scan_mosad_id(worksheet: Worksheet) -> Optional[str]:
    """Scan *worksheet* for a MosadID label/value pair.

    Strategy:
    - Iterate every cell in the worksheet (up to ``max_row`` × ``max_col``).
    - When a cell matches a known label fragment, check the cell immediately
      to the right (col + 1) and immediately to the left (col - 1), in that
      order.  Return the first non-empty adjacent value found.
    - Stop at the first match.

    Returns:
        The MosadID value as a string, or ``None`` if not found.
    """
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if not _is_label_cell(cell.value):
                continue

            # Check right neighbour first, then left.
            for delta in (1, -1):
                neighbour_col = col_idx + delta
                if neighbour_col < 1 or neighbour_col > max_col:
                    continue
                neighbour = worksheet.cell(row=row_idx, column=neighbour_col)
                value = _coerce_value(neighbour.value)
                if value is not None:
                    logger.info(
                        f"MosadID found in sheet '{worksheet.title}': "
                        f"label at ({row_idx},{col_idx}), "
                        f"value '{value}' at ({row_idx},{neighbour_col})"
                    )
                    return value

    logger.debug(f"No MosadID label found in sheet '{worksheet.title}'")
    return None
