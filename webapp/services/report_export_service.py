"""Export the user-facing processing report as a separate workbook."""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from fastapi import HTTPException

from src.excel_standardization.export.excel_safe import safe_cell_value
from webapp.models.report import WorkbookProcessingReport
from webapp.services.report_service import ReportService
from webapp.services.session_service import SessionService

logger = logging.getLogger(__name__)


class ReportExportService:
    """Write the current processing report to an internal Excel workbook."""

    def __init__(
        self,
        session_service: SessionService,
        report_service: ReportService,
        output_dir: Path,
    ) -> None:
        self.session_service = session_service
        self.report_service = report_service
        self.output_dir = Path(output_dir)

    def export(self, session_id: str) -> Path:
        record = self.session_service.get(session_id)
        report = self.report_service.build(session_id, include_details=True)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        file_name = self._build_filename(record.original_filename, report)
        output_path = (self.output_dir / file_name).resolve(strict=False)
        output_root = self.output_dir.resolve(strict=False)
        if output_root != output_path and output_root not in output_path.parents:
            raise ValueError("Refusing to write report outside the output directory.")

        temp_path = output_path.with_name(f"{output_path.name}.tmp")
        if temp_path.exists():
            temp_path.unlink()

        logger.info(
            "report_export_started",
            extra={
                "event": "report_export_started",
                "session_id": session_id,
                "output_filename": output_path.name,
                "issue_count": len(report.issues),
            },
        )

        try:
            workbook = self._build_workbook(report)
            workbook.save(temp_path)
            if output_path.exists():
                output_path.unlink()
            os.replace(temp_path, output_path)
        except Exception as exc:
            if temp_path.exists():
                temp_path.unlink()
            logger.exception(
                "report_export_failed",
                extra={
                    "event": "report_export_failed",
                    "session_id": session_id,
                    "output_filename": output_path.name,
                },
            )
            raise HTTPException(status_code=500, detail="Processing report export failed.") from exc

        logger.info(
            "report_export_completed",
            extra={
                "event": "report_export_completed",
                "session_id": session_id,
                "output_filename": output_path.name,
                "sheet_count": len(workbook.sheetnames),
            },
        )
        return output_path

    def _build_workbook(self, report: WorkbookProcessingReport) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)

        summary = wb.create_sheet("סיכום")
        self._rtl(summary)
        self._write_summary_sheet(summary, report)

        sheet_summary = wb.create_sheet("סיכום גיליונות")
        self._rtl(sheet_summary)
        self._write_sheet_summary(sheet_summary, report)

        statuses = wb.create_sheet("סטטוסים")
        self._rtl(statuses)
        self._write_status_sheet(statuses, report)

        issues = wb.create_sheet("אזהרות ושגיאות")
        self._rtl(issues)
        self._write_issue_sheet(issues, report)

        if report.manual_edits.edited_cells:
            edits = wb.create_sheet("עריכות ידניות")
            self._rtl(edits)
            self._write_manual_edits_sheet(edits, report)

        return wb

    def _write_summary_sheet(self, ws, report: WorkbookProcessingReport) -> None:
        rows = [
            ("session_id", report.session_id),
            ("file_name", report.file_name),
            ("status", report.status),
            ("export_ready", report.export_ready),
            ("dirty", report.dirty),
            ("stale", report.stale),
            ("export_blocked_reason", report.export_blocked_reason),
            ("total_sheets", report.summary.total_sheets),
            ("total_rows", report.summary.total_rows),
            ("rows_with_warnings", report.summary.rows_with_warnings),
            ("rows_with_errors", report.summary.rows_with_errors),
            ("rows_without_issues", report.summary.rows_without_issues),
            ("corrected_fields", report.summary.corrected_fields),
            ("edited_cells", report.summary.edited_cells),
            ("edited_sheets", ", ".join(report.manual_edits.edited_sheets)),
            ("edited_fields", ", ".join(report.manual_edits.edited_fields)),
            ("generated_at_utc", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ]
        self._write_kv_table(ws, "שדה", "ערך", rows, title="דוח עיבוד")

    def _write_sheet_summary(self, ws, report: WorkbookProcessingReport) -> None:
        rows = []
        for sheet in report.sheets:
            rows.append(
                [
                    sheet.sheet_name,
                    sheet.row_count,
                    sheet.column_count,
                    sheet.rows_with_warnings,
                    sheet.rows_with_errors,
                    sheet.corrected_fields,
                    sheet.issues_count,
                ]
            )
        self._write_table(
            ws,
            ["sheet_name", "row_count", "column_count", "warnings", "errors", "corrected_fields", "issues"],
            rows or [["", 0, 0, 0, 0, 0, 0]],
            title="סיכום גיליונות",
        )

    def _write_status_sheet(self, ws, report: WorkbookProcessingReport) -> None:
        rows = []
        for sheet in report.sheets:
            for status_field, counts in sheet.status_counts.items():
                for status_value, count in counts.items():
                    rows.append([sheet.sheet_name, status_field, status_value, count])
        self._write_table(
            ws,
            ["sheet_name", "status_field", "status_value", "count"],
            rows or [["", "", "", 0]],
            title="סטטוסים",
        )

    def _write_issue_sheet(self, ws, report: WorkbookProcessingReport) -> None:
        rows = []
        for issue in report.issues:
            rows.append(
                [
                    issue.severity,
                    issue.sheet_name,
                    issue.row_uid or "",
                    issue.row_number or "",
                    issue.field_name or "",
                    issue.status_field or "",
                    issue.status_message or "",
                ]
            )
        self._write_table(
            ws,
            ["severity", "sheet_name", "row_uid", "row_number", "field_name", "status_field", "status_message"],
            rows or [["", "", "", "", "", "", ""]],
            title="אזהרות ושגיאות",
        )

    def _write_manual_edits_sheet(self, ws, report: WorkbookProcessingReport) -> None:
        rows = []
        for key in self.session_service.get(report.session_id).edits:
            if not isinstance(key, tuple) or len(key) != 3:
                continue
            sheet_name, row_uid, field_name = key
            rows.append([sheet_name, row_uid, field_name])
        self._write_table(
            ws,
            ["sheet_name", "row_uid", "field_name"],
            rows or [["", "", ""]],
            title="עריכות ידניות",
        )

    def _write_kv_table(self, ws, key_header: str, value_header: str, rows: list[tuple[str, object]], title: str) -> None:
        ws["A1"] = safe_cell_value(title)
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([safe_cell_value(key_header), safe_cell_value(value_header)])
        for key, value in rows:
            ws.append([safe_cell_value(key), safe_cell_value(value)])
        self._format_header(ws, 2)
        self._apply_basic_formatting(ws)

    def _write_table(self, ws, headers: list[str], rows: list[list[object]], title: str) -> None:
        ws["A1"] = safe_cell_value(title)
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([safe_cell_value(header) for header in headers])
        for row in rows:
            ws.append([safe_cell_value(value) for value in row])
        self._format_header(ws, 2)
        self._apply_basic_formatting(ws)

    @staticmethod
    def _format_header(ws, row_number: int) -> None:
        for cell in ws[row_number]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _apply_basic_formatting(ws) -> None:
        ws.freeze_panes = "A3"
        ws.sheet_view.rightToLeft = True
        for column_cells in ws.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            width = min(max((len(value) for value in values), default=0) + 2, 48)
            ws.column_dimensions[column_cells[0].column_letter].width = max(width, 10)

    @staticmethod
    def _rtl(ws) -> None:
        ws.sheet_view.rightToLeft = True

    @staticmethod
    def _build_filename(original_filename: str, report: WorkbookProcessingReport) -> str:
        stem = Path(original_filename).stem or "report"
        safe_stem = re.sub(r'[<>:"/\\|?*]+', "_", stem).strip(" ._") or "report"
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return f"processing_report_{safe_stem}_{timestamp}.xlsx"
