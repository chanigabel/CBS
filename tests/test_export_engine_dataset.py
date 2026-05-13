from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from src.excel_standardization.export.export_engine import ExportEngine
from webapp.models.session import SessionRecord
from webapp.services.export_service import ExportService
from webapp.services.session_service import SessionService
from webapp.services.export_schema import EXPORT_MAPPING


def test_export_engine_keeps_moved_passport_value_without_source_passport_column(tmp_path):
    engine = ExportEngine()
    sheet = SheetDataset(
        sheet_name=engine.SOURCE_SHEET_SPECS[0].source_sheet_name,
        header_row=1,
        header_rows_count=1,
        field_names=["id_number", "passport_corrected"],
        rows=[
            {
                "id_number": "ABC123",
                "id_number_corrected": "",
                "passport_corrected": "ABC123",
            }
        ],
    )
    workbook = WorkbookDataset(source_file="input.xlsx", sheets=[sheet])
    output_path = tmp_path / "export.xlsx"

    engine.export_from_normalized_dataset(workbook, str(output_path))

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    darkon_col = headers.index("Darkon") + 1
    assert ws.cell(row=2, column=darkon_col).value == "ABC123"
    wb.close()


def test_export_engine_uses_corrected_standardized_fields_only(tmp_path):
    engine = ExportEngine()
    sheet = SheetDataset(
        sheet_name=engine.SOURCE_SHEET_SPECS[0].source_sheet_name,
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "father_name", "id_number", "gender"],
        rows=[
            {
                "first_name": "Original First",
                "first_name_corrected": "Corrected First",
                "last_name": "Original Last",
                "last_name_corrected": "Corrected Last",
                "father_name": "Original Father",
                "father_name_corrected": "Corrected Father",
                "id_number": "123",
                "id_number_corrected": "",
                "gender": "male",
                "gender_corrected": "",
            },
            {
                "first_name": "Visible Missing Corrected",
                "first_name_corrected": "Visible Missing Corrected",
                "gender": "female",
            },
            {
                "first_name": "Numeric Invalid Identifier",
                "first_name_corrected": "Numeric Invalid Identifier",
                "id_number": "1234567890",
                "id_number_corrected": "1234567890",
            }
        ],
    )
    workbook = WorkbookDataset(source_file="input.xlsx", sheets=[sheet])
    output_path = tmp_path / "export.xlsx"

    engine.export_from_normalized_dataset(workbook, str(output_path))

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    first_col = headers.index("ShemPrati") + 1
    last_col = headers.index("ShemMishpaha") + 1
    father_col = headers.index("ShemHaAv") + 1
    id_col = headers.index("MisparZehut") + 1
    gender_col = headers.index("Min") + 1
    assert ws.cell(row=2, column=first_col).value == "Corrected First"
    assert ws.cell(row=2, column=last_col).value == "Corrected Last"
    assert ws.cell(row=2, column=father_col).value == "Corrected Father"
    assert ws.cell(row=2, column=id_col).value is None
    assert ws.cell(row=2, column=gender_col).value is None
    assert ws.cell(row=3, column=first_col).value == "Visible Missing Corrected"
    assert ws.cell(row=3, column=gender_col).value is None
    assert ws.cell(row=4, column=first_col).value == "Numeric Invalid Identifier"
    assert ws.cell(row=4, column=id_col).value == "1234567890"
    wb.close()


def test_compatibility_and_active_export_share_standardized_column_sources():
    engine = ExportEngine()
    row = {
        "first_name": "Original First",
        "first_name_corrected": "Corrected First",
        "last_name": "Original Last",
        "last_name_corrected": "Corrected Last",
        "father_name": "Original Father",
        "father_name_corrected": "Corrected Father",
        "id_number": "123",
        "id_number_corrected": "000000123",
        "passport": "Original Passport",
        "passport_corrected": "P123",
        "gender": "female",
        "gender_corrected": 2,
        "birth_year": 80,
        "birth_year_corrected": 1980,
        "birth_month": "05",
        "birth_month_corrected": 5,
        "birth_day": "bad",
        "birth_day_corrected": "",
        "entry_year": 10,
        "entry_year_corrected": 2010,
        "entry_month": "03",
        "entry_month_corrected": 3,
        "entry_day": "01",
        "entry_day_corrected": 1,
    }

    mapped = engine._map_row_to_export_fields(
        row,
        include_dira=False,
        allow_mosad_fields=True,
    )

    for header, source_key in EXPORT_MAPPING.items():
        if header in {"MosadID", "SugMosad", "MisparDiraBeMosad", "ShnatKnisa", "HodeshKnisa"}:
            continue
        assert mapped[header] == (row.get(source_key) or "")


def test_export_engine_sanitizes_extra_sheet_names_and_values(tmp_path):
    engine = ExportEngine()
    sheet = SheetDataset(
        sheet_name="bad:name/with*chars?and a very very long suffix",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "payload"],
        rows=[{"first_name": "Visible", "payload": "Bad\x02Value"}],
    )
    workbook = WorkbookDataset(source_file="input.xlsx", sheets=[sheet])
    output_path = tmp_path / "export.xlsx"

    engine.export_from_normalized_dataset(workbook, str(output_path))

    wb = load_workbook(output_path)
    extra_name = [name for name in wb.sheetnames if name not in {"DayarimYahidim", "MeshkeyBayt", "AnasheyTzevet"}][0]
    assert len(extra_name) <= 31
    assert not any(ch in extra_name for ch in "[]:*?/\\")
    ws = wb[extra_name]
    assert ws.cell(row=2, column=2).value == "BadValue"
    wb.close()


def test_active_and_compatibility_export_paths_produce_matching_standardized_values(tmp_path):
    engine = ExportEngine()
    sheet = SheetDataset(
        sheet_name=engine.SOURCE_SHEET_SPECS[0].source_sheet_name,
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "father_name", "gender", "id_number"],
        rows=[
            {
                "first_name": "Original First",
                "first_name_corrected": "Corrected First",
                "last_name": "Original Last",
                "last_name_corrected": "Corrected Last",
                "father_name": "Original Father",
                "father_name_corrected": "Corrected Father",
                "gender": "female",
                "gender_corrected": 2,
                "id_number": "123",
                "id_number_corrected": "123",
            }
        ],
    )
    workbook = WorkbookDataset(source_file="input.xlsx", sheets=[sheet])

    compat_path = tmp_path / "compat.xlsx"
    engine.export_from_normalized_dataset(workbook, str(compat_path))

    session_service = SessionService()
    session_service.clear_all()
    record = SessionRecord(
        session_id="export-compare",
        source_file_path="uploads/export-compare.xlsx",
        working_copy_path="work/export-compare.xlsx",
        original_filename="export-compare.xlsx",
        status="standardized",
        workbook_dataset=workbook,
    )
    session_service.create(record)
    active_path = ExportService(session_service, tmp_path / "active").export("export-compare")

    compat_wb = load_workbook(compat_path)
    active_wb = load_workbook(active_path)
    compat_ws = compat_wb["DayarimYahidim"]
    active_ws = active_wb["DayarimYahidim"]

    compat_headers = [cell.value for cell in compat_ws[1]]
    active_headers = [cell.value for cell in active_ws[1]]
    for header in ["ShemPrati", "ShemMishpaha", "ShemHaAv", "Min", "MisparZehut"]:
        compat_col = compat_headers.index(header) + 1
        active_col = active_headers.index(header) + 1
        assert compat_ws.cell(row=2, column=compat_col).value == active_ws.cell(row=2, column=active_col).value

    compat_wb.close()
    active_wb.close()
