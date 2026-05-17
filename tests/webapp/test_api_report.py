from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from webapp.models.session import SessionRecord
from webapp.services.edit_service import EditService
from webapp.services.export_service import ExportService
from webapp.services.processing_report_service import ProcessingReportService
from webapp.services.report_export_service import ReportExportService
from webapp.services.report_service import ReportService
from webapp.services.session_service import SessionService
from webapp.services.standardization_service import StandardizationService
from webapp.services.upload_service import UploadService
from webapp.services.workbook_service import WorkbookService


def _patch_services(monkeypatch, tmp_path):
    import webapp.dependencies as deps

    svc = SessionService()
    svc.clear_all()
    processing_report_svc = ProcessingReportService(svc)
    report_service = ReportService(svc)
    monkeypatch.setattr(deps, "_session_service", svc)
    monkeypatch.setattr(deps, "_processing_report_service", processing_report_svc)
    monkeypatch.setattr(deps, "_report_service", report_service)
    monkeypatch.setattr(
        deps,
        "_report_export_service",
        ReportExportService(svc, report_service, tmp_path / "output"),
    )
    monkeypatch.setattr(
        deps,
        "_upload_service",
        UploadService(svc, tmp_path / "uploads", tmp_path / "work", processing_report_svc),
    )
    monkeypatch.setattr(deps, "_workbook_service", WorkbookService(svc))
    monkeypatch.setattr(deps, "_standardization_service", StandardizationService(svc, processing_report_svc))
    monkeypatch.setattr(deps, "_edit_service", EditService(svc))
    monkeypatch.setattr(deps, "_export_service", ExportService(svc, tmp_path / "output", processing_report_svc))
    return svc


def _create_session(svc):
    sheet = SheetDataset(
        sheet_name="Sheet1",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "birth_date_status"],
        rows=[{"first_name": "Alice", "birth_date_status": "שנה חסרה והושלמה"}],
    )
    record = SessionRecord(
        session_id="api-report-session",
        source_file_path="uploads/api-report-session.xlsx",
        working_copy_path="work/api-report-session.xlsx",
        original_filename="api.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="work/api-report-session.xlsx", sheets=[sheet]),
    )
    svc.create(record)
    return record


def test_report_endpoint_returns_expected_json(tmp_path, monkeypatch):
    svc = _patch_services(monkeypatch, tmp_path)
    _create_session(svc)

    from webapp.app import app

    with TestClient(app) as client:
        response = client.get("/api/workbook/api-report-session/report")

    assert response.status_code == 200
    data = response.json()
    assert data["session_id"] == "api-report-session"
    assert data["file_name"] == "api.xlsx"
    assert data["summary"]["total_rows"] == 1
    assert data["sheets"][0]["status_counts"]["birth_date_status"]["שנה חסרה והושלמה"] == 1


def test_report_endpoint_missing_session_returns_404(tmp_path, monkeypatch):
    _patch_services(monkeypatch, tmp_path)

    from webapp.app import app

    with TestClient(app) as client:
        response = client.get("/api/workbook/missing/report")

    assert response.status_code == 404


def test_report_export_endpoint_returns_xlsx_file(tmp_path, monkeypatch):
    svc = _patch_services(monkeypatch, tmp_path)
    _create_session(svc)

    from webapp.app import app

    with TestClient(app) as client:
        response = client.get("/api/workbook/api-report-session/report/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "processing_report_api_" in response.headers["content-disposition"]

    output_path = tmp_path / "report.xlsx"
    output_path.write_bytes(response.content)
    wb = load_workbook(output_path)
    assert "סיכום" in wb.sheetnames
    assert "סיכום גיליונות" in wb.sheetnames
    wb.close()


def test_report_export_endpoint_missing_session_returns_404(tmp_path, monkeypatch):
    _patch_services(monkeypatch, tmp_path)

    from webapp.app import app

    with TestClient(app) as client:
        response = client.get("/api/workbook/missing/report/export")

    assert response.status_code == 404


def test_report_endpoint_does_not_trigger_standardization(tmp_path, monkeypatch):
    _patch_services(monkeypatch, tmp_path)

    import webapp.dependencies as deps

    def fail_if_called(*args, **kwargs):
        raise AssertionError("report endpoint must not standardize")

    monkeypatch.setattr(deps._standardization_service, "standardize", fail_if_called)

    record = SessionRecord(
        session_id="uploaded-session",
        source_file_path="uploads/uploaded-session.xlsx",
        working_copy_path="work/uploaded-session.xlsx",
        original_filename="uploaded.xlsx",
        status="uploaded",
    )
    deps._session_service.create(record)

    from webapp.app import app

    with TestClient(app) as client:
        response = client.get("/api/workbook/uploaded-session/report")

    assert response.status_code == 200
    assert response.json()["export_ready"] is False
