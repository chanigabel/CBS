"""Tests for the centralized workbook loader helper."""

import io

import pytest
from openpyxl import Workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from webapp.services.workbook_loader import (
    WorkbookLoadError,
    ensure_supported_workbook,
    extract_sheet_dataset,
    extract_workbook_dataset,
    get_workbook_sheet_names,
)


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["first_name"])
    ws.append(["Alice"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_ensure_supported_workbook_accepts_uppercase_extension():
    assert ensure_supported_workbook("input.XLSX") == ".xlsx"


def test_get_workbook_sheet_names_reads_uppercase_xlsx(tmp_path):
    path = tmp_path / "Upper.XLSX"
    path.write_bytes(_xlsx_bytes())
    assert get_workbook_sheet_names(path) == ["Sheet1"]


def test_extract_workbook_dataset_routes_xls_to_legacy_reader(tmp_path, monkeypatch):
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"legacy bytes")

    called = {}

    def fake_extract(p):
        called["path"] = p
        return WorkbookDataset(source_file=p, sheets=[SheetDataset(
            sheet_name="Sheet1",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Alice"}],
        )])

    monkeypatch.setattr(
        "src.excel_standardization.io_layer.xls_reader.extract_xls_to_workbook_dataset",
        fake_extract,
    )

    dataset = extract_workbook_dataset(path)
    assert called["path"] == str(path)
    assert dataset.sheets[0].rows[0]["first_name"] == "Alice"


def test_extract_sheet_dataset_routes_xls_to_legacy_reader(tmp_path, monkeypatch):
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"legacy bytes")

    called = {}

    def fake_extract(p, sheet_name):
        called["path"] = p
        called["sheet"] = sheet_name
        return SheetDataset(
            sheet_name=sheet_name,
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Alice"}],
        )

    monkeypatch.setattr(
        "src.excel_standardization.io_layer.xls_reader.extract_xls_sheet_to_dataset",
        fake_extract,
    )

    sheet = extract_sheet_dataset(path, "Sheet1")
    assert called == {"path": str(path), "sheet": "Sheet1"}
    assert sheet.rows[0]["first_name"] == "Alice"


def test_unsupported_extension_raises_clear_loader_error():
    with pytest.raises(WorkbookLoadError):
        ensure_supported_workbook("data.csv")


def test_get_workbook_sheet_names_raises_clear_error_for_corrupt_xlsx(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a real workbook")

    with pytest.raises(WorkbookLoadError):
        get_workbook_sheet_names(path)


def test_get_workbook_sheet_names_closes_workbook_on_failure(monkeypatch, tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(_xlsx_bytes())

    closed = {"value": False}

    class FakeWorkbook:
        @property
        def sheetnames(self):
            raise RuntimeError("boom")

        def close(self):
            closed["value"] = True

    monkeypatch.setattr("webapp.services.workbook_loader.load_workbook", lambda *a, **k: FakeWorkbook())

    with pytest.raises(WorkbookLoadError):
        get_workbook_sheet_names(path)

    assert closed["value"] is True


def test_extract_sheet_dataset_closes_workbook_on_failure(monkeypatch, tmp_path):
    path = tmp_path / "sheet.xlsx"
    path.write_bytes(_xlsx_bytes())

    closed = {"value": False}

    class FakeWorksheet:
        title = "Sheet1"

    class FakeWorkbook:
        sheetnames = ["Sheet1"]

        def __getitem__(self, name):
            assert name == "Sheet1"
            return FakeWorksheet()

        def close(self):
            closed["value"] = True

    class FakeExtractor:
        def extract_sheet_to_json(self, ws):
            raise RuntimeError("extract failed")

    monkeypatch.setattr("webapp.services.workbook_loader.load_workbook", lambda *a, **k: FakeWorkbook())
    monkeypatch.setattr("webapp.services.workbook_loader._extractor", lambda: FakeExtractor())
    monkeypatch.setattr("webapp.services.workbook_loader.scan_mosad_id", lambda ws: None)

    with pytest.raises(WorkbookLoadError):
        extract_sheet_dataset(path, "Sheet1")

    assert closed["value"] is True
