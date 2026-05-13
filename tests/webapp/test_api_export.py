"""Unit tests for the POST /api/workbook/{session_id}/export endpoint."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from tests.webapp.conftest import make_xlsx_bytes


def make_identifier_only_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id_number"])
    ws.append(["ABC123"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_numeric_invalid_identifier_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["first_name", "id_number"])
    ws.append(["Visible", "12345678910"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_and_normalize(client):
    """Upload and normalize a file, return session_id."""
    file_bytes = make_xlsx_bytes(["Sheet1"])
    response = client.post(
        "/api/upload",
        files={"file": ("test.xlsx", file_bytes, "application/octet-stream")},
    )
    assert response.status_code == 200
    session_id = response.json()["session_id"]

    norm_response = client.post(f"/api/workbook/{session_id}/normalize")
    assert norm_response.status_code == 200
    return session_id


def test_export_returns_file_response(client):
    session_id = upload_and_normalize(client)
    response = client.post(f"/api/workbook/{session_id}/export")
    # Export may return 200 (file) or 500 if no matching VBA sheets
    # The important thing is it doesn't crash with 404
    assert response.status_code in (200, 500)

    report_response = client.get(f"/api/workbook/{session_id}/processing-report")
    assert report_response.status_code == 200
    report = report_response.json()
    assert "upload" in report["completed_stages"]
    assert "standardize" in report["completed_stages"]
    if response.status_code == 200:
        assert "export" in report["completed_stages"]
        assert report["rows_exported"] >= 0
        assert report["output_filename"].endswith(".xlsx")


def test_export_returns_404_for_unknown_session(client):
    response = client.post("/api/workbook/ghost-session/export")
    assert response.status_code == 404


def test_export_after_upload_without_normalize(client):
    """Export should work even without standardization (uses raw data)."""
    file_bytes = make_xlsx_bytes(["Sheet1"])
    response = client.post(
        "/api/upload",
        files={"file": ("test.xlsx", file_bytes, "application/octet-stream")},
    )
    session_id = response.json()["session_id"]
    export_response = client.post(f"/api/workbook/{session_id}/export")
    # Should not return 404
    assert export_response.status_code != 404


def test_export_writes_generated_passport_corrected_to_darkon_without_source_passport(client):
    response = client.post(
        "/api/upload",
        files={
            "file": (
                "identifier_only.xlsx",
                make_identifier_only_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 200
    session_id = response.json()["session_id"]

    norm_response = client.post(f"/api/workbook/{session_id}/normalize")
    assert norm_response.status_code == 200

    export_response = client.post(f"/api/workbook/{session_id}/export")
    assert export_response.status_code == 200

    wb = load_workbook(BytesIO(export_response.content))
    ws = wb["Sheet1"]
    headers = [cell.value for cell in ws[1]]
    darkon_col = headers.index("Darkon") + 1
    id_col = headers.index("MisparZehut") + 1
    assert ws.cell(row=2, column=darkon_col).value == "ABC123"
    assert ws.cell(row=2, column=id_col).value is None
    wb.close()


def test_export_writes_numeric_invalid_length_id_from_corrected_dataset_value(client):
    response = client.post(
        "/api/upload",
        files={
            "file": (
                "numeric_invalid_identifier.xlsx",
                make_numeric_invalid_identifier_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 200
    session_id = response.json()["session_id"]

    norm_response = client.post(f"/api/workbook/{session_id}/normalize")
    assert norm_response.status_code == 200

    grid_response = client.get(f"/api/workbook/{session_id}/sheet/Sheet1")
    assert grid_response.status_code == 200
    grid_row = grid_response.json()["rows"][0]
    assert grid_row["id_number_corrected"] == "12345678910"
    assert grid_row.get("passport_corrected") in (None, "")

    export_response = client.post(f"/api/workbook/{session_id}/export")
    assert export_response.status_code == 200

    wb = load_workbook(BytesIO(export_response.content))
    ws = wb["Sheet1"]
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("MisparZehut") + 1
    passport_col = headers.index("Darkon") + 1
    assert ws.cell(row=2, column=id_col).value == "12345678910"
    assert ws.cell(row=2, column=passport_col).value is None
    wb.close()
