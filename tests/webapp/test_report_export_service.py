import copy

import pytest
from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from webapp.models.session import SessionRecord
from webapp.services.export_service import ExportService
from webapp.services.report_export_service import ReportExportService
from webapp.services.report_service import ReportService
from webapp.services.session_service import SessionService


@pytest.fixture(autouse=True)
def clear_registry():
    svc = SessionService()
    svc.clear_all()
    yield
    svc.clear_all()


def _record(session_id="report-export-session", *, status="standardized", dirty=False):
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=[
            "first_name",
            "first_name_corrected",
            "birth_date_status",
            "_validation_status",
        ],
        rows=[
            {
                "_row_uid": "row-1",
                "first_name": "Raw",
                "first_name_corrected": "Corrected",
                "birth_date_status": "missing year corrected",
                "_validation_status": "ok",
            }
        ],
    )
    return SessionRecord(
        session_id=session_id,
        source_file_path=f"uploads/{session_id}.xlsx",
        working_copy_path=f"work/{session_id}.xlsx",
        original_filename="../unsafe:name.xlsx",
        status=status,
        workbook_dataset=WorkbookDataset(source_file=f"work/{session_id}.xlsx", sheets=[sheet]),
        edits={("DayarimYahidim", "row-1", "first_name"): "Corrected"},
        working_dataset_dirty=dirty,
    )


def _summary_values(path):
    wb = load_workbook(path)
    ws = wb["סיכום"]
    values = {ws.cell(row=row, column=1).value: ws.cell(row=row, column=2).value for row in range(3, ws.max_row + 1)}
    wb.close()
    return values


def test_report_export_returns_separate_xlsx_with_expected_sheets(tmp_path):
    svc = SessionService()
    svc.create(_record())

    output_path = ReportExportService(svc, ReportService(svc), tmp_path / "output").export("report-export-session")

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert output_path.parent == (tmp_path / "output").resolve()
    assert output_path.name.startswith("processing_report_unsafe_name_")

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["סיכום", "סיכום גיליונות", "סטטוסים", "אזהרות ושגיאות", "עריכות ידניות"]
    assert wb["סיכום"].sheet_view.rightToLeft is True
    assert wb["סיכום"]["A1"].value == "דוח עיבוד"
    wb.close()


def test_report_export_uses_report_service_data_and_does_not_reextract(tmp_path, monkeypatch):
    svc = SessionService()
    svc.create(_record())
    calls = []

    class SpyReportService(ReportService):
        def build(self, session_id, include_details=False):
            calls.append((session_id, include_details))
            return super().build(session_id, include_details=include_details)

    def fail_if_called(*args, **kwargs):
        raise AssertionError("report export must not re-read workbook files")

    monkeypatch.setattr("webapp.services.workbook_loader.extract_workbook_dataset", fail_if_called)

    ReportExportService(svc, SpyReportService(svc), tmp_path / "output").export("report-export-session")

    assert calls == [("report-export-session", True)]


def test_report_export_does_not_mutate_workbook_dataset(tmp_path):
    svc = SessionService()
    record = _record()
    svc.create(record)
    before = copy.deepcopy(record.workbook_dataset)

    ReportExportService(svc, ReportService(svc), tmp_path / "output").export("report-export-session")

    assert record.workbook_dataset == before


def test_report_export_marks_dirty_report_as_stale(tmp_path):
    svc = SessionService()
    svc.create(_record(dirty=True))

    output_path = ReportExportService(svc, ReportService(svc), tmp_path / "output").export("report-export-session")
    values = _summary_values(output_path)

    # Dirty/stale should be reported, but export_ready should be True because
    # the session was already standardized; the report should not claim the
    # user must re-run Standardization before exporting.
    assert values["export_ready"] is True
    assert values["dirty"] is True
    assert values["stale"] is True
    assert "Run Standardization again" not in str(values.get("export_blocked_reason", ""))


def test_report_export_works_before_standardization_with_loaded_dataset(tmp_path):
    svc = SessionService()
    svc.create(_record(status="uploaded"))

    output_path = ReportExportService(svc, ReportService(svc), tmp_path / "output").export("report-export-session")
    values = _summary_values(output_path)

    assert values["status"] == "uploaded"
    assert values["export_ready"] is False
    assert values["total_rows"] == 1
    assert values["export_blocked_reason"] == "Standardization has not completed yet."


def test_report_export_after_standardization_marks_export_ready(tmp_path):
    svc = SessionService()
    svc.create(_record(status="standardized", dirty=False))

    output_path = ReportExportService(svc, ReportService(svc), tmp_path / "output").export("report-export-session")
    values = _summary_values(output_path)

    assert values["status"] == "standardized"
    assert values["export_ready"] is True


def test_report_export_missing_session_returns_404(tmp_path):
    svc = SessionService()

    with pytest.raises(Exception) as exc_info:
        ReportExportService(svc, ReportService(svc), tmp_path / "output").export("missing")

    assert getattr(exc_info.value, "status_code", None) == 404


def test_main_standardized_export_does_not_include_report_sheets(tmp_path):
    svc = SessionService()
    svc.create(_record())

    output_path = ExportService(svc, tmp_path / "output").export("report-export-session")

    wb = load_workbook(output_path)
    assert "סיכום" not in wb.sheetnames
    assert "אזהרות ושגיאות" not in wb.sheetnames
    wb.close()
