import logging
from datetime import datetime

from openpyxl import Workbook, load_workbook

from src.excel_standardization.engines.date_engine import STATUS_INVALID_LENGTH
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from webapp.models.session import SessionRecord
from webapp.services.session_service import SessionService
from webapp.services.standardization_service import StandardizationService
from webapp.services.workbook_service import WorkbookService


def test_numeric_compact_date_cells_arrive_as_expected_and_parse_safely(tmp_path, caplog):
    caplog.set_level(
        logging.DEBUG,
        logger="src.excel_standardization.io_layer.excel_to_json_extractor",
    )

    workbook_path = tmp_path / "compact_dates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["name", "birth date"])

    sheet["A2"] = "Alice"
    sheet["B2"] = "1124"
    sheet["B2"].number_format = "@"

    sheet["A3"] = "Bob"
    sheet["B3"] = 1124
    sheet["B3"].number_format = "General"

    sheet["A4"] = "Carol"
    sheet["B4"] = "010224"
    sheet["B4"].number_format = "@"

    sheet["A5"] = "Dave"
    sheet["B5"] = 10224
    sheet["B5"].number_format = "General"

    sheet["A6"] = "Eve"
    sheet["B6"] = 12022001
    sheet["B6"].number_format = "General"

    sheet["A7"] = "Frank"
    sheet["B7"] = 36525
    sheet["B7"].number_format = "General"

    sheet["A8"] = "Hannah"
    sheet["B8"] = 36525
    sheet["B8"].number_format = "dd/mm/yyyy"

    workbook.save(str(workbook_path))
    workbook.close()

    workbook_loaded = load_workbook(str(workbook_path), data_only=True)
    worksheet_loaded = workbook_loaded["Sheet1"]

    assert worksheet_loaded["B2"].value == "1124"
    assert worksheet_loaded["B2"].data_type == "s"
    assert worksheet_loaded["B2"].number_format == "@"
    assert worksheet_loaded["B2"].is_date is False

    assert worksheet_loaded["B3"].value == 1124
    assert worksheet_loaded["B3"].data_type == "n"
    assert worksheet_loaded["B3"].number_format == "General"
    assert worksheet_loaded["B3"].is_date is False

    assert worksheet_loaded["B4"].value == "010224"
    assert worksheet_loaded["B4"].data_type == "s"
    assert worksheet_loaded["B4"].number_format == "@"
    assert worksheet_loaded["B4"].is_date is False

    assert worksheet_loaded["B5"].value == 10224
    assert worksheet_loaded["B5"].data_type == "n"
    assert worksheet_loaded["B5"].number_format == "General"
    assert worksheet_loaded["B5"].is_date is False

    assert worksheet_loaded["B6"].value == 12022001
    assert worksheet_loaded["B6"].data_type == "n"
    assert worksheet_loaded["B6"].number_format == "General"
    assert worksheet_loaded["B6"].is_date is False

    assert worksheet_loaded["B7"].value == 36525
    assert worksheet_loaded["B7"].data_type == "n"
    assert worksheet_loaded["B7"].number_format == "General"
    assert worksheet_loaded["B7"].is_date is False

    assert isinstance(worksheet_loaded["B8"].value, datetime)
    assert worksheet_loaded["B8"].data_type == "d"
    assert worksheet_loaded["B8"].number_format == "dd/mm/yyyy"
    assert worksheet_loaded["B8"].is_date is True

    extractor = ExcelToJsonExtractor(
        excel_reader=ExcelReader(),
        skip_empty_rows=False,
        handle_formulas=True,
        preserve_types=True,
    )

    workbook_dataset = extractor.extract_workbook_to_json(str(workbook_path))
    sheet_dataset = workbook_dataset.get_sheet_by_name("Sheet1")
    assert sheet_dataset is not None

    raw_rows = sheet_dataset.rows
    assert len(raw_rows) == 7
    assert raw_rows[0]["first_name"] == "Alice"
    assert raw_rows[0]["birth_date"] == "1124"
    assert isinstance(raw_rows[0]["birth_date"], str)
    assert raw_rows[1]["first_name"] == "Bob"
    assert raw_rows[1]["birth_date"] == 1124
    assert isinstance(raw_rows[1]["birth_date"], int)
    assert raw_rows[2]["first_name"] == "Carol"
    assert raw_rows[2]["birth_date"] == "010224"
    assert isinstance(raw_rows[2]["birth_date"], str)
    assert raw_rows[3]["first_name"] == "Dave"
    assert raw_rows[3]["birth_date"] == 10224
    assert isinstance(raw_rows[3]["birth_date"], int)
    assert raw_rows[4]["first_name"] == "Eve"
    assert raw_rows[4]["birth_date"] == 12022001
    assert isinstance(raw_rows[4]["birth_date"], int)
    assert raw_rows[5]["first_name"] == "Frank"
    assert raw_rows[5]["birth_date"] == 36525
    assert isinstance(raw_rows[5]["birth_date"], int)
    assert raw_rows[6]["first_name"] == "Hannah"
    assert isinstance(raw_rows[6]["birth_date"], datetime)

    assert "Extracted compact date field 'birth_date'" in caplog.text

    session_service = SessionService()
    session_service.clear_all()
    session_id = "compact-dates-runtime"
    session_record = SessionRecord(
        session_id=session_id,
        source_file_path=str(workbook_path),
        working_copy_path=str(workbook_path),
        original_filename="compact_dates.xlsx",
        status="uploaded",
    )
    session_service.create(session_record)

    standardization_service = StandardizationService(session_service)
    response = standardization_service.standardize(session_id, sheet_name="Sheet1")

    assert response.status == "standardized"

    normalized_sheet = session_service.get(session_id).workbook_dataset.get_sheet_by_name("Sheet1")
    assert normalized_sheet is not None
    normalized_rows = normalized_sheet.rows
    assert len(normalized_rows) == 7

    assert normalized_rows[0]["birth_year_corrected"] == 2024
    assert normalized_rows[0]["birth_month_corrected"] == 1
    assert normalized_rows[0]["birth_day_corrected"] == 1

    assert normalized_rows[1]["birth_year_corrected"] == 2024
    assert normalized_rows[1]["birth_month_corrected"] == 1
    assert normalized_rows[1]["birth_day_corrected"] == 1

    assert normalized_rows[2]["birth_year_corrected"] == 2024
    assert normalized_rows[2]["birth_month_corrected"] == 2
    assert normalized_rows[2]["birth_day_corrected"] == 1

    assert normalized_rows[3]["birth_date_status"] == STATUS_INVALID_LENGTH

    assert normalized_rows[4]["birth_year_corrected"] == 2001
    assert normalized_rows[4]["birth_month_corrected"] == 2
    assert normalized_rows[4]["birth_day_corrected"] == 12

    assert normalized_rows[5]["birth_date_status"] == STATUS_INVALID_LENGTH

    assert normalized_rows[6]["birth_year_corrected"] == 1999
    assert normalized_rows[6]["birth_month_corrected"] == 12
    assert normalized_rows[6]["birth_day_corrected"] == 31

    workbook_service = WorkbookService(session_service)
    sheet_data = workbook_service.get_sheet_data(session_id, "Sheet1")
    assert sheet_data.sheet_name == "Sheet1"
    assert "birth_year_corrected" in sheet_data.field_names
    assert len(sheet_data.rows) == 7
    assert sheet_data.rows[3]["birth_date_status"] == STATUS_INVALID_LENGTH
    assert any(row["first_name"] == "Eve" and row["birth_year_corrected"] == 2001 for row in sheet_data.rows)
