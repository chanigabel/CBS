"""Tests for POST /api/process-file."""

import io

from openpyxl import Workbook
from openpyxl import load_workbook

from tests.webapp.conftest import make_xlsx_bytes


def make_xlsx_with_invalid_split_entry_month() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([
        "first_name",
        "last_name",
        "gender",
        "id_number",
        "passport",
        "תאריך כניסה",
        "",
        "",
    ])
    ws.append([
        "",
        "",
        "",
        "",
        "",
        "שנה",
        "חודש",
        "יום",
    ])
    ws.append(["Alice", "Smith", "F", "000000000", "", 2010, "Invalid ID number", 20])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_xlsx_with_identifier_status_mix() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["first_name", "last_name", "gender", "id_number", "passport"])
    ws.append(["Alice", "Smith", "F", "000000018", "ABC123"])
    ws.append(["Bob", "Jones", "M", "000000019", "ABC123"])
    ws.append(["Cara", "Lee", "F", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_process_file_returns_exported_workbook(client):
    file_bytes = make_xlsx_bytes(["Sheet1"])

    response = client.post(
        "/api/process-file",
        files={"file": ("test.xlsx", file_bytes, "application/octet-stream")},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    assert response.headers["x-processing-report-id"]
    assert response.headers["x-processing-status"] in {
        "success",
        "partial_success",
        "failed",
    }

    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames
    wb.close()

    report_response = client.get(
        f"/api/workbook/{response.headers['x-processing-report-id']}/processing-report"
    )
    assert report_response.status_code == 200
    report = report_response.json()
    assert report["completed_stages"] == [
        "upload",
        "extract",
        "standardize",
        "validate",
        "export",
    ]
    assert report["sheets_processed"] == 1
    assert report["rows_processed"] >= 1
    assert report["rows_exported"] >= 1
    assert report["output_filename"].endswith(".xlsx")


def test_process_file_reuses_existing_services(client, monkeypatch):
    import webapp.dependencies as deps

    file_bytes = make_xlsx_bytes(["Sheet1"])
    calls = []

    original_upload = deps._upload_service.handle_upload
    original_standardize = deps._standardization_service.standardize
    original_export = deps._export_service.export

    def spy_upload(filename, uploaded_bytes):
        calls.append("upload")
        return original_upload(filename, uploaded_bytes)

    def spy_standardize(session_id, sheet_name=None):
        calls.append(("standardize", session_id, sheet_name))
        return original_standardize(session_id, sheet_name=sheet_name)

    def spy_export(session_id):
        calls.append(("export", session_id))
        return original_export(session_id)

    monkeypatch.setattr(deps._upload_service, "handle_upload", spy_upload)
    monkeypatch.setattr(deps._standardization_service, "standardize", spy_standardize)
    monkeypatch.setattr(deps._export_service, "export", spy_export)

    response = client.post(
        "/api/process-file",
        files={"file": ("test.xlsx", file_bytes, "application/octet-stream")},
    )

    assert response.status_code == 200
    assert calls[0] == "upload"
    assert calls[1][0] == "standardize"
    assert calls[1][2] is None
    assert calls[2] == ("export", calls[1][1])


def test_process_file_invalid_extension_returns_400(client):
    response = client.post(
        "/api/process-file",
        files={"file": ("test.csv", b"col1,col2\nval1,val2", "text/csv")},
    )

    assert response.status_code == 400
    assert "xlsx" in response.json()["detail"].lower()


def test_process_file_export_does_not_include_invalid_date_text(client):
    response = client.post(
        "/api/process-file",
        files={
            "file": (
                "invalid-date.xlsx",
                make_xlsx_with_invalid_split_entry_month(),
                "application/octet-stream",
            )
        },
    )

    assert response.status_code == 200

    wb = load_workbook(io.BytesIO(response.content))
    values = [
        cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    ]
    wb.close()

    assert "Invalid ID number" not in values

    report_response = client.get(
        f"/api/workbook/{response.headers['x-processing-report-id']}/processing-report"
    )
    assert report_response.status_code == 200
    report = report_response.json()

    assert report["status"] == "partial_success"
    assert report["status_reason"].startswith("partial_success because:\n- ")
    assert "invalid date values" in report["status_reason"]
    assert "invalid ID values" in report["status_reason"]
    assert report["missing_input_columns"]
    assert report["missing_required_export_fields"]

    assert report["date_summary"]
    assert report["date_summary"][0]["count"] == 1
    assert report["identifier_summary"]
    assert sum(item["count"] for item in report["identifier_summary"]) == 1
    assert all(item["message"] in {"חסר מזהים", "ת.ז. לא תקינה"} for item in report["identifier_summary"])
    assert report["missing_required_fields"]
    assert report["empty_required_columns_summary"] == report["missing_required_fields"]
    assert all("field" in item and "count" in item for item in report["missing_required_fields"])
    assert "invalid_date_values" not in report
    assert "invalid_identifier_values" not in report

    details_response = client.get(
        f"/api/workbook/{response.headers['x-processing-report-id']}/processing-report"
        "?include_details=true"
    )
    assert details_response.status_code == 200
    details = details_response.json()

    invalid_date = details["invalid_date_values"][0]
    assert invalid_date["sheet_name"] == "Sheet1"
    assert invalid_date["source_field"] == "entry_month"
    assert invalid_date["raw_value"] == "Invalid ID number"
    assert invalid_date["corrected_value"] == ""
    assert invalid_date["status_message"] == "ערך תאריך לא תקין"

    assert details["invalid_identifier_values"]
    invalid_identifier = details["invalid_identifier_values"][0]
    assert invalid_identifier["sheet_name"] == "Sheet1"
    assert invalid_identifier["source_field"] in {"id_number", "passport"}
    assert invalid_identifier["status_message"]

    per_sheet = report["per_sheet_warnings"][0]
    assert per_sheet["sheet_name"] == "Sheet1"
    assert per_sheet["rows_processed"] == 1
    assert per_sheet["rows_exported"] == 1
    assert any(w.startswith("עמודות חובה ריקות: ") for w in per_sheet["warnings"])


def test_process_file_compact_identifier_summary_excludes_valid_statuses(client):
    response = client.post(
        "/api/process-file",
        files={
            "file": (
                "identifier-mix.xlsx",
                make_xlsx_with_identifier_status_mix(),
                "application/octet-stream",
            )
        },
    )

    assert response.status_code == 200

    report_response = client.get(
        f"/api/workbook/{response.headers['x-processing-report-id']}/processing-report"
    )
    assert report_response.status_code == 200
    report = report_response.json()

    identifier_messages = [item["message"] for item in report["identifier_summary"]]
    assert "דרכון הוזן" not in identifier_messages
    assert "ת.ז. תקינה + דרכון הוזן" not in identifier_messages
    assert "חסר מזהים" in identifier_messages
    assert any("לא תקינה" in message for message in identifier_messages)

    assert "דרכון הוזן" not in report["status_reason"]
    assert "ת.ז. תקינה + דרכון הוזן" not in report["status_reason"]
    assert "rows missing identifiers" in report["status_reason"]


def test_process_file_invalid_export_returns_json_error_and_removes_fake_xlsx(
    client,
    tmp_path,
    monkeypatch,
):
    import webapp.dependencies as deps

    fake_output = tmp_path / "fake.xlsx"

    def fake_export(session_id):
        fake_output.write_bytes(b"not a real xlsx")
        return fake_output

    monkeypatch.setattr(deps._export_service, "export", fake_export)

    response = client.post(
        "/api/process-file",
        files={"file": ("test.xlsx", make_xlsx_bytes(["Sheet1"]), "application/octet-stream")},
    )

    assert response.status_code == 500
    assert response.headers["content-type"].startswith("application/json")
    assert response.json()["detail"] == "Processing produced an invalid Excel output file."
    assert not fake_output.exists()
