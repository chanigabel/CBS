"""Property-based tests for the Excel standardization Web App.

Uses Hypothesis to verify universal properties across many inputs.
"""

import io
import tempfile
import uuid
from pathlib import Path

import pytest
from hypothesis import given, settings, HealthCheck
from hypothesis import strategies as st
from openpyxl import Workbook

from webapp.services.session_service import SessionService
from webapp.services.upload_service import UploadService
from webapp.services.workbook_service import WorkbookService
from webapp.services.edit_service import EditService
from webapp.models.requests import CellEditRequest
from src.excel_standardization.data_types import SheetDataset, WorkbookDataset


# ---------------------------------------------------------------------------
# Hypothesis strategies
# ---------------------------------------------------------------------------

HEBREW_FIELD_NAMES = [
    "first_name", "last_name", "father_name",
    "gender", "id_number", "passport",
    "birth_year", "birth_month", "birth_day",
]


def _make_xlsx_bytes(sheet_names=None) -> bytes:
    """Create a minimal valid .xlsx file in memory."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in (sheet_names or ["Sheet1"]):
        ws = wb.create_sheet(name)
        ws.append(["first_name", "last_name"])
        ws.append(["Alice", "Smith"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.composite
def workbook_strategy(draw):
    """Generate valid openpyxl Workbook bytes with random sheet names."""
    n_sheets = draw(st.integers(min_value=1, max_value=3))
    sheet_names = draw(
        st.lists(
            st.text(
                alphabet=st.characters(
                    whitelist_categories=("Lu", "Ll", "Nd"),
                    whitelist_characters="_",
                ),
                min_size=1,
                max_size=20,
            ),
            min_size=n_sheets,
            max_size=n_sheets,
            unique=True,
        )
    )
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheet_names:
        ws = wb.create_sheet(name)
        ws.append(["first_name", "last_name"])
        ws.append(["Alice", "Smith"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), sheet_names


@st.composite
def sheet_dataset_strategy(draw):
    """Generate SheetDataset instances with random field names and rows."""
    field_names = draw(
        st.lists(
            st.sampled_from(HEBREW_FIELD_NAMES),
            min_size=1,
            max_size=4,
            unique=True,
        )
    )
    n_rows = draw(st.integers(min_value=1, max_value=10))
    rows = []
    for _ in range(n_rows):
        row = {f: draw(st.text(min_size=0, max_size=20)) for f in field_names}
        rows.append(row)
    return SheetDataset(
        sheet_name="TestSheet",
        header_row=1,
        header_rows_count=1,
        field_names=field_names,
        rows=rows,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_upload_service(tmp_path: Path):
    svc = SessionService()
    svc.clear_all()
    uploads = tmp_path / "uploads"
    work = tmp_path / "work"
    return UploadService(svc, uploads, work), svc


# ---------------------------------------------------------------------------
# Property 1: Upload preserves source file integrity
# Validates: Requirements 1.2, 9.2
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture], deadline=None)
def test_upload_preserves_source_file(workbook_data):
    """Property 1: Upload preserves source file integrity.

    Validates: Requirements 1.2, 9.2
    """
    file_bytes, sheet_names = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, _ = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        # Source file bytes must be identical to uploaded bytes
        source_path = Path(response.session_id)
        # Find the actual source file
        uploads_dir = tmp_path / "uploads"
        source_files = list(uploads_dir.glob(f"{response.session_id}*"))
        assert len(source_files) == 1, "Exactly one source file should exist"
        assert source_files[0].read_bytes() == file_bytes, "Source file must be byte-for-byte identical"

        # A separate working copy must exist
        work_dir = tmp_path / "work"
        work_files = list(work_dir.glob(f"{response.session_id}*"))
        assert len(work_files) == 1, "Exactly one working copy should exist"


# ---------------------------------------------------------------------------
# Property 2: Upload response reflects actual workbook structure
# Validates: Requirements 1.3, 1.6
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=30, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_upload_response_matches_workbook_structure(workbook_data):
    """Property 2: Upload response reflects actual workbook structure.

    Validates: Requirements 1.3, 1.6
    """
    file_bytes, expected_sheet_names = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, _ = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        # session_id must be a valid UUID
        parsed = uuid.UUID(response.session_id)
        assert str(parsed) == response.session_id

        # sheet_names must match the workbook's actual sheets
        assert set(response.sheet_names) == set(expected_sheet_names)


# ---------------------------------------------------------------------------
# Property 3: Invalid file extensions are always rejected
# Validates: Requirements 1.4
# ---------------------------------------------------------------------------

@given(
    st.text(min_size=1, max_size=10).filter(
        lambda x: not x.lower().endswith((".xlsx", ".xlsm"))
    )
)
@settings(max_examples=50, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_invalid_extension_rejected(extension):
    """Property 3: Invalid file extensions are always rejected.

    Validates: Requirements 1.4
    """
    from fastapi import HTTPException

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, _ = _make_upload_service(tmp_path)
        filename = f"file{extension}"
        with pytest.raises(HTTPException) as exc_info:
            upload_svc.handle_upload(filename, b"some bytes")
        assert exc_info.value.status_code == 400


# ---------------------------------------------------------------------------
# Property 5: Session initialization invariant
# Validates: Requirements 2.2
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_session_initialization_invariant(workbook_data):
    """Property 5: Session initialization invariant.

    Validates: Requirements 2.2
    """
    file_bytes, _ = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, session_svc = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        record = session_svc.get(response.session_id)
        assert record.status == "uploaded"
        assert record.source_file_path != ""
        assert record.working_copy_path != ""
        assert record.workbook_dataset is not None
        assert len(record.workbook_dataset.sheets) >= 1


# ---------------------------------------------------------------------------
# Property 8: Cell edit round-trip
# Validates: Requirements 6.2, 6.3
# ---------------------------------------------------------------------------

@given(sheet_dataset_strategy(), st.text(min_size=0, max_size=50))
@settings(max_examples=50, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_cell_edit_round_trip(sheet_dataset, new_value):
    """Property 8: Cell edit round-trip.

    Validates: Requirements 6.2, 6.3
    """
    if not sheet_dataset.rows:
        return

    session_svc = SessionService()
    session_svc.clear_all()

    # Build a minimal session with the sheet dataset
    workbook_dataset = WorkbookDataset(
        source_file="dummy.xlsx",
        sheets=[sheet_dataset],
    )
    from webapp.models.session import SessionRecord
    record = SessionRecord(
        session_id="edit-test-session",
        source_file_path="uploads/edit-test-session.xlsx",
        working_copy_path="work/edit-test-session.xlsx",
        original_filename="test.xlsx",
        status="uploaded",
        workbook_dataset=workbook_dataset,
    )
    session_svc.create(record)

    edit_svc = EditService(session_svc)
    workbook_svc = WorkbookService(session_svc)

    field_name = sheet_dataset.field_names[0]

    # Get the row_uid for the first row via get_sheet_data (which assigns UIDs)
    # We need a real file path for WorkbookService, so assign UIDs directly
    import uuid as _uuid
    first_row = sheet_dataset.rows[0]
    if "_row_uid" not in first_row:
        first_row["_row_uid"] = _uuid.uuid4().hex
    row_uid = first_row["_row_uid"]

    req = CellEditRequest(row_uid=row_uid, field_name=field_name, new_value=new_value)
    edit_svc.edit_cell("edit-test-session", "TestSheet", req)

    # Retrieve sheet and verify the edit is reflected
    sheet_response = workbook_svc.get_sheet_data("edit-test-session", "TestSheet")
    assert sheet_response.rows[0][field_name] == new_value


# ---------------------------------------------------------------------------
# Property 6: standardization updates session status and dataset
# Validates: Requirements 5.4
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=10, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_standardization_updates_session_status_and_dataset(workbook_data):
    """Property 6: standardization updates session status and dataset.

    Validates: Requirements 5.4
    """
    from webapp.services.standardization_service import standardizationService

    file_bytes, _ = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, session_svc = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        norm_svc = standardizationService(session_svc)
        norm_response = norm_svc.normalize(response.session_id)

        record = session_svc.get(response.session_id)
        assert record.status == "standardized"
        assert record.workbook_dataset is not None
        assert norm_response.status == "standardized"


# ---------------------------------------------------------------------------
# Property 7: Source file is never modified
# Validates: Requirements 5.8, 7.5, 9.2
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=10, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_source_file_never_modified(workbook_data):
    """Property 7: Source file is never modified.

    Validates: Requirements 5.8, 7.5, 9.2
    """
    import hashlib
    from webapp.services.standardization_service import standardizationService
    from webapp.services.export_service import ExportService

    file_bytes, _ = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, session_svc = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        # Record hash of source file after upload
        uploads_dir = tmp_path / "uploads"
        source_files = list(uploads_dir.glob(f"{response.session_id}*"))
        assert len(source_files) == 1
        original_hash = hashlib.sha256(source_files[0].read_bytes()).hexdigest()

        # Run standardization
        norm_svc = standardizationService(session_svc)
        norm_svc.normalize(response.session_id)

        # Source file must still be identical
        assert hashlib.sha256(source_files[0].read_bytes()).hexdigest() == original_hash

        # Run export
        output_dir = tmp_path / "output"
        export_svc = ExportService(session_svc, output_dir)
        try:
            export_svc.export(response.session_id)
        except Exception:
            pass  # Export may fail if no matching sheets; source file still must be unchanged

        # Source file must still be identical after export
        assert hashlib.sha256(source_files[0].read_bytes()).hexdigest() == original_hash


# ---------------------------------------------------------------------------
# Property 9: Web app standardization equivalence
# Validates: Requirements 12.5
# ---------------------------------------------------------------------------

@given(workbook_strategy())
@settings(max_examples=5, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_standardization_equivalence(workbook_data):
    """Property 9: Web app standardization equivalence.

    Validates: Requirements 12.5
    """
    from webapp.services.standardization_service import standardizationService
    from src.excel_standardization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
    from src.excel_standardization.io_layer.excel_reader import ExcelReader
    from src.excel_standardization.processing.standardization_pipeline import standardizationPipeline
    from src.excel_standardization.engines.name_engine import NameEngine
    from src.excel_standardization.engines.gender_engine import GenderEngine
    from src.excel_standardization.engines.date_engine import DateEngine
    from src.excel_standardization.engines.identifier_engine import IdentifierEngine
    from src.excel_standardization.engines.text_processor import TextProcessor

    file_bytes, _ = workbook_data
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        upload_svc, session_svc = _make_upload_service(tmp_path)
        response = upload_svc.handle_upload("test.xlsx", file_bytes)

        # Run via standardizationService
        norm_svc = standardizationService(session_svc)
        norm_svc.normalize(response.session_id)
        record = session_svc.get(response.session_id)
        webapp_sheets = {s.sheet_name: s.rows for s in record.workbook_dataset.sheets}

        # Run directly via pipeline
        record2 = session_svc.get(response.session_id)
        extractor = ExcelToJsonExtractor(
            excel_reader=ExcelReader(),
            skip_empty_rows=False,
            handle_formulas=True,
            preserve_types=True,
        )
        # Re-upload to get a fresh working copy
        upload_svc2, session_svc2 = _make_upload_service(tmp_path)
        response2 = upload_svc2.handle_upload("test2.xlsx", file_bytes)
        record2 = session_svc2.get(response2.session_id)

        pipeline = standardizationPipeline(
            name_engine=NameEngine(TextProcessor()),
            gender_engine=GenderEngine(),
            date_engine=DateEngine(),
            identifier_engine=IdentifierEngine(),
        )
        direct_wb = extractor.extract_workbook_to_json(record2.working_copy_path)
        direct_sheets = {}
        for sheet in direct_wb.sheets:
            normalized = pipeline.normalize_dataset(sheet)
            direct_sheets[sheet.sheet_name] = normalized.rows

        # Both paths should produce the same normalized rows
        for sheet_name in webapp_sheets:
            if sheet_name in direct_sheets:
                assert webapp_sheets[sheet_name] == direct_sheets[sheet_name], (
                    f"standardization mismatch for sheet '{sheet_name}'"
                )


# ---------------------------------------------------------------------------
# Property 4: Non-existent session always returns 404
# Validates: Requirements 2.3
# ---------------------------------------------------------------------------

@given(st.uuids())
@settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture])
def test_nonexistent_session_returns_404(session_uuid):
    """Property 4: Non-existent session always returns 404.

    Validates: Requirements 2.3
    """
    import tempfile
    from pathlib import Path
    from fastapi.testclient import TestClient
    import webapp.dependencies as deps

    session_id = str(session_uuid)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        session_svc = SessionService()
        session_svc.clear_all()

        from webapp.services.upload_service import UploadService
        from webapp.services.workbook_service import WorkbookService
        from webapp.services.standardization_service import standardizationService
        from webapp.services.edit_service import EditService
        from webapp.services.export_service import ExportService

        upload_svc = UploadService(session_svc, tmp_path / "uploads", tmp_path / "work")
        workbook_svc = WorkbookService(session_svc)
        norm_svc = standardizationService(session_svc)
        edit_svc = EditService(session_svc)
        export_svc = ExportService(session_svc, tmp_path / "output")

        # Temporarily patch the module-level service instances
        orig_session = deps._session_service
        orig_upload = deps._upload_service
        orig_workbook = deps._workbook_service
        orig_norm = deps._standardization_service
        orig_edit = deps._edit_service
        orig_export = deps._export_service

        deps._session_service = session_svc
        deps._upload_service = upload_svc
        deps._workbook_service = workbook_svc
        deps._standardization_service = norm_svc
        deps._edit_service = edit_svc
        deps._export_service = export_svc

        try:
            from webapp.app import app
            with TestClient(app, raise_server_exceptions=False) as client:
                endpoints = [
                    ("GET", f"/api/workbook/{session_id}/summary"),
                    ("GET", f"/api/workbook/{session_id}/sheet/Sheet1"),
                    ("POST", f"/api/workbook/{session_id}/normalize"),
                    ("POST", f"/api/workbook/{session_id}/export"),
                ]
                for method, url in endpoints:
                    if method == "GET":
                        resp = client.get(url)
                    else:
                        resp = client.post(url)
                    assert resp.status_code == 404, (
                        f"Expected 404 for {method} {url}, got {resp.status_code}"
                    )
        finally:
            deps._session_service = orig_session
            deps._upload_service = orig_upload
            deps._workbook_service = orig_workbook
            deps._standardization_service = orig_norm
            deps._edit_service = orig_edit
            deps._export_service = orig_export
