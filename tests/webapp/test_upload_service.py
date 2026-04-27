"""Unit tests for UploadService."""

import io
import pytest
from pathlib import Path
from openpyxl import Workbook
from fastapi import HTTPException

from webapp.services.session_service import SessionService
from webapp.services.upload_service import UploadService


def make_xlsx_bytes(sheet_names=None) -> bytes:
    """Create a minimal valid .xlsx file in memory."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in (sheet_names or ["Sheet1"]):
        ws = wb.create_sheet(name)
        ws.append(["first_name", "last_name"])
        ws.append(["Alice", "Smith"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def tmp_dirs(tmp_path):
    uploads = tmp_path / "uploads"
    work = tmp_path / "work"
    return uploads, work


@pytest.fixture
def upload_svc(tmp_dirs):
    uploads, work = tmp_dirs
    svc = SessionService()
    svc.clear_all()
    return UploadService(svc, uploads, work), svc


def test_valid_xlsx_upload_creates_session_and_returns_sheet_names(upload_svc):
    svc, session_svc = upload_svc
    file_bytes = make_xlsx_bytes(["Sheet1", "Sheet2"])
    response = svc.handle_upload("data.xlsx", file_bytes)

    assert response.session_id
    assert set(response.sheet_names) == {"Sheet1", "Sheet2"}

    # Session should be stored
    record = session_svc.get(response.session_id)
    assert record.status == "uploaded"
    assert record.original_filename == "data.xlsx"
    # workbook_dataset is lazily loaded on first sheet access
    assert record.workbook_dataset is None


def test_xlsm_extension_is_accepted(upload_svc, tmp_dirs):
    svc, _ = upload_svc
    # Create a minimal xlsm-like file (openpyxl can write .xlsm)
    wb = Workbook()
    ws = wb.active
    ws.append(["first_name"])
    ws.append(["Alice"])
    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    # Should not raise
    response = svc.handle_upload("data.xlsm", file_bytes)
    assert response.session_id


def test_csv_extension_raises_400(upload_svc):
    svc, _ = upload_svc
    with pytest.raises(HTTPException) as exc_info:
        svc.handle_upload("data.csv", b"col1,col2\nval1,val2")
    assert exc_info.value.status_code == 400
    assert ".xlsx" in exc_info.value.detail


def test_txt_extension_raises_400(upload_svc):
    svc, _ = upload_svc
    with pytest.raises(HTTPException) as exc_info:
        svc.handle_upload("data.txt", b"some text")
    assert exc_info.value.status_code == 400


def test_corrupted_bytes_raises_422(upload_svc):
    svc, _ = upload_svc
    with pytest.raises(HTTPException) as exc_info:
        svc.handle_upload("data.xlsx", b"this is not a valid xlsx file at all")
    assert exc_info.value.status_code == 422


def test_source_file_is_byte_identical_to_upload(upload_svc, tmp_dirs):
    svc, _ = upload_svc
    uploads, _ = tmp_dirs
    file_bytes = make_xlsx_bytes(["MySheet"])
    response = svc.handle_upload("test.xlsx", file_bytes)

    source_files = list(uploads.glob(f"{response.session_id}*"))
    assert len(source_files) == 1
    assert source_files[0].read_bytes() == file_bytes


def test_working_copy_is_separate_from_source(upload_svc, tmp_dirs):
    svc, _ = upload_svc
    uploads, work = tmp_dirs
    file_bytes = make_xlsx_bytes()
    response = svc.handle_upload("test.xlsx", file_bytes)

    source_files = list(uploads.glob(f"{response.session_id}*"))
    work_files = list(work.glob(f"{response.session_id}*"))
    assert len(source_files) == 1
    assert len(work_files) == 1
    # They should be different file objects (separate copies)
    assert source_files[0] != work_files[0]
