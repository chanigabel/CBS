"""Unit tests for ExportService."""

import pytest
from pathlib import Path
from fastapi import HTTPException
from unittest.mock import patch
from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from webapp.models.session import SessionRecord
from webapp.services.export_service import ExportService
from webapp.services.session_service import SessionService


def make_session_with_workbook(session_id="export-session"):
    svc = SessionService()
    svc.clear_all()
    # Use the actual VBA sheet names that ExportEngine expects
    sheet = SheetDataset(
        sheet_name="דיירים יחידים",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name"],
        rows=[
            {"first_name": "Alice", "last_name": "Smith",
             "first_name_corrected": "Alice", "last_name_corrected": "Smith"},
        ],
    )
    wb = WorkbookDataset(source_file="test.xlsx", sheets=[sheet])
    record = SessionRecord(
        session_id=session_id,
        source_file_path="uploads/export-session.xlsx",
        working_copy_path="work/export-session.xlsx",
        original_filename="test.xlsx",
        status="standardized",
        workbook_dataset=wb,
    )
    svc.create(record)
    return svc, record


@pytest.fixture(autouse=True)
def clear_registry():
    svc = SessionService()
    svc.clear_all()
    yield
    svc.clear_all()


def test_successful_export_returns_path_with_normalized_suffix(tmp_path):
    svc, _ = make_session_with_workbook()
    export_svc = ExportService(svc, tmp_path / "output")
    output_path = export_svc.export("export-session")
    assert "_standardized_" in output_path.name
    assert output_path.suffix == ".xlsx"
    assert output_path.exists()


def test_export_failure_raises_500_and_preserves_session(tmp_path):
    svc, record = make_session_with_workbook()
    original_dataset = record.workbook_dataset
    export_svc = ExportService(svc, tmp_path / "output")

    with patch(
        "webapp.services.export_service.Workbook",
        side_effect=RuntimeError("disk full"),
    ):
        with pytest.raises(HTTPException) as exc_info:
            export_svc.export("export-session")
        assert exc_info.value.status_code == 500

    # Session state must be preserved
    record_after = svc.get("export-session")
    assert record_after.workbook_dataset is original_dataset
    assert not list((tmp_path / "output").glob("*.xlsx"))


def test_export_cleans_up_partial_temp_file_on_failure(tmp_path, monkeypatch):
    svc, _ = make_session_with_workbook("temp-failure-session")
    export_svc = ExportService(svc, tmp_path / "output")

    def fake_write_export_workbook(record, output_path, workbook_factory=None):
        output_path.write_bytes(b"partial")
        raise RuntimeError("save failed")

    monkeypatch.setattr("webapp.services.export_service.write_export_workbook", fake_write_export_workbook)

    with pytest.raises(HTTPException) as exc_info:
        export_svc.export("temp-failure-session")
    assert exc_info.value.status_code == 500
    assert not list((tmp_path / "output").glob("*.tmp"))
    assert not list((tmp_path / "output").glob("*.xlsx"))


def test_export_raises_404_for_unknown_session(tmp_path):
    svc = SessionService()
    export_svc = ExportService(svc, tmp_path / "output")
    with pytest.raises(HTTPException) as exc_info:
        export_svc.export("ghost-session")
    assert exc_info.value.status_code == 404


def test_export_raises_500_when_no_workbook_dataset(tmp_path):
    svc = SessionService()
    record = SessionRecord(
        session_id="no-wb-session",
        source_file_path="uploads/no-wb.xlsx",
        working_copy_path="work/no-wb.xlsx",
        original_filename="test.xlsx",
        status="uploaded",
        workbook_dataset=None,
    )
    svc.create(record)
    export_svc = ExportService(svc, tmp_path / "output")
    with pytest.raises(HTTPException) as exc_info:
        export_svc.export("no-wb-session")
    assert exc_info.value.status_code == 500


def test_export_keeps_moved_passport_value_without_source_passport_column(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["id_number", "passport_corrected"],
        rows=[
            {
                "id_number": "ABC123",
                "id_number_corrected": "",
                "passport_corrected": "ABC123",
                "identifier_status": "moved",
            }
        ],
    )
    record = SessionRecord(
        session_id="passport-export-session",
        source_file_path="uploads/passport-export-session.xlsx",
        working_copy_path="work/passport-export-session.xlsx",
        original_filename="test.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="test.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("passport-export-session")

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    darkon_col = headers.index("Darkon") + 1
    assert ws.cell(row=2, column=darkon_col).value == "ABC123"
    wb.close()


def test_export_lazy_loads_xls_with_xls_reader(tmp_path, monkeypatch):
    svc = SessionService()
    working_path = tmp_path / "work" / "legacy.xls"
    working_path.parent.mkdir(parents=True, exist_ok=True)
    working_path.write_bytes(b"legacy bytes")
    record = SessionRecord(
        session_id="xls-export-session",
        source_file_path=str(tmp_path / "uploads" / "legacy.xls"),
        working_copy_path=str(working_path),
        original_filename="legacy.xls",
        status="uploaded",
        workbook_dataset=None,
    )
    svc.create(record)

    def fake_extract(path):
        assert path == str(working_path)
        sheet = SheetDataset(
            sheet_name="DayarimYahidim",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Raw", "first_name_corrected": "Corrected"}],
        )
        return WorkbookDataset(source_file=path, sheets=[sheet])

    monkeypatch.setattr(
        "src.excel_standardization.io_layer.xls_reader.extract_xls_to_workbook_dataset",
        fake_extract,
    )

    output_path = ExportService(svc, tmp_path / "output").export("xls-export-session")

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    first_col = headers.index("ShemPrati") + 1
    assert ws.cell(row=2, column=first_col).value == "Corrected"
    wb.close()


def test_export_keeps_numeric_invalid_length_id_out_of_id_and_passport(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "id_number"],
        rows=[
            {
                "first_name": "Visible",
                "first_name_corrected": "Visible",
                "id_number": "1234567890",
                "id_number_corrected": "1234567890",
                "identifier_status": "ת.ז. לא תקינה",
            }
        ],
    )
    record = SessionRecord(
        session_id="identifier-export-session",
        source_file_path="uploads/identifier-export-session.xlsx",
        working_copy_path="work/identifier-export-session.xlsx",
        original_filename="test.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="test.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("identifier-export-session")

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("MisparZehut") + 1
    passport_col = headers.index("Darkon") + 1
    assert ws.cell(row=2, column=id_col).value == "1234567890"
    assert ws.cell(row=2, column=passport_col).value is None
    wb.close()


def test_export_uses_corrected_name_fields_only(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "father_name"],
        rows=[
            {
                "first_name": "Original First",
                "first_name_corrected": "Corrected First",
                "last_name": "Original Last",
                "last_name_corrected": "Corrected Last",
                "father_name": "Original Father",
                "father_name_corrected": "Corrected Father",
            },
            {
                "first_name": "Fallback First",
                "first_name_corrected": "",
                "last_name": "Fallback Last",
                "last_name_corrected": "",
                "father_name": "Fallback Father",
                "father_name_corrected": "",
            },
        ],
    )
    record = SessionRecord(
        session_id="name-export-session",
        source_file_path="uploads/name-export-session.xlsx",
        working_copy_path="work/name-export-session.xlsx",
        original_filename="test.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="test.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("name-export-session")

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    first_col = headers.index("ShemPrati") + 1
    last_col = headers.index("ShemMishpaha") + 1
    father_col = headers.index("ShemHaAv") + 1
    assert ws.cell(row=2, column=first_col).value == "Corrected First"
    assert ws.cell(row=2, column=last_col).value == "Corrected Last"
    assert ws.cell(row=2, column=father_col).value == "Corrected Father"
    assert ws.cell(row=3, column=first_col).value is None
    assert ws.cell(row=3, column=last_col).value is None
    assert ws.cell(row=3, column=father_col).value is None
    wb.close()


def test_export_uses_corrected_gender_field_only(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["gender"],
        rows=[
            {"gender": "Original Female", "gender_corrected": 2},
            {"gender": "Fallback Male", "gender_corrected": ""},
            {"gender": "Missing Corrected"},
        ],
    )
    record = SessionRecord(
        session_id="gender-export-session",
        source_file_path="uploads/gender-export-session.xlsx",
        working_copy_path="work/gender-export-session.xlsx",
        original_filename="test.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="test.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("gender-export-session")

    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    gender_col = headers.index("Min") + 1
    assert ws.cell(row=2, column=gender_col).value == 2
    assert ws.cell(row=3, column=gender_col).value is None
    assert ws.cell(row=4, column=gender_col).value is None
    wb.close()


def test_export_sanitizes_sheet_names_and_cell_values_for_openable_xlsx(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="bad:name/with*chars?and a very very long suffix",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "father_name"],
        rows=[
            {
                "first_name": "Original",
                "first_name_corrected": "=NOT_A_REAL_FORMULA(",
                "last_name": "Original",
                "last_name_corrected": "Bad\x01Name",
                "father_name": "Original",
                "father_name_corrected": {"nested": ["value"]},
            }
        ],
    )
    record = SessionRecord(
        session_id="safe-export-session",
        source_file_path="uploads/safe-export-session.xlsx",
        working_copy_path="work/safe-export-session.xlsx",
        original_filename="safe.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="safe.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("safe-export-session")

    wb = load_workbook(output_path)
    assert len(wb.sheetnames[0]) <= 31
    assert not any(ch in wb.sheetnames[0] for ch in "[]:*?/\\")
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    first_col = headers.index("ShemPrati") + 1
    last_col = headers.index("ShemMishpaha") + 1
    father_col = headers.index("ShemHaAv") + 1
    assert ws.cell(row=2, column=first_col).value == "'=NOT_A_REAL_FORMULA("
    assert ws.cell(row=2, column=last_col).value == "BadName"
    assert ws.cell(row=2, column=father_col).value == '{"nested": ["value"]}'
    wb.close()


def test_export_sanitizes_colliding_sheet_names(tmp_path):
    svc = SessionService()
    sheets = [
        SheetDataset(
            sheet_name="bad:name",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Raw", "first_name_corrected": "One"}],
        ),
        SheetDataset(
            sheet_name="bad/name",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Raw", "first_name_corrected": "Two"}],
        ),
    ]
    record = SessionRecord(
        session_id="sheet-collision-export-session",
        source_file_path="uploads/sheet-collision.xlsx",
        working_copy_path="work/sheet-collision.xlsx",
        original_filename="collision.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="collision.xlsx", sheets=sheets),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export(record.session_id)

    wb = load_workbook(output_path)
    assert "bad_name" in wb.sheetnames
    assert "bad_name_1" in wb.sheetnames
    wb.close()


def test_export_does_not_mutate_dataset_rows_when_injecting_institution_metadata(tmp_path):
    svc, record = make_session_with_workbook("non-mutating-export-session")
    record.mosad_id = "999"
    record.mosad_types = ["123"]
    source_row = record.workbook_dataset.sheets[0].rows[0]
    assert "MosadID" not in source_row
    assert "SugMosad" not in source_row

    ExportService(svc, tmp_path / "output").export("non-mutating-export-session")

    assert "MosadID" not in source_row
    assert "SugMosad" not in source_row


def test_export_logs_unsupported_cell_value_type(tmp_path, caplog):
    class UnsupportedCellValue:
        pass

    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name"],
        rows=[{"first_name": "Raw", "first_name_corrected": UnsupportedCellValue()}],
    )
    record = SessionRecord(
        session_id="unsupported-cell-export-session",
        source_file_path="uploads/unsupported.xlsx",
        working_copy_path="work/unsupported.xlsx",
        original_filename="unsupported.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="unsupported.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    with caplog.at_level("WARNING", logger="src.excel_standardization.export.excel_safe"):
        ExportService(svc, tmp_path / "output").export(record.session_id)

    assert "unsupported_export_cell_value_type" in caplog.text


def test_export_sanitizes_output_filename_special_characters(tmp_path):
    svc = SessionService()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name"],
        rows=[{"first_name": "Raw", "first_name_corrected": "Corrected"}],
    )
    record = SessionRecord(
        session_id="filename-export-session",
        source_file_path="uploads/filename-export-session.xlsx",
        working_copy_path="work/filename-export-session.xlsx",
        original_filename='שם:קובץ?מקור.xlsx',
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="safe.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    output_path = ExportService(svc, tmp_path / "output").export("filename-export-session")

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert not any(ch in output_path.name for ch in '<>:"/\\|?*')
