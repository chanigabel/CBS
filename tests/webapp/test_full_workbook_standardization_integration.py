"""Integration test: full-workbook standardization loads missing sheets.

Scenario:
1. Upload workbook with two sheets: SheetA and SheetB.
2. Load only SheetA so the in-memory Working Dataset contains SheetA only.
3. Edit a value in SheetA to ensure edits are preserved.
4. Call full-workbook standardization (no sheet param).
5. Assert SheetB was added to the Working Dataset and was standardized.
6. Assert SheetA edits were not overwritten.
7. Assert export contains both sheets with standardized data.
"""

from io import BytesIO

from openpyxl import load_workbook

from tests.webapp.conftest import make_xlsx_bytes
from webapp.services.session_service import SessionService


def test_full_workbook_standardization_loads_missing_sheets(client):
    """Verify that a full-workbook normalize loads and standardizes missing sheets."""
    # client fixture yields a TestClient instance patched for tmp dirs
    test_client = client

    # 1) Upload workbook with two sheets
    file_bytes = make_xlsx_bytes(["SheetA", "SheetB"])
    resp = test_client.post(
        "/api/upload",
        files={"file": ("two_sheets.xlsx", file_bytes, "application/octet-stream")},
    )
    assert resp.status_code == 200
    session_id = resp.json()["session_id"]

    # 2) Load only SheetA
    sheet_resp = test_client.get(f"/api/workbook/{session_id}/sheet/SheetA")
    assert sheet_resp.status_code == 200
    sheet_data = sheet_resp.json()
    assert sheet_data["sheet_name"] == "SheetA"
    assert sheet_data["rows"]

    # Inspect session: should only have SheetA loaded in workbook_dataset
    svc = SessionService()
    record = svc.get(session_id)
    assert record.workbook_dataset is not None
    loaded_names = [s.sheet_name for s in record.workbook_dataset.sheets]
    assert loaded_names == ["SheetA"]

    # 3) Edit a cell in SheetA so we can ensure it's preserved
    first_row = sheet_data["rows"][0]
    row_uid = first_row["_row_uid"]
    # edit a known source column present in the test workbook
    field_name = "first_name"

    edit_resp = test_client.patch(
        f"/api/workbook/{session_id}/cell",
        json={
            "sheet_name": "SheetA",
            "row_uid": row_uid,
            "field": field_name,
            "value": "EditedValue",
        },
    )
    assert edit_resp.status_code == 200

    # Confirm in-memory row updated
    record = svc.get(session_id)
    sheet_a = record.workbook_dataset.get_sheet_by_name("SheetA")
    assert sheet_a is not None
    assert any(r.get(field_name) == "EditedValue" for r in sheet_a.rows)

    # 4) Run full-workbook standardization (no sheet param)
    norm_resp = test_client.post(f"/api/workbook/{session_id}/normalize")
    assert norm_resp.status_code == 200
    norm_data = norm_resp.json()
    assert norm_data.get("status") == "standardized"

    # 5) Assert SheetB was added to workbook_dataset
    record = svc.get(session_id)
    names_after = [s.sheet_name for s in record.workbook_dataset.sheets]
    assert "SheetA" in names_after and "SheetB" in names_after

    sheet_b = record.workbook_dataset.get_sheet_by_name("SheetB")
    assert sheet_b is not None

    # 6) Assert SheetB was standardized (corrected field present)
    # Name standardization populates "first_name_corrected" for our sample rows
    assert any("first_name_corrected" in row for row in sheet_b.rows)

    # 7) Assert existing SheetA edits were not overwritten
    sheet_a_after = record.workbook_dataset.get_sheet_by_name("SheetA")
    assert any(r.get(field_name) == "EditedValue" for r in sheet_a_after.rows)

    # 8) Export and verify exported workbook contains both sheets
    export_resp = test_client.post(f"/api/workbook/{session_id}/export")
    assert export_resp.status_code == 200
    wb = load_workbook(BytesIO(export_resp.content))
    assert "SheetA" in wb.sheetnames and "SheetB" in wb.sheetnames
    wb.close()
