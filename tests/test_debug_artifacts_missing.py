"""Bug Condition Exploration Test - Debugging Artifacts Missing

**Validates: Requirements 1.1, 1.2, 1.3, 1.4**

This test verifies that the pipeline exports raw_dataset.json alongside the
output Excel file. The normalized_dataset.json artifact was part of the old
JSON-based pipeline and is no longer produced by the VBA-parity pipeline.
"""

import os
import tempfile
import pytest
from hypothesis import given, strategies as st, settings
from openpyxl import Workbook

from src.excel_normalization.orchestrator import NormalizationOrchestrator


class TestDebugArtifactsMissing:
    """Test that raw_dataset.json debug artifact is exported by the pipeline."""

    def test_debugging_artifacts_missing_simple(self):
        """Verify raw_dataset.json is exported alongside the output Excel file.
        
        **Validates: Requirements 1.1**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create paths
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            raw_json_path = os.path.join(tmpdir, "raw_dataset.json")
            
            # Create test Excel file with sample data
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add headers
            headers = ["שם פרטי", "שם משפחה", "מין", "תאריך לידה"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add sample data
            test_data = [
                ["יוסי", "כהן", "ז", "15/05/1980"],
                ["שרה", "לוי", "נ", "20/12/1985"],
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save input file
            wb.save(input_path)
            
            original_cwd = os.getcwd()
            try:
                os.chdir(tmpdir)
                
                orchestrator = NormalizationOrchestrator()
                orchestrator.process_workbook_json(input_path, output_path)
                
                # Requirement 1.1: raw_dataset.json should exist after extraction
                assert os.path.exists(raw_json_path), (
                    f"raw_dataset.json not found at {raw_json_path}."
                )
                
            finally:
                os.chdir(original_cwd)

    @settings(max_examples=2, deadline=None)
    @given(
        first_name=st.text(min_size=1, max_size=20, alphabet=st.characters(whitelist_categories=('L',))),
        last_name=st.text(min_size=1, max_size=20, alphabet=st.characters(whitelist_categories=('L',))),
        gender=st.sampled_from(["ז", "נ", "1", "2"]),
    )
    def test_debugging_artifacts_missing_property(self, first_name, last_name, gender):
        """Property 1: raw_dataset.json is exported for any valid input.
        
        **Validates: Requirements 1.1**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            raw_json_path = os.path.join(tmpdir, "raw_dataset.json")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            headers = ["שם פרטי", "שם משפחה", "מין"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            ws.cell(row=2, column=1, value=first_name)
            ws.cell(row=2, column=2, value=last_name)
            ws.cell(row=2, column=3, value=gender)
            
            wb.save(input_path)
            
            original_cwd = os.getcwd()
            try:
                os.chdir(tmpdir)
                
                orchestrator = NormalizationOrchestrator()
                orchestrator.process_workbook_json(input_path, output_path)
                
                assert os.path.exists(raw_json_path), (
                    f"raw_dataset.json not found at {raw_json_path}."
                )
                
            finally:
                os.chdir(original_cwd)
