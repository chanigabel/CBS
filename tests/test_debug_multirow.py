"""Debug script to understand multi-row header detection."""

from openpyxl import Workbook
from src.excel_normalization.io_layer.excel_reader import ExcelReader


def create_complex_multirow_workbook():
    """Create a workbook with multiple multi-row header groups."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex Multi-Row"

    # Row 1: Main headers
    ws['A1'] = "Personal Info"
    ws['D1'] = "תאריך לידה"  # Birth Date (Hebrew)
    ws['G1'] = "תאריך כניסה"  # Entry Date (Hebrew)
    
    # Merge cells for parent headers
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:F1')
    ws.merge_cells('G1:I1')
    
    # Row 2: Sub-headers
    ws['A2'] = "First Name"
    ws['B2'] = "Last Name"
    ws['C2'] = "Father Name"
    ws['D2'] = "שנה"  # Year
    ws['E2'] = "חודש"  # Month
    ws['F2'] = "יום"  # Day
    ws['G2'] = "שנה"  # Year
    ws['H2'] = "חודש"  # Month
    ws['I2'] = "יום"  # Day
    
    # Add data
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = "אברהם"
    ws['D3'] = 1990
    ws['E3'] = 5
    ws['F3'] = 15
    ws['G3'] = 2020
    ws['H3'] = 3
    ws['I3'] = 10
    
    return wb


def debug_complex_multirow():
    """Debug complex multi-row headers."""
    wb = create_complex_multirow_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    print("=" * 80)
    print("DEBUG: Complex Multi-Row Headers")
    print("=" * 80)
    
    # Check what's in each cell
    print("\nRow 1 (Parent headers):")
    for col in range(1, 10):
        cell_value = ws.cell(row=1, column=col).value
        print(f"  Column {col}: {cell_value}")
    
    print("\nRow 2 (Sub-headers):")
    for col in range(1, 10):
        cell_value = ws.cell(row=2, column=col).value
        print(f"  Column {col}: {cell_value}")
    
    # Test normalization
    print("\nNormalized text:")
    for col in range(1, 10):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            normalized = reader._normalize_text(str(cell_value))
            print(f"  Column {col}: '{normalized}'")
    
    # Test field matching
    print("\nField matching for row 1:")
    for col in range(1, 10):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            normalized = reader._normalize_text(str(cell_value))
            matched_field = reader._match_field(normalized)
            print(f"  Column {col}: '{normalized}' → {matched_field}")
    
    # Test row scoring
    print("\nRow scoring:")
    for row in range(1, 3):
        score = reader._score_header_row(ws, row_idx=row, max_col=9)
        print(f"  Row {row}: {score}")
    
    # Test table region detection
    print("\nTable region detection:")
    table_region = reader.detect_table_region(ws)
    if table_region:
        print(f"  Start row: {table_region.start_row}")
        print(f"  Header rows: {table_region.header_rows}")
        print(f"  Data start row: {table_region.data_start_row}")
    
    # Test column detection
    print("\nColumn detection:")
    column_mapping = reader.detect_columns(ws)
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"  {field_name:20s} → Column {col_info.col}")


if __name__ == "__main__":
    debug_complex_multirow()
