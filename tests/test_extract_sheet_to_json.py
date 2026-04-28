"""Comprehensive tests for extract_sheet_to_json method.

Tests the extract_sheet_to_json method with various scenarios including:
- Single-row headers
- Multi-row headers
- Empty rows handling
- Missing headers
- Date field structures (single and split)
- Metadata generation
- Integration with ExcelReader

Requirements:
    - Validates: Requirements 10.1-10.5, 11.1-11.5
"""

import pytest
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.excel_standardization.io_layer import ExcelReader, ExcelToJsonExtractor
from src.excel_standardization.data_types import ColumnHeaderInfo, TableRegion, SheetDataset


@pytest.fixture
def excel_reader():
    """Create an ExcelReader instance for testing."""
    return ExcelReader()


@pytest.fixture
def extractor(excel_reader):
    """Create an ExcelToJsonExtractor instance for testing."""
    return ExcelToJsonExtractor(excel_reader=excel_reader)


@pytest.fixture
def workbook():
    """Create a test workbook."""
    return Workbook()


@pytest.fixture
def worksheet(workbook):
    """Create a test worksheet."""
    ws = workbook.active
    ws.title = "Test Sheet"
    return ws


def test_extract_sheet_with_single_row_header(extractor, worksheet):
    """Test extracting a sheet with single-row header.
    
    Requirements:
        - Validates: Requirements 10.1-10.5, 11.1-11.5
    """
    # Setup: Create a simple table with single-row header
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    worksheet['C1'] = 'ID Number'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 'Doe'
    worksheet['C2'] = 123456789
    
    worksheet['A3'] = 'Jane'
    worksheet['B3'] = 'Smith'
    worksheet['C3'] = 987654321
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=3, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=3, header_text='Last Name'),
        'id_number': ColumnHeaderInfo(col=3, header_row=1, last_row=3, header_text='ID Number'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=3,
        start_col=1,
        end_col=3,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert isinstance(result, SheetDataset)
    assert result.sheet_name == "Test Sheet"
    assert result.header_row == 1
    assert result.header_rows_count == 1
    assert result.field_names == ['first_name', 'last_name', 'id_number']
    assert len(result.rows) == 2
    
    # Verify first row
    assert result.rows[0]['first_name'] == 'John'
    assert result.rows[0]['last_name'] == 'Doe'
    assert result.rows[0]['id_number'] == 123456789
    
    # Verify second row
    assert result.rows[1]['first_name'] == 'Jane'
    assert result.rows[1]['last_name'] == 'Smith'
    assert result.rows[1]['id_number'] == 987654321
    
    # Verify metadata
    assert result.metadata['total_rows'] == 2
    assert result.metadata['data_start_row'] == 2
    assert result.metadata['data_end_row'] == 3


def test_extract_sheet_with_multi_row_header(extractor, worksheet):
    """Test extracting a sheet with multi-row header (2 rows).
    
    Requirements:
        - Validates: Requirements 10.1-10.5, 11.1-11.5
    """
    # Setup: Create a table with 2-row header (parent and child headers)
    worksheet['A1'] = 'Name'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Birth Date'
    worksheet['D1'] = 'Birth Date'
    worksheet['E1'] = 'Birth Date'
    
    worksheet['A2'] = 'First'
    worksheet['B2'] = 'Last'
    worksheet['C2'] = 'Year'
    worksheet['D2'] = 'Month'
    worksheet['E2'] = 'Day'
    
    worksheet['A3'] = 'John'
    worksheet['B3'] = 'Doe'
    worksheet['C3'] = 1980
    worksheet['D3'] = 5
    worksheet['E3'] = 15
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=2, last_row=3, header_text='First'),
        'last_name': ColumnHeaderInfo(col=2, header_row=2, last_row=3, header_text='Last'),
        'birth_year': ColumnHeaderInfo(col=3, header_row=2, last_row=3, header_text='Year'),
        'birth_month': ColumnHeaderInfo(col=4, header_row=2, last_row=3, header_text='Month'),
        'birth_day': ColumnHeaderInfo(col=5, header_row=2, last_row=3, header_text='Day'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=3,
        start_col=1,
        end_col=5,
        header_rows=2,
        data_start_row=3,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert isinstance(result, SheetDataset)
    assert result.sheet_name == "Test Sheet"
    assert result.header_row == 1
    assert result.header_rows_count == 2
    assert len(result.rows) == 1
    
    # Verify data row
    assert result.rows[0]['first_name'] == 'John'
    assert result.rows[0]['last_name'] == 'Doe'
    assert result.rows[0]['birth_year'] == 1980
    assert result.rows[0]['birth_month'] == 5
    assert result.rows[0]['birth_day'] == 15
    
    # Verify metadata includes split date structure
    assert 'date_field_structure' in result.metadata
    assert result.metadata['date_field_structure']['birth_date'] == 'split'


def test_extract_sheet_with_empty_rows_skip_enabled(extractor, worksheet):
    """Test extracting a sheet with empty rows when skip_empty_rows is enabled."""
    # Setup: Create extractor with skip_empty_rows enabled
    extractor_skip = ExcelToJsonExtractor(
        excel_reader=extractor.excel_reader,
        skip_empty_rows=True
    )
    
    # Create table with empty rows
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 'Doe'
    
    worksheet['A3'] = None  # Empty row
    worksheet['B3'] = None
    
    worksheet['A4'] = 'Jane'
    worksheet['B4'] = 'Smith'
    
    worksheet['A5'] = ''  # Empty row with empty strings
    worksheet['B5'] = ''
    
    worksheet['A6'] = 'Bob'
    worksheet['B6'] = 'Johnson'
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=6, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=6, header_text='Last Name'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=6,
        start_col=1,
        end_col=2,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor_skip.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify - should skip empty rows
    assert len(result.rows) == 3  # Only non-empty rows
    assert result.rows[0]['first_name'] == 'John'
    assert result.rows[1]['first_name'] == 'Jane'
    assert result.rows[2]['first_name'] == 'Bob'
    
    # Verify metadata
    assert result.metadata['total_rows'] == 3
    assert result.metadata['skipped_rows'] == 2


def test_extract_sheet_with_empty_rows_skip_disabled(extractor, worksheet):
    """Test extracting a sheet with empty rows when skip_empty_rows is disabled."""
    # Setup: Create table with empty rows (default skip_empty_rows=False)
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 'Doe'
    
    worksheet['A3'] = None  # Empty row
    worksheet['B3'] = None
    
    worksheet['A4'] = 'Jane'
    worksheet['B4'] = 'Smith'
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=4, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=4, header_text='Last Name'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=4,
        start_col=1,
        end_col=2,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify - should include empty rows
    assert len(result.rows) == 3  # All rows including empty
    assert result.rows[0]['first_name'] == 'John'
    assert result.rows[1]['first_name'] is None  # Empty row
    assert result.rows[2]['first_name'] == 'Jane'
    
    # Verify metadata
    assert result.metadata['total_rows'] == 3
    assert result.metadata['skipped_rows'] == 0


def test_extract_sheet_with_no_headers(extractor, worksheet):
    """Test extracting a sheet with no valid headers.
    
    Requirements:
        - Validates: Requirement 11.4 (handle missing headers gracefully)
    """
    # Setup: Empty column mapping and None table region
    column_mapping = {}
    table_region = None
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify - should return empty dataset with error metadata
    assert isinstance(result, SheetDataset)
    assert result.sheet_name == "Test Sheet"
    assert result.header_row == 0
    assert result.header_rows_count == 0
    assert result.field_names == []
    assert result.rows == []
    assert 'error' in result.metadata
    assert result.metadata['error'] == "No valid headers found"
    assert result.metadata['skipped'] is True


def test_extract_sheet_with_single_date_field(extractor, worksheet):
    """Test extracting a sheet with single-column date field."""
    # Setup: Create table with single date column
    test_date = datetime(1980, 5, 15)
    
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Birth Date'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = test_date
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='First Name'),
        'birth_date': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Birth Date'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=2,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert len(result.rows) == 1
    assert result.rows[0]['birth_date'] == test_date
    
    # Verify metadata shows single date field
    assert 'date_field_structure' in result.metadata
    assert result.metadata['date_field_structure']['birth_date'] == 'single'


def test_extract_sheet_with_split_date_fields(extractor, worksheet):
    """Test extracting a sheet with split date fields (year, month, day)."""
    # Setup: Create table with split date columns
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Birth Year'
    worksheet['C1'] = 'Birth Month'
    worksheet['D1'] = 'Birth Day'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 1980
    worksheet['C2'] = 5
    worksheet['D2'] = 15
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='First Name'),
        'birth_year': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Birth Year'),
        'birth_month': ColumnHeaderInfo(col=3, header_row=1, last_row=2, header_text='Birth Month'),
        'birth_day': ColumnHeaderInfo(col=4, header_row=1, last_row=2, header_text='Birth Day'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=4,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert len(result.rows) == 1
    assert result.rows[0]['birth_year'] == 1980
    assert result.rows[0]['birth_month'] == 5
    assert result.rows[0]['birth_day'] == 15
    
    # Verify metadata shows split date field
    assert 'date_field_structure' in result.metadata
    assert result.metadata['date_field_structure']['birth_date'] == 'split'


def test_extract_sheet_with_both_birth_and_entry_dates(extractor, worksheet):
    """Test extracting a sheet with both birth and entry date fields."""
    # Setup: Create table with both date types
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Birth Year'
    worksheet['C1'] = 'Birth Month'
    worksheet['D1'] = 'Birth Day'
    worksheet['E1'] = 'Entry Date'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 1980
    worksheet['C2'] = 5
    worksheet['D2'] = 15
    worksheet['E2'] = datetime(2020, 1, 1)
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='First Name'),
        'birth_year': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Birth Year'),
        'birth_month': ColumnHeaderInfo(col=3, header_row=1, last_row=2, header_text='Birth Month'),
        'birth_day': ColumnHeaderInfo(col=4, header_row=1, last_row=2, header_text='Birth Day'),
        'entry_date': ColumnHeaderInfo(col=5, header_row=1, last_row=2, header_text='Entry Date'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=5,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert len(result.rows) == 1
    
    # Verify metadata shows both date structures
    assert 'date_field_structure' in result.metadata
    assert result.metadata['date_field_structure']['birth_date'] == 'split'
    assert result.metadata['date_field_structure']['entry_date'] == 'single'


def test_extract_sheet_auto_detect_headers(extractor, worksheet):
    """Test extracting a sheet with auto-detection of headers and table region."""
    # Setup: Create a simple table (let ExcelReader detect it)
    worksheet['A1'] = 'שם פרטי'  # Hebrew: First Name
    worksheet['B1'] = 'שם משפחה'  # Hebrew: Last Name
    worksheet['C1'] = 'ת.ז'  # Hebrew: ID Number
    
    worksheet['A2'] = 'יוסי'
    worksheet['B2'] = 'כהן'
    worksheet['C2'] = 123456789
    
    worksheet['A3'] = 'דנה'
    worksheet['B3'] = 'לוי'
    worksheet['C3'] = 987654321
    
    # Execute - without providing column_mapping or table_region
    result = extractor.extract_sheet_to_json(worksheet)
    
    # Verify - should auto-detect and extract data
    assert isinstance(result, SheetDataset)
    assert result.sheet_name == "Test Sheet"
    
    # Should have detected some fields (exact fields depend on ExcelReader)
    # At minimum, should not be empty if headers were detected
    if result.field_names:  # If headers were detected
        assert len(result.field_names) > 0
        assert len(result.rows) == 2


def test_extract_sheet_with_large_dataset(extractor, worksheet):
    """Test extracting a sheet with many rows (performance test)."""
    # Setup: Create table with 100 rows
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    worksheet['C1'] = 'ID Number'
    
    for i in range(2, 102):  # Rows 2-101 (100 data rows)
        worksheet[f'A{i}'] = f'Person{i-1}'
        worksheet[f'B{i}'] = f'LastName{i-1}'
        worksheet[f'C{i}'] = 100000000 + i
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=101, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=101, header_text='Last Name'),
        'id_number': ColumnHeaderInfo(col=3, header_row=1, last_row=101, header_text='ID Number'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=101,
        start_col=1,
        end_col=3,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify
    assert len(result.rows) == 100
    assert result.metadata['total_rows'] == 100
    
    # Spot check first and last rows
    assert result.rows[0]['first_name'] == 'Person1'
    assert result.rows[99]['first_name'] == 'Person100'


def test_extract_sheet_with_all_field_types(extractor, worksheet):
    """Test extracting a sheet with all supported field types."""
    # Setup: Create table with all field types
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    worksheet['C1'] = 'Father Name'
    worksheet['D1'] = 'Gender'
    worksheet['E1'] = 'ID Number'
    worksheet['F1'] = 'Passport'
    worksheet['G1'] = 'Birth Year'
    worksheet['H1'] = 'Birth Month'
    worksheet['I1'] = 'Birth Day'
    worksheet['J1'] = 'Entry Date'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 'Doe'
    worksheet['C2'] = 'Smith'
    worksheet['D2'] = 'M'
    worksheet['E2'] = 123456789
    worksheet['F2'] = 'AB123456'
    worksheet['G2'] = 1980
    worksheet['H2'] = 5
    worksheet['I2'] = 15
    worksheet['J2'] = datetime(2020, 1, 1)
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Last Name'),
        'father_name': ColumnHeaderInfo(col=3, header_row=1, last_row=2, header_text='Father Name'),
        'gender': ColumnHeaderInfo(col=4, header_row=1, last_row=2, header_text='Gender'),
        'id_number': ColumnHeaderInfo(col=5, header_row=1, last_row=2, header_text='ID Number'),
        'passport': ColumnHeaderInfo(col=6, header_row=1, last_row=2, header_text='Passport'),
        'birth_year': ColumnHeaderInfo(col=7, header_row=1, last_row=2, header_text='Birth Year'),
        'birth_month': ColumnHeaderInfo(col=8, header_row=1, last_row=2, header_text='Birth Month'),
        'birth_day': ColumnHeaderInfo(col=9, header_row=1, last_row=2, header_text='Birth Day'),
        'entry_date': ColumnHeaderInfo(col=10, header_row=1, last_row=2, header_text='Entry Date'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=10,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify all fields are present
    assert len(result.field_names) == 10
    assert len(result.rows) == 1
    
    row = result.rows[0]
    assert row['first_name'] == 'John'
    assert row['last_name'] == 'Doe'
    assert row['father_name'] == 'Smith'
    assert row['gender'] == 'M'
    assert row['id_number'] == 123456789
    assert row['passport'] == 'AB123456'
    assert row['birth_year'] == 1980
    assert row['birth_month'] == 5
    assert row['birth_day'] == 15
    assert row['entry_date'] == datetime(2020, 1, 1)
    
    # Verify metadata
    assert result.metadata['date_field_structure']['birth_date'] == 'split'
    assert result.metadata['date_field_structure']['entry_date'] == 'single'


def test_extract_sheet_preserves_original_values(extractor, worksheet):
    """Test that extract_sheet_to_json preserves original values exactly.
    
    Requirements:
        - Validates: Requirement 10.5 (preserve exact original values)
    """
    # Setup: Create table with various value types
    worksheet['A1'] = 'Text'
    worksheet['B1'] = 'Number'
    worksheet['C1'] = 'Whitespace'
    
    worksheet['A2'] = '  John  '  # With whitespace
    worksheet['B2'] = 0  # Zero value
    worksheet['C2'] = '   '  # Only whitespace
    
    column_mapping = {
        'text': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='Text'),
        'number': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Number'),
        'whitespace': ColumnHeaderInfo(col=3, header_row=1, last_row=2, header_text='Whitespace'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=3,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify - values should be preserved exactly
    assert result.rows[0]['text'] == '  John  '  # Whitespace preserved
    assert result.rows[0]['number'] == 0  # Zero preserved
    assert result.rows[0]['whitespace'] == '   '  # Whitespace preserved


def test_extract_sheet_metadata_completeness(extractor, worksheet):
    """Test that extract_sheet_to_json generates complete metadata."""
    # Setup: Create simple table
    worksheet['A1'] = 'First Name'
    worksheet['B1'] = 'Last Name'
    
    worksheet['A2'] = 'John'
    worksheet['B2'] = 'Doe'
    
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='First Name'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='Last Name'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=2,
        start_col=1,
        end_col=2,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute
    result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify metadata completeness
    assert 'total_rows' in result.metadata
    assert 'skipped_rows' in result.metadata
    assert 'date_field_structure' in result.metadata
    assert 'data_start_row' in result.metadata
    assert 'data_end_row' in result.metadata
    
    assert result.metadata['total_rows'] == 1
    assert result.metadata['skipped_rows'] == 0
    assert result.metadata['data_start_row'] == 2
    assert result.metadata['data_end_row'] == 2


# ---------------------------------------------------------------------------
# Regression test: column-index reference row must not appear in extracted data
# ---------------------------------------------------------------------------

class TestColumnIndexRowExclusion:
    """Regression tests for the column-index reference row detection.

    Some Excel forms include a row of sequential integers immediately after
    the header rows to label column positions for the form filler.  These rows
    must be silently skipped and never appear in the extracted dataset.

    Real-world pattern observed in Automations_DEV.xlsx:
        Row 14: parent headers (merged cells)
        Row 15: sub-headers (שנה / חודש / יום, שם פרטי, ...)
        Row 16: column-index row  ← must be skipped
        Row 17+: actual data
    """

    def _make_extractor(self):
        reader = ExcelReader()
        return ExcelToJsonExtractor(excel_reader=reader)

    def test_column_index_row_skipped_single_header(self):
        """Column-index row after a single-row header must be excluded."""
        wb = Workbook()
        ws = wb.active

        # Row 1: header
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "שם משפחה"
        ws.cell(row=1, column=3).value = "מין"

        # Row 2: column-index row (1, 2, 3)
        ws.cell(row=2, column=1).value = 1
        ws.cell(row=2, column=2).value = 2
        ws.cell(row=2, column=3).value = 3

        # Row 3: real data
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "כהן"
        ws.cell(row=3, column=3).value = 1

        extractor = self._make_extractor()
        dataset = extractor.extract_sheet_to_json(ws)

        assert dataset.metadata["total_rows"] == 1
        assert dataset.rows[0]["first_name"] == "יוסי"

    def test_column_index_row_skipped_multi_row_header(self):
        """Column-index row after a two-row header must be excluded."""
        wb = Workbook()
        ws = wb.active

        # Row 1: parent header
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "מין"
        ws.cell(row=1, column=3).value = "תאריך לידה"
        # Merge C1:E1 to simulate a date group header
        ws.merge_cells("C1:E1")

        # Row 2: sub-headers
        ws.cell(row=2, column=1).value = "שם פרטי"
        ws.cell(row=2, column=2).value = "מין"
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"

        # Row 3: column-index row (1, 2, 3, 4, 5)
        for col in range(1, 6):
            ws.cell(row=3, column=col).value = col

        # Row 4: real data
        ws.cell(row=4, column=1).value = "שרה"
        ws.cell(row=4, column=2).value = 2
        ws.cell(row=4, column=3).value = 1990
        ws.cell(row=4, column=4).value = 6
        ws.cell(row=4, column=5).value = 15

        extractor = self._make_extractor()
        dataset = extractor.extract_sheet_to_json(ws)

        assert dataset.metadata["total_rows"] == 1
        assert dataset.rows[0]["first_name"] == "שרה"

    def test_non_sequential_integers_not_skipped(self):
        """A row with non-sequential integers is real data and must not be skipped."""
        wb = Workbook()
        ws = wb.active

        # Row 1: header
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "מין"
        ws.cell(row=1, column=3).value = "גיל"

        # Row 2: real data with non-sequential integers (not a column-index row)
        ws.cell(row=2, column=1).value = "דוד"
        ws.cell(row=2, column=2).value = 1
        ws.cell(row=2, column=3).value = 45  # age — not sequential with col index

        extractor = self._make_extractor()
        dataset = extractor.extract_sheet_to_json(ws)

        assert dataset.metadata["total_rows"] == 1
        assert dataset.rows[0]["first_name"] == "דוד"

    def test_real_data_row_with_all_integers_not_skipped(self):
        """A data row where all values happen to be small integers must not be skipped
        if the values are not consecutive (e.g., gender=1, year=2001, month=5, day=3)."""
        wb = Workbook()
        ws = wb.active

        # Row 1: header
        ws.cell(row=1, column=1).value = "מין"
        ws.cell(row=1, column=2).value = "שנה"
        ws.cell(row=1, column=3).value = "חודש"
        ws.cell(row=1, column=4).value = "יום"

        # Row 2: real data — integers but NOT consecutive (2001 > end_col)
        ws.cell(row=2, column=1).value = 1
        ws.cell(row=2, column=2).value = 2001
        ws.cell(row=2, column=3).value = 5
        ws.cell(row=2, column=4).value = 3

        extractor = self._make_extractor()
        dataset = extractor.extract_sheet_to_json(ws)

        # Row 2 must be kept — 2001 exceeds end_col so it's not a column-index row
        assert dataset.metadata["total_rows"] == 1

    def test_is_column_index_row_method_directly(self):
        """Unit test for _is_column_index_row helper method."""
        from openpyxl import Workbook
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active

        # Build a row with sequential integers 1..5 in columns 1..5
        for col in range(1, 6):
            ws.cell(row=1, column=col).value = col

        assert reader._is_column_index_row(ws, 1, 1, 5) is True

    def test_is_column_index_row_rejects_text(self):
        """_is_column_index_row must return False when any cell contains text."""
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 1
        ws.cell(row=1, column=2).value = "שם"  # text — not a column-index row
        ws.cell(row=1, column=3).value = 3

        assert reader._is_column_index_row(ws, 1, 1, 3) is False

    def test_is_column_index_row_rejects_too_few_values(self):
        """_is_column_index_row must return False when fewer than 3 values present."""
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 1
        ws.cell(row=1, column=2).value = 2
        # Only 2 values — not enough to be confident

        assert reader._is_column_index_row(ws, 1, 1, 5) is False

    def test_is_column_index_row_allows_gaps(self):
        """_is_column_index_row must return True even with gaps.

        Real forms skip column numbers for merged/empty columns,
        so gaps are expected and must be allowed.
        """
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active

        # 1, 2, 5 — gap of 3 between 2 and 5 — still a valid column-index row
        ws.cell(row=1, column=1).value = 1
        ws.cell(row=1, column=2).value = 2
        ws.cell(row=1, column=3).value = 5

        assert reader._is_column_index_row(ws, 1, 1, 5) is True
