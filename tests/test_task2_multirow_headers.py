"""Test script for Task 2: Multi-row header detection enhancements.

This script tests the enhanced _score_subheader_row and _detect_date_subcolumns
methods to verify they properly detect parent-child relationships in multi-row
headers with merged cells.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from src.excel_standardization.io_layer.excel_reader import ExcelReader


def create_multirow_header_workbook():
    """Create a workbook with multi-row headers and merged parent cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Row Headers"

    # Row 1: Main header
    ws['A1'] = "Census Data"
    
    # Row 2: Column headers with merged parent for date
    ws['A2'] = "ID"
    ws['B2'] = "First Name"
    ws['C2'] = "Last Name"
    ws['D2'] = "תאריך לידה"  # Birth Date - parent header (merged)
    ws['G2'] = "Gender"
    
    # Merge cells D2:F2 to create parent header spanning 3 columns
    ws.merge_cells('D2:F2')
    
    # Row 3: Sub-headers for date columns
    ws['D3'] = "שנה"  # Year
    ws['E3'] = "חודש"  # Month
    ws['F3'] = "יום"  # Day
    
    # Add data rows
    ws['A4'] = 1
    ws['B4'] = "יוסי"
    ws['C4'] = "כהן"
    ws['D4'] = 1990
    ws['E4'] = 5
    ws['F4'] = 15
    ws['G4'] = 1
    
    ws['A5'] = 2
    ws['B5'] = "שרה"
    ws['C5'] = "לוי"
    ws['D5'] = 1985
    ws['E5'] = 12
    ws['F5'] = 25
    ws['G5'] = 2
    
    return wb


def create_complex_multirow_workbook():
    """Create a workbook with multiple multi-row header groups."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex Multi-Row"

    # Row 1: Main headers
    ws['A1'] = "Personal Info"
    ws['D1'] = "תאריך לידה"  # Birth Date (Hebrew)
    ws['G1'] = "תאריך כניסה"  # Entry Date (Hebrew)
    
    # Merge cells for parent headers
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:F1')
    ws.merge_cells('G1:I1')
    
    # Row 2: Sub-headers
    ws['A2'] = "First Name"
    ws['B2'] = "Last Name"
    ws['C2'] = "Father Name"
    ws['D2'] = "שנה"  # Year
    ws['E2'] = "חודש"  # Month
    ws['F2'] = "יום"  # Day
    ws['G2'] = "שנה"  # Year
    ws['H2'] = "חודש"  # Month
    ws['I2'] = "יום"  # Day
    
    # Add data
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = "אברהם"
    ws['D3'] = 1990
    ws['E3'] = 5
    ws['F3'] = 15
    ws['G3'] = 2020
    ws['H3'] = 3
    ws['I3'] = 10
    
    return wb


def create_non_merged_multirow_workbook():
    """Create a workbook with multi-row headers but no merged cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Merged Multi-Row"

    # Row 1: Parent headers (not merged)
    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws['C1'] = "תאריך לידה"  # Birth Date parent
    ws['F1'] = "Gender"
    
    # Row 2: Sub-headers
    ws['A2'] = "ID"
    ws['B2'] = "Name"
    ws['C2'] = "שנה"  # Year
    ws['D2'] = "חודש"  # Month
    ws['E2'] = "יום"  # Day
    ws['F2'] = "Gender"
    
    # Add data
    ws['A3'] = 1
    ws['B3'] = "יוסי"
    ws['C3'] = 1990
    ws['D3'] = 5
    ws['E3'] = 15
    ws['F3'] = 1
    
    return wb


def test_score_subheader_row_with_merged_parent():
    """Test _score_subheader_row with merged parent headers."""
    print("=" * 80)
    print("TEST 1: _score_subheader_row with Merged Parent Headers")
    print("=" * 80)
    
    wb = create_multirow_header_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Score row 3 (sub-header row with year/month/day)
    score = reader._score_subheader_row(ws, row_idx=3, max_col=7)
    
    print(f"\nScoring row 3 (sub-header row with year/month/day):")
    print(f"  Score: {score}")
    print(f"  Expected: > 0 (should detect parent-child relationship)")
    
    if score > 0:
        print(f"  ✓ PASS: Sub-header row correctly scored")
    else:
        print(f"  ✗ FAIL: Sub-header row not detected")
    
    # Verify the score is high enough to be considered a sub-header
    if score >= 2:
        print(f"  ✓ PASS: Score is high enough to indicate sub-header row")
    else:
        print(f"  ✗ FAIL: Score is too low")
    
    print()


def test_detect_date_subcolumns_with_merged_parent():
    """Test _detect_date_subcolumns with merged parent headers."""
    print("=" * 80)
    print("TEST 2: _detect_date_subcolumns with Merged Parent Headers")
    print("=" * 80)
    
    wb = create_multirow_header_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Detect date sub-columns starting from column D (merged parent)
    date_columns = reader._detect_date_subcolumns(
        ws, start_col=4, subheader_row=3, max_col=7
    )
    
    print(f"\nDetecting date sub-columns from merged parent (column D):")
    print(f"  Found columns: {date_columns}")
    print(f"  Expected: {{'year': 4, 'month': 5, 'day': 6}}")
    
    if date_columns == {'year': 4, 'month': 5, 'day': 6}:
        print(f"  ✓ PASS: All three date components detected correctly")
    else:
        print(f"  ✗ FAIL: Date columns not detected correctly")
    
    print()


def test_detect_table_region_with_multirow():
    """Test detect_table_region with multi-row headers."""
    print("=" * 80)
    print("TEST 3: detect_table_region with Multi-Row Headers")
    print("=" * 80)
    
    wb = create_multirow_header_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Detect table region
    table_region = reader.detect_table_region(ws)
    
    print(f"\nDetecting table region:")
    if table_region:
        print(f"  Start row: {table_region.start_row}")
        print(f"  Header rows: {table_region.header_rows}")
        print(f"  Data start row: {table_region.data_start_row}")
        print(f"  Expected header_rows: 2 (main + sub-headers)")
        
        if table_region.header_rows == 2:
            print(f"  ✓ PASS: Multi-row header correctly detected")
        else:
            print(f"  ✗ FAIL: Header rows not detected correctly")
    else:
        print(f"  ✗ FAIL: Table region not detected")
    
    print()


def test_detect_columns_with_multirow():
    """Test detect_columns with multi-row headers."""
    print("=" * 80)
    print("TEST 4: detect_columns with Multi-Row Headers")
    print("=" * 80)
    
    wb = create_multirow_header_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Detect columns
    column_mapping = reader.detect_columns(ws)
    
    print(f"\nDetecting columns with multi-row headers:")
    print(f"  Found {len(column_mapping)} columns:")
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"    {field_name:20s} → Column {col_info.col}")
    
    # Check for split date fields
    has_birth_year = 'birth_year' in column_mapping
    has_birth_month = 'birth_month' in column_mapping
    has_birth_day = 'birth_day' in column_mapping
    
    if has_birth_year and has_birth_month and has_birth_day:
        print(f"  ✓ PASS: Split date fields detected correctly")
    else:
        print(f"  ✗ FAIL: Split date fields not detected")
        print(f"    birth_year: {has_birth_year}")
        print(f"    birth_month: {has_birth_month}")
        print(f"    birth_day: {has_birth_day}")
    
    print()


def test_complex_multirow_headers():
    """Test with multiple multi-row header groups."""
    print("=" * 80)
    print("TEST 5: Complex Multi-Row Headers (Multiple Groups)")
    print("=" * 80)
    
    wb = create_complex_multirow_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Detect columns
    column_mapping = reader.detect_columns(ws)
    
    print(f"\nDetecting columns with multiple multi-row groups:")
    print(f"  Found {len(column_mapping)} columns:")
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"    {field_name:20s} → Column {col_info.col}")
    
    # Check for both birth and entry date fields
    has_birth_fields = any(f.startswith('birth_') for f in column_mapping)
    has_entry_fields = any(f.startswith('entry_') for f in column_mapping)
    
    if has_birth_fields:
        print(f"  ✓ PASS: Birth date fields detected")
    else:
        print(f"  ✗ FAIL: Birth date fields not detected")
    
    if has_entry_fields:
        print(f"  ✓ PASS: Entry date fields detected")
    else:
        print(f"  ✗ FAIL: Entry date fields not detected")
    
    print()


def test_non_merged_multirow_headers():
    """Test with non-merged multi-row headers."""
    print("=" * 80)
    print("TEST 6: Non-Merged Multi-Row Headers")
    print("=" * 80)
    
    wb = create_non_merged_multirow_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Detect columns
    column_mapping = reader.detect_columns(ws)
    
    print(f"\nDetecting columns with non-merged multi-row headers:")
    print(f"  Found {len(column_mapping)} columns:")
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"    {field_name:20s} → Column {col_info.col}")
    
    # Check for split date fields
    has_birth_year = 'birth_year' in column_mapping
    has_birth_month = 'birth_month' in column_mapping
    has_birth_day = 'birth_day' in column_mapping
    
    if has_birth_year and has_birth_month and has_birth_day:
        print(f"  ✓ PASS: Split date fields detected even without merged cells")
    else:
        print(f"  ✗ FAIL: Split date fields not detected")
    
    print()


def test_merged_cell_detection():
    """Test that merged cells are properly detected."""
    print("=" * 80)
    print("TEST 7: Merged Cell Detection")
    print("=" * 80)
    
    wb = create_multirow_header_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    # Test _is_merged_cell
    is_merged_d2 = reader._is_merged_cell(ws, row=2, col=4)  # D2 (part of merged range)
    is_merged_e2 = reader._is_merged_cell(ws, row=2, col=5)  # E2 (part of merged range)
    is_merged_f2 = reader._is_merged_cell(ws, row=2, col=6)  # F2 (part of merged range)
    is_merged_g2 = reader._is_merged_cell(ws, row=2, col=7)  # G2 (not merged)
    
    print(f"\nTesting _is_merged_cell:")
    print(f"  D2 (should be merged): {is_merged_d2}")
    print(f"  E2 (should be merged): {is_merged_e2}")
    print(f"  F2 (should be merged): {is_merged_f2}")
    print(f"  G2 (should not be merged): {is_merged_g2}")
    
    if is_merged_d2 and is_merged_e2 and is_merged_f2 and not is_merged_g2:
        print(f"  ✓ PASS: Merged cells correctly detected")
    else:
        print(f"  ✗ FAIL: Merged cell detection failed")
    
    # Test _get_merged_cell_range
    merge_range = reader._get_merged_cell_range(ws, row=2, col=4)
    print(f"\nTesting _get_merged_cell_range:")
    print(f"  Merge range for D2: {merge_range}")
    print(f"  Expected: (2, 2, 4, 6)")
    
    if merge_range == (2, 2, 4, 6):
        print(f"  ✓ PASS: Merged cell range correctly retrieved")
    else:
        print(f"  ✗ FAIL: Merged cell range not correct")
    
    print()


if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("TASK 2: MULTI-ROW HEADER DETECTION TESTS")
    print("=" * 80 + "\n")
    
    test_merged_cell_detection()
    test_score_subheader_row_with_merged_parent()
    test_detect_date_subcolumns_with_merged_parent()
    test_detect_table_region_with_multirow()
    test_detect_columns_with_multirow()
    test_complex_multirow_headers()
    test_non_merged_multirow_headers()
    
    print("=" * 80)
    print("ALL TESTS COMPLETED")
    print("=" * 80 + "\n")
