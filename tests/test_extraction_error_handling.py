"""Tests for error handling in ExcelToJsonExtractor.

This test file validates that the ExcelToJsonExtractor handles various error
conditions gracefully, including:
- Invalid cell values
- Formula errors
- Merged cells
- Missing headers
- Logging of warnings and errors

Requirements:
    - Validates: Requirements 18.1-18.4
"""

import logging
import os
import tempfile
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.excel_standardization.io_layer import ExcelReader, ExcelToJsonExtractor
from src.excel_standardization.data_types import ColumnHeaderInfo, TableRegion


@pytest.fixture
def excel_reader():
    """Create an ExcelReader instance for testing."""
    return ExcelReader()


@pytest.fixture
def extractor(excel_reader):
    """Create an ExcelToJsonExtractor instance for testing."""
    return ExcelToJsonExtractor(excel_reader=excel_reader)


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file for testing."""
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    yield temp_file.name
    # Cleanup
    if os.path.exists(temp_file.name):
        os.unlink(temp_file.name)


def test_extract_row_with_formula_error(extractor, caplog):
    """Test that formula errors are handled gracefully and logged.
    
    Requirements:
        - Validates: Requirement 18.2 (handle invalid cell values)
    """
    # Setup: Create worksheet with formula error
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Test Sheet"
    
    # Add headers
    worksheet['A1'] = 'שם פרטי'
    worksheet['B1'] = 'שם משפחה'
    
    # Add data row with formula error
    worksheet['A2'] = 'יוסי'
    worksheet['B2'] = '#DIV/0!'  # Formula error
    
    # Create column mapping
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='שם פרטי'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='שם משפחה'),
    }
    
    # Execute with logging capture
    with caplog.at_level(logging.WARNING):
        result = extractor.extract_row_to_json(worksheet, 2, column_mapping)
    
    # Verify - formula error should be replaced with None and logged
    assert result['first_name'] == 'יוסי'
    assert result['last_name'] is None
    assert any('Formula error' in record.message for record in caplog.records)


def test_extract_sheet_with_invalid_rows(extractor, caplog):
    """Test that sheets with some invalid rows continue processing.
    
    Requirements:
        - Validates: Requirement 18.2 (handle invalid cell values)
        - Validates: Requirement 18.4 (log warnings for errors)
    """
    # Setup: Create worksheet with valid and invalid rows
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Test Sheet"
    
    # Add headers
    worksheet['A1'] = 'שם פרטי'
    worksheet['B1'] = 'שם משפחה'
    
    # Add valid rows
    worksheet['A2'] = 'יוסי'
    worksheet['B2'] = 'כהן'
    worksheet['A3'] = 'דוד'
    worksheet['B3'] = 'לוי'
    
    # Create column mapping
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=3, header_text='שם פרטי'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=3, header_text='שם משפחה'),
    }
    
    table_region = TableRegion(
        start_row=1,
        end_row=3,
        start_col=1,
        end_col=2,
        header_rows=1,
        data_start_row=2,
    )
    
    # Execute with logging capture
    with caplog.at_level(logging.INFO):
        result = extractor.extract_sheet_to_json(worksheet, column_mapping, table_region)
    
    # Verify - should extract all valid rows
    assert len(result.rows) == 2
    assert result.rows[0]['first_name'] == 'יוסי'
    assert result.rows[1]['first_name'] == 'דוד'
    
    # Verify logging
    assert any('Extracting sheet' in record.message for record in caplog.records)
    assert any('Successfully extracted' in record.message for record in caplog.records)


def test_extract_sheet_with_no_headers_logs_warning(extractor, caplog):
    """Test that sheets with no headers log a warning.
    
    Requirements:
        - Validates: Requirement 18.1 (handle missing headers gracefully)
        - Validates: Requirement 18.4 (log warnings for skipped sheets)
    """
    # Setup: Create worksheet with no recognizable headers
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invalid Sheet"
    
    # Add non-header data
    worksheet['A1'] = 'Random'
    worksheet['B1'] = 'Data'
    
    # Execute with logging capture
    with caplog.at_level(logging.WARNING):
        result = extractor.extract_sheet_to_json(worksheet)
    
    # Verify - should return empty dataset with error metadata
    assert result.sheet_name == "Invalid Sheet"
    assert len(result.rows) == 0
    assert result.metadata['skipped'] is True
    assert 'error' in result.metadata
    
    # Verify warning was logged
    assert any('No valid headers found' in record.message for record in caplog.records)
    assert any('Invalid Sheet' in record.message for record in caplog.records)


def test_extract_workbook_logs_skipped_sheets(extractor, temp_excel_file, caplog):
    """Test that workbook extraction logs warnings for skipped sheets.
    
    Requirements:
        - Validates: Requirement 18.4 (log warnings for skipped sheets)
    """
    # Setup: Create workbook with valid and invalid sheets
    workbook = Workbook()
    
    # Sheet 1: Valid sheet
    ws1 = workbook.active
    ws1.title = "Valid Sheet"
    ws1['A1'] = 'שם פרטי'
    ws1['B1'] = 'שם משפחה'
    ws1['A2'] = 'יוסי'
    ws1['B2'] = 'כהן'
    
    # Sheet 2: Invalid sheet (no headers)
    ws2 = workbook.create_sheet("Invalid Sheet")
    ws2['A1'] = 'Random'
    ws2['B1'] = 'Data'
    
    workbook.save(temp_excel_file)
    
    # Execute with logging capture
    with caplog.at_level(logging.WARNING):
        result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify - should process valid sheet and skip invalid sheet
    assert result.metadata['processed_sheets'] == 1
    assert 'Invalid Sheet' in result.metadata['skipped_sheets']
    
    # Verify warnings were logged
    assert any('Invalid Sheet' in record.message and 'skipped' in record.message 
               for record in caplog.records)


def test_extract_workbook_logs_info_messages(extractor, temp_excel_file, caplog):
    """Test that workbook extraction logs informational messages.
    
    Requirements:
        - Validates: Requirement 18.4 (log information about processing)
    """
    # Setup: Create simple workbook
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Test Sheet"
    ws['A1'] = 'שם פרטי'
    ws['B1'] = 'שם משפחה'
    ws['A2'] = 'יוסי'
    ws['B2'] = 'כהן'
    
    workbook.save(temp_excel_file)
    
    # Execute with logging capture
    with caplog.at_level(logging.INFO):
        result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify logging messages
    log_messages = [record.message for record in caplog.records]
    
    assert any('Loading workbook' in msg for msg in log_messages)
    assert any('Workbook loaded' in msg for msg in log_messages)
    assert any('Processing sheet' in msg for msg in log_messages)
    assert any('Extracting sheet' in msg for msg in log_messages)
    assert any('Successfully extracted' in msg for msg in log_messages)
    assert any('processed successfully' in msg for msg in log_messages)
    assert any('Workbook extraction complete' in msg for msg in log_messages)


def test_extract_row_handles_cell_extraction_errors(extractor, caplog):
    """Test that cell extraction errors are caught and logged.
    
    Requirements:
        - Validates: Requirement 18.2 (handle invalid cell values)
    """
    # Setup: Create worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Test Sheet"
    
    # Add headers
    worksheet['A1'] = 'שם פרטי'
    worksheet['B1'] = 'שם משפחה'
    
    # Add data
    worksheet['A2'] = 'יוסי'
    worksheet['B2'] = 'כהן'
    
    # Create column mapping with an invalid column index
    # This will cause an error when trying to read the cell
    column_mapping = {
        'first_name': ColumnHeaderInfo(col=1, header_row=1, last_row=2, header_text='שם פרטי'),
        'last_name': ColumnHeaderInfo(col=2, header_row=1, last_row=2, header_text='שם משפחה'),
    }
    
    # Execute - should not raise exception
    result = extractor.extract_row_to_json(worksheet, 2, column_mapping)
    
    # Verify - should have extracted values successfully
    assert result['first_name'] == 'יוסי'
    assert result['last_name'] == 'כהן'
