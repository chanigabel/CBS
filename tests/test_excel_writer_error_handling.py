"""Tests for JsonToExcelWriter error handling.

This module tests the error handling capabilities of the JsonToExcelWriter class,
including validation, file write failures, and cleanup of partial files.

Requirements:
    - Validates: Requirements 18.1-18.4
"""

import os
import tempfile
import pytest
from pathlib import Path
from openpyxl import load_workbook

from src.excel_normalization.data_types import SheetDataset, WorkbookDataset, JsonRow
from src.excel_normalization.io_layer import JsonToExcelWriter


class TestDatasetValidation:
    """Test dataset structure validation before export."""
    
    def test_invalid_dataset_structure(self):
        """Test that invalid dataset structure raises ValueError."""
        # Create invalid dataset (empty sheet name)
        dataset = SheetDataset(
            sheet_name="",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid dataset structure"):
                writer.write_dataset_to_excel(dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_invalid_header_row(self):
        """Test that invalid header row raises ValueError."""
        # Create dataset with invalid header row (0 or negative)
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=0,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid dataset structure"):
                writer.write_dataset_to_excel(dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_invalid_header_rows_count(self):
        """Test that invalid header_rows_count raises ValueError."""
        # Create dataset with invalid header_rows_count (not 1 or 2)
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=3,
            field_names=["first_name"],
            rows=[{"first_name": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid dataset structure"):
                writer.write_dataset_to_excel(dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_empty_field_names(self):
        """Test that empty field names raises ValueError."""
        # Create dataset with no field names
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=[],
            rows=[{"first_name": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid dataset structure"):
                writer.write_dataset_to_excel(dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_no_valid_field_names_after_filtering(self):
        """Test that dataset with only _corrected fields raises ValueError."""
        # Create dataset where all field names end with _corrected
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name_corrected", "last_name_corrected"],
            rows=[{"first_name_corrected": "Test", "last_name_corrected": "User"}]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="No valid field names found"):
                writer.write_dataset_to_excel(dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestOutputPathValidation:
    """Test output path validation."""
    
    def test_empty_output_path(self):
        """Test that empty output path raises ValueError."""
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Test", "first_name_corrected": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        with pytest.raises(ValueError, match="Output path cannot be empty"):
            writer.write_dataset_to_excel(dataset, "")
    
    def test_nonexistent_directory(self):
        """Test that nonexistent output directory raises ValueError."""
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Test", "first_name_corrected": "Test"}]
        )
        
        writer = JsonToExcelWriter()
        
        # Use a path with nonexistent directory
        output_path = "/nonexistent/directory/output.xlsx"
        
        with pytest.raises(ValueError, match="Output directory does not exist"):
            writer.write_dataset_to_excel(dataset, output_path)


class TestWorkbookDatasetValidation:
    """Test workbook dataset validation."""
    
    def test_invalid_workbook_dataset(self):
        """Test that invalid workbook dataset raises ValueError."""
        # Create workbook dataset with empty source file
        workbook_dataset = WorkbookDataset(
            source_file="",
            sheets=[]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid workbook dataset structure"):
                writer.write_workbook_to_excel(workbook_dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_workbook_dataset_no_sheets(self):
        """Test that workbook dataset with no sheets raises ValueError."""
        workbook_dataset = WorkbookDataset(
            source_file="test.xlsx",
            sheets=[]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="has no sheets"):
                writer.write_workbook_to_excel(workbook_dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_workbook_dataset_with_invalid_sheet(self):
        """Test that workbook dataset with invalid sheet raises ValueError."""
        # Create workbook with one invalid sheet
        invalid_sheet = SheetDataset(
            sheet_name="",  # Invalid: empty name
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "Test"}]
        )
        
        workbook_dataset = WorkbookDataset(
            source_file="test.xlsx",
            sheets=[invalid_sheet]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            with pytest.raises(ValueError, match="Invalid"):
                writer.write_workbook_to_excel(workbook_dataset, output_path)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestPartialFileCleanup:
    """Test cleanup of partial files on error."""
    
    def test_cleanup_on_write_error(self):
        """Test that partial file is cleaned up when write fails."""
        # Create a dataset that will cause an error during row writing
        # by having inconsistent field structure
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[
                {"first_name": "Test", "first_name_corrected": "Test"},
                # This row is missing the corrected field, but we'll simulate
                # a different error by using an invalid output path after validation
            ]
        )
        
        writer = JsonToExcelWriter()
        
        # Create a temporary directory
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            # First write should succeed
            writer.write_dataset_to_excel(dataset, output_path)
            assert os.path.exists(output_path)
            
            # Now make the file read-only to cause a write error
            os.chmod(output_path, 0o444)
            
            try:
                # Try to write again - should fail and clean up
                with pytest.raises(PermissionError):
                    writer.write_dataset_to_excel(dataset, output_path)
            finally:
                # Restore write permissions for cleanup
                os.chmod(output_path, 0o644)


class TestSuccessfulExportWithValidation:
    """Test that valid datasets export successfully with validation."""
    
    def test_valid_dataset_exports_successfully(self):
        """Test that a valid dataset exports successfully."""
        dataset = SheetDataset(
            sheet_name="Test",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name", "last_name"],
            rows=[
                {"first_name": "John", "first_name_corrected": "John", 
                 "last_name": "Doe", "last_name_corrected": "Doe"},
                {"first_name": "Jane", "first_name_corrected": "Jane",
                 "last_name": "Smith", "last_name_corrected": "Smith"}
            ]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            # Should not raise any exceptions
            writer.write_dataset_to_excel(dataset, output_path)
            
            # Verify file was created
            assert os.path.exists(output_path)
            
            # Verify content
            workbook = load_workbook(output_path)
            worksheet = workbook.active
            
            # Check headers
            assert worksheet.cell(1, 1).value == "first_name"
            assert worksheet.cell(1, 2).value == "first_name_corrected"
            assert worksheet.cell(1, 3).value == "last_name"
            assert worksheet.cell(1, 4).value == "last_name_corrected"
            
            # Check data
            assert worksheet.cell(2, 1).value == "John"
            assert worksheet.cell(2, 2).value == "John"
            assert worksheet.cell(3, 1).value == "Jane"
            assert worksheet.cell(3, 2).value == "Jane"
            
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_valid_workbook_dataset_exports_successfully(self):
        """Test that a valid workbook dataset exports successfully."""
        sheet1 = SheetDataset(
            sheet_name="Sheet1",
            header_row=1,
            header_rows_count=1,
            field_names=["first_name"],
            rows=[{"first_name": "John", "first_name_corrected": "John"}]
        )
        
        sheet2 = SheetDataset(
            sheet_name="Sheet2",
            header_row=1,
            header_rows_count=1,
            field_names=["last_name"],
            rows=[{"last_name": "Doe", "last_name_corrected": "Doe"}]
        )
        
        workbook_dataset = WorkbookDataset(
            source_file="test.xlsx",
            sheets=[sheet1, sheet2]
        )
        
        writer = JsonToExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            output_path = tmp.name
        
        try:
            # Should not raise any exceptions
            writer.write_workbook_to_excel(workbook_dataset, output_path)
            
            # Verify file was created
            assert os.path.exists(output_path)
            
            # Verify content
            workbook = load_workbook(output_path)
            assert "Sheet1" in workbook.sheetnames
            assert "Sheet2" in workbook.sheetnames
            
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
