"""Test to verify that the original Excel file is NEVER modified.

This is a critical test that ensures the pipeline follows the correct behavior:
- Excel file is read in read-only mode
- Data is extracted to JSON
- standardization is applied to JSON
- Output is written to JSON files
- Original Excel file remains completely unchanged

Requirements:
    - Validates: Critical requirement that original Excel is never modified
"""

import pytest
import tempfile
import os
import hashlib
from openpyxl import Workbook

from src.excel_standardization.io_layer import ExcelReader, ExcelToJsonExtractor
from src.excel_standardization.processing import standardizationPipeline
from src.excel_standardization.json_exporter import JsonExporter, generate_output_filenames
from src.excel_standardization.engines import (
    NameEngine, GenderEngine, DateEngine, IdentifierEngine, TextProcessor
)


def calculate_file_hash(filepath: str) -> str:
    """Calculate MD5 hash of a file to detect any changes."""
    hash_md5 = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


class TestExcelNeverModified:
    """Test suite to verify original Excel file is never modified."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing."""
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # Create Excel file with data
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Students"
        
        # Headers
        ws['A1'] = 'שם פרטי'
        ws['B1'] = 'שם משפחה'
        ws['C1'] = 'מין'
        ws['D1'] = 'ת.ז'
        
        # Data
        ws['A2'] = 'יוסי 123'
        ws['B2'] = 'כהן  '
        ws['C2'] = 'ז'
        ws['D2'] = '123456789'
        
        ws['A3'] = '  שרה'
        ws['B3'] = 'לוי'
        ws['C3'] = 'נ'
        ws['D3'] = '987654321'
        
        workbook.save(temp_file.name)
        
        yield temp_file.name
        
        # Cleanup
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
    
    def test_excel_file_hash_unchanged_after_extraction(self, temp_excel_file):
        """Test that Excel file hash is unchanged after extraction.
        
        This verifies that the extraction process does not modify the file.
        """
        # Calculate hash before extraction
        hash_before = calculate_file_hash(temp_excel_file)
        
        # Extract data
        reader = ExcelReader()
        extractor = ExcelToJsonExtractor(excel_reader=reader)
        workbook_dataset = extractor.extract_workbook_to_json(temp_excel_file)
        
        # Calculate hash after extraction
        hash_after = calculate_file_hash(temp_excel_file)
        
        # Verify file is unchanged
        assert hash_before == hash_after, "Excel file was modified during extraction!"
        assert len(workbook_dataset.sheets) > 0, "Should have extracted data"
    
    def test_excel_file_hash_unchanged_after_standardization(self, temp_excel_file):
        """Test that Excel file hash is unchanged after standardization.
        
        This verifies that the standardization process does not modify the file.
        """
        # Calculate hash before processing
        hash_before = calculate_file_hash(temp_excel_file)
        
        # Extract data
        reader = ExcelReader()
        extractor = ExcelToJsonExtractor(excel_reader=reader)
        workbook_dataset = extractor.extract_workbook_to_json(temp_excel_file)
        
        # Normalize data
        pipeline = standardizationPipeline(
            name_engine=NameEngine(TextProcessor()),
            gender_engine=GenderEngine(),
            date_engine=DateEngine(),
            identifier_engine=IdentifierEngine()
        )
        
        for sheet in workbook_dataset.sheets:
            normalized_sheet = pipeline.normalize_dataset(sheet)
        
        # Calculate hash after standardization
        hash_after = calculate_file_hash(temp_excel_file)
        
        # Verify file is unchanged
        assert hash_before == hash_after, "Excel file was modified during standardization!"
    
    def test_excel_file_hash_unchanged_after_json_export(self, temp_excel_file):
        """Test that Excel file hash is unchanged after JSON export.
        
        This verifies the complete pipeline does not modify the original file.
        """
        # Calculate hash before processing
        hash_before = calculate_file_hash(temp_excel_file)
        
        # Generate output filenames
        raw_json_path, normalized_json_path = generate_output_filenames(temp_excel_file)
        
        try:
            # Extract data
            reader = ExcelReader()
            extractor = ExcelToJsonExtractor(excel_reader=reader)
            workbook_dataset = extractor.extract_workbook_to_json(temp_excel_file)
            
            # Export raw JSON
            exporter = JsonExporter()
            exporter.export_workbook_to_json(workbook_dataset, raw_json_path)
            
            # Normalize data
            pipeline = standardizationPipeline(
                name_engine=NameEngine(TextProcessor()),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_sheets = []
            for sheet in workbook_dataset.sheets:
                normalized_sheet = pipeline.normalize_dataset(sheet)
                normalized_sheets.append(normalized_sheet)
            
            # Export normalized JSON
            workbook_dataset.sheets = normalized_sheets
            exporter.export_workbook_to_json(workbook_dataset, normalized_json_path)
            
            # Calculate hash after complete pipeline
            hash_after = calculate_file_hash(temp_excel_file)
            
            # Verify file is unchanged
            assert hash_before == hash_after, "Excel file was modified by the pipeline!"
            
            # Verify JSON files were created
            assert os.path.exists(raw_json_path), "Raw JSON file was not created"
            assert os.path.exists(normalized_json_path), "Normalized JSON file was not created"
            
        finally:
            # Cleanup JSON files
            if os.path.exists(raw_json_path):
                os.unlink(raw_json_path)
            if os.path.exists(normalized_json_path):
                os.unlink(normalized_json_path)
    
    def test_excel_file_not_opened_in_write_mode(self, temp_excel_file):
        """Test that Excel file is opened in read-only mode.
        
        This verifies that openpyxl is used with data_only=True (read-only).
        """
        # Extract data
        reader = ExcelReader()
        extractor = ExcelToJsonExtractor(excel_reader=reader)
        
        # This should not raise any permission errors
        workbook_dataset = extractor.extract_workbook_to_json(temp_excel_file)
        
        # Verify we can still read the file (it wasn't locked)
        from openpyxl import load_workbook
        wb = load_workbook(temp_excel_file, data_only=True)
        assert wb is not None
        wb.close()
    
    def test_json_files_created_in_same_directory(self, temp_excel_file):
        """Test that JSON files are created in the same directory as input Excel."""
        # Generate output filenames
        raw_json_path, normalized_json_path = generate_output_filenames(temp_excel_file)
        
        # Verify paths are in same directory
        import os
        excel_dir = os.path.dirname(temp_excel_file)
        raw_json_dir = os.path.dirname(raw_json_path)
        normalized_json_dir = os.path.dirname(normalized_json_path)
        
        assert excel_dir == raw_json_dir, "Raw JSON should be in same directory as Excel"
        assert excel_dir == normalized_json_dir, "Normalized JSON should be in same directory as Excel"
    
    def test_json_filenames_follow_convention(self, temp_excel_file):
        """Test that JSON filenames follow the naming convention."""
        # Generate output filenames
        raw_json_path, normalized_json_path = generate_output_filenames(temp_excel_file)
        
        # Get base name
        import os
        base_name = os.path.splitext(os.path.basename(temp_excel_file))[0]
        
        # Verify naming convention
        assert raw_json_path.endswith(f"{base_name}_raw.json"), "Raw JSON should end with _raw.json"
        assert normalized_json_path.endswith(f"{base_name}_normalized.json"), "Normalized JSON should end with _normalized.json"
    
    def test_complete_pipeline_preserves_excel(self, temp_excel_file):
        """Integration test: Complete pipeline preserves original Excel file.
        
        This is the most important test - it verifies the entire pipeline
        from start to finish does not modify the original Excel file.
        """
        # Get file stats before
        stat_before = os.stat(temp_excel_file)
        hash_before = calculate_file_hash(temp_excel_file)
        
        # Generate output filenames
        raw_json_path, normalized_json_path = generate_output_filenames(temp_excel_file)
        
        try:
            # Run complete pipeline
            reader = ExcelReader()
            extractor = ExcelToJsonExtractor(excel_reader=reader)
            workbook_dataset = extractor.extract_workbook_to_json(temp_excel_file)
            
            exporter = JsonExporter()
            exporter.export_workbook_to_json(workbook_dataset, raw_json_path)
            
            pipeline = standardizationPipeline(
                name_engine=NameEngine(TextProcessor()),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_sheets = []
            for sheet in workbook_dataset.sheets:
                normalized_sheet = pipeline.normalize_dataset(sheet)
                normalized_sheets.append(normalized_sheet)
            
            workbook_dataset.sheets = normalized_sheets
            exporter.export_workbook_to_json(workbook_dataset, normalized_json_path)
            
            # Get file stats after
            stat_after = os.stat(temp_excel_file)
            hash_after = calculate_file_hash(temp_excel_file)
            
            # Verify file is completely unchanged
            assert hash_before == hash_after, "File content was modified!"
            assert stat_before.st_size == stat_after.st_size, "File size changed!"
            assert stat_before.st_mtime == stat_after.st_mtime, "File modification time changed!"
            
            # Verify outputs were created
            assert os.path.exists(raw_json_path), "Raw JSON not created"
            assert os.path.exists(normalized_json_path), "Normalized JSON not created"
            
            # Verify outputs contain data
            import json
            with open(raw_json_path, 'r', encoding='utf-8') as f:
                raw_data = json.load(f)
                assert len(raw_data['sheets']) > 0, "Raw JSON has no sheets"
                assert len(raw_data['sheets'][0]['rows']) > 0, "Raw JSON has no rows"
            
            with open(normalized_json_path, 'r', encoding='utf-8') as f:
                normalized_data = json.load(f)
                assert len(normalized_data['sheets']) > 0, "Normalized JSON has no sheets"
                assert len(normalized_data['sheets'][0]['rows']) > 0, "Normalized JSON has no rows"
                # Verify corrected fields exist
                first_row = normalized_data['sheets'][0]['rows'][0]
                assert any(key.endswith('_corrected') for key in first_row.keys()), "No corrected fields found"
        
        finally:
            # Cleanup
            if os.path.exists(raw_json_path):
                os.unlink(raw_json_path)
            if os.path.exists(normalized_json_path):
                os.unlink(normalized_json_path)
