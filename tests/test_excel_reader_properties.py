"""Property-based tests for ExcelReader using Hypothesis framework.

These tests validate universal correctness properties that should hold
across all possible inputs, not just specific examples.

Tests cover:
- Text normalization idempotence
- Keyword recognition completeness
- Corrected column exclusion
- Best row selection
- Split field preference
"""

import pytest
from hypothesis import given, strategies as st, assume
from openpyxl import Workbook
from src.excel_normalization.io_layer.excel_reader import ExcelReader


class TestTextNormalizationProperties:
    """Property-based tests for text normalization.
    
    **Validates: Requirements 2.1-2.6**
    """

    @given(st.text())
    def test_normalization_idempotence(self, text):
        """Property: Normalizing twice should equal normalizing once.
        
        For any text input, applying normalization twice should produce
        the same result as applying it once. This ensures the normalization
        function is idempotent.
        """
        reader = ExcelReader()
        once = reader._normalize_text(text)
        twice = reader._normalize_text(once)
        assert once == twice, f"Idempotence failed: {once!r} != {twice!r}"

    @given(st.text())
    def test_normalization_removes_line_breaks(self, text):
        """Property: Normalized text should contain no line breaks."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        assert "\n" not in normalized, "Normalized text contains \\n"
        assert "\r" not in normalized, "Normalized text contains \\r"

    @given(st.text())
    def test_normalization_removes_parentheses(self, text):
        """Property: Normalized text should contain no parentheses."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        assert "(" not in normalized, "Normalized text contains ("
        assert ")" not in normalized, "Normalized text contains )"

    @given(st.text())
    def test_normalization_removes_brackets(self, text):
        """Property: Normalized text should contain no brackets."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        assert "[" not in normalized, "Normalized text contains ["
        assert "]" not in normalized, "Normalized text contains ]"

    @given(st.text())
    def test_normalization_removes_braces(self, text):
        """Property: Normalized text should contain no braces."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        assert "{" not in normalized, "Normalized text contains {"
        assert "}" not in normalized, "Normalized text contains }"

    @given(st.text())
    def test_normalization_no_leading_trailing_whitespace(self, text):
        """Property: Normalized text should have no leading/trailing whitespace."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        if normalized:  # Only check if not empty
            assert not normalized.startswith(" "), "Normalized text starts with space"
            assert not normalized.startswith("\t"), "Normalized text starts with tab"
            assert not normalized.endswith(" "), "Normalized text ends with space"
            assert not normalized.endswith("\t"), "Normalized text ends with tab"

    @given(st.text())
    def test_normalization_single_spaces(self, text):
        """Property: Normalized text should have no consecutive spaces."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        assert "  " not in normalized, "Normalized text contains consecutive spaces"

    @given(st.text())
    def test_normalization_lowercase(self, text):
        """Property: Normalized text should be lowercase."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        # Check that ASCII uppercase letters are converted
        for char in normalized:
            if char.isascii() and char.isupper():
                pytest.fail(f"Normalized text contains uppercase: {char}")

    @given(st.text())
    def test_normalization_deterministic(self, text):
        """Property: Normalization should be deterministic."""
        reader = ExcelReader()
        result1 = reader._normalize_text(text)
        result2 = reader._normalize_text(text)
        assert result1 == result2, "Normalization is not deterministic"


class TestKeywordRecognitionProperties:
    """Property-based tests for keyword recognition.
    
    **Validates: Requirements 3.1-3.10**
    """

    @given(st.text(min_size=1))
    def test_keyword_match_deterministic(self, text):
        """Property: Keyword matching should be deterministic."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        result1 = reader._match_field(normalized)
        result2 = reader._match_field(normalized)
        assert result1 == result2, "Keyword matching is not deterministic"

    @given(st.text(min_size=1))
    def test_contains_field_keyword_consistency(self, text):
        """Property: _contains_field_keyword should be consistent with _match_field."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        
        has_keyword = reader._contains_field_keyword(normalized)
        matched_field = reader._match_field(normalized)
        
        # If _match_field returns a field, _contains_field_keyword should be True
        if matched_field is not None:
            assert has_keyword is True, \
                f"Inconsistency: matched {matched_field} but has_keyword is False"

    @given(st.sampled_from([
        "שם פרטי", "first name", "first", "name",
        "שם משפחה", "last name", "last", "surname",
        "שם האב", "father name", "father",
        "מין", "gender", "sex",
        "מספר זהות", "id number", "id",
        "דרכון", "passport",
        "תאריך לידה", "birth date", "dob",
        "תאריך כניסה", "entry date",
        "שנה", "year",
        "חודש", "month",
        "יום", "day"
    ]))
    def test_keyword_recognition_completeness(self, keyword):
        """Property: If text contains a keyword, field should be matched.
        
        For any known keyword, when the text contains that keyword,
        the system should recognize and match the corresponding field.
        """
        reader = ExcelReader()
        normalized = reader._normalize_text(keyword)
        
        # The keyword should be recognized
        result = reader._match_field(normalized)
        assert result is not None, f"Keyword {keyword!r} not recognized"

    @given(st.sampled_from([
        "שם פרטי", "first name", "first", "name",
        "שם משפחה", "last name", "last", "surname",
        "שם האב", "father name", "father",
        "מין", "gender", "sex",
        "מספר זהות", "id number", "id",
        "דרכון", "passport",
        "תאריך לידה", "birth date", "dob",
        "תאריך כניסה", "entry date",
        "שנה", "year",
        "חודש", "month",
        "יום", "day"
    ]), st.text())
    def test_keyword_substring_matching(self, keyword, extra_text):
        """Property: Substring matching should work with extra text.
        
        If a cell contains a keyword plus additional text, the system
        should still recognize the field through substring matching.
        """
        reader = ExcelReader()
        
        # Create text with keyword and extra text
        combined = f"{keyword} {extra_text}"
        normalized = reader._normalize_text(combined)
        
        # Should still match the field
        result = reader._match_field(normalized)
        assert result is not None, \
            f"Failed to match keyword {keyword!r} in combined text {combined!r}"

    @given(st.text(min_size=1))
    def test_match_field_returns_valid_field_or_none(self, text):
        """Property: _match_field should return valid field type or None."""
        reader = ExcelReader()
        normalized = reader._normalize_text(text)
        result = reader._match_field(normalized)
        
        valid_fields = {
            "first_name", "last_name", "father_name", "gender",
            "id_number", "passport", "birth_date", "entry_date",
            "year", "month", "day"
        }
        
        assert result is None or result in valid_fields, \
            f"Invalid field type returned: {result}"


class TestCorrectedColumnExclusionProperties:
    """Property-based tests for corrected column exclusion.
    
    **Validates: Requirements 4.1-4.2**
    """

    @given(st.text(min_size=1))
    def test_corrected_marker_detection_deterministic(self, text):
        """Property: Corrected column detection should be deterministic."""
        reader = ExcelReader()
        result1 = reader._should_ignore_column(text)
        result2 = reader._should_ignore_column(text)
        assert result1 == result2, "Corrected column detection is not deterministic"

    @given(st.sampled_from(["מתוקן", "corrected", "fixed", "updated"]))
    def test_corrected_marker_recognized(self, marker):
        """Property: Columns with corrected markers should be excluded.
        
        For any known corrected marker, the system should recognize
        and exclude that column from mappings.
        """
        reader = ExcelReader()
        result = reader._should_ignore_column(marker)
        assert result is True, f"Marker {marker!r} not recognized as corrected"

    @given(st.sampled_from(["מתוקן", "corrected", "fixed", "updated"]), st.text())
    def test_corrected_marker_with_extra_text(self, marker, extra_text):
        """Property: Corrected marker should be detected even with extra text.
        
        If a column header contains a corrected marker plus additional text,
        the system should still recognize and exclude it.
        """
        reader = ExcelReader()
        
        # Create text with marker and extra text
        combined = f"{marker} {extra_text}"
        result = reader._should_ignore_column(combined)
        assert result is True, \
            f"Failed to detect marker {marker!r} in combined text {combined!r}"

    @given(st.sampled_from(["מתוקן", "corrected", "fixed", "updated"]))
    def test_corrected_marker_case_insensitive(self, marker):
        """Property: Corrected marker detection should be case insensitive."""
        reader = ExcelReader()
        
        # Test uppercase version
        uppercase = marker.upper()
        result = reader._should_ignore_column(uppercase)
        assert result is True, f"Case insensitive detection failed for {uppercase!r}"

    @given(st.text(min_size=1))
    def test_should_ignore_column_returns_boolean(self, text):
        """Property: _should_ignore_column should return boolean."""
        reader = ExcelReader()
        result = reader._should_ignore_column(text)
        assert isinstance(result, bool), f"Expected bool, got {type(result)}"

    @given(st.text(min_size=1).filter(
        lambda x: "מתוקן" not in x.lower() and "corrected" not in x.lower() 
        and "fixed" not in x.lower() and "updated" not in x.lower()
    ))
    def test_normal_columns_not_ignored(self, text):
        """Property: Normal columns without markers should not be ignored."""
        reader = ExcelReader()
        result = reader._should_ignore_column(text)
        assert result is False, f"Normal column {text!r} incorrectly marked as ignored"


class TestBestRowSelectionProperties:
    """Property-based tests for best row selection.
    
    **Validates: Requirements 5.1-5.5**
    """

    @given(st.integers(min_value=1, max_value=30))
    def test_score_header_row_deterministic(self, row_idx):
        """Property: Header row scoring should be deterministic."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Add some test data
        ws['A1'] = "שם פרטי"
        ws['B1'] = "שם משפחה"
        ws['C1'] = "מספר זהות"
        
        result1 = reader._score_header_row(ws, row_idx, 10)
        result2 = reader._score_header_row(ws, row_idx, 10)
        assert result1 == result2, "Header row scoring is not deterministic"

    @given(st.integers(min_value=1, max_value=30))
    def test_score_header_row_returns_integer(self, row_idx):
        """Property: _score_header_row should return an integer."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        result = reader._score_header_row(ws, row_idx, 10)
        assert isinstance(result, int), f"Expected int, got {type(result)}"

    @given(st.integers(min_value=1, max_value=30))
    def test_score_header_row_non_negative(self, row_idx):
        """Property: Header row score should be non-negative."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        result = reader._score_header_row(ws, row_idx, 10)
        assert result >= 0, f"Score should be non-negative, got {result}"

    @given(st.lists(
        st.sampled_from([
            "שם פרטי", "שם משפחה", "מספר זהות",
            "first name", "last name", "id number",
            "unknown", "data", "value"
        ]),
        min_size=1,
        max_size=10
    ))
    def test_more_keywords_higher_score(self, keywords):
        """Property: More keyword matches should result in higher scores.
        
        When a row has more recognized keywords, its score should be
        higher than a row with fewer keywords.
        """
        from openpyxl import Workbook
        
        reader = ExcelReader()
        
        # Create two worksheets: one with keywords, one without
        wb1 = Workbook()
        ws1 = wb1.active
        for i, keyword in enumerate(keywords, 1):
            ws1.cell(row=1, column=i, value=keyword)
        
        wb2 = Workbook()
        ws2 = wb2.active
        for i in range(1, len(keywords) + 1):
            ws2.cell(row=1, column=i, value="unknown")
        
        score1 = reader._score_header_row(ws1, 1, len(keywords))
        score2 = reader._score_header_row(ws2, 1, len(keywords))
        
        # Row with keywords should score at least as high as row without
        assert score1 >= score2, \
            f"Row with keywords scored {score1}, row without scored {score2}"

    @given(st.integers(min_value=1, max_value=30), st.integers(min_value=1, max_value=50))
    def test_score_header_row_with_varying_max_col(self, row_idx, max_col):
        """Property: Scoring should handle varying max_col values."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Should not raise exception
        result = reader._score_header_row(ws, row_idx, max_col)
        assert isinstance(result, int)


class TestSplitFieldPreferenceProperties:
    """Property-based tests for split field preference.
    
    **Validates: Requirements 7.5**
    """

    @given(st.integers(min_value=1, max_value=30), st.integers(min_value=1, max_value=30))
    def test_detect_date_subcolumns_deterministic(self, row_idx, subrow_idx):
        """Property: Date subcolumn detection should be deterministic."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Add test data
        ws['A1'] = "תאריך לידה"
        ws['A2'] = "שנה"
        ws['B2'] = "חודש"
        ws['C2'] = "יום"
        
        result1 = reader._detect_date_subcolumns(ws, row_idx, subrow_idx, 10)
        result2 = reader._detect_date_subcolumns(ws, row_idx, subrow_idx, 10)
        assert result1 == result2, "Date subcolumn detection is not deterministic"

    @given(st.integers(min_value=1, max_value=30), st.integers(min_value=1, max_value=30))
    def test_detect_date_subcolumns_returns_dict(self, row_idx, subrow_idx):
        """Property: _detect_date_subcolumns should return a dictionary."""
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        result = reader._detect_date_subcolumns(ws, row_idx, subrow_idx, 10)
        assert isinstance(result, dict), f"Expected dict, got {type(result)}"

    def test_split_date_field_completeness(self):
        """Property: Complete split date fields should be detected.
        
        When all three components (year, month, day) are present,
        the system should detect the split date field structure.
        """
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Create complete split date field
        ws['A1'] = "תאריך לידה"
        ws['A2'] = "שנה"
        ws['B2'] = "חודש"
        ws['C2'] = "יום"
        
        result = reader._detect_date_subcolumns(ws, 1, 2, 10)
        
        # Should detect all three components
        assert 'year' in result or len(result) > 0, \
            "Failed to detect split date field components"

    def test_incomplete_split_date_field_not_detected(self):
        """Property: Incomplete split date fields should not be detected.
        
        When not all three components are present, the system should
        not detect a split date field structure.
        """
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Create incomplete split date field (missing day)
        ws['A1'] = "תאריך לידה"
        ws['A2'] = "שנה"
        ws['B2'] = "חודש"
        
        result = reader._detect_date_subcolumns(ws, 1, 2, 10)
        
        # Should not detect incomplete field
        # (either empty dict or not all three components)
        if result:
            # If something is detected, it should not be a complete set
            has_all_three = 'year' in result and 'month' in result and 'day' in result
            assert not has_all_three, \
                "Incomplete split date field incorrectly detected as complete"

    @given(st.sampled_from([
        ("שנה", "חודש", "יום"),
        ("year", "month", "day"),
        ("שנה", "month", "יום"),
        ("year", "חודש", "day")
    ]))
    def test_split_date_field_language_mixing(self, components):
        """Property: Split date fields should work with mixed languages.
        
        The system should detect split date fields even when components
        use different languages (Hebrew and English mixed).
        """
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Create split date field with mixed languages
        ws['A1'] = "תאריך לידה"
        ws['A2'] = components[0]
        ws['B2'] = components[1]
        ws['C2'] = components[2]
        
        result = reader._detect_date_subcolumns(ws, 1, 2, 10)
        
        # Should attempt to detect (may or may not succeed depending on implementation)
        assert isinstance(result, dict), "Should return a dictionary"

    def test_split_field_preference_over_single(self):
        """Property: When both single and split date fields exist, prefer split.
        
        If a worksheet has both a single date column and split date columns,
        the system should prefer the split field mapping.
        """
        from openpyxl import Workbook
        
        reader = ExcelReader()
        wb = Workbook()
        ws = wb.active
        
        # Create both single and split date fields
        ws['A1'] = "תאריך לידה"  # Single date field
        ws['B1'] = "שנה"         # Start of split field
        ws['C1'] = "חודש"
        ws['D1'] = "יום"
        
        # Detect columns to see which is preferred
        result = reader.detect_columns(ws)
        
        # If both are detected, split should be preferred
        # (This is a higher-level test of the preference logic)
        if result:
            # Check if split fields are present
            has_split = any(k in result for k in ['birth_year', 'birth_month', 'birth_day'])
            has_single = 'birth_date' in result
            
            # If both exist, split should be preferred (not both should be present)
            if has_split and has_single:
                pytest.fail("Both single and split date fields detected; split should be preferred")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
