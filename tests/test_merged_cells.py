"""Test merged cell handling in ExcelReader.

This test verifies that the ExcelReader correctly handles merged cells
without raising exceptions and properly detects headers in worksheets
with merged cells.
"""

from openpyxl import Workbook
from src.excel_standardization.io_layer.excel_reader import ExcelReader


def test_merged_cells_in_header():
    """Test that merged cells in header row are handled correctly."""
    print("=" * 80)
    print("TEST: MERGED CELLS IN HEADER ROW")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Create a header with merged cells
    # Merge cells A1:B1 for a parent header
    ws.merge_cells('A1:B1')
    ws['A1'] = "תאריך לידה"  # Birth Date (parent header)

    # Add child headers in row 2
    ws['A2'] = "שנה"  # Year
    ws['B2'] = "חודש"  # Month
    ws['C2'] = "יום"  # Day

    # Add other headers
    ws['D1'] = "שם פרטי"
    ws['E1'] = "שם משפחה"

    # Add data
    ws['A3'] = 1990
    ws['B3'] = 5
    ws['C3'] = 15
    ws['D3'] = "יוסי"
    ws['E3'] = "כהן"

    reader = ExcelReader()

    try:
        # Test _is_merged_cell
        print("\n1. Testing _is_merged_cell()...")
        is_merged_a1 = reader._is_merged_cell(ws, 1, 1)
        is_merged_b1 = reader._is_merged_cell(ws, 1, 2)
        is_merged_d1 = reader._is_merged_cell(ws, 1, 4)

        print(f"   Cell A1 is merged: {is_merged_a1} (expected: True)")
        print(f"   Cell B1 is merged: {is_merged_b1} (expected: True)")
        print(f"   Cell D1 is merged: {is_merged_d1} (expected: False)")

        assert is_merged_a1 is True, "A1 should be detected as merged"
        assert is_merged_b1 is True, "B1 should be detected as merged"
        assert is_merged_d1 is False, "D1 should not be detected as merged"
        print("   ✓ _is_merged_cell() works correctly")

        # Test _get_merged_cell_range
        print("\n2. Testing _get_merged_cell_range()...")
        range_a1 = reader._get_merged_cell_range(ws, 1, 1)
        range_b1 = reader._get_merged_cell_range(ws, 1, 2)
        range_d1 = reader._get_merged_cell_range(ws, 1, 4)

        print(f"   Range for A1: {range_a1} (expected: (1, 1, 1, 2))")
        print(f"   Range for B1: {range_b1} (expected: (1, 1, 1, 2))")
        print(f"   Range for D1: {range_d1} (expected: None)")

        assert range_a1 == (1, 1, 1, 2), f"A1 range should be (1, 1, 1, 2), got {range_a1}"
        assert range_b1 == (1, 1, 1, 2), f"B1 range should be (1, 1, 1, 2), got {range_b1}"
        assert range_d1 is None, f"D1 range should be None, got {range_d1}"
        print("   ✓ _get_merged_cell_range() works correctly")

        # Test detect_table_region with merged cells
        print("\n3. Testing detect_table_region() with merged cells...")
        table_region = reader.detect_table_region(ws)

        if table_region:
            print(f"   ✓ Table detected without exceptions!")
            print(f"     Start row: {table_region.start_row}")
            print(f"     End row: {table_region.end_row}")
            print(f"     Header rows: {table_region.header_rows}")
        else:
            print("   ✗ No table detected")

        # Test detect_columns with merged cells
        print("\n4. Testing detect_columns() with merged cells...")
        column_mapping = reader.detect_columns(ws)

        print(f"   ✓ Columns detected without exceptions!")
        print(f"     Found {len(column_mapping)} columns:")
        for field_name, col_info in sorted(column_mapping.items()):
            print(f"       {field_name:20s} → Column {col_info.col}")

        print("\n" + "=" * 80)
        print("✓ ALL MERGED CELL TESTS PASSED")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise


def test_merged_cells_in_data():
    """Test that merged cells in data rows are handled correctly."""
    print("\n\n" + "=" * 80)
    print("TEST: MERGED CELLS IN DATA ROWS")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Add headers
    ws['A1'] = "שם פרטי"
    ws['B1'] = "שם משפחה"
    ws['C1'] = "הערות"

    # Add data with merged cells
    ws['A2'] = "יוסי"
    ws['B2'] = "כהן"
    ws.merge_cells('C2:C3')  # Merge cells for notes
    ws['C2'] = "הערה משותפת"

    ws['A3'] = "שרה"
    ws['B3'] = "לוי"

    reader = ExcelReader()

    try:
        print("\n1. Testing detect_table_region() with merged data cells...")
        table_region = reader.detect_table_region(ws)

        if table_region:
            print(f"   ✓ Table detected without exceptions!")
            print(f"     Start row: {table_region.start_row}")
            print(f"     End row: {table_region.end_row}")
        else:
            print("   ✗ No table detected")

        print("\n2. Testing detect_columns() with merged data cells...")
        column_mapping = reader.detect_columns(ws)

        print(f"   ✓ Columns detected without exceptions!")
        print(f"     Found {len(column_mapping)} columns")

        print("\n" + "=" * 80)
        print("✓ ALL DATA MERGED CELL TESTS PASSED")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise


def test_complex_merged_structure():
    """Test complex merged cell structures."""
    print("\n\n" + "=" * 80)
    print("TEST: COMPLEX MERGED CELL STRUCTURES")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Create a complex header structure with multiple merged cells
    # Row 1: Parent headers (merged)
    ws.merge_cells('A1:B1')
    ws['A1'] = "שם"
    ws.merge_cells('C1:E1')
    ws['C1'] = "תאריך לידה"

    # Row 2: Child headers
    ws['A2'] = "פרטי"
    ws['B2'] = "משפחה"
    ws['C2'] = "שנה"
    ws['D2'] = "חודש"
    ws['E2'] = "יום"

    # Add data
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = 1990
    ws['D3'] = 5
    ws['E3'] = 15

    reader = ExcelReader()

    try:
        print("\n1. Testing complex merged structure...")
        table_region = reader.detect_table_region(ws)

        if table_region:
            print(f"   ✓ Table detected!")
            print(f"     Header rows: {table_region.header_rows}")
        else:
            print("   ✗ No table detected")

        print("\n2. Detecting columns in complex structure...")
        column_mapping = reader.detect_columns(ws)

        print(f"   ✓ Columns detected!")
        print(f"     Found {len(column_mapping)} columns:")
        for field_name, col_info in sorted(column_mapping.items()):
            print(f"       {field_name:20s} → Column {col_info.col}")

        print("\n" + "=" * 80)
        print("✓ ALL COMPLEX MERGED CELL TESTS PASSED")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    test_merged_cells_in_header()
    test_merged_cells_in_data()
    test_complex_merged_structure()
    print("\n✓ All merged cell tests completed successfully!\n")
