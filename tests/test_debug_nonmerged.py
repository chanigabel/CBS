"""Debug script for non-merged multi-row headers."""

from openpyxl import Workbook
from src.excel_standardization.io_layer.excel_reader import ExcelReader


def create_non_merged_multirow_workbook():
    """Create a workbook with multi-row headers but no merged cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Merged Multi-Row"

    # Row 1: Parent headers (not merged)
    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws['C1'] = "תאריך לידה"  # Birth Date parent
    ws['F1'] = "Gender"
    
    # Row 2: Sub-headers
    ws['A2'] = "ID"
    ws['B2'] = "Name"
    ws['C2'] = "שנה"  # Year
    ws['D2'] = "חודש"  # Month
    ws['E2'] = "יום"  # Day
    ws['F2'] = "Gender"
    
    # Add data
    ws['A3'] = 1
    ws['B3'] = "יוסי"
    ws['C3'] = 1990
    ws['D3'] = 5
    ws['E3'] = 15
    ws['F3'] = 1
    
    return wb


def debug_non_merged():
    """Debug non-merged multi-row headers."""
    wb = create_non_merged_multirow_workbook()
    ws = wb.active
    reader = ExcelReader()
    
    print("=" * 80)
    print("DEBUG: Non-Merged Multi-Row Headers")
    print("=" * 80)
    
    # Check what's in each cell
    print("\nRow 1 (Parent headers):")
    for col in range(1, 7):
        cell_value = ws.cell(row=1, column=col).value
        print(f"  Column {col}: {cell_value}")
    
    print("\nRow 2 (Sub-headers):")
    for col in range(1, 7):
        cell_value = ws.cell(row=2, column=col).value
        print(f"  Column {col}: {cell_value}")
    
    # Test row scoring
    print("\nRow scoring:")
    for row in range(1, 3):
        score = reader._score_header_row(ws, row_idx=row, max_col=6)
        print(f"  Row {row}: {score}")
    
    # Test sub-header scoring
    print("\nSub-header scoring:")
    for row in range(2, 3):
        score = reader._score_subheader_row(ws, row_idx=row, max_col=6)
        print(f"  Row {row}: {score}")
    
    # Test table region detection
    print("\nTable region detection:")
    table_region = reader.detect_table_region(ws)
    if table_region:
        print(f"  Start row: {table_region.start_row}")
        print(f"  Header rows: {table_region.header_rows}")
        print(f"  Data start row: {table_region.data_start_row}")
    
    # Test column detection
    print("\nColumn detection:")
    column_mapping = reader.detect_columns(ws)
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"  {field_name:20s} → Column {col_info.col}")


if __name__ == "__main__":
    debug_non_merged()
