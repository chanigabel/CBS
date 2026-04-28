"""Tests for standardizationOrchestrator JSON pipeline paths.

Validates process_workbook_json, export_normalized_json, and export_raw_json
end-to-end using real temp Excel files. Confirms all engines are wired and
corrected fields appear in output.
"""

import os
import json
import tempfile
import pytest
from openpyxl import Workbook, load_workbook

from src.excel_standardization.orchestrator import standardizationOrchestrator


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_excel(path: str, rows: list, headers: list, sheet_name: str = "Sheet1") -> None:
    """Write a minimal Excel file with one header row and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, start=2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
    wb.save(path)


HEBREW_HEADERS = ["שם פרטי", "שם משפחה", "מין", "מספר זהות", "דרכון"]
SAMPLE_ROWS = [
    ["  יוסי  ", "כהן  ", "ז", "123456782", "A1234567"],
    ["שרה123",   "לוי",   "נ", "234567891", "B2345678"],
]


# ---------------------------------------------------------------------------
# process_workbook_json
# ---------------------------------------------------------------------------

class TestProcessWorkbookJson:
    def test_output_excel_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "output.xlsx")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            orch = standardizationOrchestrator()
            orch.process_workbook_json(inp, out)

            assert os.path.exists(out)

    def test_output_has_corrected_columns(self):
        """Output workbook must have original Hebrew headers plus '- מתוקן' corrected columns."""
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "output.xlsx")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().process_workbook_json(inp, out)

            wb = load_workbook(out)
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, 20) if ws.cell(row=1, column=c).value]

            # Original columns must be preserved
            assert "שם פרטי" in headers
            assert "שם משפחה" in headers
            assert "מין" in headers
            # Corrected columns must be inserted beside originals
            assert "שם פרטי - מתוקן" in headers
            assert "שם משפחה - מתוקן" in headers

    def test_name_trimming_applied(self):
        """Corrected name column must contain trimmed value; original column must be unchanged."""
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "output.xlsx")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().process_workbook_json(inp, out)

            wb = load_workbook(out)
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, 20) if ws.cell(row=1, column=c).value]

            orig_col = headers.index("שם פרטי") + 1
            corr_col = headers.index("שם פרטי - מתוקן") + 1

            # Original value must be untouched
            assert ws.cell(row=2, column=orig_col).value == "  יוסי  "
            # Corrected value must be trimmed
            assert ws.cell(row=2, column=corr_col).value == "יוסי"

    def test_gender_standardization_applied(self):
        """Corrected gender column must contain numeric codes; original column must be unchanged."""
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "output.xlsx")
            # Use the full multi-line gender header that the processor expects
            headers = ["שם פרטי", "שם משפחה", "מין\n1=זכר\n2+נקבה", "מספר זהות", "דרכון"]
            make_excel(inp, SAMPLE_ROWS, headers)

            standardizationOrchestrator().process_workbook_json(inp, out)

            wb = load_workbook(out)
            ws = wb.active
            headers_out = [ws.cell(row=1, column=c).value for c in range(1, 20) if ws.cell(row=1, column=c).value]

            orig_col = headers_out.index("מין\n1=זכר\n2+נקבה") + 1
            corr_col = headers_out.index("מין - מתוקן") + 1

            assert ws.cell(row=2, column=orig_col).value == "ז"
            assert ws.cell(row=2, column=corr_col).value == 1   # male
            assert ws.cell(row=3, column=corr_col).value == 2   # female

    def test_file_not_found_raises(self):
        orch = standardizationOrchestrator()
        with pytest.raises((FileNotFoundError, IOError)):
            orch.process_workbook_json("/nonexistent/path.xlsx", "/tmp/out.xlsx")

    def test_raw_json_side_file_created(self):
        """raw_dataset.json must be written alongside the output Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "output.xlsx")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().process_workbook_json(inp, out)

            assert os.path.exists(os.path.join(tmpdir, "raw_dataset.json"))


# ---------------------------------------------------------------------------
# export_normalized_json
# ---------------------------------------------------------------------------

class TestExportNormalizedJson:
    def test_json_file_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "normalized.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_normalized_json(inp, out)

            assert os.path.exists(out)

    def test_json_contains_corrected_fields(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "normalized.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_normalized_json(inp, out)

            with open(out, encoding="utf-8") as f:
                data = json.load(f)

            # Find first row across sheets
            sheets = data.get("sheets", [data])  # handle both wrapped and flat formats
            first_row = None
            for sheet in sheets:
                rows = sheet.get("rows", [])
                if rows:
                    first_row = rows[0]
                    break

            assert first_row is not None
            assert "first_name_corrected" in first_row
            assert "gender_corrected" in first_row

    def test_json_name_trimmed(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "normalized.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_normalized_json(inp, out)

            with open(out, encoding="utf-8") as f:
                data = json.load(f)

            sheets = data.get("sheets", [data])
            first_row = sheets[0]["rows"][0]

            assert first_row["first_name"] == "  יוסי  "
            assert first_row["first_name_corrected"] == "יוסי"

    def test_file_not_found_raises(self):
        orch = standardizationOrchestrator()
        with pytest.raises((FileNotFoundError, IOError)):
            orch.export_normalized_json("/nonexistent/path.xlsx", "/tmp/out.json")


# ---------------------------------------------------------------------------
# export_raw_json
# ---------------------------------------------------------------------------

class TestExportRawJson:
    def test_json_file_created(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "raw.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_raw_json(inp, out)

            assert os.path.exists(out)

    def test_raw_json_has_no_corrected_fields(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "raw.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_raw_json(inp, out)

            with open(out, encoding="utf-8") as f:
                data = json.load(f)

            sheets = data.get("sheets", [data])
            first_row = sheets[0]["rows"][0]

            assert "first_name_corrected" not in first_row
            assert "gender_corrected" not in first_row

    def test_raw_json_preserves_original_values(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            inp = os.path.join(tmpdir, "input.xlsx")
            out = os.path.join(tmpdir, "raw.json")
            make_excel(inp, SAMPLE_ROWS, HEBREW_HEADERS)

            standardizationOrchestrator().export_raw_json(inp, out)

            with open(out, encoding="utf-8") as f:
                data = json.load(f)

            sheets = data.get("sheets", [data])
            first_row = sheets[0]["rows"][0]

            assert first_row["first_name"] == "  יוסי  "

    def test_file_not_found_raises(self):
        orch = standardizationOrchestrator()
        with pytest.raises((FileNotFoundError, IOError)):
            orch.export_raw_json("/nonexistent/path.xlsx", "/tmp/out.json")
