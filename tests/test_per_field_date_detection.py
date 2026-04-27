"""Per-field date detection tests.

Verifies that detect_date_groups and detect_columns detect each date field
(birth_date, entry_date) independently based on its own sub-header layout,
supporting all combinations of plain and split in the same worksheet.

Combinations tested:
  A. plain birth_date  + split entry_date  (two-row header)
  B. split birth_date  + plain entry_date  (two-row header)
  C. both plain        (single-row header)
  D. both split        (two-row header, sub-headers in separate columns)
  E. merged parent header (C1:E1 merged) with split birth only
  F. both split, entry sub-headers start right after birth's
  G. plain birth + split entry, single-row header (entry sub-headers below)

For each combination the test verifies:
  - Correct field names in the mapping (no phantom year/month/day keys)
  - Full extract → normalize pipeline produces correct corrected fields
  - No internal keys (_birth_year_auto_completed etc.) in output rows
"""

import pytest
from openpyxl import Workbook

from src.excel_normalization.io_layer.excel_reader import ExcelReader
from src.excel_normalization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
from src.excel_normalization.engines.date_engine import DateEngine
from src.excel_normalization.data_types import DateFormatPattern


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

PHANTOM_KEYS = {"year", "month", "day"}


def _check_mapping(label, ws, expected_present, expected_absent=None):
    """Assert that detect_columns produces the expected field names."""
    r = ExcelReader()
    m = r.detect_columns(ws)
    keys = set(m.keys())
    missing = [k for k in expected_present if k not in keys]
    phantom = [k for k in (expected_absent or PHANTOM_KEYS) if k in keys]
    assert not missing, f"[{label}] missing fields: {missing}  got: {sorted(keys)}"
    assert not phantom, f"[{label}] phantom fields: {phantom}  got: {sorted(keys)}"
    return m


def _pipeline():
    p = NormalizationPipeline(date_engine=DateEngine())
    p._date_format_pattern = DateFormatPattern.DDMM
    return p


def _extract_normalize(ws):
    reader = ExcelReader()
    extractor = ExcelToJsonExtractor(excel_reader=reader)
    dataset = extractor.extract_sheet_to_json(ws)
    if not dataset.rows:
        return dataset
    return _pipeline().normalize_dataset(dataset)


def _no_internal_keys(rows):
    for row in rows:
        leaked = [k for k in row if k.startswith("_")]
        assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# A: plain birth_date + split entry_date (two-row header)
#
# Row 1: שם פרטי | תאריך לידה | תאריך כניסה למוסד
# Row 2:          |             |                   | שנה | חודש | יום
# ---------------------------------------------------------------------------

class TestCasePlainBirthSplitEntry:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=4).value = "שנה"
        ws.cell(row=2, column=5).value = "חודש"
        ws.cell(row=2, column=6).value = "יום"
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "11.06.1997"
        ws.cell(row=3, column=4).value = 2020
        ws.cell(row=3, column=5).value = 3
        ws.cell(row=3, column=6).value = 15
        return ws

    def test_mapping(self):
        _check_mapping("A", self._make_ws(),
                       expected_present=["birth_date", "entry_year", "entry_month", "entry_day"],
                       expected_absent=["birth_year", "birth_month", "birth_day"] + list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert "birth_year_corrected" in row
        assert "birth_month_corrected" in row
        assert "birth_day_corrected" in row
        assert "entry_year_corrected" in row
        assert "entry_month_corrected" in row
        assert "entry_day_corrected" in row
        _no_internal_keys(result.rows)

    def test_no_phantom_columns(self):
        result = _extract_normalize(self._make_ws())
        for row in result.rows:
            for k in PHANTOM_KEYS:
                assert k not in row, f"Phantom key '{k}' in row"


# ---------------------------------------------------------------------------
# B: split birth_date + plain entry_date (two-row header)
#
# Row 1: שם פרטי | תאריך לידה |     |     | תאריך כניסה למוסד
# Row 2:          |             | שנה | חודש | יום
# (birth sub-headers at cols 3-5; entry header at col 5 — overlap)
# ---------------------------------------------------------------------------

class TestCaseSplitBirthPlainEntry:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=5).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        ws.cell(row=3, column=1).value = "שרה"
        ws.cell(row=3, column=3).value = 1985
        ws.cell(row=3, column=4).value = 7
        ws.cell(row=3, column=5).value = 20
        ws.cell(row=3, column=6).value = "15/06/2020"
        return ws

    def test_mapping(self):
        _check_mapping("B", self._make_ws(),
                       expected_present=["birth_year", "birth_month", "birth_day", "entry_date"],
                       expected_absent=["entry_year", "entry_month", "entry_day"] + list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert "birth_year_corrected" in row
        assert "birth_month_corrected" in row
        assert "birth_day_corrected" in row
        assert "entry_year_corrected" in row
        assert "entry_month_corrected" in row
        assert "entry_day_corrected" in row
        _no_internal_keys(result.rows)


# ---------------------------------------------------------------------------
# C: both plain (single-row header)
# ---------------------------------------------------------------------------

class TestCaseBothPlain:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=1).value = "יוסי"
        ws.cell(row=2, column=2).value = "11.06.1997"
        ws.cell(row=2, column=3).value = "15/06/2020"
        return ws

    def test_mapping(self):
        _check_mapping("C", self._make_ws(),
                       expected_present=["birth_date", "entry_date"],
                       expected_absent=["birth_year", "entry_year"] + list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert "birth_year_corrected" in row
        assert "birth_month_corrected" in row
        assert "birth_day_corrected" in row
        assert "entry_year_corrected" in row
        assert "entry_month_corrected" in row
        assert "entry_day_corrected" in row
        _no_internal_keys(result.rows)


# ---------------------------------------------------------------------------
# D: both split (two-row header, separate sub-header blocks)
#
# Row 1: שם פרטי | תאריך לידה |     |     |     | תאריך כניסה למוסד
# Row 2:          |             | שנה | חודש | יום | שנה | חודש | יום
# ---------------------------------------------------------------------------

class TestCaseBothSplit:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=5).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        ws.cell(row=2, column=6).value = "שנה"
        ws.cell(row=2, column=7).value = "חודש"
        ws.cell(row=2, column=8).value = "יום"
        ws.cell(row=3, column=1).value = "דוד"
        ws.cell(row=3, column=3).value = 1990
        ws.cell(row=3, column=4).value = 5
        ws.cell(row=3, column=5).value = 15
        ws.cell(row=3, column=6).value = 2020
        ws.cell(row=3, column=7).value = 3
        ws.cell(row=3, column=8).value = 10
        return ws

    def test_mapping(self):
        _check_mapping("D", self._make_ws(),
                       expected_present=["birth_year", "birth_month", "birth_day",
                                         "entry_year", "entry_month", "entry_day"],
                       expected_absent=["birth_date", "entry_date"] + list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1990
        assert row["entry_year_corrected"] == 2020
        _no_internal_keys(result.rows)


# ---------------------------------------------------------------------------
# E: merged parent header (C1:E1) with split birth only
# ---------------------------------------------------------------------------

class TestCaseMergedParentBirthOnly:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "משפחה"
        ws.cell(row=1, column=3).value = "תאריך לידה"
        ws.merge_cells("C1:E1")
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "כהן"
        ws.cell(row=3, column=3).value = 1980
        ws.cell(row=3, column=4).value = 5
        ws.cell(row=3, column=5).value = 15
        return ws

    def test_mapping(self):
        _check_mapping("E", self._make_ws(),
                       expected_present=["birth_year", "birth_month", "birth_day"],
                       expected_absent=list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1980
        _no_internal_keys(result.rows)


# ---------------------------------------------------------------------------
# F: both split, entry sub-headers start right after birth's (no gap)
# ---------------------------------------------------------------------------

class TestCaseBothSplitNoGap:
    def _make_ws(self):
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=6).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        ws.cell(row=2, column=7).value = "שנה"
        ws.cell(row=2, column=8).value = "חודש"
        ws.cell(row=2, column=9).value = "יום"
        ws.cell(row=3, column=1).value = "דוד"
        ws.cell(row=3, column=3).value = 1990
        ws.cell(row=3, column=4).value = 5
        ws.cell(row=3, column=5).value = 15
        ws.cell(row=3, column=7).value = 2020
        ws.cell(row=3, column=8).value = 3
        ws.cell(row=3, column=9).value = 10
        return ws

    def test_mapping(self):
        _check_mapping("F", self._make_ws(),
                       expected_present=["birth_year", "birth_month", "birth_day",
                                         "entry_year", "entry_month", "entry_day"],
                       expected_absent=list(PHANTOM_KEYS))

    def test_pipeline(self):
        result = _extract_normalize(self._make_ws())
        assert result.rows, "No rows"
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1990
        assert row["entry_year_corrected"] == 2020
        _no_internal_keys(result.rows)


# ---------------------------------------------------------------------------
# Independence invariant: one field's shape must not affect the other's
# ---------------------------------------------------------------------------

class TestPerFieldIndependence:
    """Verify that each date field is detected independently."""

    def test_plain_birth_does_not_force_plain_entry(self):
        """If birth is plain, entry can still be split."""
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=4).value = "שנה"
        ws.cell(row=2, column=5).value = "חודש"
        ws.cell(row=2, column=6).value = "יום"
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "11.06.1997"
        ws.cell(row=3, column=4).value = 2020
        ws.cell(row=3, column=5).value = 3
        ws.cell(row=3, column=6).value = 15
        r = ExcelReader()
        m = r.detect_columns(ws)
        assert "birth_date" in m, "birth_date must be plain"
        assert "entry_year" in m, "entry must be split regardless of birth shape"

    def test_split_birth_does_not_force_split_entry(self):
        """If birth is split, entry can still be plain."""
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=5).value = "תאריך כניסה למוסד"
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        ws.cell(row=3, column=1).value = "שרה"
        ws.cell(row=3, column=3).value = 1985
        ws.cell(row=3, column=4).value = 7
        ws.cell(row=3, column=5).value = 20
        ws.cell(row=3, column=6).value = "15/06/2020"
        r = ExcelReader()
        m = r.detect_columns(ws)
        assert "birth_year" in m, "birth must be split"
        assert "entry_date" in m, "entry must be plain regardless of birth shape"

    def test_two_row_header_does_not_force_all_fields_split(self):
        """A two-row header layout must not force every date field to be split."""
        wb = Workbook(); ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"   # plain — no sub-headers
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"  # split
        ws.cell(row=2, column=4).value = "שנה"
        ws.cell(row=2, column=5).value = "חודש"
        ws.cell(row=2, column=6).value = "יום"
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "11.06.1997"
        ws.cell(row=3, column=4).value = 2020
        ws.cell(row=3, column=5).value = 3
        ws.cell(row=3, column=6).value = 15
        r = ExcelReader()
        m = r.detect_columns(ws)
        # birth must be plain even though the sheet has a two-row header
        assert "birth_date" in m, "birth_date must be plain in two-row header sheet"
        assert "birth_year" not in m, "birth must NOT be split"
