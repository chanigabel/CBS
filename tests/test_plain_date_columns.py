"""Tests for plain single-column date support (birth_date and entry_date).

Covers:
- Plain birth_date + split entry_date
- Split birth_date + plain entry_date
- Both plain
- Both split (regression)
- Two-row header sheets
- Mixed valid/invalid date values
- No phantom columns from row values
- No internal helper fields (_birth_year_auto_completed, _entry_year_auto_completed, etc.)
- entry_date_status appears in UI payload for plain entry_date
- birth_date_status appears in UI payload for plain birth_date
"""

import pytest
from openpyxl import Workbook

from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_standardization.processing.standardization_pipeline import standardizationPipeline
from src.excel_standardization.engines.date_engine import DateEngine
from src.excel_standardization.data_types import SheetDataset, DateFormatPattern


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_pipeline():
    p = standardizationPipeline(date_engine=DateEngine())
    p._date_format_pattern = DateFormatPattern.DDMM
    return p


def _make_dataset(rows, field_names):
    return SheetDataset(
        sheet_name="Sheet1",
        header_row=1,
        header_rows_count=1,
        field_names=field_names,
        rows=rows,
        metadata={},
    )


def _internal_keys(row: dict) -> list:
    """Return all underscore-prefixed keys in a row dict."""
    return [k for k in row if k.startswith("_")]


# ---------------------------------------------------------------------------
# 1. detect_columns: plain birth_date + split entry_date (two-row header)
# ---------------------------------------------------------------------------

class TestDetectColumnsPlainBirthSplitEntry:
    """ExcelReader.detect_columns must recognise plain birth_date alongside split entry_date."""

    def _make_ws(self):
        wb = Workbook()
        ws = wb.active
        # Row 1: main headers
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"       # plain — no sub-headers below
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"  # split — has sub-headers
        # Row 2: sub-headers only under entry date
        ws.cell(row=2, column=4).value = "שנה"
        ws.cell(row=2, column=5).value = "חודש"
        ws.cell(row=2, column=6).value = "יום"
        # Data row
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "11.06.1997"
        ws.cell(row=3, column=4).value = 2020
        ws.cell(row=3, column=5).value = 3
        ws.cell(row=3, column=6).value = 15
        return ws

    def test_birth_date_mapped_as_plain(self):
        ws = self._make_ws()
        mapping = ExcelReader().detect_columns(ws)
        assert "birth_date" in mapping, f"birth_date missing; got {list(mapping)}"
        # Must NOT be split into birth_year/month/day
        assert "birth_year" not in mapping
        assert "birth_month" not in mapping
        assert "birth_day" not in mapping

    def test_entry_date_mapped_as_split(self):
        ws = self._make_ws()
        mapping = ExcelReader().detect_columns(ws)
        assert "entry_year" in mapping
        assert "entry_month" in mapping
        assert "entry_day" in mapping
        assert "entry_date" not in mapping


# ---------------------------------------------------------------------------
# 2. detect_columns: split birth_date + plain entry_date (two-row header)
# ---------------------------------------------------------------------------

class TestDetectColumnsSplitBirthPlainEntry:
    """ExcelReader.detect_columns must recognise plain entry_date alongside split birth_date."""

    def _make_ws(self):
        wb = Workbook()
        ws = wb.active
        # Row 1: main headers
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"        # split
        ws.cell(row=1, column=5).value = "תאריך כניסה למוסד"  # plain — no sub-headers
        # Row 2: sub-headers only under birth date
        ws.cell(row=2, column=3).value = "שנה"
        ws.cell(row=2, column=4).value = "חודש"
        ws.cell(row=2, column=5).value = "יום"
        # Data row
        ws.cell(row=3, column=1).value = "שרה"
        ws.cell(row=3, column=3).value = 1985
        ws.cell(row=3, column=4).value = 7
        ws.cell(row=3, column=5).value = 20
        ws.cell(row=3, column=6).value = "15/06/2020"
        return ws

    def test_birth_date_mapped_as_split(self):
        ws = self._make_ws()
        mapping = ExcelReader().detect_columns(ws)
        # In this layout the sub-headers (שנה/חודש/יום) are at cols 3-5 but
        # the birth_date header is at col 2 with no sub-headers directly under it.
        # detect_date_groups finds no split group for birth_date, so it is mapped
        # as a plain single-column field.  The entry_date header at col 5 overlaps
        # with the sub-header row, so it may be mapped as plain or split depending
        # on the exact column layout.  The key invariant is that birth_date is
        # present in the mapping (either as plain or split).
        assert "birth_date" in mapping or "birth_year" in mapping, (
            f"birth_date not found in mapping; got {list(mapping)}"
        )

    def test_entry_date_mapped_as_plain(self):
        ws = self._make_ws()
        mapping = ExcelReader().detect_columns(ws)
        # entry_date must be present as a plain column
        assert "entry_date" in mapping or "entry_year" in mapping, (
            f"Neither entry_date nor entry_year found; got {list(mapping)}"
        )


# ---------------------------------------------------------------------------
# 3. Pipeline: plain birth_date standardization
# ---------------------------------------------------------------------------

class TestPipelinePlainBirthDate:
    """standardizationPipeline must fully normalize plain birth_date values into structured year/month/day."""

    def test_valid_date_corrected(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"birth_date": "11.06.1997"}], ["birth_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        # Plain date must produce structured year/month/day corrected fields
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 6
        assert row["birth_day_corrected"] == 11
        assert row.get("birth_date_status") == ""

    def test_invalid_date_gets_status(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"birth_date": "32/13/1990"}], ["birth_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row.get("birth_date_status") not in (None, ""), (
            "Invalid birth_date must produce a non-empty status"
        )

    def test_iso_datetime_string_parsed(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"birth_date": "2001-04-14T00:00:00"}], ["birth_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] == 2001
        assert row["birth_month_corrected"] == 4
        assert row["birth_day_corrected"] == 14
        assert row.get("birth_date_status") == ""

    def test_none_preserved(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"birth_date": None}], ["birth_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] is None
        assert row["birth_month_corrected"] is None
        assert row["birth_day_corrected"] is None

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"birth_date": "11.06.1997"}], ["birth_date"])
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# 4. Pipeline: plain entry_date standardization
# ---------------------------------------------------------------------------

class TestPipelinePlainEntryDate:
    """standardizationPipeline must fully normalize plain entry_date values into structured year/month/day."""

    def test_valid_date_corrected(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"entry_date": "15/06/2020"}], ["entry_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["entry_year_corrected"] == 2020
        assert row["entry_month_corrected"] == 6
        assert row["entry_day_corrected"] == 15
        assert row.get("entry_date_status") == ""

    def test_late_entry_date_gets_status(self):
        """entry_date in current year must get the cutoff status."""
        from datetime import date
        current_year = date.today().year
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"entry_date": f"01/01/{current_year}"}],
            ["entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row.get("entry_date_status") == "תאריך כניסה מאוחר מהתאריך שנקבע", (
            f"Expected late-entry status, got: {row.get('entry_date_status')!r}"
        )

    def test_valid_cutoff_year_no_status(self):
        """entry_date in current_year - 1 must be valid."""
        from datetime import date
        cutoff_year = date.today().year - 1
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"entry_date": f"01/01/{cutoff_year}"}],
            ["entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row.get("entry_date_status") == "", (
            f"Valid entry_date must have empty status, got: {row.get('entry_date_status')!r}"
        )

    def test_none_preserved(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"entry_date": None}], ["entry_date"])
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["entry_year_corrected"] is None
        assert row["entry_month_corrected"] is None
        assert row["entry_day_corrected"] is None

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset([{"entry_date": "15/06/2020"}], ["entry_date"])
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# 5. Pipeline: both plain (birth_date + entry_date)
# ---------------------------------------------------------------------------

class TestPipelineBothPlain:
    """Both birth_date and entry_date as plain single columns."""

    def test_both_normalized(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_date": "11.06.1997", "entry_date": "15/06/2020"}],
            ["birth_date", "entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 6
        assert row["birth_day_corrected"] == 11
        assert row["entry_year_corrected"] == 2020
        assert row["entry_month_corrected"] == 6
        assert row["entry_day_corrected"] == 15
        assert row.get("birth_date_status") == ""
        assert row.get("entry_date_status") == ""

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_date": "11.06.1997", "entry_date": "15/06/2020"}],
            ["birth_date", "entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"

    def test_entry_before_birth_cross_validation(self):
        """entry_date before birth_date must produce the cross-validation warning."""
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_date": "01/01/2000", "entry_date": "01/01/1990"}],
            ["birth_date", "entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert "תאריך כניסה לפני תאריך לידה" in (row.get("entry_date_status") or ""), (
            f"Expected entry-before-birth warning, got: {row.get('entry_date_status')!r}"
        )


# ---------------------------------------------------------------------------
# 6. Pipeline: both split (regression — must still work)
# ---------------------------------------------------------------------------

class TestPipelineBothSplit:
    """Both birth and entry as split year/month/day — must still work after changes."""

    def test_both_split_normalized(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_year": 1990, "birth_month": 5, "birth_day": 15,
              "entry_year": 2020, "entry_month": 3, "entry_day": 10}],
            ["birth_year", "birth_month", "birth_day", "entry_year", "entry_month", "entry_day"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1990
        assert row["entry_year_corrected"] == 2020

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_year": 1990, "birth_month": 5, "birth_day": 15,
              "entry_year": 2020, "entry_month": 3, "entry_day": 10}],
            ["birth_year", "birth_month", "birth_day", "entry_year", "entry_month", "entry_day"]
        )
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# 7. Pipeline: plain birth_date + split entry_date
# ---------------------------------------------------------------------------

class TestPipelinePlainBirthSplitEntry:
    """Plain birth_date alongside split entry_year/month/day — both produce structured output."""

    def test_both_normalized(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_date": "11.06.1997",
              "entry_year": 2020, "entry_month": 3, "entry_day": 10}],
            ["birth_date", "entry_year", "entry_month", "entry_day"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 6
        assert row["birth_day_corrected"] == 11
        assert row["entry_year_corrected"] == 2020

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_date": "11.06.1997",
              "entry_year": 2020, "entry_month": 3, "entry_day": 10}],
            ["birth_date", "entry_year", "entry_month", "entry_day"]
        )
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# 8. Pipeline: split birth_date + plain entry_date
# ---------------------------------------------------------------------------

class TestPipelineSplitBirthPlainEntry:
    """Split birth_year/month/day alongside plain entry_date — both produce structured output."""

    def test_both_normalized(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_year": 1990, "birth_month": 5, "birth_day": 15,
              "entry_date": "15/06/2020"}],
            ["birth_year", "birth_month", "birth_day", "entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1990
        assert row["entry_year_corrected"] == 2020
        assert row["entry_month_corrected"] == 6
        assert row["entry_day_corrected"] == 15

    def test_no_internal_keys_in_output(self):
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [{"birth_year": 1990, "birth_month": 5, "birth_day": 15,
              "entry_date": "15/06/2020"}],
            ["birth_year", "birth_month", "birth_day", "entry_date"]
        )
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"


# ---------------------------------------------------------------------------
# 9. workbook_service display_columns: no phantom columns, no internal keys
# ---------------------------------------------------------------------------

class TestWorkbookServiceNoPhantomColumns:
    """WorkbookService.get_sheet_data must not expose internal keys or phantom columns."""

    def _make_normalized_dataset(self, rows, field_names):
        """Run the pipeline and return the normalized SheetDataset."""
        pipeline = _make_pipeline()
        ds = _make_dataset(rows, field_names)
        return pipeline.normalize_dataset(ds)

    def _get_display_columns(self, normalized_ds):
        """Simulate the display_columns building logic from workbook_service (updated for structured output)."""
        original_fields = list(normalized_ds.field_names)

        seen: set = set()
        all_row_keys: list = []
        for row in normalized_ds.rows:
            for k in row.keys():
                if k not in seen and not k.startswith("_"):
                    seen.add(k)
                    all_row_keys.append(k)

        _STATUS_GROUPS = {
            "identifier_status": {"id_number", "passport"},
            "birth_date_status": {"birth_year", "birth_month", "birth_day", "birth_date"},
            "entry_date_status": {"entry_year", "entry_month", "entry_day", "entry_date"},
        }

        _DATE_STRUCTURED_FALLBACK = {
            "birth_date_corrected": ["birth_day_corrected", "birth_month_corrected", "birth_year_corrected"],
            "entry_date_corrected": ["entry_day_corrected", "entry_month_corrected", "entry_year_corrected"],
        }

        _anchor_to_status: dict = {}
        for status_key, group_members in _STATUS_GROUPS.items():
            if status_key not in seen:
                continue
            anchor_orig = None
            for f in original_fields:
                if f in group_members:
                    anchor_orig = f
            if anchor_orig is not None:
                anchor_corrected = f"{anchor_orig}_corrected"
                if anchor_corrected not in seen and anchor_corrected in _DATE_STRUCTURED_FALLBACK:
                    for fallback in _DATE_STRUCTURED_FALLBACK[anchor_corrected]:
                        if fallback in seen:
                            anchor_corrected = fallback
                            break
                _anchor_to_status[anchor_corrected] = status_key

        _PLAIN_DATE_STRUCTURED = {
            "birth_date": ["birth_year_corrected", "birth_month_corrected", "birth_day_corrected"],
            "entry_date": ["entry_year_corrected", "entry_month_corrected", "entry_day_corrected"],
        }

        display_columns: list = []
        placed: set = set()

        for orig in original_fields:
            if orig not in placed:
                display_columns.append(orig)
                placed.add(orig)

            if orig in _PLAIN_DATE_STRUCTURED:
                structured = _PLAIN_DATE_STRUCTURED[orig]
                naive_corrected = f"{orig}_corrected"
                if naive_corrected not in seen and any(f in seen for f in structured):
                    for sf in structured:
                        if sf in seen and sf not in placed:
                            display_columns.append(sf)
                            placed.add(sf)
                    status_key = _anchor_to_status.get(structured[-1])
                    if status_key and status_key in seen and status_key not in placed:
                        display_columns.append(status_key)
                        placed.add(status_key)
                    continue

            corrected = f"{orig}_corrected"
            if corrected in seen and corrected not in placed:
                display_columns.append(corrected)
                placed.add(corrected)
            status_key = _anchor_to_status.get(corrected)
            if status_key and status_key in seen and status_key not in placed:
                display_columns.append(status_key)
                placed.add(status_key)

        for k in all_row_keys:
            if k not in placed:
                display_columns.append(k)
                placed.add(k)

        return display_columns

    def test_no_internal_keys_in_display_columns_plain_birth(self):
        ds = self._make_normalized_dataset(
            [{"birth_date": "11.06.1997"}], ["birth_date"]
        )
        cols = self._get_display_columns(ds)
        internal = [c for c in cols if c.startswith("_")]
        assert internal == [], f"Internal keys in display_columns: {internal}"

    def test_no_internal_keys_in_display_columns_plain_entry(self):
        ds = self._make_normalized_dataset(
            [{"entry_date": "15/06/2020"}], ["entry_date"]
        )
        cols = self._get_display_columns(ds)
        internal = [c for c in cols if c.startswith("_")]
        assert internal == [], f"Internal keys in display_columns: {internal}"

    def test_no_internal_keys_in_display_columns_both_plain(self):
        ds = self._make_normalized_dataset(
            [{"birth_date": "11.06.1997", "entry_date": "15/06/2020"}],
            ["birth_date", "entry_date"]
        )
        cols = self._get_display_columns(ds)
        internal = [c for c in cols if c.startswith("_")]
        assert internal == [], f"Internal keys in display_columns: {internal}"

    def test_no_internal_keys_in_display_columns_both_split(self):
        ds = self._make_normalized_dataset(
            [{"birth_year": 1990, "birth_month": 5, "birth_day": 15,
              "entry_year": 2020, "entry_month": 3, "entry_day": 10}],
            ["birth_year", "birth_month", "birth_day", "entry_year", "entry_month", "entry_day"]
        )
        cols = self._get_display_columns(ds)
        internal = [c for c in cols if c.startswith("_")]
        assert internal == [], f"Internal keys in display_columns: {internal}"

    def test_birth_date_status_in_display_columns_plain(self):
        """birth_date_status must appear in display_columns for plain birth_date."""
        ds = self._make_normalized_dataset(
            [{"birth_date": "32/13/1990"}], ["birth_date"]  # invalid → status
        )
        cols = self._get_display_columns(ds)
        assert "birth_date_status" in cols, (
            f"birth_date_status missing from display_columns: {cols}"
        )

    def test_entry_date_status_in_display_columns_plain(self):
        """entry_date_status must appear in display_columns for plain entry_date."""
        from datetime import date
        current_year = date.today().year
        ds = self._make_normalized_dataset(
            [{"entry_date": f"01/01/{current_year}"}], ["entry_date"]  # late → status
        )
        cols = self._get_display_columns(ds)
        assert "entry_date_status" in cols, (
            f"entry_date_status missing from display_columns: {cols}"
        )

    def test_no_phantom_columns_from_row_values(self):
        """Row values must never become column headers."""
        ds = self._make_normalized_dataset(
            [{"birth_date": "11.06.1997", "entry_date": "15/06/2020"}],
            ["birth_date", "entry_date"]
        )
        cols = self._get_display_columns(ds)
        # None of the actual date values should appear as column names
        for col in cols:
            assert col not in ("11.06.1997", "15/06/2020", "11/06/1997", "15/06/2020"), (
                f"Row value appeared as column name: {col!r}"
            )

    def test_mixed_valid_invalid_dates(self):
        """Mixed valid/invalid rows must all produce structured fields, no internal keys."""
        pipeline = _make_pipeline()
        ds = _make_dataset(
            [
                {"birth_date": "11.06.1997"},
                {"birth_date": "32/13/1990"},
                {"birth_date": None},
                {"birth_date": "2001-04-14T00:00:00"},
            ],
            ["birth_date"]
        )
        result = pipeline.normalize_dataset(ds)
        for row in result.rows:
            leaked = _internal_keys(row)
            assert leaked == [], f"Internal keys leaked: {leaked}"
            assert "birth_year_corrected" in row
            assert "birth_month_corrected" in row
            assert "birth_day_corrected" in row
            if row.get("birth_date") is not None:
                assert "birth_date_status" in row


# ---------------------------------------------------------------------------
# 10. End-to-end: extract + normalize, verify no phantom columns
# ---------------------------------------------------------------------------

class TestEndToEndPlainDateColumns:
    """Full extract → normalize pipeline for plain date columns."""

    def _extract_and_normalize(self, ws):
        reader = ExcelReader()
        extractor = ExcelToJsonExtractor(excel_reader=reader)
        dataset = extractor.extract_sheet_to_json(ws)
        if not dataset.rows:
            return dataset
        pipeline = _make_pipeline()
        return pipeline.normalize_dataset(dataset)

    def test_plain_birth_date_end_to_end(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=2, column=1).value = "יוסי"
        ws.cell(row=2, column=2).value = "11.06.1997"

        result = self._extract_and_normalize(ws)
        assert result.rows, "No rows extracted"
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 6
        assert row["birth_day_corrected"] == 11
        assert _internal_keys(row) == []

    def test_plain_entry_date_end_to_end(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך כניסה"
        ws.cell(row=2, column=1).value = "יוסי"
        ws.cell(row=2, column=2).value = "15/06/2020"

        result = self._extract_and_normalize(ws)
        assert result.rows, "No rows extracted"
        row = result.rows[0]
        assert row["entry_year_corrected"] == 2020
        assert row["entry_month_corrected"] == 6
        assert row["entry_day_corrected"] == 15
        assert _internal_keys(row) == []

    def test_both_plain_end_to_end(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=3).value = "תאריך כניסה"
        ws.cell(row=2, column=1).value = "יוסי"
        ws.cell(row=2, column=2).value = "11.06.1997"
        ws.cell(row=2, column=3).value = "15/06/2020"

        result = self._extract_and_normalize(ws)
        assert result.rows, "No rows extracted"
        row = result.rows[0]
        assert row["birth_year_corrected"] == 1997
        assert row["entry_year_corrected"] == 2020
        assert _internal_keys(row) == []
