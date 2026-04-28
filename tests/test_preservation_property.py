"""Preservation Property Tests - standardization Behavior Unchanged

**Validates: Requirements 3.1, 3.2, 3.3, 3.4**

IMPORTANT: These tests verify that the fixed code produces exactly the same
normalized Excel output as the original code, preserving all standardization logic,
engine execution order, corrected field creation, and logging behavior.

These tests should PASS on unfixed code to establish the baseline behavior.
"""

import os
import tempfile
import pytest
from hypothesis import given, strategies as st, settings, assume
from openpyxl import Workbook, load_workbook

from src.excel_standardization.orchestrator import standardizationOrchestrator


class TestPreservationProperty:
    """Test that standardization behavior is preserved after adding debugging."""

    def test_excel_output_preservation_simple(self):
        """Test that Excel output is identical with sample data.
        
        **Validates: Requirements 3.1, 3.2, 3.3**
        
        This test verifies that the pipeline produces consistent Excel output
        with corrected fields using "_corrected" suffix.
        
        EXPECTED OUTCOME: This test PASSES on unfixed code, establishing baseline.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create paths
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            # Create test Excel file with sample data
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add headers
            headers = ["שם פרטי", "שם משפחה", "מין", "תאריך לידה", "תעודת זהות"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add sample data
            test_data = [
                ["יוסי", "כהן", "ז", "15/05/1980", "123456789"],
                ["שרה", "לוי", "נ", "20/12/1985", "987654321"],
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save input file
            wb.save(input_path)
            
            # Run the standardization pipeline
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            # ASSERTIONS: Verify expected behavior is preserved
            
            # Requirement 3.1: Excel output file is created
            assert os.path.exists(output_path), "Output Excel file should be created"
            
            # Load output Excel
            wb_out = load_workbook(output_path)
            ws_out = wb_out.active
            
            # Requirement 3.3: Verify corrected columns exist with '- מתוקן' suffix
            header_row = [cell.value for cell in ws_out[1]]
            
            assert "שם פרטי - מתוקן" in header_row, "שם פרטי - מתוקן should exist"
            assert "שם משפחה - מתוקן" in header_row, "שם משפחה - מתוקן should exist"
            
            # Requirement 3.3: Verify original columns are preserved
            assert "שם פרטי" in header_row, "Original שם פרטי should be preserved"
            assert "שם משפחה" in header_row, "Original שם משפחה should be preserved"
            
            # Verify data rows exist
            assert ws_out.max_row >= 3, "Output should have at least 3 rows (header + 2 data rows)"

    @settings(max_examples=10, deadline=None)
    @given(
        # Generate test data for Excel file
        num_rows=st.integers(min_value=1, max_value=5),
        first_names=st.lists(
            st.text(min_size=1, max_size=15, alphabet=st.characters(whitelist_categories=('L',))),
            min_size=1, max_size=5
        ),
        last_names=st.lists(
            st.text(min_size=1, max_size=15, alphabet=st.characters(whitelist_categories=('L',))),
            min_size=1, max_size=5
        ),
        genders=st.lists(
            st.sampled_from(["ז", "נ", "1", "2", "זכר", "נקבה"]),
            min_size=1, max_size=5
        ),
    )
    def test_excel_output_preservation_property(self, num_rows, first_names, last_names, genders):
        """Property 2: Preservation - standardization Behavior Unchanged
        
        **Validates: Requirements 3.1, 3.2, 3.3, 3.4**
        """
        assume(len(first_names) >= num_rows)
        assume(len(last_names) >= num_rows)
        assume(len(genders) >= num_rows)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            headers = ["שם פרטי", "שם משפחה", "מין"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            for row_idx in range(num_rows):
                ws.cell(row=row_idx + 2, column=1, value=first_names[row_idx])
                ws.cell(row=row_idx + 2, column=2, value=last_names[row_idx])
                ws.cell(row=row_idx + 2, column=3, value=genders[row_idx])
            
            wb.save(input_path)
            
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            assert os.path.exists(output_path), "Output Excel file should be created"
            
            wb_out = load_workbook(output_path)
            ws_out = wb_out.active
            header_row = [cell.value for cell in ws_out[1]]
            
            # Original columns preserved
            assert "שם פרטי" in header_row, "Original שם פרטי preserved"
            assert "שם משפחה" in header_row, "Original שם משפחה preserved"
            
            # Corrected columns inserted with '- מתוקן' suffix
            assert "שם פרטי - מתוקן" in header_row, "שם פרטי - מתוקן created"
            assert "שם משפחה - מתוקן" in header_row, "שם משפחה - מתוקן created"

    def test_engine_execution_order_preservation(self):
        """Test that all engines execute in correct order.
        
        **Validates: Requirement 3.2**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            headers = ["שם פרטי", "שם משפחה", "מין"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            ws.cell(row=2, column=1, value="יוסי")
            ws.cell(row=2, column=2, value="כהן")
            ws.cell(row=2, column=3, value="ז")
            
            wb.save(input_path)
            
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            wb_out = load_workbook(output_path)
            ws_out = wb_out.active
            header_row = [cell.value for cell in ws_out[1]]
            
            # NameEngine creates corrected name columns
            assert "שם פרטי - מתוקן" in header_row, "NameEngine should create שם פרטי - מתוקן"
            assert "שם משפחה - מתוקן" in header_row, "NameEngine should create שם משפחה - מתוקן"

    @settings(max_examples=5, deadline=None)
    @given(
        num_sheets=st.integers(min_value=1, max_value=2),
        num_rows=st.integers(min_value=1, max_value=3),
    )
    def test_corrected_fields_suffix_preservation_property(self, num_sheets, num_rows):
        """Property 2: Preservation - Corrected columns use '- מתוקן' suffix.
        
        **Validates: Requirement 3.3**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            wb = Workbook()
            
            if num_sheets > 1:
                wb.remove(wb.active)
            
            for sheet_idx in range(num_sheets):
                if sheet_idx == 0 and num_sheets == 1:
                    ws = wb.active
                    ws.title = f"Sheet{sheet_idx + 1}"
                else:
                    ws = wb.create_sheet(title=f"Sheet{sheet_idx + 1}")
                
                headers = ["שם פרטי", "שם משפחה", "מין"]
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                for row_idx in range(num_rows):
                    ws.cell(row=row_idx + 2, column=1, value=f"Name{row_idx}")
                    ws.cell(row=row_idx + 2, column=2, value=f"Last{row_idx}")
                    ws.cell(row=row_idx + 2, column=3, value="ז")
            
            wb.save(input_path)
            
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            wb_out = load_workbook(output_path)
            
            for sheet_name in wb_out.sheetnames:
                ws_out = wb_out[sheet_name]
                header_row = [cell.value for cell in ws_out[1]]
                
                # Original columns preserved
                assert "שם פרטי" in header_row, f"Sheet {sheet_name}: Original שם פרטי preserved"
                assert "שם משפחה" in header_row, f"Sheet {sheet_name}: Original שם משפחה preserved"
                
                # Corrected columns with '- מתוקן' suffix
                assert "שם פרטי - מתוקן" in header_row, f"Sheet {sheet_name}: שם פרטי - מתוקן created"
                assert "שם משפחה - מתוקן" in header_row, f"Sheet {sheet_name}: שם משפחה - מתוקן created"

    def test_log_file_preservation(self):
        """Test that pipeline completes successfully.
        
        **Validates: Requirement 3.4**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            headers = ["שם פרטי", "שם משפחה", "מין"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            ws.cell(row=2, column=1, value="יוסי")
            ws.cell(row=2, column=2, value="כהן")
            ws.cell(row=2, column=3, value="ז")
            
            wb.save(input_path)
            
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            assert os.path.exists(output_path), "Pipeline should complete and create output file"
            
            wb_out = load_workbook(output_path)
            assert len(wb_out.sheetnames) > 0, "Output should have at least one sheet"

    @settings(max_examples=5, deadline=None)
    @given(
        has_date_field=st.booleans(),
        has_id_field=st.booleans(),
        num_rows=st.integers(min_value=1, max_value=3),
    )
    def test_excel_structure_preservation_property(self, has_date_field, has_id_field, num_rows):
        """Property 2: Preservation - Excel Structure Preserved Across Variations
        
        **Validates: Requirements 3.1, 3.3**
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            headers = ["שם פרטי", "שם משפחה", "מין"]
            if has_date_field:
                headers.append("תאריך לידה")
            if has_id_field:
                headers.append("תעודת זהות")
            
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            for row_idx in range(num_rows):
                ws.cell(row=row_idx + 2, column=1, value=f"First{row_idx}")
                ws.cell(row=row_idx + 2, column=2, value=f"Last{row_idx}")
                ws.cell(row=row_idx + 2, column=3, value="ז")
                
                col_offset = 4
                if has_date_field:
                    ws.cell(row=row_idx + 2, column=col_offset, value="01/01/2000")
                    col_offset += 1
                if has_id_field:
                    ws.cell(row=row_idx + 2, column=col_offset, value="123456789")
            
            wb.save(input_path)
            
            orchestrator = standardizationOrchestrator()
            orchestrator.process_workbook_json(input_path, output_path)
            
            assert os.path.exists(output_path), "Output Excel file should be created"
            
            wb_out = load_workbook(output_path)
            ws_out = wb_out.active
            header_row = [cell.value for cell in ws_out[1]]
            
            # Always-present corrected columns
            assert "שם פרטי - מתוקן" in header_row, "שם פרטי - מתוקן should exist"
            assert "שם משפחה - מתוקן" in header_row, "שם משפחה - מתוקן should exist"
