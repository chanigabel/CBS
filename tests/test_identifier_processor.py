"""Tests for IdentifierFieldProcessor.

This module tests the IdentifierFieldProcessor class that processes
Israeli ID and passport fields together.
"""

import pytest
from openpyxl import Workbook
from src.excel_standardization.processing.identifier_processor import IdentifierFieldProcessor
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_writer import ExcelWriter
from src.excel_standardization.engines.identifier_engine import IdentifierEngine


@pytest.fixture
def workbook():
    """Create a test workbook."""
    wb = Workbook()
    ws = wb.active
    return wb


@pytest.fixture
def processor():
    """Create an IdentifierFieldProcessor instance."""
    reader = ExcelReader()
    writer = ExcelWriter()
    engine = IdentifierEngine()
    return IdentifierFieldProcessor(reader, writer, engine)


def test_find_headers_both_found(workbook, processor):
    """Test finding both ID and passport headers."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add some data
    ws['A2'] = '123456782'
    ws['B2'] = 'P123456'
    
    result = processor.find_headers(ws)
    
    assert result is True
    assert processor.id_header_info is not None
    assert processor.id_header_info.col == 1
    assert processor.passport_header_info is not None
    assert processor.passport_header_info.col == 2


def test_find_headers_id_missing(workbook, processor):
    """Test when ID header is missing."""
    ws = workbook.active
    
    # Only passport header
    ws['A1'] = 'מספר דרכון'
    
    result = processor.find_headers(ws)
    
    assert result is False


def test_find_headers_passport_missing(workbook, processor):
    """Test when passport header is missing."""
    ws = workbook.active
    
    # Only ID header
    ws['A1'] = 'מספר זהות'
    
    result = processor.find_headers(ws)
    
    assert result is False


def test_prepare_output_columns(workbook, processor):
    """Test preparing output columns."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Find headers first
    processor.find_headers(ws)
    
    # Prepare output columns
    processor.prepare_output_columns(ws)
    
    # Check that columns were inserted
    assert processor.corrected_id_col is not None
    assert processor.corrected_passport_col is not None
    assert processor.corrected_status_col is not None
    
    # Check headers
    assert ws.cell(1, processor.corrected_id_col).value == 'ת.ז. - מתוקן'
    assert ws.cell(1, processor.corrected_passport_col).value == 'דרכון - מתוקן'
    assert ws.cell(1, processor.corrected_status_col).value == 'סטטוס מזהה'


def test_process_data_valid_id(workbook, processor):
    """Test processing data with valid ID."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add data - valid ID
    ws['A2'] = '123456782'  # Valid checksum
    ws['B2'] = ''
    
    # Process
    processor.find_headers(ws)
    processor.prepare_output_columns(ws)
    processor.process_data(ws)
    
    # Check results
    corrected_id = ws.cell(2, processor.corrected_id_col).value
    corrected_passport = ws.cell(2, processor.corrected_passport_col).value
    status = ws.cell(2, processor.corrected_status_col).value
    
    assert corrected_id == '123456782'
    assert corrected_passport == ''
    assert status == 'ת.ז. תקינה'


def test_process_data_invalid_id(workbook, processor):
    """Test processing data with invalid ID."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add data - invalid ID
    ws['A2'] = '123456789'  # Invalid checksum
    ws['B2'] = ''
    
    # Process
    processor.find_headers(ws)
    processor.prepare_output_columns(ws)
    processor.process_data(ws)
    
    # Check results
    corrected_id = ws.cell(2, processor.corrected_id_col).value
    status = ws.cell(2, processor.corrected_status_col).value
    
    assert corrected_id == '123456789'
    assert status == 'ת.ז. לא תקינה'


def test_process_data_id_with_passport(workbook, processor):
    """Test processing data with both ID and passport."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add data
    ws['A2'] = '123456782'  # Valid ID
    ws['B2'] = 'P123456'
    
    # Process
    processor.find_headers(ws)
    processor.prepare_output_columns(ws)
    processor.process_data(ws)
    
    # Check results
    corrected_id = ws.cell(2, processor.corrected_id_col).value
    corrected_passport = ws.cell(2, processor.corrected_passport_col).value
    status = ws.cell(2, processor.corrected_status_col).value
    
    assert corrected_id == '123456782'
    assert corrected_passport == 'P123456'
    assert status == 'ת.ז. תקינה + דרכון הוזן'


def test_process_data_reads_to_max_last_row(workbook, processor):
    """Test that processing reads both columns to maximum last row."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add data - ID column has 3 rows, passport has 5 rows
    ws['A2'] = '123456782'
    ws['A3'] = '234567893'
    ws['A4'] = '345678904'
    
    ws['B2'] = 'P1'
    ws['B3'] = 'P2'
    ws['B4'] = 'P3'
    ws['B5'] = 'P4'
    ws['B6'] = 'P5'
    
    # Process
    processor.find_headers(ws)
    processor.prepare_output_columns(ws)
    processor.process_data(ws)
    
    # Check that all 5 rows were processed
    # Row 2
    assert ws.cell(2, processor.corrected_id_col).value == '123456782'
    assert ws.cell(2, processor.corrected_passport_col).value == 'P1'
    
    # Row 3
    assert ws.cell(3, processor.corrected_id_col).value == '234567893'
    assert ws.cell(3, processor.corrected_passport_col).value == 'P2'
    
    # Row 4
    assert ws.cell(4, processor.corrected_id_col).value == '345678904'
    assert ws.cell(4, processor.corrected_passport_col).value == 'P3'
    
    # Row 5 - only passport
    assert ws.cell(5, processor.corrected_id_col).value == ''
    assert ws.cell(5, processor.corrected_passport_col).value == 'P4'
    assert ws.cell(5, processor.corrected_status_col).value == 'דרכון הוזן'
    
    # Row 6 - only passport
    assert ws.cell(6, processor.corrected_id_col).value == ''
    assert ws.cell(6, processor.corrected_passport_col).value == 'P5'
    assert ws.cell(6, processor.corrected_status_col).value == 'דרכון הוזן'


def test_process_field_integration(workbook, processor):
    """Test the full process_field template method."""
    ws = workbook.active
    
    # Set up headers
    ws['A1'] = 'מספר זהות'
    ws['B1'] = 'מספר דרכון'
    
    # Add data
    ws['A2'] = '123456782'  # Valid ID
    ws['B2'] = 'P123456'
    ws['A3'] = '000000000'  # All zeros
    ws['B3'] = ''
    ws['A4'] = 'ABC123'  # Invalid format - should move to passport
    ws['B4'] = ''
    
    # Process using template method
    processor.process_field(ws)
    
    # Check that all steps were executed
    assert processor.corrected_id_col is not None
    assert processor.corrected_passport_col is not None
    assert processor.corrected_status_col is not None
    
    # Check row 2 - valid ID
    assert ws.cell(2, processor.corrected_id_col).value == '123456782'
    assert ws.cell(2, processor.corrected_status_col).value == 'ת.ז. תקינה + דרכון הוזן'
    
    # Check row 3 - all zeros
    assert ws.cell(3, processor.corrected_id_col).value == '000000000'
    assert ws.cell(3, processor.corrected_status_col).value == 'ת.ז. לא תקינה'
    
    # Check row 4 - moved to passport
    # 'ABC123' has no hyphens, so clean_id_number leaves it as 'ABC123'.
    # _process_id_value sees 'A' (non-digit/non-dash) and moves the whole
    # value to passport via clean_passport, which keeps letters → 'ABC123'.
    assert ws.cell(4, processor.corrected_id_col).value == ''
    assert ws.cell(4, processor.corrected_passport_col).value == 'ABC123'
    assert ws.cell(4, processor.corrected_status_col).value == 'ת.ז. הועברה לדרכון'
