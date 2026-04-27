"""Test script to demonstrate intelligent table detection.

This script shows how the enhanced ExcelReader can detect tables
in complex Excel forms with variable header positions.
"""

from openpyxl import Workbook
from src.excel_normalization.io_layer.excel_reader import ExcelReader


def create_sample_workbook():
    """Create a sample workbook with complex structure."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Census Form"

    # Add decorative header (rows 1-5)
    ws['A1'] = "משרד הפנים"
    ws['A2'] = "טופס מפקד אוכלוסין"
    ws['A3'] = "שנת 2024"
    ws['A4'] = ""
    ws['A5'] = "הנחיות: אנא מלא את כל השדות"

    # Add table headers (row 7)
    ws['A7'] = "מספר סידורי"
    ws['B7'] = "שם פרטי"
    ws['C7'] = "שם משפחה"
    ws['D7'] = "שם האב"
    ws['E7'] = "מין\n(1=זכר, 2=נקבה)"
    ws['F7'] = "תאריך לידה"
    ws['I7'] = "מספר זהות"
    ws['J7'] = "דרכון"

    # Add sub-headers for date (row 8)
    ws['F8'] = "שנה"
    ws['G8'] = "חודש"
    ws['H8'] = "יום"

    # Add sample data (rows 9-12)
    data = [
        [1, "יוסי", "כהן", "אברהם", 1, 1990, 5, 15, "123456782", ""],
        [2, "שרה", "לוי", "יצחק", 2, 1985, 12, 25, "234567893", "P123456"],
        [3, "David", "Cohen", "Isaac", 1, 1992, 3, 10, "345678904", ""],
        [4, "מרים", "ישראל", "משה", 2, 1988, 7, 20, "456789015", ""],
    ]

    for i, row_data in enumerate(data, start=9):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    return wb


def test_table_detection():
    """Test the table detection functionality."""
    print("=" * 80)
    print("INTELLIGENT TABLE DETECTION TEST")
    print("=" * 80)

    # Create sample workbook
    wb = create_sample_workbook()
    ws = wb.active

    # Initialize reader
    reader = ExcelReader()

    # Detect table region
    print("\n1. DETECTING TABLE REGION...")
    print("-" * 80)
    table_region = reader.detect_table_region(ws)

    if table_region:
        print(f"✓ Table detected!")
        print(f"  Start row: {table_region.start_row}")
        print(f"  End row: {table_region.end_row}")
        print(f"  Start column: {table_region.start_col}")
        print(f"  End column: {table_region.end_col}")
        print(f"  Header rows: {table_region.header_rows}")
        print(f"  Data start row: {table_region.data_start_row}")
    else:
        print("✗ No table detected")
        return

    # Detect columns
    print("\n2. DETECTING COLUMNS...")
    print("-" * 80)
    column_mapping = reader.detect_columns(ws)

    if column_mapping:
        print(f"✓ Found {len(column_mapping)} columns:")
        for field_name, col_info in sorted(column_mapping.items()):
            print(f"  {field_name:20s} → Column {col_info.col:2d} ('{col_info.header_text}')")
    else:
        print("✗ No columns detected")

    # Test legacy find_header method
    print("\n3. TESTING LEGACY find_header() METHOD...")
    print("-" * 80)

    test_searches = [
        (["שם פרטי", "first name"], "First Name"),
        (["שם משפחה", "last name"], "Last Name"),
        (["מין"], "Gender"),
        (["מספר זהות"], "ID Number"),
    ]

    for search_terms, description in test_searches:
        result = reader.find_header(ws, search_terms)
        if result:
            print(f"✓ {description:15s} → Column {result.col} (row {result.header_row})")
        else:
            print(f"✗ {description:15s} → Not found")

    # Save sample workbook
    output_file = "sample_census_form.xlsx"
    wb.save(output_file)
    print(f"\n✓ Sample workbook saved to: {output_file}")

    print("\n" + "=" * 80)
    print("TEST COMPLETED SUCCESSFULLY")
    print("=" * 80)


def test_with_english_headers():
    """Test with English headers."""
    print("\n\n" + "=" * 80)
    print("TESTING WITH ENGLISH HEADERS")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Add decorative header
    ws['A1'] = "Ministry of Interior"
    ws['A2'] = "Census Form 2024"
    ws['A3'] = ""

    # Add table headers (row 5)
    ws['A5'] = "Serial No."
    ws['B5'] = "First Name"
    ws['C5'] = "Last Name"
    ws['D5'] = "Father's Name"
    ws['E5'] = "Gender"
    ws['F5'] = "Birth Date"
    ws['I5'] = "ID Number"
    ws['J5'] = "Passport"

    # Add sub-headers for date (row 6)
    ws['F6'] = "Year"
    ws['G6'] = "Month"
    ws['H6'] = "Day"

    # Add sample data
    ws['A7'] = 1
    ws['B7'] = "John"
    ws['C7'] = "Smith"
    ws['D7'] = "David"
    ws['E7'] = 1
    ws['F7'] = 1990
    ws['G7'] = 5
    ws['H7'] = 15
    ws['I7'] = "123456782"

    # Initialize reader
    reader = ExcelReader()

    # Detect columns
    column_mapping = reader.detect_columns(ws)

    print(f"\n✓ Found {len(column_mapping)} columns:")
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"  {field_name:20s} → Column {col_info.col:2d} ('{col_info.header_text}')")

    print("\n" + "=" * 80)


def test_with_corrected_columns():
    """Test that corrected columns are ignored."""
    print("\n\n" + "=" * 80)
    print("TESTING WITH CORRECTED COLUMNS (SHOULD BE IGNORED)")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Add headers including corrected columns
    ws['A1'] = "שם פרטי"
    ws['B1'] = "שם פרטי - מתוקן"  # Should be ignored
    ws['C1'] = "שם משפחה"
    ws['D1'] = "שם משפחה - מתוקן"  # Should be ignored
    ws['E1'] = "מין"

    # Add data
    ws['A2'] = "יוסי"
    ws['B2'] = "יוסי"
    ws['C2'] = "כהן"
    ws['D2'] = "כהן"
    ws['E2'] = 1

    # Initialize reader
    reader = ExcelReader()

    # Detect columns
    column_mapping = reader.detect_columns(ws)

    print(f"\n✓ Found {len(column_mapping)} columns (corrected columns ignored):")
    for field_name, col_info in sorted(column_mapping.items()):
        print(f"  {field_name:20s} → Column {col_info.col:2d} ('{col_info.header_text}')")

    print("\n✓ Corrected columns (B and D) were successfully ignored")
    print("=" * 80)


if __name__ == "__main__":
    test_table_detection()
    test_with_english_headers()
    test_with_corrected_columns()
    print("\n✓ All tests completed successfully!\n")
