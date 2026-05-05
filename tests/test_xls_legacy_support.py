"""Tests for legacy .xls file support.

Covers:
- XlsReader: sheet name discovery, content extraction, header detection
- UploadService: accepts .xls, rejects unsupported extensions
- Full pipeline: .xls → SheetDataset → normalization → export as .xlsx
- Error handling: corrupt/unreadable .xls files
"""

import io
import os
import tempfile
from pathlib import Path

import pytest

# ---------------------------------------------------------------------------
# Fixture path
# ---------------------------------------------------------------------------

FIXTURE_XLS = Path(__file__).parent / "fixtures" / "sample_legacy.xls"


# ---------------------------------------------------------------------------
# 1. XlsReader — sheet name discovery
# ---------------------------------------------------------------------------

class TestGetXlsSheetNames:
    def test_returns_sheet_names(self):
        from src.excel_standardization.io_layer.xls_reader import get_xls_sheet_names
        names = get_xls_sheet_names(str(FIXTURE_XLS))
        assert isinstance(names, list)
        assert len(names) >= 1
        assert "דיירים יחידים" in names

    def test_raises_on_nonexistent_file(self):
        from src.excel_standardization.io_layer.xls_reader import get_xls_sheet_names, XLS_ERROR_HE
        with pytest.raises(ValueError) as exc_info:
            get_xls_sheet_names("/nonexistent/path/file.xls")
        assert XLS_ERROR_HE in str(exc_info.value)

    def test_raises_on_corrupt_bytes(self, tmp_path):
        from src.excel_standardization.io_layer.xls_reader import get_xls_sheet_names, XLS_ERROR_HE
        bad_file = tmp_path / "corrupt.xls"
        bad_file.write_bytes(b"this is not an xls file at all")
        with pytest.raises(ValueError) as exc_info:
            get_xls_sheet_names(str(bad_file))
        assert XLS_ERROR_HE in str(exc_info.value)


# ---------------------------------------------------------------------------
# 2. XlsReader — content extraction into SheetDataset
# ---------------------------------------------------------------------------

class TestExtractXlsToWorkbookDataset:
    def test_returns_workbook_dataset(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        assert wbd is not None
        assert len(wbd.sheets) >= 1

    def test_sheet_has_field_names(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        sheet = wbd.sheets[0]
        assert len(sheet.field_names) > 0

    def test_sheet_has_rows(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        sheet = wbd.sheets[0]
        assert len(sheet.rows) >= 2

    def test_detects_name_fields(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        sheet = wbd.sheets[0]
        # ExcelReader should map Hebrew headers to internal field names
        assert "first_name" in sheet.field_names
        assert "last_name" in sheet.field_names

    def test_detects_id_field(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        sheet = wbd.sheets[0]
        assert "id_number" in sheet.field_names

    def test_row_values_are_correct(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        sheet = wbd.sheets[0]
        first_row = sheet.rows[0]
        assert first_row.get("first_name") == "יוסי"
        assert first_row.get("last_name") == "כהן"

    def test_source_format_metadata(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        assert wbd.get_metadata("source_format") == "xls"


class TestExtractXlsSheetToDataset:
    def test_extracts_named_sheet(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_sheet_to_dataset
        ds = extract_xls_sheet_to_dataset(str(FIXTURE_XLS), "דיירים יחידים")
        assert ds.sheet_name == "דיירים יחידים"
        assert len(ds.rows) >= 2

    def test_raises_key_error_for_missing_sheet(self):
        from src.excel_standardization.io_layer.xls_reader import extract_xls_sheet_to_dataset
        with pytest.raises(KeyError):
            extract_xls_sheet_to_dataset(str(FIXTURE_XLS), "NonExistentSheet")


# ---------------------------------------------------------------------------
# 3. UploadService — accepts .xls, rejects unsupported extensions
# ---------------------------------------------------------------------------

class TestUploadServiceXls:
    def _make_service(self, tmp_path):
        from webapp.services.session_service import SessionService
        from webapp.services.upload_service import UploadService
        ss = SessionService()
        return UploadService(
            session_service=ss,
            uploads_dir=tmp_path / "uploads",
            work_dir=tmp_path / "work",
        )

    def test_accepts_xls_file(self, tmp_path):
        svc = self._make_service(tmp_path)
        file_bytes = FIXTURE_XLS.read_bytes()
        response = svc.handle_upload("sample_legacy.xls", file_bytes)
        assert response.session_id
        assert "דיירים יחידים" in response.sheet_names

    def test_accepts_xlsx_file(self, tmp_path):
        """Existing .xlsx flow must still work."""
        from fastapi import HTTPException
        svc = self._make_service(tmp_path)
        # Use any existing xlsx from the uploads folder if available,
        # otherwise skip (we just verify no regression in extension check).
        xlsx_files = list(Path("uploads").glob("*.xlsx"))
        if not xlsx_files:
            pytest.skip("No .xlsx fixture available for regression test")
        file_bytes = xlsx_files[0].read_bytes()
        response = svc.handle_upload(xlsx_files[0].name, file_bytes)
        assert response.session_id

    def test_rejects_unsupported_extension(self, tmp_path):
        from fastapi import HTTPException
        svc = self._make_service(tmp_path)
        with pytest.raises(HTTPException) as exc_info:
            svc.handle_upload("data.csv", b"col1,col2\n1,2")
        assert exc_info.value.status_code == 400

    def test_rejects_corrupt_xls(self, tmp_path):
        from fastapi import HTTPException
        svc = self._make_service(tmp_path)
        with pytest.raises(HTTPException) as exc_info:
            svc.handle_upload("bad.xls", b"not an xls file")
        assert exc_info.value.status_code == 422
        # Hebrew error message
        from src.excel_standardization.io_layer.xls_reader import XLS_ERROR_HE
        assert XLS_ERROR_HE in exc_info.value.detail


# ---------------------------------------------------------------------------
# 4. Full pipeline: .xls → normalization → export as .xlsx
# ---------------------------------------------------------------------------

class TestXlsPipeline:
    def test_normalization_runs_on_xls_dataset(self):
        """SheetDataset from .xls goes through the standardization pipeline."""
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        from src.excel_standardization.processing.standardization_pipeline import (
            standardizationPipeline,
        )
        from src.excel_standardization.engines.name_engine import NameEngine
        from src.excel_standardization.engines.gender_engine import GenderEngine
        from src.excel_standardization.engines.date_engine import DateEngine
        from src.excel_standardization.engines.identifier_engine import IdentifierEngine
        from src.excel_standardization.engines.text_processor import TextProcessor

        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        assert wbd.sheets, "No sheets extracted from .xls fixture"

        pipeline = standardizationPipeline(
            name_engine=NameEngine(TextProcessor()),
            gender_engine=GenderEngine(),
            date_engine=DateEngine(),
            identifier_engine=IdentifierEngine(),
        )

        for sheet in wbd.sheets:
            normalized = pipeline.normalize_dataset(sheet)
            assert normalized is not None
            assert len(normalized.rows) == len(sheet.rows)
            # Corrected fields should be present
            if "first_name" in sheet.field_names and sheet.rows:
                assert "first_name_corrected" in normalized.rows[0]

    def test_export_produces_xlsx(self, tmp_path):
        """Export from .xls-sourced dataset writes a valid .xlsx file."""
        from src.excel_standardization.io_layer.xls_reader import extract_xls_to_workbook_dataset
        from src.excel_standardization.processing.standardization_pipeline import (
            standardizationPipeline,
        )
        from src.excel_standardization.engines.name_engine import NameEngine
        from src.excel_standardization.engines.gender_engine import GenderEngine
        from src.excel_standardization.engines.date_engine import DateEngine
        from src.excel_standardization.engines.identifier_engine import IdentifierEngine
        from src.excel_standardization.engines.text_processor import TextProcessor
        from src.excel_standardization.export.export_engine import ExportEngine

        wbd = extract_xls_to_workbook_dataset(str(FIXTURE_XLS))
        pipeline = standardizationPipeline(
            name_engine=NameEngine(TextProcessor()),
            gender_engine=GenderEngine(),
            date_engine=DateEngine(),
            identifier_engine=IdentifierEngine(),
        )
        wbd.sheets = [pipeline.normalize_dataset(s) for s in wbd.sheets]

        output_path = str(tmp_path / "output.xlsx")
        engine = ExportEngine()
        result = engine.export_from_normalized_dataset(wbd, output_path)

        assert Path(result).exists()
        assert Path(result).suffix == ".xlsx"
        # Verify it's a valid xlsx
        from openpyxl import load_workbook as _lw
        wb = _lw(result)
        assert len(wb.sheetnames) >= 0  # just verify it opens


# ---------------------------------------------------------------------------
# 5. _XlsWorksheet shim — unit tests
# ---------------------------------------------------------------------------

class TestXlsWorksheetShim:
    def test_cell_returns_correct_value(self):
        from src.excel_standardization.io_layer.xls_reader import _XlsWorksheet
        data = [["A", "B"], ["1", "2"]]
        ws = _XlsWorksheet("Sheet1", data)
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 2).value == "B"
        assert ws.cell(2, 1).value == "1"

    def test_cell_out_of_bounds_returns_none(self):
        from src.excel_standardization.io_layer.xls_reader import _XlsWorksheet
        ws = _XlsWorksheet("Sheet1", [["A"]])
        assert ws.cell(99, 99).value is None

    def test_max_row_and_col(self):
        from src.excel_standardization.io_layer.xls_reader import _XlsWorksheet
        data = [["A", "B", "C"], ["1", "2", "3"]]
        ws = _XlsWorksheet("Sheet1", data)
        assert ws.max_row == 2
        assert ws.max_column == 3

    def test_title(self):
        from src.excel_standardization.io_layer.xls_reader import _XlsWorksheet
        ws = _XlsWorksheet("MySheet", [])
        assert ws.title == "MySheet"

    def test_merged_cells_empty(self):
        from src.excel_standardization.io_layer.xls_reader import _XlsWorksheet
        ws = _XlsWorksheet("Sheet1", [["A"]])
        assert len(ws.merged_cells) == 0
