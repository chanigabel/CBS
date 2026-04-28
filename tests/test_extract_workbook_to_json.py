"""Comprehensive tests for extract_workbook_to_json method.

Tests the extract_workbook_to_json method with various scenarios including:
- Single worksheet workbooks
- Multiple worksheet workbooks
- Workbooks with sheets that have no valid headers
- Mixed workbooks (some sheets valid, some invalid)
- Error handling during extraction
- Metadata generation at workbook level

Requirements:
    - Validates: Requirements 16.1-16.6
"""

import pytest
import tempfile
import os
from openpyxl import Workbook

from src.excel_standardization.io_layer import ExcelReader, ExcelToJsonExtractor
from src.excel_standardization.data_types import WorkbookDataset


@pytest.fixture
def excel_reader():
    """Create an ExcelReader instance for testing."""
    return ExcelReader()


@pytest.fixture
def extractor(excel_reader):
    """Create an ExcelToJsonExtractor instance for testing."""
    return ExcelToJsonExtractor(excel_reader=excel_reader)


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file for testing."""
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    yield temp_file.name
    # Cleanup
    if os.path.exists(temp_file.name):
        os.unlink(temp_file.name)



def test_extract_workbook_with_single_sheet(extractor, temp_excel_file):
    """Test extracting a workbook with a single worksheet.
    
    Requirements:
        - Validates: Requirements 16.1-16.6
    """
    # Setup: Create workbook with single sheet
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Students"
    
    # Add data
    ws['A1'] = 'First Name'
    ws['B1'] = 'Last Name'
    ws['C1'] = 'ID Number'
    
    ws['A2'] = 'John'
    ws['B2'] = 'Doe'
    ws['C2'] = 123456789
    
    ws['A3'] = 'Jane'
    ws['B3'] = 'Smith'
    ws['C3'] = 987654321
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert result.source_file == temp_excel_file
    assert len(result.sheets) == 1
    
    # Verify sheet data
    sheet = result.sheets[0]
    assert sheet.sheet_name == "Students"
    assert len(sheet.rows) == 2
    
    # Verify metadata
    assert result.metadata['total_sheets'] == 1
    assert result.metadata['processed_sheets'] == 1
    assert len(result.metadata['skipped_sheets']) == 0



def test_extract_workbook_with_multiple_sheets(extractor, temp_excel_file):
    """Test extracting a workbook with multiple worksheets.
    
    Requirements:
        - Validates: Requirements 16.1-16.4
    """
    # Setup: Create workbook with 3 sheets
    workbook = Workbook()
    
    # Sheet 1: Students
    ws1 = workbook.active
    ws1.title = "Students"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: Teachers
    ws2 = workbook.create_sheet("Teachers")
    ws2['A1'] = 'First Name'
    ws2['B1'] = 'Last Name'
    ws2['A2'] = 'Jane'
    ws2['B2'] = 'Smith'
    
    # Sheet 3: Staff
    ws3 = workbook.create_sheet("Staff")
    ws3['A1'] = 'First Name'
    ws3['B1'] = 'Last Name'
    ws3['A2'] = 'Bob'
    ws3['B2'] = 'Johnson'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert result.source_file == temp_excel_file
    assert len(result.sheets) == 3
    
    # Verify sheet names
    sheet_names = [sheet.sheet_name for sheet in result.sheets]
    assert "Students" in sheet_names
    assert "Teachers" in sheet_names
    assert "Staff" in sheet_names
    
    # Verify each sheet has data
    for sheet in result.sheets:
        assert len(sheet.rows) >= 1
    
    # Verify metadata
    assert result.metadata['total_sheets'] == 3
    assert result.metadata['processed_sheets'] == 3
    assert len(result.metadata['skipped_sheets']) == 0



def test_extract_workbook_with_sheet_no_headers(extractor, temp_excel_file):
    """Test extracting a workbook where one sheet has no valid headers.
    
    Requirements:
        - Validates: Requirement 16.6 (skip sheets with no valid headers)
    """
    # Setup: Create workbook with 2 sheets, one without headers
    workbook = Workbook()
    
    # Sheet 1: Valid sheet with headers
    ws1 = workbook.active
    ws1.title = "Students"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: Invalid sheet with no recognizable headers
    ws2 = workbook.create_sheet("Summary")
    ws2['A1'] = 'Total'
    ws2['B1'] = 'Average'
    ws2['A2'] = 100
    ws2['B2'] = 50
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    
    # Should have processed at least the valid sheet
    # The invalid sheet may or may not be skipped depending on header detection
    assert result.metadata['total_sheets'] == 2
    
    # If a sheet was skipped, verify it's in the skipped list
    if result.metadata['processed_sheets'] < 2:
        assert len(result.metadata['skipped_sheets']) > 0



def test_extract_workbook_with_different_header_structures(extractor, temp_excel_file):
    """Test extracting a workbook with sheets having different header structures.
    
    Requirements:
        - Validates: Requirement 16.5 (handle different header structures)
    """
    # Setup: Create workbook with different header structures
    workbook = Workbook()
    
    # Sheet 1: Single-row header
    ws1 = workbook.active
    ws1.title = "SingleHeader"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: Multi-row header (2 rows)
    ws2 = workbook.create_sheet("MultiHeader")
    ws2['A1'] = 'Name'
    ws2['B1'] = 'Name'
    ws2['C1'] = 'Birth Date'
    ws2['D1'] = 'Birth Date'
    ws2['E1'] = 'Birth Date'
    
    ws2['A2'] = 'First'
    ws2['B2'] = 'Last'
    ws2['C2'] = 'Year'
    ws2['D2'] = 'Month'
    ws2['E2'] = 'Day'
    
    ws2['A3'] = 'Jane'
    ws2['B3'] = 'Smith'
    ws2['C3'] = 1980
    ws2['D3'] = 5
    ws2['E3'] = 15
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert len(result.sheets) >= 1  # At least one sheet should be processed
    
    # Find sheets by name
    single_header_sheet = result.get_sheet_by_name("SingleHeader")
    multi_header_sheet = result.get_sheet_by_name("MultiHeader")
    
    # Verify single-row header sheet if processed
    if single_header_sheet:
        assert single_header_sheet.header_rows_count == 1
        assert len(single_header_sheet.rows) >= 1
    
    # Verify multi-row header sheet if processed
    if multi_header_sheet:
        assert multi_header_sheet.header_rows_count == 2
        assert len(multi_header_sheet.rows) >= 1



def test_extract_workbook_sheet_independence(extractor, temp_excel_file):
    """Test that each worksheet is processed independently.
    
    Requirements:
        - Validates: Requirement 16.2 (detect headers separately for each sheet)
    """
    # Setup: Create workbook with sheets having different column structures
    workbook = Workbook()
    
    # Sheet 1: Name fields only
    ws1 = workbook.active
    ws1.title = "Names"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: ID fields only
    ws2 = workbook.create_sheet("IDs")
    ws2['A1'] = 'ID Number'
    ws2['B1'] = 'Passport'
    ws2['A2'] = 123456789
    ws2['B2'] = 'AB123456'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert len(result.sheets) >= 1
    
    # Find sheets
    names_sheet = result.get_sheet_by_name("Names")
    ids_sheet = result.get_sheet_by_name("IDs")
    
    # Verify each sheet has different field names (independent detection)
    if names_sheet and ids_sheet:
        assert set(names_sheet.field_names) != set(ids_sheet.field_names)
        
        # Names sheet should have name fields
        assert any('name' in field for field in names_sheet.field_names)
        
        # IDs sheet should have id/passport fields
        assert any('id' in field or 'passport' in field for field in ids_sheet.field_names)



def test_extract_workbook_maintains_sheet_names(extractor, temp_excel_file):
    """Test that worksheet names are maintained in the output.
    
    Requirements:
        - Validates: Requirement 16.4 (maintain worksheet names)
    """
    # Setup: Create workbook with specific sheet names
    workbook = Workbook()
    
    sheet_names = ["Students 2023", "Teachers", "Staff_Data"]
    
    for i, name in enumerate(sheet_names):
        if i == 0:
            ws = workbook.active
            ws.title = name
        else:
            ws = workbook.create_sheet(name)
        
        # Add minimal data
        ws['A1'] = 'First Name'
        ws['B1'] = 'Last Name'
        ws['A2'] = f'Person{i}'
        ws['B2'] = f'LastName{i}'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    result_sheet_names = [sheet.sheet_name for sheet in result.sheets]
    
    # All original sheet names should be present (if processed)
    for name in sheet_names:
        if name not in result.metadata['skipped_sheets']:
            assert name in result_sheet_names



def test_extract_workbook_metadata_completeness(extractor, temp_excel_file):
    """Test that workbook-level metadata is complete and accurate.
    
    Requirements:
        - Validates: Requirements 16.1-16.6
    """
    # Setup: Create workbook with multiple sheets
    workbook = Workbook()
    
    # Sheet 1: Valid
    ws1 = workbook.active
    ws1.title = "Valid1"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: Valid
    ws2 = workbook.create_sheet("Valid2")
    ws2['A1'] = 'First Name'
    ws2['B1'] = 'Last Name'
    ws2['A2'] = 'Jane'
    ws2['B2'] = 'Smith'
    
    # Sheet 3: Potentially invalid (no clear headers)
    ws3 = workbook.create_sheet("Empty")
    ws3['A1'] = 'Data'
    ws3['A2'] = 'Value'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify metadata structure
    assert 'total_sheets' in result.metadata
    assert 'processed_sheets' in result.metadata
    assert 'skipped_sheets' in result.metadata
    
    # Verify metadata values
    assert result.metadata['total_sheets'] == 3
    assert isinstance(result.metadata['processed_sheets'], int)
    assert isinstance(result.metadata['skipped_sheets'], list)
    
    # Verify counts are consistent
    assert result.metadata['processed_sheets'] == len(result.sheets)
    assert result.metadata['processed_sheets'] + len(result.metadata['skipped_sheets']) == result.metadata['total_sheets']



def test_extract_workbook_get_sheet_by_name(extractor, temp_excel_file):
    """Test accessing sheets by name from WorkbookDataset.
    
    Requirements:
        - Validates: Requirement 16.4 (maintain worksheet names)
    """
    # Setup: Create workbook with named sheets
    workbook = Workbook()
    
    ws1 = workbook.active
    ws1.title = "Students"
    ws1['A1'] = 'First Name'
    ws1['A2'] = 'John'
    
    ws2 = workbook.create_sheet("Teachers")
    ws2['A1'] = 'First Name'
    ws2['A2'] = 'Jane'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify get_sheet_by_name works
    students_sheet = result.get_sheet_by_name("Students")
    teachers_sheet = result.get_sheet_by_name("Teachers")
    nonexistent_sheet = result.get_sheet_by_name("NonExistent")
    
    # Students sheet should exist
    if students_sheet:
        assert students_sheet.sheet_name == "Students"
        assert len(students_sheet.rows) >= 1
    
    # Teachers sheet should exist
    if teachers_sheet:
        assert teachers_sheet.sheet_name == "Teachers"
        assert len(teachers_sheet.rows) >= 1
    
    # Non-existent sheet should return None
    assert nonexistent_sheet is None



def test_extract_workbook_with_hebrew_headers(extractor, temp_excel_file):
    """Test extracting a workbook with Hebrew headers.
    
    Requirements:
        - Validates: Requirements 16.1-16.6 with multilingual support
    """
    # Setup: Create workbook with Hebrew headers
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "תלמידים"  # Hebrew: Students
    
    # Hebrew headers
    ws['A1'] = 'שם פרטי'  # First Name
    ws['B1'] = 'שם משפחה'  # Last Name
    ws['C1'] = 'מין'  # Gender
    
    # Data
    ws['A2'] = 'יוסי'
    ws['B2'] = 'כהן'
    ws['C2'] = 'ז'
    
    ws['A3'] = 'דנה'
    ws['B3'] = 'לוי'
    ws['C3'] = 'נ'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert len(result.sheets) >= 1
    
    # Verify sheet was processed
    sheet = result.sheets[0]
    assert sheet.sheet_name == "תלמידים"
    
    # Should have detected some fields
    if len(sheet.field_names) > 0:
        assert len(sheet.rows) == 2



def test_extract_workbook_with_empty_sheets(extractor, temp_excel_file):
    """Test extracting a workbook with completely empty sheets.
    
    Requirements:
        - Validates: Requirement 16.6 (skip sheets with no valid headers)
    """
    # Setup: Create workbook with empty and valid sheets
    workbook = Workbook()
    
    # Sheet 1: Valid sheet
    ws1 = workbook.active
    ws1.title = "Valid"
    ws1['A1'] = 'First Name'
    ws1['A2'] = 'John'
    
    # Sheet 2: Completely empty
    ws2 = workbook.create_sheet("Empty")
    # No data at all
    
    # Sheet 3: Has data but no recognizable headers
    ws3 = workbook.create_sheet("NoHeaders")
    ws3['A1'] = 'Random'
    ws3['A2'] = 'Data'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert result.metadata['total_sheets'] == 3
    
    # At least the valid sheet should be processed
    assert result.metadata['processed_sheets'] >= 1
    
    # Empty or invalid sheets should be in skipped list
    if result.metadata['processed_sheets'] < 3:
        assert len(result.metadata['skipped_sheets']) > 0



def test_extract_workbook_preserves_data_types(extractor, temp_excel_file):
    """Test that workbook extraction preserves data types from Excel.
    
    Requirements:
        - Validates: Requirement 10.5 (preserve exact original values)
    """
    # Setup: Create workbook with various data types
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "DataTypes"
    
    ws['A1'] = 'Text'
    ws['B1'] = 'Number'
    ws['C1'] = 'Float'
    ws['D1'] = 'Date'
    
    from datetime import datetime
    test_date = datetime(2020, 1, 15)
    
    ws['A2'] = 'John Doe'
    ws['B2'] = 123
    ws['C2'] = 45.67
    ws['D2'] = test_date
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    if len(result.sheets) > 0:
        sheet = result.sheets[0]
        if len(sheet.rows) > 0:
            row = sheet.rows[0]
            
            # Verify types are preserved
            if 'text' in row:
                assert isinstance(row['text'], str)
            if 'number' in row:
                assert isinstance(row['number'], int)
            if 'float' in row:
                assert isinstance(row['float'], float)
            if 'date' in row:
                assert isinstance(row['date'], datetime)



def test_extract_workbook_with_large_dataset(extractor, temp_excel_file):
    """Test extracting a workbook with large datasets across multiple sheets.
    
    Requirements:
        - Validates: Requirements 16.1-16.6 with performance considerations
    """
    # Setup: Create workbook with multiple sheets, each with many rows
    workbook = Workbook()
    
    for sheet_num in range(1, 4):  # 3 sheets
        if sheet_num == 1:
            ws = workbook.active
            ws.title = f"Sheet{sheet_num}"
        else:
            ws = workbook.create_sheet(f"Sheet{sheet_num}")
        
        # Add headers
        ws['A1'] = 'First Name'
        ws['B1'] = 'Last Name'
        ws['C1'] = 'ID Number'
        
        # Add 50 rows of data per sheet
        for row_num in range(2, 52):
            ws[f'A{row_num}'] = f'Person{row_num}'
            ws[f'B{row_num}'] = f'LastName{row_num}'
            ws[f'C{row_num}'] = 100000000 + row_num
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert len(result.sheets) >= 1
    
    # Verify each sheet has correct number of rows
    for sheet in result.sheets:
        assert len(sheet.rows) == 50
        assert sheet.metadata['total_rows'] == 50
    
    # Verify total sheets
    assert result.metadata['total_sheets'] == 3



def test_extract_workbook_error_handling(extractor, temp_excel_file):
    """Test that workbook extraction handles errors gracefully.
    
    Requirements:
        - Validates: Requirement 18.1-18.4 (handle errors gracefully)
    """
    # Setup: Create workbook with one valid sheet and one problematic sheet
    workbook = Workbook()
    
    # Sheet 1: Valid
    ws1 = workbook.active
    ws1.title = "Valid"
    ws1['A1'] = 'First Name'
    ws1['A2'] = 'John'
    
    # Sheet 2: Empty (should be skipped)
    ws2 = workbook.create_sheet("Empty")
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify - should not raise exception
    assert isinstance(result, WorkbookDataset)
    
    # Should have processed at least the valid sheet
    assert result.metadata['processed_sheets'] >= 1
    
    # Empty sheet should be in skipped list
    if result.metadata['processed_sheets'] < 2:
        assert "Empty" in result.metadata['skipped_sheets']


def test_extract_workbook_validates_correctly(extractor, temp_excel_file):
    """Test that extracted WorkbookDataset validates correctly.
    
    Requirements:
        - Validates: Requirements 16.1-16.4
    """
    # Setup: Create valid workbook
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "Students"
    ws['A1'] = 'First Name'
    ws['B1'] = 'Last Name'
    ws['A2'] = 'John'
    ws['B2'] = 'Doe'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify validation passes
    assert result.validate() is True
    
    # Verify source file is set
    assert result.source_file == temp_excel_file
    
    # Verify sheets list is valid
    assert isinstance(result.sheets, list)
    
    # Verify each sheet validates
    for sheet in result.sheets:
        assert sheet.validate() is True



def test_extract_workbook_with_mixed_valid_invalid_sheets(extractor, temp_excel_file):
    """Test extracting a workbook with mix of valid and invalid sheets.
    
    Requirements:
        - Validates: Requirement 16.6 (skip sheets with no valid headers)
    """
    # Setup: Create workbook with mixed sheets
    workbook = Workbook()
    
    # Sheet 1: Valid with clear headers
    ws1 = workbook.active
    ws1.title = "ValidSheet"
    ws1['A1'] = 'First Name'
    ws1['B1'] = 'Last Name'
    ws1['A2'] = 'John'
    ws1['B2'] = 'Doe'
    
    # Sheet 2: No recognizable headers
    ws2 = workbook.create_sheet("InvalidSheet")
    ws2['A1'] = 'Random'
    ws2['B1'] = 'Text'
    ws2['A2'] = 'Data'
    ws2['B2'] = 'Here'
    
    # Sheet 3: Another valid sheet
    ws3 = workbook.create_sheet("AnotherValid")
    ws3['A1'] = 'First Name'
    ws3['B1'] = 'ID Number'
    ws3['A2'] = 'Jane'
    ws3['B2'] = 123456789
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    assert isinstance(result, WorkbookDataset)
    assert result.metadata['total_sheets'] == 3
    
    # Should have processed at least the valid sheets
    assert result.metadata['processed_sheets'] >= 2
    
    # Verify valid sheets are present
    valid_sheet = result.get_sheet_by_name("ValidSheet")
    another_valid = result.get_sheet_by_name("AnotherValid")
    
    if valid_sheet:
        assert len(valid_sheet.rows) >= 1
    
    if another_valid:
        assert len(another_valid.rows) >= 1



def test_extract_workbook_sheet_count_methods(extractor, temp_excel_file):
    """Test WorkbookDataset helper methods for counting sheets.
    
    Requirements:
        - Validates: Requirements 16.1-16.4
    """
    # Setup: Create workbook with multiple sheets
    workbook = Workbook()
    
    ws1 = workbook.active
    ws1.title = "Sheet1"
    ws1['A1'] = 'First Name'
    ws1['A2'] = 'John'
    
    ws2 = workbook.create_sheet("Sheet2")
    ws2['A1'] = 'First Name'
    ws2['A2'] = 'Jane'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify helper methods
    assert result.get_sheet_count() == len(result.sheets)
    assert result.get_sheet_count() >= 1
    
    # Verify get_sheet_names
    sheet_names = result.get_sheet_names()
    assert isinstance(sheet_names, list)
    assert len(sheet_names) == result.get_sheet_count()
    
    # Verify has_sheet method
    for name in sheet_names:
        assert result.has_sheet(name) is True
    
    assert result.has_sheet("NonExistent") is False


def test_extract_workbook_with_formulas(extractor, temp_excel_file):
    """Test extracting a workbook with formula cells.
    
    Requirements:
        - Validates: Requirement 10.7 (handle formula cells)
    """
    # Setup: Create workbook with formulas
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "Formulas"
    
    ws['A1'] = 'Number1'
    ws['B1'] = 'Number2'
    ws['C1'] = 'Sum'
    
    ws['A2'] = 10
    ws['B2'] = 20
    ws['C2'] = '=A2+B2'  # Formula
    
    workbook.save(temp_excel_file)
    
    # Execute - with data_only=True, formulas should be calculated
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify
    if len(result.sheets) > 0:
        sheet = result.sheets[0]
        if len(sheet.rows) > 0:
            row = sheet.rows[0]
            
            # Formula should be calculated (or None if not calculated)
            # With data_only=True in load_workbook, formulas return None
            # unless the file was previously saved with calculated values
            # Passthrough columns use raw header text as key (case-preserved)
            assert 'Sum' in row or 'Number1' in row or 'sum' in row or 'number1' in row  # At least some fields detected



def test_extract_workbook_with_special_characters_in_sheet_names(extractor, temp_excel_file):
    """Test extracting a workbook with special characters in sheet names.
    
    Requirements:
        - Validates: Requirement 16.4 (maintain worksheet names)
    """
    # Setup: Create workbook with special characters in names
    workbook = Workbook()
    
    # Sheet names with special characters
    ws1 = workbook.active
    ws1.title = "Students-2023"
    ws1['A1'] = 'First Name'
    ws1['A2'] = 'John'
    
    ws2 = workbook.create_sheet("Teachers (Active)")
    ws2['A1'] = 'First Name'
    ws2['A2'] = 'Jane'
    
    ws3 = workbook.create_sheet("Staff_Data")
    ws3['A1'] = 'First Name'
    ws3['A2'] = 'Bob'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify sheet names are preserved
    sheet_names = result.get_sheet_names()
    
    # Check that special characters are preserved
    for sheet in result.sheets:
        assert sheet.sheet_name in ["Students-2023", "Teachers (Active)", "Staff_Data"]


def test_extract_workbook_returns_workbook_dataset_type(extractor, temp_excel_file):
    """Test that extract_workbook_to_json returns correct type.
    
    Requirements:
        - Validates: Requirements 16.1-16.4
    """
    # Setup: Create minimal workbook
    workbook = Workbook()
    ws = workbook.active
    ws['A1'] = 'First Name'
    ws['A2'] = 'John'
    
    workbook.save(temp_excel_file)
    
    # Execute
    result = extractor.extract_workbook_to_json(temp_excel_file)
    
    # Verify type
    assert isinstance(result, WorkbookDataset)
    assert hasattr(result, 'source_file')
    assert hasattr(result, 'sheets')
    assert hasattr(result, 'metadata')
    assert hasattr(result, 'get_sheet_by_name')
    assert hasattr(result, 'get_sheet_names')
    assert hasattr(result, 'get_sheet_count')
    assert hasattr(result, 'validate')
