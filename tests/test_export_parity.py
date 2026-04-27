from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.excel_normalization.orchestrator import NormalizationOrchestrator


def _make_source_workbook() -> Workbook:
    """Create a minimal but representative source workbook for parity checking.

    It intentionally uses raw (non "- מתוקן") headers so BOTH pipelines start from
    the same baseline input workbook.
    """
    wb = Workbook()
    wb.remove(wb[wb.sheetnames[0]])

    for name in ["דיירים יחידים", "מתגוררים במשקי בית", "אנשי צוות ובני משפחותיהם"]:
        ws = wb.create_sheet(name)

        # Header row (raw)
        ws.cell(1, 1).value = "שם פרטי"
        ws.cell(1, 2).value = "שם משפחה"
        ws.cell(1, 3).value = "שם האב"
        ws.cell(1, 4).value = "מספר זהות"
        ws.cell(1, 5).value = "מספר דרכון"
        ws.cell(1, 6).value = "מין\n1=זכר\n2+נקבה"
        ws.cell(1, 7).value = "תאריך לידה"

        # Subheader row for split date group
        ws.cell(2, 7).value = "שנה"
        ws.cell(2, 8).value = "חודש"
        ws.cell(2, 9).value = "יום"

        # Data row
        ws.cell(3, 1).value = "יוסי"
        ws.cell(3, 2).value = "כהן"
        ws.cell(3, 3).value = "ישראל"
        ws.cell(3, 4).value = "123456782"
        ws.cell(3, 5).value = ""
        ws.cell(3, 6).value = "1"
        ws.cell(3, 7).value = 1990
        ws.cell(3, 8).value = 5
        ws.cell(3, 9).value = 15

        # Blank row that must be ignored by export validity rule
        ws.cell(4, 1).value = ""
        ws.cell(4, 2).value = ""
        ws.cell(4, 3).value = ""
        ws.cell(4, 4).value = ""
        ws.cell(4, 5).value = ""
        ws.cell(4, 6).value = ""

    return wb


def _read_export_table(path: Path) -> dict[str, list[list]]:
    """Read export workbook sheets into comparable 2D arrays."""
    wb = load_workbook(path, data_only=True)
    out: dict[str, list[list]] = {}

    for name in ["DayarimYahidim", "MeshkeyBayt", "AnasheyTzevet"]:
        ws = wb[name]

        # Determine width from row 1 headers to avoid trailing empty columns
        last_header_col = 0
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip() != "":
                last_header_col = c
        width = last_header_col or 1

        values: list[list] = []
        for r in range(1, (ws.max_row or 0) + 1):
            row_vals = []
            for c in range(1, width + 1):
                v = ws.cell(r, c).value
                if v == "":
                    v = None
                row_vals.append(v)
            values.append(row_vals)

        # Trim trailing completely-empty rows (keep header)
        while len(values) > 1 and all(v is None for v in values[-1]):
            values.pop()

        out[name] = values

    return out


def test_export_parity_processors_vs_json(tmp_path: Path):
    src_wb = _make_source_workbook()
    input_path = tmp_path / "input.xlsx"
    src_wb.save(input_path)

    orch = NormalizationOrchestrator()

    out_proc = tmp_path / "out_proc.xlsx"
    orch.export_vba_parity_workbook_from_processors(str(input_path), str(out_proc))

    out_json = tmp_path / "out_json.xlsx"
    orch.export_vba_parity_workbook_from_json(str(input_path), str(out_json))

    proc_tables = _read_export_table(out_proc)
    json_tables = _read_export_table(out_json)

    if proc_tables != json_tables:
        for sheet_name in ["DayarimYahidim", "MeshkeyBayt", "AnasheyTzevet"]:
            p = proc_tables[sheet_name]
            j = json_tables[sheet_name]
            if p == j:
                continue

            max_rows = max(len(p), len(j))
            for r in range(max_rows):
                prow = p[r] if r < len(p) else []
                jrow = j[r] if r < len(j) else []
                if prow == jrow:
                    continue

                max_cols = max(len(prow), len(jrow))
                for c in range(max_cols):
                    pv = prow[c] if c < len(prow) else None
                    jv = jrow[c] if c < len(jrow) else None
                    if pv != jv:
                        raise AssertionError(
                            f"Mismatch in {sheet_name} at row {r+1}, col {c+1}: "
                            f"processors={pv!r} json={jv!r}. "
                            f"Row(proc)={prow!r} Row(json)={jrow!r}"
                        )

    assert proc_tables == json_tables

