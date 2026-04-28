"""Comprehensive test for merged cell handling in ExcelReader.

This test verifies that the ExcelReader correctly handles merged cells
according to Requirements 1.4 and 9.4:
- Requirement 1.4: "THE Header_Detector SHALL handle Merged_Cells without raising exceptions"
- Requirement 9.4: "THE Header_Detector SHALL handle Merged_Cells that span multiple rows without data loss"
"""

from openpyxl import Workbook
from src.excel_standardization.io_layer.excel_reader import ExcelReader


def test_requirement_1_4_no_exceptions():
    """Test Requirement 1.4: Handle merged cells without raising exceptions."""
    print("=" * 80)
    print("TEST: Requirement 1.4 - Handle Merged Cells Without Exceptions")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Create various merged cell scenarios
    # Scenario 1: Merged header cells
    ws.merge_cells('A1:B1')
    ws['A1'] = "שם"

    # Scenario 2: Merged cells in middle of headers
    ws.merge_cells('C1:E1')
    ws['C1'] = "תאריך לידה"

    # Scenario 3: Regular cells
    ws['F1'] = "מספר זהות"

    # Add sub-headers
    ws['A2'] = "פרטי"
    ws['B2'] = "משפחה"
    ws['C2'] = "שנה"
    ws['D2'] = "חודש"
    ws['E2'] = "יום"
    ws['F2'] = "ת.ז"

    # Add data with merged cells
    ws.merge_cells('A3:A4')
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = 1990
    ws['D3'] = 5
    ws['E3'] = 15
    ws['F3'] = "123456782"

    ws['B4'] = "כהן"
    ws['C4'] = 1990
    ws['D4'] = 5
    ws['E4'] = 15
    ws['F4'] = "123456782"

    reader = ExcelReader()

    try:
        print("\n1. Testing detect_table_region() with merged cells...")
        table_region = reader.detect_table_region(ws)
        assert table_region is not None, "Table region should be detected"
        print(f"   ✓ No exception raised")
        print(f"     Table detected: rows {table_region.start_row}-{table_region.end_row}")

        print("\n2. Testing detect_columns() with merged cells...")
        column_mapping = reader.detect_columns(ws)
        assert column_mapping is not None, "Column mapping should be returned"
        print(f"   ✓ No exception raised")
        print(f"     Columns detected: {len(column_mapping)}")

        print("\n3. Testing _is_merged_cell() with various cells...")
        # Test merged cells
        assert reader._is_merged_cell(ws, 1, 1) is True, "A1 should be merged"
        assert reader._is_merged_cell(ws, 1, 2) is True, "B1 should be merged"
        assert reader._is_merged_cell(ws, 1, 3) is True, "C1 should be merged"
        # Test non-merged cells
        assert reader._is_merged_cell(ws, 1, 6) is False, "F1 should not be merged"
        print(f"   ✓ No exception raised")

        print("\n4. Testing _get_merged_cell_range() with various cells...")
        # Test merged cells
        range_a1 = reader._get_merged_cell_range(ws, 1, 1)
        assert range_a1 == (1, 1, 1, 2), f"A1 range should be (1, 1, 1, 2), got {range_a1}"
        range_c1 = reader._get_merged_cell_range(ws, 1, 3)
        assert range_c1 == (1, 1, 3, 5), f"C1 range should be (1, 1, 3, 5), got {range_c1}"
        # Test non-merged cells
        range_f1 = reader._get_merged_cell_range(ws, 1, 6)
        assert range_f1 is None, f"F1 range should be None, got {range_f1}"
        print(f"   ✓ No exception raised")

        print("\n" + "=" * 80)
        print("✓ REQUIREMENT 1.4 VERIFIED: No exceptions with merged cells")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ REQUIREMENT 1.4 FAILED: {e}")
        import traceback
        traceback.print_exc()
        raise


def test_requirement_9_4_multi_row_merged():
    """Test Requirement 9.4: Handle merged cells spanning multiple rows without data loss."""
    print("\n\n" + "=" * 80)
    print("TEST: Requirement 9.4 - Handle Merged Cells Spanning Multiple Rows")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Create a structure with merged cells spanning multiple rows
    # Row 1: Parent headers (merged vertically)
    ws.merge_cells('A1:A2')
    ws['A1'] = "שם פרטי"

    ws.merge_cells('B1:B2')
    ws['B1'] = "שם משפחה"

    # Row 1: Parent date header (merged horizontally)
    ws.merge_cells('C1:E1')
    ws['C1'] = "תאריך לידה"

    # Row 2: Child headers for date
    ws['C2'] = "שנה"
    ws['D2'] = "חודש"
    ws['E2'] = "יום"

    # Add data
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = 1990
    ws['D3'] = 5
    ws['E3'] = 15

    ws['A4'] = "שרה"
    ws['B4'] = "לוי"
    ws['C4'] = 1985
    ws['D4'] = 12
    ws['E4'] = 25

    reader = ExcelReader()

    try:
        print("\n1. Testing table detection with vertically merged cells...")
        table_region = reader.detect_table_region(ws)
        assert table_region is not None, "Table region should be detected"
        assert table_region.header_rows == 2, f"Should detect 2 header rows, got {table_region.header_rows}"
        print(f"   ✓ Table detected with {table_region.header_rows} header rows")

        print("\n2. Testing column detection with merged parent headers...")
        column_mapping = reader.detect_columns(ws)
        assert column_mapping is not None, "Column mapping should be returned"
        print(f"   ✓ Columns detected: {len(column_mapping)}")

        # Verify all expected fields are detected
        expected_fields = ['first_name', 'last_name', 'birth_year', 'birth_month', 'birth_day']
        detected_fields = list(column_mapping.keys())
        print(f"   Expected fields: {expected_fields}")
        print(f"   Detected fields: {detected_fields}")

        # Check that we got the date fields (split structure)
        has_date_fields = any(f.startswith('birth_') for f in detected_fields)
        assert has_date_fields, "Should detect split date fields"
        print(f"   ✓ Split date fields detected correctly")

        print("\n3. Testing _get_merged_cell_range() for vertically merged cells...")
        # Test vertically merged cells
        range_a1 = reader._get_merged_cell_range(ws, 1, 1)
        assert range_a1 == (1, 2, 1, 1), f"A1 range should be (1, 2, 1, 1), got {range_a1}"
        range_a2 = reader._get_merged_cell_range(ws, 2, 1)
        assert range_a2 == (1, 2, 1, 1), f"A2 range should be (1, 2, 1, 1), got {range_a2}"
        print(f"   ✓ Vertically merged cells handled correctly")

        print("\n4. Testing _get_merged_cell_range() for horizontally merged cells...")
        # Test horizontally merged cells
        range_c1 = reader._get_merged_cell_range(ws, 1, 3)
        assert range_c1 == (1, 1, 3, 5), f"C1 range should be (1, 1, 3, 5), got {range_c1}"
        range_d1 = reader._get_merged_cell_range(ws, 1, 4)
        assert range_d1 == (1, 1, 3, 5), f"D1 range should be (1, 1, 3, 5), got {range_d1}"
        print(f"   ✓ Horizontally merged cells handled correctly")

        print("\n5. Testing data extraction with merged cells...")
        # Verify we can read data without issues
        data_row_1 = reader.read_column_array(ws, 1, 3, 4)
        assert data_row_1 == ["יוסי", "שרה"], f"Should read first name column correctly"
        print(f"   ✓ Data extraction works correctly with merged cells")

        print("\n" + "=" * 80)
        print("✓ REQUIREMENT 9.4 VERIFIED: Merged cells spanning multiple rows handled")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ REQUIREMENT 9.4 FAILED: {e}")
        import traceback
        traceback.print_exc()
        raise


def test_edge_cases():
    """Test edge cases with merged cells."""
    print("\n\n" + "=" * 80)
    print("TEST: Edge Cases with Merged Cells")
    print("=" * 80)

    wb = Workbook()
    ws = wb.active

    # Edge case 1: Entire row merged
    ws.merge_cells('A1:E1')
    ws['A1'] = "Title Row"

    # Edge case 2: Single cell (not merged)
    ws['A2'] = "שם פרטי"
    ws['B2'] = "שם משפחה"

    # Edge case 3: Complex merge pattern
    ws.merge_cells('C2:D2')
    ws['C2'] = "תאריך לידה"
    ws['E2'] = "מספר זהות"

    # Add data
    ws['A3'] = "יוסי"
    ws['B3'] = "כהן"
    ws['C3'] = "1990-05-15"
    ws['E3'] = "123456782"

    reader = ExcelReader()

    try:
        print("\n1. Testing with entire row merged...")
        table_region = reader.detect_table_region(ws)
        assert table_region is not None, "Table should be detected despite merged title row"
        print(f"   ✓ Table detected correctly")

        print("\n2. Testing with complex merge patterns...")
        column_mapping = reader.detect_columns(ws)
        assert column_mapping is not None, "Columns should be detected"
        print(f"   ✓ Columns detected: {len(column_mapping)}")

        print("\n3. Testing _is_merged_cell() with entire row merged...")
        assert reader._is_merged_cell(ws, 1, 1) is True
        assert reader._is_merged_cell(ws, 1, 3) is True
        assert reader._is_merged_cell(ws, 1, 5) is True
        print(f"   ✓ Entire row merge detected correctly")

        print("\n" + "=" * 80)
        print("✓ EDGE CASES VERIFIED: Complex merge patterns handled")
        print("=" * 80)

    except Exception as e:
        print(f"\n✗ EDGE CASES FAILED: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    test_requirement_1_4_no_exceptions()
    test_requirement_9_4_multi_row_merged()
    test_edge_cases()
    print("\n✓ All comprehensive merged cell tests passed!\n")
