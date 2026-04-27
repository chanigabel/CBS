"""End-to-end integration test for the complete normalization pipeline.

This test validates the complete flow:
1. Excel → JSON extraction (ExcelToJsonExtractor)
2. JSON normalization (NormalizationPipeline)
3. JSON → Excel export (JsonToExcelWriter)

The test creates a sample Excel file with Hebrew data, processes it through
the entire pipeline, and verifies that the output contains both original and
corrected columns with proper normalization applied.
"""

import os
import tempfile
import pytest
from openpyxl import Workbook, load_workbook

from src.excel_normalization.io_layer.excel_reader import ExcelReader
from src.excel_normalization.io_layer.excel_to_json_extractor import ExcelToJsonExtractor
from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
from src.excel_normalization.io_layer.excel_writer import JsonToExcelWriter
from src.excel_normalization.engines.name_engine import NameEngine
from src.excel_normalization.engines.gender_engine import GenderEngine
from src.excel_normalization.engines.date_engine import DateEngine
from src.excel_normalization.engines.identifier_engine import IdentifierEngine
from src.excel_normalization.engines.text_processor import TextProcessor


class TestEndToEndPipeline:
    """Test the complete Excel normalization pipeline end-to-end."""

    def test_end_to_end_pipeline(self):
        """Test complete pipeline: Excel → JSON → Normalize → Excel.
        
        This test validates:
        1. ExcelToJsonExtractor extracts data correctly
        2. NormalizationPipeline normalizes all fields
        3. JsonToExcelWriter exports with original and corrected columns
        4. Corrected values differ from originals where normalization occurred
        """
        # Create temporary files
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.xlsx")
            output_path = os.path.join(tmpdir, "output.xlsx")
            
            # Step 1: Create test Excel file with sample data
            wb = Workbook()
            ws = wb.active
            ws.title = "TestData"
            
            # Add headers
            headers = ["שם פרטי", "שם משפחה", "שם האב", "מין", "תאריך לידה", "מספר זהות", "דרכון"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add sample data with values that need normalization
            test_data = [
                # Row 2: Name with extra spaces, male gender, valid date, valid ID
                ["  יוסי  ", "כהן  ", "דוד", "ז", "15/05/1980", "123456782", "A1234567"],
                # Row 3: Name with English chars, female gender, valid date, valid ID
                ["שרה123", "לוי", "משה", "נ", "20/12/1985", "234567891", "B2345678"],
                # Row 4: Clean name, numeric gender, valid date, valid ID
                ["דוד", "אברהם", "יעקב", "2", "01/01/1990", "345678909", "C3456789"],
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save input file
            wb.save(input_path)
            
            # Step 2: Extract Excel to JSON
            excel_reader = ExcelReader()
            extractor = ExcelToJsonExtractor(
                excel_reader=excel_reader,
                skip_empty_rows=True
            )
            
            dataset = extractor.extract_workbook_to_json(input_path)
            
            # Verify extraction
            assert dataset is not None
            assert len(dataset.sheets) == 1
            sheet_dataset = dataset.sheets[0]
            assert sheet_dataset.sheet_name == "TestData"
            assert len(sheet_dataset.rows) == 3
            
            # Step 3: Normalize the data
            text_processor = TextProcessor()
            pipeline = NormalizationPipeline(
                name_engine=NameEngine(text_processor),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_dataset = pipeline.normalize_dataset(sheet_dataset)
            
            # Verify normalization created corrected fields
            assert len(normalized_dataset.rows) == 3
            first_row = normalized_dataset.rows[0]
            
            # Check that corrected fields exist
            assert "first_name_corrected" in first_row
            assert "last_name_corrected" in first_row
            assert "father_name_corrected" in first_row
            assert "gender_corrected" in first_row
            # Date corrected fields are now always structured year/month/day
            assert "birth_year_corrected" in first_row or "birth_date_corrected" in first_row
            assert "id_number_corrected" in first_row
            assert "passport_corrected" in first_row
            
            # Step 4: Export back to Excel
            writer = JsonToExcelWriter()
            writer.write_dataset_to_excel(normalized_dataset, output_path)
            
            # Step 5: Verify output Excel file
            output_wb = load_workbook(output_path)
            output_ws = output_wb.active
            
            # Verify headers exist (original and corrected columns)
            header_row = 1
            headers_found = []
            for col_idx in range(1, 20):  # Check first 20 columns
                cell_value = output_ws.cell(row=header_row, column=col_idx).value
                if cell_value:
                    headers_found.append(cell_value)
            
            # Verify both original and corrected columns exist
            assert "first_name" in headers_found
            assert "first_name_corrected" in headers_found
            assert "last_name" in headers_found
            assert "last_name_corrected" in headers_found
            assert "gender" in headers_found
            assert "gender_corrected" in headers_found
            assert "id_number" in headers_found
            assert "id_number_corrected" in headers_found
            
            # Step 6: Verify corrected values differ from originals where normalization occurred
            # Read first data row (row 2)
            data_row_idx = 2
            
            # Find column indices for first_name and first_name_corrected
            first_name_col = headers_found.index("first_name") + 1
            first_name_corrected_col = headers_found.index("first_name_corrected") + 1
            
            original_first_name = output_ws.cell(row=data_row_idx, column=first_name_col).value
            corrected_first_name = output_ws.cell(row=data_row_idx, column=first_name_corrected_col).value
            
            # Original had extra spaces "  יוסי  ", corrected should be trimmed
            assert original_first_name == "  יוסי  "
            assert corrected_first_name == "יוסי"
            assert original_first_name != corrected_first_name
            
            # Find column indices for last_name
            last_name_col = headers_found.index("last_name") + 1
            last_name_corrected_col = headers_found.index("last_name_corrected") + 1
            
            original_last_name = output_ws.cell(row=data_row_idx, column=last_name_col).value
            corrected_last_name = output_ws.cell(row=data_row_idx, column=last_name_corrected_col).value
            
            # Original had trailing space "כהן  ", corrected should be trimmed
            assert original_last_name == "כהן  "
            assert corrected_last_name == "כהן"
            
            # Verify gender normalization (row 2: "ז" should become 1 for male)
            gender_col = headers_found.index("gender") + 1
            gender_corrected_col = headers_found.index("gender_corrected") + 1
            
            original_gender = output_ws.cell(row=data_row_idx, column=gender_col).value
            corrected_gender = output_ws.cell(row=data_row_idx, column=gender_corrected_col).value
            
            assert original_gender == "ז"
            assert corrected_gender == 1  # 1 = male, 2 = female
            
            # Verify second row (row 3): Name with numbers should be cleaned
            data_row_2_idx = 3
            original_first_name_2 = output_ws.cell(row=data_row_2_idx, column=first_name_col).value
            corrected_first_name_2 = output_ws.cell(row=data_row_2_idx, column=first_name_corrected_col).value
            
            # Original had numbers "שרה123", corrected should remove them
            assert original_first_name_2 == "שרה123"
            assert corrected_first_name_2 == "שרה"
            assert original_first_name_2 != corrected_first_name_2
            
            # Verify third row (row 4): Numeric gender should remain numeric
            data_row_3_idx = 4
            original_gender_3 = output_ws.cell(row=data_row_3_idx, column=gender_col).value
            corrected_gender_3 = output_ws.cell(row=data_row_3_idx, column=gender_corrected_col).value
            
            assert original_gender_3 == "2"
            assert corrected_gender_3 == 2
            
            # Verify all 3 data rows are present
            assert output_ws.max_row >= 4  # Header + 3 data rows

    def test_end_to_end_with_split_date_fields(self):
        """Test pipeline with split date fields (year/month/day columns).
        
        This test validates that the pipeline correctly handles multi-row headers
        with split date fields and normalizes them properly.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input_split.xlsx")
            output_path = os.path.join(tmpdir, "output_split.xlsx")
            
            # Create test Excel with split date fields
            wb = Workbook()
            ws = wb.active
            ws.title = "SplitDates"
            
            # Add parent header for birth date
            ws.cell(row=1, column=1, value="שם פרטי")
            ws.cell(row=1, column=2, value="משפחה")
            ws.cell(row=1, column=3, value="תאריך לידה")
            ws.merge_cells('C1:E1')
            ws.cell(row=1, column=6, value="מין")
            
            # Add sub-headers for date components
            ws.cell(row=2, column=3, value="שנה")
            ws.cell(row=2, column=4, value="חודש")
            ws.cell(row=2, column=5, value="יום")
            
            # Add sample data
            test_data = [
                ["יוסי", "כהן", 1980, 5, 15, "ז"],
                ["שרה", "לוי", 1985, 12, 20, "נ"],
            ]
            
            for row_idx, row_data in enumerate(test_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(input_path)
            
            # Extract to JSON
            excel_reader = ExcelReader()
            extractor = ExcelToJsonExtractor(excel_reader=excel_reader)
            dataset = extractor.extract_workbook_to_json(input_path)
            
            sheet_dataset = dataset.sheets[0]
            
            # Verify split date fields were detected
            assert "birth_year" in sheet_dataset.field_names
            assert "birth_month" in sheet_dataset.field_names
            assert "birth_day" in sheet_dataset.field_names
            
            # Normalize
            text_processor = TextProcessor()
            pipeline = NormalizationPipeline(
                name_engine=NameEngine(text_processor),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_dataset = pipeline.normalize_dataset(sheet_dataset)
            
            # Verify corrected date fields exist
            first_row = normalized_dataset.rows[0]
            assert "birth_year_corrected" in first_row
            assert "birth_month_corrected" in first_row
            assert "birth_day_corrected" in first_row
            
            # Export to Excel
            writer = JsonToExcelWriter()
            writer.write_dataset_to_excel(normalized_dataset, output_path)
            
            # Verify output
            output_wb = load_workbook(output_path)
            output_ws = output_wb.active
            
            # Verify headers
            headers_found = []
            for col_idx in range(1, 20):
                cell_value = output_ws.cell(row=1, column=col_idx).value
                if cell_value:
                    headers_found.append(cell_value)
            
            assert "birth_year" in headers_found
            assert "birth_year_corrected" in headers_found
            assert "birth_month" in headers_found
            assert "birth_month_corrected" in headers_found
            assert "birth_day" in headers_found
            assert "birth_day_corrected" in headers_found
            
            # Verify data is present
            assert output_ws.max_row >= 3  # Header + 2 data rows

    def test_end_to_end_empty_and_none_values(self):
        """Test pipeline handles empty and None values correctly.
        
        Validates that the pipeline preserves None/empty values through
        the entire flow without errors.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input_empty.xlsx")
            output_path = os.path.join(tmpdir, "output_empty.xlsx")
            
            # Create test Excel with empty values
            wb = Workbook()
            ws = wb.active
            ws.title = "EmptyValues"
            
            # Add headers
            headers = ["שם פרטי", "שם משפחה", "מין", "מספר זהות"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add data with some empty cells
            test_data = [
                ["יוסי", "כהן", "ז", "123456782"],
                ["", "לוי", "נ", "234567891"],  # Empty first name
                ["דוד", "", "ז", ""],  # Empty last name and ID
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    if value:  # Only write non-empty values
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(input_path)
            
            # Extract to JSON
            excel_reader = ExcelReader()
            extractor = ExcelToJsonExtractor(excel_reader=excel_reader)
            dataset = extractor.extract_workbook_to_json(input_path)
            
            sheet_dataset = dataset.sheets[0]
            assert len(sheet_dataset.rows) == 3
            
            # Normalize
            text_processor = TextProcessor()
            pipeline = NormalizationPipeline(
                name_engine=NameEngine(text_processor),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_dataset = pipeline.normalize_dataset(sheet_dataset)
            
            # Verify empty values are preserved
            row_with_empty = normalized_dataset.rows[1]  # Second row has empty first name
            assert row_with_empty["first_name"] in [None, ""]
            assert row_with_empty["first_name_corrected"] in [None, ""]
            
            # Export to Excel
            writer = JsonToExcelWriter()
            writer.write_dataset_to_excel(normalized_dataset, output_path)
            
            # Verify output file exists and is valid
            assert os.path.exists(output_path)
            output_wb = load_workbook(output_path)
            output_ws = output_wb.active
            
            # Verify structure
            assert output_ws.max_row >= 4  # Header + 3 data rows

    def test_end_to_end_multisheet_workbook(self):
        """Test pipeline with multiple sheets in workbook.
        
        Validates that the pipeline correctly processes multiple sheets
        and exports them all to the output workbook.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input_multi.xlsx")
            output_path = os.path.join(tmpdir, "output_multi.xlsx")
            
            # Create test Excel with multiple sheets
            wb = Workbook()
            
            # Sheet 1: Students
            ws1 = wb.active
            ws1.title = "Students"
            headers1 = ["שם פרטי", "שם משפחה", "מין"]
            for col_idx, header in enumerate(headers1, 1):
                ws1.cell(row=1, column=col_idx, value=header)
            ws1.cell(row=2, column=1, value="יוסי")
            ws1.cell(row=2, column=2, value="כהן")
            ws1.cell(row=2, column=3, value="ז")
            
            # Sheet 2: Teachers
            ws2 = wb.create_sheet("Teachers")
            headers2 = ["שם פרטי", "שם משפחה", "מספר זהות"]
            for col_idx, header in enumerate(headers2, 1):
                ws2.cell(row=1, column=col_idx, value=header)
            ws2.cell(row=2, column=1, value="שרה")
            ws2.cell(row=2, column=2, value="לוי")
            ws2.cell(row=2, column=3, value="123456782")
            
            wb.save(input_path)
            
            # Extract to JSON
            excel_reader = ExcelReader()
            extractor = ExcelToJsonExtractor(excel_reader=excel_reader)
            workbook_dataset = extractor.extract_workbook_to_json(input_path)
            
            # Verify extraction
            assert len(workbook_dataset.sheets) == 2
            assert workbook_dataset.sheets[0].sheet_name == "Students"
            assert workbook_dataset.sheets[1].sheet_name == "Teachers"
            
            # Normalize each sheet
            text_processor = TextProcessor()
            pipeline = NormalizationPipeline(
                name_engine=NameEngine(text_processor),
                gender_engine=GenderEngine(),
                date_engine=DateEngine(),
                identifier_engine=IdentifierEngine()
            )
            
            normalized_sheets = []
            for sheet in workbook_dataset.sheets:
                normalized_sheet = pipeline.normalize_dataset(sheet)
                normalized_sheets.append(normalized_sheet)
            
            # Create normalized workbook dataset
            from src.excel_normalization.data_types import WorkbookDataset
            normalized_workbook = WorkbookDataset(
                source_file=input_path,
                sheets=normalized_sheets,
                metadata=workbook_dataset.metadata
            )
            
            # Export to Excel
            writer = JsonToExcelWriter()
            writer.write_workbook_to_excel(normalized_workbook, output_path)
            
            # Verify output
            output_wb = load_workbook(output_path)
            assert len(output_wb.sheetnames) == 2
            assert "Students" in output_wb.sheetnames
            assert "Teachers" in output_wb.sheetnames
            
            # Verify each sheet has correct structure
            students_ws = output_wb["Students"]
            teachers_ws = output_wb["Teachers"]
            
            # Check Students sheet
            students_headers = []
            for col_idx in range(1, 10):
                cell_value = students_ws.cell(row=1, column=col_idx).value
                if cell_value:
                    students_headers.append(cell_value)
            
            assert "first_name" in students_headers
            assert "first_name_corrected" in students_headers
            assert "gender" in students_headers
            assert "gender_corrected" in students_headers
            
            # Check Teachers sheet
            teachers_headers = []
            for col_idx in range(1, 10):
                cell_value = teachers_ws.cell(row=1, column=col_idx).value
                if cell_value:
                    teachers_headers.append(cell_value)
            
            assert "first_name" in teachers_headers
            assert "first_name_corrected" in teachers_headers
            assert "id_number" in teachers_headers
            assert "id_number_corrected" in teachers_headers
