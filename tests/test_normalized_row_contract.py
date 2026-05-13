from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset
from src.excel_standardization.normalized_row_contract import (
    build_grid_field_metadata,
    build_standard_export_row,
    grid_group_for_source_field,
    select_corrected_or_original,
)
from webapp.models.session import SessionRecord
from webapp.services.export_service import ExportService
from webapp.services.session_service import SessionService
from webapp.services.workbook_service import WorkbookService


def test_validation_source_selection_prefers_corrected_then_original():
    row = {"first_name": "Original", "first_name_corrected": "Corrected"}
    assert select_corrected_or_original(row, "first_name") == "Corrected"

    row["first_name_corrected"] = ""
    assert select_corrected_or_original(row, "first_name") == "Original"


def test_grid_contract_metadata_exposes_shared_grouping_and_status_maps():
    metadata = build_grid_field_metadata()

    assert metadata.groups
    assert metadata.source_to_corrected["first_name"] == "first_name_corrected"
    assert metadata.source_to_corrected["birth_date"] == "birth_date_corrected"
    assert metadata.source_to_status["gender"] == "gender_status"
    assert metadata.source_to_status["passport"] == "identifier_status"
    assert metadata.status_to_sources["gender_status"] == ("gender",)
    assert "passport" in metadata.status_to_sources["identifier_status"]
    assert metadata.structured_date_fallbacks["birth_date_corrected"][0] == "birth_day_corrected"
    assert grid_group_for_source_field("entry_year").name == "entry_date"
    assert grid_group_for_source_field("last_name").corrected_fields[1] == "last_name_corrected"


def test_corrected_values_survive_dataset_to_grid_to_export(tmp_path):
    svc = SessionService()
    svc.clear_all()
    sheet = SheetDataset(
        sheet_name="DayarimYahidim",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "gender"],
        rows=[
            {
                "first_name": "Original First",
                "first_name_corrected": "Corrected First",
                "last_name": "Original Last",
                "last_name_corrected": "Corrected Last",
                "gender": "Female",
                "gender_corrected": 2,
            }
        ],
    )
    record = SessionRecord(
        session_id="contract-session",
        source_file_path="uploads/contract.xlsx",
        working_copy_path="work/contract.xlsx",
        original_filename="contract.xlsx",
        status="standardized",
        workbook_dataset=WorkbookDataset(source_file="contract.xlsx", sheets=[sheet]),
    )
    svc.create(record)

    grid = WorkbookService(svc).get_sheet_data(record.session_id, "DayarimYahidim")
    assert grid.rows[0]["first_name"] == "Original First"
    assert grid.rows[0]["first_name_corrected"] == "Corrected First"

    mapped = build_standard_export_row(grid.rows[0], include_dira=False)
    assert mapped["ShemPrati"] == "Corrected First"
    assert mapped["ShemMishpaha"] == "Corrected Last"
    assert mapped["Min"] == 2

    output_path = ExportService(svc, tmp_path / "output").export(record.session_id)
    wb = load_workbook(output_path)
    ws = wb["DayarimYahidim"]
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(row=2, column=headers.index("ShemPrati") + 1).value == "Corrected First"
    assert ws.cell(row=2, column=headers.index("ShemMishpaha") + 1).value == "Corrected Last"
    assert ws.cell(row=2, column=headers.index("Min") + 1).value == 2
    wb.close()
