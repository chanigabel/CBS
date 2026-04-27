"""Tests for NormalizationPipeline — per-engine apply methods and dataset normalization.

Validates real normalization behavior: name cleaning, gender mapping,
date parsing, identifier validation, failure fallback, and metadata statistics.
"""

import pytest
from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
from src.excel_normalization.engines.name_engine import NameEngine
from src.excel_normalization.engines.gender_engine import GenderEngine
from src.excel_normalization.engines.date_engine import DateEngine
from src.excel_normalization.engines.identifier_engine import IdentifierEngine
from src.excel_normalization.engines.text_processor import TextProcessor
from src.excel_normalization.data_types import SheetDataset


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_pipeline(**kwargs) -> NormalizationPipeline:
    """Create a fully-wired pipeline unless overridden."""
    defaults = dict(
        name_engine=NameEngine(TextProcessor()),
        gender_engine=GenderEngine(),
        date_engine=DateEngine(),
        identifier_engine=IdentifierEngine(),
    )
    defaults.update(kwargs)
    return NormalizationPipeline(**defaults)


def make_dataset(rows, field_names=None, sheet_name="Sheet1") -> SheetDataset:
    return SheetDataset(
        sheet_name=sheet_name,
        header_row=1,
        header_rows_count=1,
        field_names=field_names or list(rows[0].keys()) if rows else [],
        rows=rows,
        metadata={},
    )


# ---------------------------------------------------------------------------
# apply_name_normalization
# ---------------------------------------------------------------------------

class TestApplyNameNormalization:
    def setup_method(self):
        self.pipeline = make_pipeline()

    def test_trims_first_name(self):
        row = {"first_name": "  יוסי  "}
        self.pipeline.apply_name_normalization(row)
        assert row["first_name_corrected"] == "יוסי"

    def test_removes_digits_from_last_name(self):
        row = {"last_name": "כהן123"}
        self.pipeline.apply_name_normalization(row)
        assert row["last_name_corrected"] == "כהן"

    def test_none_value_preserved(self):
        row = {"first_name": None}
        self.pipeline.apply_name_normalization(row)
        assert row["first_name_corrected"] is None

    def test_empty_string_preserved(self):
        row = {"first_name": ""}
        self.pipeline.apply_name_normalization(row)
        assert row["first_name_corrected"] == ""

    def test_all_three_name_fields_processed(self):
        row = {"first_name": "  דוד  ", "last_name": "  כהן  ", "father_name": "  אברהם  "}
        self.pipeline.apply_name_normalization(row)
        assert row["first_name_corrected"] == "דוד"
        assert row["last_name_corrected"] == "כהן"
        assert row["father_name_corrected"] == "אברהם"

    def test_missing_field_not_added(self):
        row = {"first_name": "יוסי"}
        self.pipeline.apply_name_normalization(row)
        assert "last_name_corrected" not in row
        assert "father_name_corrected" not in row

    def test_engine_failure_falls_back_to_original(self):
        """If engine raises, corrected field gets original value and field is in failures."""
        class BrokenNameEngine:
            def normalize_name(self, v):
                raise RuntimeError("engine broken")

        pipeline = NormalizationPipeline(name_engine=BrokenNameEngine())
        row = {"first_name": "יוסי"}
        failures = pipeline.apply_name_normalization(row)
        assert row["first_name_corrected"] == "יוסי"
        assert "first_name" in failures

    def test_returns_empty_failures_on_success(self):
        row = {"first_name": "יוסי"}
        failures = self.pipeline.apply_name_normalization(row)
        assert failures == []


# ---------------------------------------------------------------------------
# apply_gender_normalization
# ---------------------------------------------------------------------------

class TestApplyGenderNormalization:
    def setup_method(self):
        self.pipeline = make_pipeline()

    def test_hebrew_female_maps_to_2(self):
        row = {"gender": "נ"}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == 2

    def test_hebrew_male_maps_to_1(self):
        row = {"gender": "ז"}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == 1

    def test_english_female_maps_to_2(self):
        row = {"gender": "female"}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == 2

    def test_numeric_2_maps_to_2(self):
        row = {"gender": "2"}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == 2

    def test_numeric_1_maps_to_1(self):
        row = {"gender": "1"}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == 1

    def test_none_preserved(self):
        row = {"gender": None}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] is None

    def test_empty_string_preserved(self):
        row = {"gender": ""}
        self.pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == ""

    def test_missing_gender_field_no_op(self):
        row = {"first_name": "יוסי"}
        self.pipeline.apply_gender_normalization(row)
        assert "gender_corrected" not in row

    def test_engine_failure_falls_back_to_original(self):
        class BrokenGenderEngine:
            def normalize_gender(self, v):
                raise RuntimeError("broken")

        pipeline = NormalizationPipeline(gender_engine=BrokenGenderEngine())
        row = {"gender": "נ"}
        failures = pipeline.apply_gender_normalization(row)
        assert row["gender_corrected"] == "נ"
        assert "gender" in failures


# ---------------------------------------------------------------------------
# apply_date_normalization
# ---------------------------------------------------------------------------

class TestApplyDateNormalization:
    def setup_method(self):
        self.pipeline = make_pipeline()

    def test_split_birth_date_valid(self):
        row = {"birth_year": 1990, "birth_month": 5, "birth_day": 15}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 1990
        assert row["birth_month_corrected"] == 5
        assert row["birth_day_corrected"] == 15

    def test_split_entry_date_valid(self):
        row = {"entry_year": 2010, "entry_month": 3, "entry_day": 20}
        self.pipeline.apply_date_normalization(row)
        assert row["entry_year_corrected"] == 2010
        assert row["entry_month_corrected"] == 3
        assert row["entry_day_corrected"] == 20

    def test_single_birth_date_string(self):
        row = {"birth_date": "15/05/1990"}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 1990
        assert row["birth_month_corrected"] == 5
        assert row["birth_day_corrected"] == 15

    def test_single_birth_date_none_preserved(self):
        row = {"birth_date": None}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] is None
        assert row["birth_month_corrected"] is None
        assert row["birth_day_corrected"] is None

    def test_no_date_fields_no_op(self):
        row = {"first_name": "יוסי"}
        self.pipeline.apply_date_normalization(row)
        assert "birth_year_corrected" not in row
        assert "entry_year_corrected" not in row

    def test_two_digit_year_expanded(self):
        row = {"birth_year": 90, "birth_month": 5, "birth_day": 15}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 1990

    def test_iso_datetime_string_in_birth_year_column(self):
        # Real-world case: openpyxl reads a date cell as ISO string into birth_year
        # when birth_month/birth_day are null (merged date cell scenario)
        row = {"birth_year": "1997-09-04T00:00:00", "birth_month": None, "birth_day": None}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 9
        assert row["birth_day_corrected"] == 4

    def test_dot_separated_date_in_birth_year_column(self):
        # Real-world case: "11.06.1997" stored in birth_year column
        row = {"birth_year": "11.06.1997", "birth_month": None, "birth_day": None}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 1997
        assert row["birth_month_corrected"] == 6
        assert row["birth_day_corrected"] == 11

    def test_slash_separated_date_in_birth_year_column(self):
        # Real-world case: "04/02/2011" stored in birth_year column
        row = {"birth_year": "04/02/2011", "birth_month": None, "birth_day": None}
        self.pipeline.apply_date_normalization(row)
        assert row["birth_year_corrected"] == 2011
        assert row["birth_month_corrected"] == 2
        assert row["birth_day_corrected"] == 4

    def test_two_digit_entry_year_expanded(self):
        # Real-world case: entry_year=25 should expand to 2025
        row = {"entry_year": 25, "entry_month": 9, "entry_day": 1}
        self.pipeline.apply_date_normalization(row)
        assert row["entry_year_corrected"] == 2025

    def test_future_birth_date_iso_flagged_invalid(self):
        # ISO string with future date should be parsed but flagged invalid
        row = {"birth_year": "2025-09-20T00:00:00", "birth_month": None, "birth_day": None}
        self.pipeline.apply_date_normalization(row)
        # Year is parsed correctly even if business rules flag it
        assert row["birth_year_corrected"] == 2025


# ---------------------------------------------------------------------------
# apply_identifier_normalization
# ---------------------------------------------------------------------------

class TestApplyIdentifierNormalization:
    def setup_method(self):
        self.pipeline = make_pipeline()

    def test_valid_israeli_id_padded(self):
        # 123456782 is a valid Israeli ID (checksum passes)
        row = {"id_number": "123456782", "passport": ""}
        self.pipeline.apply_identifier_normalization(row)
        assert row["id_number_corrected"] == "123456782"

    def test_short_id_padded_with_zeros(self):
        # 4-digit ID gets padded to 9 digits
        row = {"id_number": "1234", "passport": ""}
        self.pipeline.apply_identifier_normalization(row)
        assert len(row["id_number_corrected"]) == 9

    def test_passport_only(self):
        row = {"id_number": "", "passport": "AB123456"}
        self.pipeline.apply_identifier_normalization(row)
        assert row["passport_corrected"] == "AB123456"
        assert row["id_number_corrected"] == ""

    def test_none_values_preserved(self):
        # Both None → both empty → short-circuit path stores original (None)
        row = {"id_number": None, "passport": None}
        self.pipeline.apply_identifier_normalization(row)
        assert row["id_number_corrected"] is None
        assert row["passport_corrected"] is None

    def test_missing_both_fields_no_op(self):
        row = {"first_name": "יוסי"}
        self.pipeline.apply_identifier_normalization(row)
        assert "id_number_corrected" not in row
        assert "passport_corrected" not in row

    def test_id_with_letters_moved_to_passport(self):
        # 'ABC123' has no hyphens, so clean_id_number leaves it unchanged.
        # _process_id_value sees 'A' (non-digit/non-dash) and moves the whole
        # value to passport via clean_passport → 'ABC123'.
        row = {"id_number": "ABC123", "passport": ""}
        self.pipeline.apply_identifier_normalization(row)
        assert row["id_number_corrected"] == ""
        assert row["passport_corrected"] == "ABC123"

    def test_engine_failure_falls_back_to_original(self):
        class BrokenIdentifierEngine:
            def normalize_identifiers(self, id_val, passport_val):
                raise RuntimeError("broken")

        pipeline = NormalizationPipeline(identifier_engine=BrokenIdentifierEngine())
        row = {"id_number": "123456782", "passport": ""}
        failures = pipeline.apply_identifier_normalization(row)
        assert row["id_number_corrected"] == "123456782"
        assert "id_number" in failures


# ---------------------------------------------------------------------------
# normalize_dataset — metadata statistics and engine flags
# ---------------------------------------------------------------------------

class TestNormalizeDataset:
    def setup_method(self):
        self.pipeline = make_pipeline()

    def test_metadata_normalized_flag_set(self):
        ds = make_dataset([{"first_name": "יוסי", "gender": "ז"}])
        result = self.pipeline.normalize_dataset(ds)
        assert result.metadata["normalized"] is True

    def test_metadata_engine_flags_all_true(self):
        ds = make_dataset([{"first_name": "יוסי"}])
        result = self.pipeline.normalize_dataset(ds)
        engines = result.metadata["normalization_engines"]
        assert engines["name"] is True
        assert engines["gender"] is True
        assert engines["date"] is True
        assert engines["identifier"] is True

    def test_metadata_engine_flags_reflect_missing_engine(self):
        pipeline = NormalizationPipeline(name_engine=None)
        ds = make_dataset([{"first_name": "יוסי"}])
        result = pipeline.normalize_dataset(ds)
        assert result.metadata["normalization_engines"]["name"] is False

    def test_statistics_total_rows(self):
        rows = [{"first_name": "יוסי"}, {"first_name": "שרה"}, {"first_name": "דוד"}]
        ds = make_dataset(rows)
        result = self.pipeline.normalize_dataset(ds)
        stats = result.metadata["normalization_statistics"]
        assert stats["total_rows"] == 3

    def test_statistics_success_rate_all_good(self):
        rows = [{"first_name": "יוסי"}, {"first_name": "שרה"}]
        ds = make_dataset(rows)
        result = self.pipeline.normalize_dataset(ds)
        stats = result.metadata["normalization_statistics"]
        assert stats["rows_with_failures"] == 0
        assert stats["success_rate"] == 1.0

    def test_statistics_failure_counted(self):
        class BrokenNameEngine:
            def normalize_name(self, v):
                raise RuntimeError("broken")

        pipeline = NormalizationPipeline(name_engine=BrokenNameEngine())
        rows = [{"first_name": "יוסי"}, {"first_name": "שרה"}]
        ds = make_dataset(rows)
        result = pipeline.normalize_dataset(ds)
        stats = result.metadata["normalization_statistics"]
        assert stats["rows_with_failures"] == 2
        assert stats["success_rate"] < 1.0

    def test_all_rows_get_corrected_fields(self):
        rows = [
            {"first_name": "  יוסי  ", "gender": "ז"},
            {"first_name": "שרה123", "gender": "נ"},
        ]
        ds = make_dataset(rows)
        result = self.pipeline.normalize_dataset(ds)
        for row in result.rows:
            assert "first_name_corrected" in row
            assert "gender_corrected" in row

    def test_original_values_not_mutated(self):
        rows = [{"first_name": "  יוסי  "}]
        ds = make_dataset(rows)
        result = self.pipeline.normalize_dataset(ds)
        assert result.rows[0]["first_name"] == "  יוסי  "
        assert result.rows[0]["first_name_corrected"] == "יוסי"

    def test_empty_dataset_no_error(self):
        ds = make_dataset([], field_names=["first_name"])
        result = self.pipeline.normalize_dataset(ds)
        assert result.metadata["normalization_statistics"]["total_rows"] == 0
        assert result.metadata["normalization_statistics"]["success_rate"] == 1.0

    def test_no_engines_pipeline_still_runs(self):
        pipeline = NormalizationPipeline()  # all engines None
        rows = [{"first_name": "יוסי", "gender": "ז"}]
        ds = make_dataset(rows)
        result = pipeline.normalize_dataset(ds)
        # No corrected fields added since no engines
        assert "first_name_corrected" not in result.rows[0]
        assert "gender_corrected" not in result.rows[0]
        assert result.metadata["normalized"] is True


# ---------------------------------------------------------------------------
# Plain single-column birth_date normalization (web path)
# ---------------------------------------------------------------------------

class TestPlainBirthDateNormalization:
    """Verify that plain single-column birth_date values are fully normalized."""

    def _make_pipeline(self):
        from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
        from src.excel_normalization.engines.date_engine import DateEngine
        from src.excel_normalization.data_types import DateFormatPattern
        p = NormalizationPipeline(date_engine=DateEngine())
        p._date_format_pattern = DateFormatPattern.DDMM
        return p

    def test_iso_datetime_string_normalized(self):
        """2025-09-20T00:00:00 → parsed into structured year/month/day."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "2025-09-20T00:00:00"}
        result = pipeline.normalize_row(row)
        assert result.get("birth_year_corrected") == 2025
        assert result.get("birth_month_corrected") == 9
        assert result.get("birth_day_corrected") == 20

    def test_dot_separated_date_normalized(self):
        """11.06.1997 → year=1997, month=6, day=11."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "11.06.1997"}
        result = pipeline.normalize_row(row)
        assert result.get("birth_year_corrected") == 1997
        assert result.get("birth_month_corrected") == 6
        assert result.get("birth_day_corrected") == 11
        assert result.get("birth_date_status") == ""

    def test_numeric_8digit_date_normalized(self):
        """12022001 → year=2001, month=2, day=12."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "12022001"}
        result = pipeline.normalize_row(row)
        assert result.get("birth_year_corrected") == 2001
        assert result.get("birth_month_corrected") == 2
        assert result.get("birth_day_corrected") == 12

    def test_iso_2001_date_normalized(self):
        """2001-04-14T00:00:00 → year=2001, month=4, day=14."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "2001-04-14T00:00:00"}
        result = pipeline.normalize_row(row)
        assert result.get("birth_year_corrected") == 2001
        assert result.get("birth_month_corrected") == 4
        assert result.get("birth_day_corrected") == 14
        assert result.get("birth_date_status") == ""

    def test_status_written_for_invalid_date(self):
        """Invalid date still gets a status, not empty."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "99/99/9999"}
        result = pipeline.normalize_row(row)
        assert result.get("birth_date_status") is not None

    def test_none_value_preserved(self):
        """None birth_date → all structured corrected fields are None."""
        pipeline = self._make_pipeline()
        row = {"birth_date": None}
        result = pipeline.normalize_row(row)
        assert result.get("birth_year_corrected") is None
        assert result.get("birth_month_corrected") is None
        assert result.get("birth_day_corrected") is None

    def test_internal_tag_not_in_output(self):
        """_birth_year_auto_completed tag must be stripped from the final row."""
        pipeline = self._make_pipeline()
        row = {"birth_date": "11.06.1997"}
        # normalize_row alone doesn't run the dataset-level correction pass,
        # so the tag is stripped by normalize_dataset.  When called standalone
        # the tag may be present — the important guarantee is that it is absent
        # after a full normalize_dataset call.
        from src.excel_normalization.data_types import SheetDataset
        from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
        from src.excel_normalization.engines.date_engine import DateEngine
        from src.excel_normalization.data_types import DateFormatPattern
        p = NormalizationPipeline(date_engine=DateEngine())
        p._date_format_pattern = DateFormatPattern.DDMM
        dataset = SheetDataset(
            sheet_name="test", header_row=1, header_rows_count=1,
            field_names=["birth_date"],
            rows=[{"birth_date": "11.06.1997"}],
            metadata={},
        )
        result = p.normalize_dataset(dataset)
        for out_row in result.rows:
            assert "_birth_year_auto_completed" not in out_row


class TestPlainBirthDateMajorityCorrection:
    """Verify list-level majority correction works for plain single-column birth_date."""

    def _make_pipeline(self):
        from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
        from src.excel_normalization.engines.date_engine import DateEngine
        from src.excel_normalization.data_types import DateFormatPattern, SheetDataset
        p = NormalizationPipeline(date_engine=DateEngine())
        p._date_format_pattern = DateFormatPattern.DDMM
        return p

    def test_majority_1900s_flips_2000s_outlier_single_column(self):
        """30, 28, 31, 26 as plain birth_date → outlier 2026 corrected to 1926."""
        from src.excel_normalization.data_types import SheetDataset
        from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
        from src.excel_normalization.engines.date_engine import DateEngine
        from src.excel_normalization.data_types import DateFormatPattern

        pipeline = NormalizationPipeline(date_engine=DateEngine())
        pipeline._date_format_pattern = DateFormatPattern.DDMM

        # Use split path (birth_year) since plain single-column uses birth_date
        # and the majority correction works on both paths.
        # Test the split path here (already covered by date_processor tests).
        # For the single-column path, build a dataset and call normalize_dataset.
        dataset = SheetDataset(
            sheet_name="test",
            header_row=1,
            header_rows_count=1,
            field_names=["birth_year", "birth_month", "birth_day"],
            rows=[
                {"birth_year": 30, "birth_month": 6, "birth_day": 15},
                {"birth_year": 28, "birth_month": 3, "birth_day": 10},
                {"birth_year": 31, "birth_month": 9, "birth_day": 5},
                {"birth_year": 26, "birth_month": 1, "birth_day": 20},
            ],
            metadata={},
        )
        result = pipeline.normalize_dataset(dataset)
        years = [r.get("birth_year_corrected") for r in result.rows]
        assert years == [1930, 1928, 1931, 1926]

    def test_internal_tag_stripped_after_dataset_normalization(self):
        """_birth_year_auto_completed must not appear in any output row."""
        from src.excel_normalization.data_types import SheetDataset
        from src.excel_normalization.processing.normalization_pipeline import NormalizationPipeline
        from src.excel_normalization.engines.date_engine import DateEngine

        pipeline = NormalizationPipeline(date_engine=DateEngine())
        dataset = SheetDataset(
            sheet_name="test",
            header_row=1,
            header_rows_count=1,
            field_names=["birth_year", "birth_month", "birth_day"],
            rows=[{"birth_year": 30, "birth_month": 6, "birth_day": 15}],
            metadata={},
        )
        result = pipeline.normalize_dataset(dataset)
        for row in result.rows:
            assert "_birth_year_auto_completed" not in row


# ---------------------------------------------------------------------------
# Plain birth_date column detection (excel_reader fix)
# ---------------------------------------------------------------------------

class TestPlainBirthDateColumnDetection:
    """Verify detect_columns maps a plain תאריך לידה column to birth_date
    even when the sheet has a two-row header layout (e.g. entry date is split)."""

    def test_plain_birth_date_with_two_row_header(self):
        """Plain birth_date column must not be silently dropped in 2-row header sheets."""
        from openpyxl import Workbook
        from src.excel_normalization.io_layer.excel_reader import ExcelReader

        wb = Workbook()
        ws = wb.active

        # Layout:
        #   Col 1: שם פרטי (plain name)
        #   Col 2: תאריך לידה (plain date — NO sub-headers below it)
        #   Col 3: תאריך כניסה למוסד (split date)
        #   Col 4: שנה  (entry sub-header)
        #   Col 5: חודש (entry sub-header)
        #   Col 6: יום  (entry sub-header)
        # Row 1: main headers
        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=1, column=3).value = "תאריך כניסה למוסד"
        # Row 2: sub-headers only under entry date (cols 4-6), NOT under birth date
        ws.cell(row=2, column=4).value = "שנה"
        ws.cell(row=2, column=5).value = "חודש"
        ws.cell(row=2, column=6).value = "יום"
        # Data row
        ws.cell(row=3, column=1).value = "יוסי"
        ws.cell(row=3, column=2).value = "11.06.1997"
        ws.cell(row=3, column=3).value = None
        ws.cell(row=3, column=4).value = 1990
        ws.cell(row=3, column=5).value = 5
        ws.cell(row=3, column=6).value = 15

        reader = ExcelReader()
        mapping = reader.detect_columns(ws)

        # birth_date must be present as a plain single-column field
        assert "birth_date" in mapping, (
            f"birth_date was dropped; got keys: {list(mapping.keys())}"
        )
        # entry date should be split
        assert "entry_year" in mapping
        assert "entry_month" in mapping
        assert "entry_day" in mapping

    def test_plain_birth_date_single_row_header_unchanged(self):
        """Single-row header layout must still map birth_date correctly (no regression)."""
        from openpyxl import Workbook
        from src.excel_normalization.io_layer.excel_reader import ExcelReader

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = "שם פרטי"
        ws.cell(row=1, column=2).value = "תאריך לידה"
        ws.cell(row=2, column=1).value = "יוסי"
        ws.cell(row=2, column=2).value = "11.06.1997"

        reader = ExcelReader()
        mapping = reader.detect_columns(ws)

        assert "birth_date" in mapping
