"""Unit tests for NameFieldProcessor.

Tests the name field processing logic including header detection,
column preparation, and name standardization.
"""

import pytest
from openpyxl import Workbook
from src.excel_standardization.processing.name_processor import NameFieldProcessor
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_writer import ExcelWriter
from src.excel_standardization.engines.name_engine import NameEngine
from src.excel_standardization.engines.text_processor import TextProcessor
from src.excel_standardization.data_types import FatherNamePattern


class TestNameFieldProcessor:
    """Tests for NameFieldProcessor."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.reader = ExcelReader()
        self.writer = ExcelWriter()
        self.text_processor = TextProcessor()
        self.name_engine = NameEngine(self.text_processor)
        self.processor = NameFieldProcessor(self.reader, self.writer, self.name_engine)
    
    def test_find_headers_hebrew_first_name(self):
        """Should find Hebrew first name header."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'שם פרטי'
        ws['A2'] = 'יוסי'
        
        result = self.processor.find_headers(ws)
        
        assert result is True
        assert self.processor.first_name_info is not None
        assert self.processor.first_name_info.col == 1
        assert self.processor.first_name_info.header_row == 1
    
    def test_find_headers_english_last_name(self):
        """Should find English last name header."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'last name'
        ws['A2'] = 'Cohen'
        
        result = self.processor.find_headers(ws)
        
        assert result is True
        assert self.processor.last_name_info is not None
        assert self.processor.last_name_info.col == 1
    
    def test_find_headers_father_name(self):
        """Should find father's name header."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'שם האב'
        ws['A2'] = 'אברהם'
        
        result = self.processor.find_headers(ws)
        
        assert result is True
        assert self.processor.father_name_info is not None
        assert self.processor.father_name_info.col == 1
    
    def test_find_headers_no_headers(self):
        """Should return False when no name headers found."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Other Column'
        
        result = self.processor.find_headers(ws)
        
        assert result is False
    
    def test_detect_father_name_pattern_none(self):
        """Should return NONE when fewer than 3 contain last name."""
        father_names = ['אברהם', 'יצחק', 'יעקב', 'משה', 'אהרון']
        last_names = ['כהן', 'לוי', 'ישראל', 'כהן', 'לוי']
        
        pattern = self.processor.detect_father_name_pattern(father_names, last_names)
        
        assert pattern == FatherNamePattern.NONE
    
    def test_detect_father_name_pattern_remove_first(self):
        """Should return REMOVE_FIRST when last name is in first position."""
        father_names = ['כהן אברהם', 'לוי יצחק', 'ישראל יעקב', 'כהן משה', 'לוי אהרון']
        last_names = ['כהן', 'לוי', 'ישראל', 'כהן', 'לוי']
        
        pattern = self.processor.detect_father_name_pattern(father_names, last_names)
        
        assert pattern == FatherNamePattern.REMOVE_FIRST
    
    def test_detect_father_name_pattern_remove_last(self):
        """Should return REMOVE_LAST when last name is in last position."""
        father_names = ['אברהם כהן', 'יצחק לוי', 'יעקב ישראל', 'משה כהן', 'אהרון לוי']
        last_names = ['כהן', 'לוי', 'ישראל', 'כהן', 'לוי']
        
        pattern = self.processor.detect_father_name_pattern(father_names, last_names)
        
        assert pattern == FatherNamePattern.REMOVE_LAST
    
    def test_process_field_integration(self):
        """Integration test for complete name processing."""
        wb = Workbook()
        ws = wb.active
        
        # Set up headers
        ws['A1'] = 'שם פרטי'
        ws['B1'] = 'שם משפחה'
        
        # Set up data
        ws['A2'] = '  יוסי  '  # With extra spaces
        ws['B2'] = 'כהן'
        ws['A3'] = 'משה'
        ws['B3'] = '  לוי  '
        
        # Process the field
        self.processor.process_field(ws)
        
        # After processing:
        # Column A: שם פרטי (original)
        # Column B: שם פרטי - מתוקן (inserted after A)
        # Column C: שם משפחה (original, shifted from B)
        # Column D: שם משפחה - מתוקן (inserted after C)
        
        # Verify corrected columns were created
        assert ws['B1'].value == 'שם פרטי - מתוקן'
        assert ws['D1'].value == 'שם משפחה - מתוקן'
        
        # Verify normalized values
        assert ws['B2'].value == 'יוסי'  # Trimmed
        assert ws['D2'].value == 'כהן'
        assert ws['B3'].value == 'משה'
        assert ws['D3'].value == 'לוי'  # Trimmed


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
