"""Tests for JsonToExcelWriter class.

This module tests the JsonToExcelWriter class which exports JSON datasets
back to Excel format with original and corrected columns.
"""

import os
import tempfile
from openpyxl import load_workbook

from src.excel_standardization.data_types import SheetDataset, WorkbookDataset, JsonRow
from src.excel_standardization.io_layer import JsonToExcelWriter


def test_write_dataset_to_excel_basic():
    """Test writing a basic dataset to Excel."""
    # Create test dataset
    rows = [
        {
            "first_name": "יוסי",
            "first_name_corrected": "יוסי",
            "last_name": "כהן",
            "last_name_corrected": "כהן",
            "gender": "ז",
            "gender_corrected": "2"
        },
        {
            "first_name": "שרה",
            "first_name_corrected": "שרה",
            "last_name": "לוי",
            "last_name_corrected": "לוי",
            "gender": "נ",
            "gender_corrected": "1"
        }
    ]
    
    dataset = SheetDataset(
        sheet_name="Test Sheet",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "gender"],
        rows=rows,
        metadata={}
    )
    
    # Write to temporary file
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_path = f.name
    
    try:
        writer = JsonToExcelWriter()
        writer.write_dataset_to_excel(dataset, output_path)
        
        # Verify file was created
        assert os.path.exists(output_path)
        
        # Load and verify content
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check sheet name
        assert worksheet.title == "Test Sheet"
        
        # Check headers
        assert worksheet.cell(1, 1).value == "first_name"
        assert worksheet.cell(1, 2).value == "first_name_corrected"
        assert worksheet.cell(1, 3).value == "last_name"
        assert worksheet.cell(1, 4).value == "last_name_corrected"
        assert worksheet.cell(1, 5).value == "gender"
        assert worksheet.cell(1, 6).value == "gender_corrected"
        
        # Check first data row
        assert worksheet.cell(2, 1).value == "יוסי"
        assert worksheet.cell(2, 2).value == "יוסי"
        assert worksheet.cell(2, 3).value == "כהן"
        assert worksheet.cell(2, 4).value == "כהן"
        assert worksheet.cell(2, 5).value == "ז"
        assert worksheet.cell(2, 6).value == "2"
        
        # Check second data row
        assert worksheet.cell(3, 1).value == "שרה"
        assert worksheet.cell(3, 2).value == "שרה"
        assert worksheet.cell(3, 3).value == "לוי"
        assert worksheet.cell(3, 4).value == "לוי"
        assert worksheet.cell(3, 5).value == "נ"
        assert worksheet.cell(3, 6).value == "1"
        
    finally:
        # Clean up
        if os.path.exists(output_path):
            os.remove(output_path)


def test_write_dataset_with_none_values():
    """Test writing dataset with None values."""
    rows = [
        {
            "first_name": "יוסי",
            "first_name_corrected": "יוסי",
            "last_name": None,
            "last_name_corrected": None,
            "gender": "ז",
            "gender_corrected": "2"
        }
    ]
    
    dataset = SheetDataset(
        sheet_name="Test Sheet",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name", "gender"],
        rows=rows,
        metadata={}
    )
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_path = f.name
    
    try:
        writer = JsonToExcelWriter()
        writer.write_dataset_to_excel(dataset, output_path)
        
        # Verify file was created
        assert os.path.exists(output_path)
        
        # Load and verify content
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check None values are handled
        assert worksheet.cell(2, 1).value == "יוסי"
        assert worksheet.cell(2, 3).value is None
        assert worksheet.cell(2, 4).value is None
        
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)


def test_write_workbook_to_excel_multiple_sheets():
    """Test writing workbook with multiple sheets."""
    # Create first sheet
    sheet1_rows = [
        {
            "first_name": "יוסי",
            "first_name_corrected": "יוסי",
            "gender": "ז",
            "gender_corrected": "2"
        }
    ]
    
    sheet1 = SheetDataset(
        sheet_name="Students",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "gender"],
        rows=sheet1_rows,
        metadata={}
    )
    
    # Create second sheet
    sheet2_rows = [
        {
            "last_name": "כהן",
            "last_name_corrected": "כהן",
            "id_number": "123456789",
            "id_number_corrected": "123456789"
        }
    ]
    
    sheet2 = SheetDataset(
        sheet_name="Teachers",
        header_row=1,
        header_rows_count=1,
        field_names=["last_name", "id_number"],
        rows=sheet2_rows,
        metadata={}
    )
    
    # Create workbook dataset
    workbook_dataset = WorkbookDataset(
        source_file="test.xlsx",
        sheets=[sheet1, sheet2],
        metadata={}
    )
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_path = f.name
    
    try:
        writer = JsonToExcelWriter()
        writer.write_workbook_to_excel(workbook_dataset, output_path)
        
        # Verify file was created
        assert os.path.exists(output_path)
        
        # Load and verify content
        workbook = load_workbook(output_path)
        
        # Check sheet names
        assert "Students" in workbook.sheetnames
        assert "Teachers" in workbook.sheetnames
        
        # Check first sheet
        ws1 = workbook["Students"]
        assert ws1.cell(1, 1).value == "first_name"
        assert ws1.cell(1, 2).value == "first_name_corrected"
        assert ws1.cell(1, 3).value == "gender"
        assert ws1.cell(1, 4).value == "gender_corrected"
        assert ws1.cell(2, 1).value == "יוסי"
        assert ws1.cell(2, 4).value == "2"
        
        # Check second sheet
        ws2 = workbook["Teachers"]
        assert ws2.cell(1, 1).value == "last_name"
        assert ws2.cell(1, 2).value == "last_name_corrected"
        assert ws2.cell(1, 3).value == "id_number"
        assert ws2.cell(1, 4).value == "id_number_corrected"
        assert ws2.cell(2, 1).value == "כהן"
        assert ws2.cell(2, 4).value == "123456789"
        
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)


def test_header_formatting():
    """Test that header formatting is applied correctly."""
    rows = [
        {
            "first_name": "יוסי",
            "first_name_corrected": "יוסי"
        }
    ]
    
    dataset = SheetDataset(
        sheet_name="Test",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name"],
        rows=rows,
        metadata={}
    )
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_path = f.name
    
    try:
        writer = JsonToExcelWriter(apply_header_formatting=True)
        writer.write_dataset_to_excel(dataset, output_path)
        
        # Load and verify formatting
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check headers are bold
        assert worksheet.cell(1, 1).font.bold is True
        assert worksheet.cell(1, 2).font.bold is True
        
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)


def test_empty_dataset():
    """Test writing dataset with no rows."""
    dataset = SheetDataset(
        sheet_name="Empty",
        header_row=1,
        header_rows_count=1,
        field_names=["first_name", "last_name"],
        rows=[],
        metadata={}
    )
    
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as f:
        output_path = f.name
    
    try:
        writer = JsonToExcelWriter()
        writer.write_dataset_to_excel(dataset, output_path)
        
        # Verify file was created
        assert os.path.exists(output_path)
        
        # Load and verify content
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check headers exist
        assert worksheet.cell(1, 1).value == "first_name"
        assert worksheet.cell(1, 2).value == "first_name_corrected"
        
        # Check no data rows
        assert worksheet.cell(2, 1).value is None
        
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)
