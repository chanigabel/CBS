"""Unit tests for DateFieldProcessor.

This module tests the DateFieldProcessor class which processes birth date
and entry date fields with split year/month/day columns.
"""

import pytest
from openpyxl import Workbook
from src.excel_standardization.processing.date_processor import DateFieldProcessor
from src.excel_standardization.io_layer.excel_reader import ExcelReader
from src.excel_standardization.io_layer.excel_writer import ExcelWriter
from src.excel_standardization.engines.date_engine import DateEngine
from src.excel_standardization.data_types import DateFormatPattern, DateFieldType


@pytest.fixture
def date_processor():
    """Create a DateFieldProcessor instance for testing."""
    reader = ExcelReader()
    writer = ExcelWriter()
    engine = DateEngine()
    return DateFieldProcessor(reader, writer, engine)


@pytest.fixture
def workbook_with_birth_date():
    """Create a workbook with birth date headers and sub-headers."""
    wb = Workbook()
    ws = wb.active
    
    # Set up headers
    ws.cell(row=1, column=1).value = "תאריך לידה"
    ws.cell(row=2, column=1).value = "שנה"
    ws.cell(row=2, column=2).value = "חודש"
    ws.cell(row=2, column=3).value = "יום"
    
    # Add some data
    ws.cell(row=3, column=1).value = 1990
    ws.cell(row=3, column=2).value = 5
    ws.cell(row=3, column=3).value = 15
    
    ws.cell(row=4, column=1).value = 1985
    ws.cell(row=4, column=2).value = 12
    ws.cell(row=4, column=3).value = 25
    
    return wb


@pytest.fixture
def workbook_with_entry_date():
    """Create a workbook with entry date headers and sub-headers."""
    wb = Workbook()
    ws = wb.active
    
    # Set up headers
    ws.cell(row=1, column=1).value = "תאריך כניסה למוסד"
    ws.cell(row=2, column=1).value = "שנה"
    ws.cell(row=2, column=2).value = "חודש"
    ws.cell(row=2, column=3).value = "יום"
    
    # Add some data
    ws.cell(row=3, column=1).value = 2020
    ws.cell(row=3, column=2).value = 3
    ws.cell(row=3, column=3).value = 10
    
    return wb


def test_find_headers_birth_date(date_processor, workbook_with_birth_date):
    """Test finding birth date headers with sub-headers."""
    ws = workbook_with_birth_date.active
    
    result = date_processor.find_headers(ws)
    
    assert result is True
    assert DateFieldType.BIRTH_DATE in date_processor.date_fields
    
    birth_date_groups = date_processor.date_fields[DateFieldType.BIRTH_DATE]
    assert len(birth_date_groups) == 1
    birth_date_info = birth_date_groups[0]
    assert birth_date_info['year_col'] == 1
    assert birth_date_info['month_col'] == 2
    assert birth_date_info['day_col'] == 3
    assert birth_date_info['sub_header_row'] == 2


def test_find_headers_entry_date(date_processor, workbook_with_entry_date):
    """Test finding entry date headers with sub-headers."""
    ws = workbook_with_entry_date.active
    
    result = date_processor.find_headers(ws)
    
    assert result is True
    assert DateFieldType.ENTRY_DATE in date_processor.date_fields
    
    entry_date_groups = date_processor.date_fields[DateFieldType.ENTRY_DATE]
    assert len(entry_date_groups) == 1
    entry_date_info = entry_date_groups[0]
    assert entry_date_info['year_col'] == 1
    assert entry_date_info['month_col'] == 2
    assert entry_date_info['day_col'] == 3
    assert entry_date_info['sub_header_row'] == 2


def test_find_headers_no_date_fields(date_processor):
    """Test finding headers when no date fields exist."""
    wb = Workbook()
    ws = wb.active
    
    # Set up non-date headers
    ws.cell(row=1, column=1).value = "שם פרטי"
    ws.cell(row=1, column=2).value = "שם משפחה"
    
    result = date_processor.find_headers(ws)
    
    assert result is False
    # date_fields is initialized with empty lists for both types
    assert all(len(v) == 0 for v in date_processor.date_fields.values())


def test_detect_date_format_pattern_ddmm(date_processor):
    """Test detecting DDMM pattern."""
    # Dates with day > 12 in first position
    date_values = ["15/05/1990", "25/12/1985", "20/03/2000"]
    
    pattern = date_processor.detect_date_format_pattern(date_values)
    
    assert pattern == DateFormatPattern.DDMM


def test_detect_date_format_pattern_mmdd(date_processor):
    """Test detecting MMDD pattern."""
    # Dates with day > 12 in second position
    date_values = ["05/15/1990", "12/25/1985", "03/20/2000"]
    
    pattern = date_processor.detect_date_format_pattern(date_values)
    
    assert pattern == DateFormatPattern.MMDD


def test_detect_date_format_pattern_mixed(date_processor):
    """Test detecting pattern with mixed dates (defaults to DDMM)."""
    # Mix of ambiguous dates
    date_values = ["05/10/1990", "10/05/1985", "03/08/2000"]
    
    pattern = date_processor.detect_date_format_pattern(date_values)
    
    # Should default to DDMM when counts are equal
    assert pattern == DateFormatPattern.DDMM


def test_prepare_output_columns(date_processor, workbook_with_birth_date):
    """Test preparing output columns for date fields."""
    ws = workbook_with_birth_date.active
    
    # Find headers first
    date_processor.find_headers(ws)
    
    # Prepare output columns
    date_processor.prepare_output_columns(ws)
    
    # Check that corrected columns were created
    birth_date_groups = date_processor.date_fields[DateFieldType.BIRTH_DATE]
    assert len(birth_date_groups) == 1
    birth_date_info = birth_date_groups[0]
    assert 'corrected_year_col' in birth_date_info
    assert 'corrected_month_col' in birth_date_info
    assert 'corrected_day_col' in birth_date_info
    assert 'corrected_status_col' in birth_date_info
    
    # Check that headers were set correctly
    assert ws.cell(row=2, column=birth_date_info['corrected_year_col']).value == "שנה - מתוקן"
    assert ws.cell(row=2, column=birth_date_info['corrected_month_col']).value == "חודש - מתוקן"
    assert ws.cell(row=2, column=birth_date_info['corrected_day_col']).value == "יום - מתוקן"
    assert ws.cell(row=2, column=birth_date_info['corrected_status_col']).value == "סטטוס תאריך"


def test_process_data_valid_dates(date_processor, workbook_with_birth_date):
    """Test processing valid date data."""
    ws = workbook_with_birth_date.active
    
    # Find headers, prepare columns, and process data
    date_processor.find_headers(ws)
    date_processor.prepare_output_columns(ws)
    date_processor.process_data(ws)
    
    # Check that corrected values were written
    birth_date_info = date_processor.date_fields[DateFieldType.BIRTH_DATE][0]
    
    # First row (1990-05-15)
    assert ws.cell(row=3, column=birth_date_info['corrected_year_col']).value == 1990
    assert ws.cell(row=3, column=birth_date_info['corrected_month_col']).value == 5
    assert ws.cell(row=3, column=birth_date_info['corrected_day_col']).value == 15
    
    # Second row (1985-12-25)
    assert ws.cell(row=4, column=birth_date_info['corrected_year_col']).value == 1985
    assert ws.cell(row=4, column=birth_date_info['corrected_month_col']).value == 12
    assert ws.cell(row=4, column=birth_date_info['corrected_day_col']).value == 25


def test_process_field_integration(date_processor, workbook_with_birth_date):
    """Test the full process_field template method."""
    ws = workbook_with_birth_date.active
    
    # Process the field using the template method
    date_processor.process_field(ws)
    
    # Verify that all steps were executed
    assert DateFieldType.BIRTH_DATE in date_processor.date_fields
    
    birth_date_info = date_processor.date_fields[DateFieldType.BIRTH_DATE][0]
    
    # Check that corrected columns exist
    assert 'corrected_year_col' in birth_date_info
    
    # Check that data was processed
    assert ws.cell(row=3, column=birth_date_info['corrected_year_col']).value == 1990


# ---------------------------------------------------------------------------
# List-level majority century correction
# ---------------------------------------------------------------------------

class TestApplyMajorityCenturyCorrection:
    """Verify the one-way 1900s majority correction in _apply_majority_century_correction."""

    def _make_processor(self):
        return DateFieldProcessor(ExcelReader(), ExcelWriter(), DateEngine())

    def _auto_result(self, year, month=6, day=15, is_valid=True):
        """Build a DateParseResult that looks like an auto-completed year."""
        from src.excel_standardization.data_types import DateParseResult
        r = DateParseResult(
            year=year, month=month, day=day,
            is_valid=is_valid, status_text="",
            year_was_auto_completed=True,
        )
        return r

    def _explicit_result(self, year, month=6, day=15, is_valid=True):
        """Build a DateParseResult that looks like an explicit 4-digit year."""
        from src.excel_standardization.data_types import DateParseResult
        r = DateParseResult(
            year=year, month=month, day=day,
            is_valid=is_valid, status_text="",
            year_was_auto_completed=False,
        )
        return r

    # ------------------------------------------------------------------
    # Core majority-correction cases
    # ------------------------------------------------------------------

    def test_majority_1900s_flips_2000s_outliers(self):
        """30, 28, 31, 26 → 1930, 1928, 1931, 2026 → after correction: 1926."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(1928),
            self._auto_result(1931),
            self._auto_result(2026),  # outlier
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [1930, 1928, 1931, 1926]

    def test_majority_2000s_does_not_flip_1900s(self):
        """05, 07, 12, 98 → 2005, 2007, 2012, 1998 → no change."""
        proc = self._make_processor()
        results = [
            self._auto_result(2005),
            self._auto_result(2007),
            self._auto_result(2012),
            self._auto_result(1998),  # minority 1900s — must NOT be flipped
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [2005, 2007, 2012, 1998]

    def test_explicit_4_digit_year_never_touched(self):
        """Explicit 2005 in an otherwise old list must remain 2005."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(1928),
            self._auto_result(1931),
            self._explicit_result(2005),  # explicit — must not change
        ]
        corrected = proc._apply_majority_century_correction(results)
        # The three auto-completed 1900s are the majority; the explicit 2005
        # is not auto-completed so it must remain 2005.
        assert corrected[3].year == 2005
        assert corrected[3].year_was_auto_completed is False

    def test_single_row_no_majority_correction(self):
        """Single row — only base algorithm applies, no list correction."""
        proc = self._make_processor()
        results = [self._auto_result(2026)]
        corrected = proc._apply_majority_century_correction(results)
        # 1 auto-2000s, 0 auto-1900s → 2000s are not minority → no flip
        assert corrected[0].year == 2026

    def test_equal_split_no_correction(self):
        """Equal 1900s and 2000s — 1900s are not strict majority → no flip."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(2026),
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [1930, 2026]

    def test_all_1900s_no_change(self):
        """All auto-completed 1900s — nothing to flip."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(1945),
            self._auto_result(1960),
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [1930, 1945, 1960]

    def test_all_2000s_no_change(self):
        """All auto-completed 2000s — no 1900s majority → no flip."""
        proc = self._make_processor()
        results = [
            self._auto_result(2005),
            self._auto_result(2010),
            self._auto_result(2015),
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [2005, 2010, 2015]

    def test_no_auto_completed_rows_no_change(self):
        """All explicit 4-digit years — correction never fires."""
        proc = self._make_processor()
        results = [
            self._explicit_result(1930),
            self._explicit_result(2005),
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [1930, 2005]

    def test_multiple_2000s_outliers_all_flipped(self):
        """Multiple 2000s outliers in a mostly-1900s list — all flipped."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(1928),
            self._auto_result(1931),
            self._auto_result(1945),
            self._auto_result(2026),  # outlier 1
            self._auto_result(2012),  # outlier 2
        ]
        corrected = proc._apply_majority_century_correction(results)
        years = [r.year for r in corrected]
        assert years == [1930, 1928, 1931, 1945, 1926, 1912]

    def test_auto_completed_flag_preserved_after_correction(self):
        """Flipped rows must still carry year_was_auto_completed=True."""
        proc = self._make_processor()
        results = [
            self._auto_result(1930),
            self._auto_result(1928),
            self._auto_result(1931),
            self._auto_result(2026),
        ]
        corrected = proc._apply_majority_century_correction(results)
        assert corrected[3].year_was_auto_completed is True

    def test_invalid_rows_ignored_in_counting(self):
        """Rows with year=None (parse failures) don't affect the majority count."""
        from src.excel_standardization.data_types import DateParseResult
        proc = self._make_processor()
        invalid = DateParseResult(
            year=None, month=None, day=None,
            is_valid=False, status_text="תוכן לא ניתן לפריקה",
            year_was_auto_completed=True,
        )
        results = [
            self._auto_result(1930),
            self._auto_result(1928),
            self._auto_result(1931),
            self._auto_result(2026),
            invalid,
        ]
        corrected = proc._apply_majority_century_correction(results)
        # 2026 should still be flipped; invalid row unchanged
        assert corrected[3].year == 1926
        assert corrected[4].year is None


class TestMajorityCorrectionEndToEnd:
    """Integration: full process_data with majority correction via worksheet."""

    def _make_processor(self):
        return DateFieldProcessor(ExcelReader(), ExcelWriter(), DateEngine())

    def test_mostly_1900s_outlier_corrected_in_worksheet(self):
        """End-to-end: 30, 28, 31, 26 → output years 1930, 1928, 1931, 1926."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = "תאריך לידה"
        ws.cell(row=2, column=1).value = "שנה"
        ws.cell(row=2, column=2).value = "חודש"
        ws.cell(row=2, column=3).value = "יום"

        # Shortened years: 30, 28, 31, 26
        data = [(30, 6, 15), (28, 3, 10), (31, 9, 5), (26, 1, 20)]
        for i, (y, m, d) in enumerate(data, start=3):
            ws.cell(row=i, column=1).value = y
            ws.cell(row=i, column=2).value = m
            ws.cell(row=i, column=3).value = d

        proc = self._make_processor()
        proc.find_headers(ws)
        proc.prepare_output_columns(ws)
        proc.process_data(ws)

        info = proc.date_fields[DateFieldType.BIRTH_DATE][0]
        out_years = [
            ws.cell(row=r, column=info["corrected_year_col"]).value
            for r in range(3, 7)
        ]
        assert out_years == [1930, 1928, 1931, 1926]

    def test_mostly_2000s_1900s_not_changed(self):
        """End-to-end: 05, 07, 12, 98 → output years 2005, 2007, 2012, 1998."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = "תאריך לידה"
        ws.cell(row=2, column=1).value = "שנה"
        ws.cell(row=2, column=2).value = "חודש"
        ws.cell(row=2, column=3).value = "יום"

        data = [(5, 6, 15), (7, 3, 10), (12, 9, 5), (98, 1, 20)]
        for i, (y, m, d) in enumerate(data, start=3):
            ws.cell(row=i, column=1).value = y
            ws.cell(row=i, column=2).value = m
            ws.cell(row=i, column=3).value = d

        proc = self._make_processor()
        proc.find_headers(ws)
        proc.prepare_output_columns(ws)
        proc.process_data(ws)

        info = proc.date_fields[DateFieldType.BIRTH_DATE][0]
        out_years = [
            ws.cell(row=r, column=info["corrected_year_col"]).value
            for r in range(3, 7)
        ]
        assert out_years == [2005, 2007, 2012, 1998]

    def test_explicit_4_digit_2000s_in_old_list_unchanged(self):
        """Explicit 2005 in an otherwise old list must remain 2005."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = "תאריך לידה"
        ws.cell(row=2, column=1).value = "שנה"
        ws.cell(row=2, column=2).value = "חודש"
        ws.cell(row=2, column=3).value = "יום"

        # Three shortened old years + one explicit 4-digit 2005
        data = [(30, 6, 15), (28, 3, 10), (31, 9, 5), (2005, 1, 20)]
        for i, (y, m, d) in enumerate(data, start=3):
            ws.cell(row=i, column=1).value = y
            ws.cell(row=i, column=2).value = m
            ws.cell(row=i, column=3).value = d

        proc = self._make_processor()
        proc.find_headers(ws)
        proc.prepare_output_columns(ws)
        proc.process_data(ws)

        info = proc.date_fields[DateFieldType.BIRTH_DATE][0]
        out_years = [
            ws.cell(row=r, column=info["corrected_year_col"]).value
            for r in range(3, 7)
        ]
        # First three auto-completed → 1930, 1928, 1931
        # Last one explicit → must stay 2005
        assert out_years[0] == 1930
        assert out_years[1] == 1928
        assert out_years[2] == 1931
        assert out_years[3] == 2005
