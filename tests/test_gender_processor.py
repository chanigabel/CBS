"""Tests for GenderFieldProcessor.

This module tests the GenderFieldProcessor class to ensure it correctly
finds gender headers with line break normalization, prepares output columns,
and processes gender data.
"""

import pytest
from openpyxl import Workbook
from src.excel_normalization.processing.gender_processor import GenderFieldProcessor
from src.excel_normalization.io_layer.excel_reader import ExcelReader
from src.excel_normalization.io_layer.excel_writer import ExcelWriter
from src.excel_normalization.engines.gender_engine import GenderEngine


def test_find_headers_with_line_breaks():
    """Test that gender header is found with line break normalization."""
    # Create a workbook with gender header containing line breaks
    wb = Workbook()
    ws = wb.active
    
    # Set gender header with actual line breaks
    ws.cell(row=1, column=1).value = "מין\n1=זכר\n2+נקבה"
    ws.cell(row=2, column=1).value = "1"
    ws.cell(row=3, column=1).value = "2"
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Test find_headers
    result = processor.find_headers(ws)
    
    assert result is True
    assert processor.gender_info is not None
    assert processor.gender_info.col == 1
    assert processor.gender_info.header_row == 1


def test_find_headers_with_different_line_break_variants():
    """Test that gender header is found with different line break variants."""
    # Create a workbook with gender header containing Windows line breaks
    wb = Workbook()
    ws = wb.active
    
    # Set gender header with \r\n line breaks
    ws.cell(row=1, column=1).value = "מין\r\n1=זכר\r\n2+נקבה"
    ws.cell(row=2, column=1).value = "1"
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Test find_headers
    result = processor.find_headers(ws)
    
    assert result is True
    assert processor.gender_info is not None


def test_find_headers_not_found():
    """Test that find_headers returns False when gender header is not present."""
    # Create a workbook without gender header
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1).value = "שם פרטי"
    ws.cell(row=2, column=1).value = "John"
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Test find_headers
    result = processor.find_headers(ws)
    
    assert result is False
    assert processor.gender_info is None


def test_prepare_output_columns():
    """Test that corrected column is inserted with correct header."""
    # Create a workbook with gender header
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1).value = "מין\n1=זכר\n2+נקבה"
    ws.cell(row=2, column=1).value = "1"
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Find headers first
    processor.find_headers(ws)
    
    # Test prepare_output_columns
    processor.prepare_output_columns(ws)
    
    assert processor.corrected_col == 2
    assert ws.cell(row=1, column=2).value == "מין - מתוקן"


def test_process_data():
    """Test that gender values are normalized correctly."""
    # Create a workbook with gender data
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1).value = "מין\n1=זכר\n2+נקבה"
    ws.cell(row=2, column=1).value = "1"
    ws.cell(row=3, column=1).value = "female"
    ws.cell(row=4, column=1).value = "נ"
    ws.cell(row=5, column=1).value = ""
    ws.cell(row=6, column=1).value = "זכר"
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Process field
    processor.process_field(ws)
    
    # Verify corrected values
    assert ws.cell(row=2, column=2).value == 1  # "1" -> 1 (male)
    assert ws.cell(row=3, column=2).value == 2  # "female" -> 2 (female)
    assert ws.cell(row=4, column=2).value == 2  # "נ" -> 2 (female)
    assert ws.cell(row=5, column=2).value == 1  # "" -> 1 (male, default)
    assert ws.cell(row=6, column=2).value == 1  # "זכר" -> 1 (male)


def test_process_data_with_highlighting():
    """Test that changed cells are highlighted in pink."""
    # Create a workbook with gender data
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1).value = "מין\n1=זכר\n2+נקבה"
    ws.cell(row=2, column=1).value = "female"  # Will change to 2
    ws.cell(row=3, column=1).value = "2"       # Will stay 2
    
    # Create processor
    reader = ExcelReader()
    writer = ExcelWriter()
    gender_engine = GenderEngine()
    processor = GenderFieldProcessor(reader, writer, gender_engine)
    
    # Process field
    processor.process_field(ws)
    
    # Verify highlighting
    # Row 2: "female" -> 2, should be highlighted (values differ)
    cell_2 = ws.cell(row=2, column=2)
    assert cell_2.value == 2
    assert cell_2.fill.start_color.rgb == "FFC7CE" or cell_2.fill.start_color.rgb == "00FFC7CE"
    
    # Row 3: "2" -> 2, should not be highlighted (values same)
    cell_3 = ws.cell(row=3, column=2)
    assert cell_3.value == 2
    # No fill or default fill (not pink)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
